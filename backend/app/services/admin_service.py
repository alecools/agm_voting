"""
Service layer for admin portal operations.
"""
from __future__ import annotations

import csv
import io
import uuid
import zipfile
from datetime import UTC, datetime, timezone

import bleach
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from fastapi import HTTPException
from sqlalchemy import Integer, case, delete, func, literal, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.logging_config import get_logger
from app.models import (
    GeneralMeeting,
    GeneralMeetingLotWeight,
    GeneralMeetingStatus,
    BallotSubmission,
    Building,
    EmailDelivery,
    EmailDeliveryStatus,
    FinancialPosition,
    FinancialPositionSnapshot,
    Lot,
    LotOwner,  # backward-compat alias for Lot
    LotProxy,
    Motion,
    MotionOption,
    Person,
    Vote,
    VoteChoice,
    VoteStatus,
    get_effective_status,
    lot_persons,
)
from app.models.tenant_settings import TenantSettings
from app.schemas.admin import (
    AdminVoteEntryRequest,
    BuildingUpdate,
    GeneralMeetingCreate,
    LotOwnerCreate,
    LotOwnerUpdate,
    MotionAddRequest,
    MotionUpdateRequest,
    UpdatePersonRequest,
)

logger = get_logger(__name__)


def _sanitise_description(desc: str | None) -> str | None:
    """Strip all HTML tags from a motion description and return None if blank."""
    if desc is None:
        return None
    return bleach.clean(desc, tags=[], strip=True).strip() or None


def _sanitise_option_text(text: str) -> str:
    """Strip all HTML tags from a motion option text."""
    return bleach.clean(text, tags=[], strip=True).strip()


def _parse_name(raw: str) -> tuple[str | None, str | None]:
    """Split a full name into (given_name, surname).

    Last space-delimited token → surname.
    Everything before → given_name.
    Single token (e.g. company name) → (None, token).
    Blank / whitespace-only → (None, None).
    """
    value = raw.strip()
    if not value:
        return None, None
    parts = value.split()
    if len(parts) == 1:
        return None, parts[0]
    return " ".join(parts[:-1]), parts[-1]


# ---------------------------------------------------------------------------
# Buildings
# ---------------------------------------------------------------------------


async def import_buildings_from_csv(
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    """
    Parse a CSV of buildings and upsert records.
    Returns {"created": int, "updated": int}.
    Raises HTTPException 422 on validation errors.
    """
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    required_headers = {"building_name", "manager_email"}
    if reader.fieldnames is None:
        raise HTTPException(status_code=422, detail="CSV has no headers")

    fieldnames_lower = {f.strip().lower() for f in reader.fieldnames}
    missing = required_headers - fieldnames_lower
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required CSV headers: {sorted(missing)}",
        )

    rows = list(reader)

    # Validate all rows first, collect errors
    errors: list[str] = []
    for i, row in enumerate(rows, start=2):  # row 1 is header
        building_name = row.get("building_name", "").strip()
        manager_email = row.get("manager_email", "").strip()
        if not building_name:
            errors.append(f"Row {i}: building_name is empty")
        if not manager_email:
            errors.append(f"Row {i}: manager_email is empty")

    if errors:
        raise HTTPException(status_code=422, detail=errors)

    # Bulk-load all matching buildings in a single query (N+1 fix).
    all_names_lower = [row["building_name"].strip().lower() for row in rows]
    bulk_result = await db.execute(
        select(Building).where(func.lower(Building.name).in_(all_names_lower))
    )
    buildings_by_lower_name: dict[str, Building] = {
        b.name.lower(): b for b in bulk_result.scalars().all()
    }

    created = 0
    updated = 0

    try:
        for row in rows:
            building_name = row["building_name"].strip()
            manager_email = row["manager_email"].strip()

            existing = buildings_by_lower_name.get(building_name.lower())

            if existing is None:
                new_building = Building(name=building_name, manager_email=manager_email)
                db.add(new_building)
                created += 1
            else:
                existing.manager_email = manager_email
                updated += 1

        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Concurrent import conflict — please retry",
        ) from exc
    return {"created": created, "updated": updated}


async def import_buildings_from_excel(
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    """
    Parse an Excel file of buildings and upsert records.
    Returns {"created": int, "updated": int}.
    Raises HTTPException 422 on validation errors.
    """
    # RR4-26: Narrow exception catch to file-format errors only; re-raise unexpected ones.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    if header_row is None or all(v is None for v in header_row):  # pragma: no cover  # openpyxl never yields None as a header row when the sheet has rows; StopIteration handles the empty case above
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    required_headers = {"building_name", "manager_email"}
    missing = required_headers - set(headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required Excel headers: {sorted(missing)}",
        )

    building_name_idx = headers.index("building_name")
    manager_email_idx = headers.index("manager_email")

    data_rows = list(rows_iter)
    wb.close()

    errors: list[str] = []
    parsed: list[dict] = []

    row_num = 0
    for raw_row in data_rows:
        # Skip completely blank rows
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row_num += 1

        def _cell(idx: int) -> str:
            if idx < len(raw_row) and raw_row[idx] is not None:
                return str(raw_row[idx]).strip()
            return ""

        building_name = _cell(building_name_idx)
        manager_email = _cell(manager_email_idx)

        row_errors = []
        if not building_name:
            row_errors.append(f"Row {row_num}: building_name is empty")
        if not manager_email:
            row_errors.append(f"Row {row_num}: manager_email is empty")

        if row_errors:
            errors.extend(row_errors)
        else:
            parsed.append({"building_name": building_name, "manager_email": manager_email})

    if errors:
        raise HTTPException(status_code=422, detail=errors)

    # Bulk-load all matching buildings in a single query (N+1 fix).
    all_names_lower = [r["building_name"].lower() for r in parsed]
    bulk_result = await db.execute(
        select(Building).where(func.lower(Building.name).in_(all_names_lower))
    )
    buildings_by_lower_name: dict[str, Building] = {
        b.name.lower(): b for b in bulk_result.scalars().all()
    }

    created = 0
    updated = 0

    try:
        for row_data in parsed:
            building_name = row_data["building_name"]
            manager_email = row_data["manager_email"]

            existing = buildings_by_lower_name.get(building_name.lower())

            if existing is None:
                new_building = Building(name=building_name, manager_email=manager_email)
                db.add(new_building)
                created += 1
            else:
                existing.manager_email = manager_email
                updated += 1

        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Concurrent import conflict — please retry",
        ) from exc
    return {"created": created, "updated": updated}


# ---------------------------------------------------------------------------
# Subscription / tenant settings
# ---------------------------------------------------------------------------


async def _count_active_buildings(db: AsyncSession) -> int:
    """Return the count of non-archived buildings."""
    result = await db.execute(
        select(func.count()).select_from(Building).where(Building.is_archived == False)  # noqa: E712
    )
    return result.scalar_one()


async def get_subscription(db: AsyncSession):
    """Return the current subscription settings plus active building count.

    Returns defaults (None plan, None limit) when no settings row exists yet.
    """
    from app.schemas.admin import SubscriptionResponse

    result = await db.execute(select(TenantSettings).where(TenantSettings.id == 1))
    settings = result.scalar_one_or_none()

    active_count = await _count_active_buildings(db)

    if settings is None:
        return SubscriptionResponse(
            tier_name=None,
            building_limit=None,
            active_building_count=active_count,
        )
    return SubscriptionResponse(
        tier_name=settings.tier_name,
        building_limit=settings.building_limit,
        active_building_count=active_count,
    )


async def upsert_subscription(
    db: AsyncSession,
    tier_name: str | None,
    building_limit: int | None,
):
    """Upsert the single tenant_settings row (id=1)."""
    from app.schemas.admin import SubscriptionResponse

    result = await db.execute(select(TenantSettings).where(TenantSettings.id == 1))
    settings = result.scalar_one_or_none()

    if settings is None:
        settings = TenantSettings(id=1, tier_name=tier_name, building_limit=building_limit)
        db.add(settings)
    else:
        settings.tier_name = tier_name
        settings.building_limit = building_limit

    await db.commit()
    await db.refresh(settings)

    active_count = await _count_active_buildings(db)
    return SubscriptionResponse(
        tier_name=settings.tier_name,
        building_limit=settings.building_limit,
        active_building_count=active_count,
    )


async def send_subscription_change_request(
    db: AsyncSession,
    *,
    origin: str,
    current_tier: str | None,
    requested_tier: str,
    requester_email: str,
) -> None:
    """Send a tier-change request email to support@ocss.tech.

    Raises SmtpNotConfiguredError if SMTP is not configured.
    """
    from app.services.email_service import SmtpNotConfiguredError
    from app.services.smtp_config_service import get_smtp_config, get_decrypted_password
    import aiosmtplib
    from email.mime.text import MIMEText

    smtp_config = await get_smtp_config(db)
    if (
        not smtp_config.smtp_host
        or not smtp_config.smtp_username
        or not smtp_config.smtp_from_email
        or smtp_config.smtp_password_enc is None
    ):
        raise SmtpNotConfiguredError("SMTP not configured")

    smtp_password = get_decrypted_password(smtp_config)

    tier_display = current_tier or "No plan set"
    body = (
        f"A tier change has been requested.\n\n"
        f"Deployment: {origin}\n"
        f"Current tier: {tier_display}\n"
        f"Requested tier: {requested_tier}\n"
        f"Requested by: {requester_email}\n"
    )

    msg = MIMEText(body, "plain")
    msg["Subject"] = f"Tier change request — {origin}"
    msg["From"] = smtp_config.smtp_from_email
    msg["To"] = "support@ocss.tech"

    await aiosmtplib.send(
        msg,
        hostname=smtp_config.smtp_host,
        port=smtp_config.smtp_port,
        username=smtp_config.smtp_username,
        password=smtp_password,
        start_tls=True,
    )


async def unarchive_building(building_id: uuid.UUID, db: AsyncSession) -> Building:
    """Set is_archived=False on the given building. Raises 404 if not found."""
    result = await db.execute(select(Building).where(Building.id == building_id))
    building = result.scalar_one_or_none()
    if building is None:
        raise HTTPException(status_code=404, detail="Building not found")
    building.is_archived = False
    building.unarchive_count += 1
    await db.commit()
    await db.refresh(building)
    return building


# ---------------------------------------------------------------------------
# Buildings (CRUD)
# ---------------------------------------------------------------------------


async def create_building(name: str, manager_email: str, db: AsyncSession) -> Building:
    # Enforce subscription building limit (if set).
    sub_result = await db.execute(select(TenantSettings).where(TenantSettings.id == 1))
    settings = sub_result.scalar_one_or_none()
    if settings is not None and settings.building_limit is not None:
        active_count = await _count_active_buildings(db)
        if active_count >= settings.building_limit:
            tier = settings.tier_name or "current"
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Building limit reached. You have {active_count} of "
                    f"{settings.building_limit} active buildings on the {tier} plan. "
                    "Contact support to upgrade."
                ),
            )

    result = await db.execute(
        select(Building).where(func.lower(Building.name) == func.lower(name))
    )
    if result.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=409,
            detail=f"A building named '{name}' already exists",
        )
    building = Building(name=name, manager_email=manager_email)
    db.add(building)
    await db.commit()
    await db.refresh(building)
    return building


_BUILDINGS_TEXT_SORT_COLUMNS = {"name", "manager_email"}
_BUILDINGS_SORT_COLUMNS = {
    "name": Building.name,
    "manager_email": Building.manager_email,
    "created_at": Building.created_at,
}


def _buildings_order_clause(sort_by: str | None, sort_dir: str | None):
    key = sort_by or "created_at"
    col = _BUILDINGS_SORT_COLUMNS.get(key, Building.created_at)
    # Use func.lower() for text columns to make sorting case-insensitive
    effective_col = func.lower(col) if key in _BUILDINGS_TEXT_SORT_COLUMNS else col
    if (sort_dir or "desc") == "asc":
        return effective_col.asc()
    return effective_col.desc()


async def list_buildings(
    db: AsyncSession,
    limit: int = 100,
    offset: int = 0,
    name: str | None = None,
    is_archived: bool | None = None,
    sort_by: str | None = None,
    sort_dir: str | None = None,
) -> list[Building]:
    q = select(Building).order_by(_buildings_order_clause(sort_by, sort_dir))
    if name is not None:
        q = q.where(func.lower(Building.name).contains(name.lower()))
    if is_archived is not None:
        q = q.where(Building.is_archived == is_archived)
    result = await db.execute(q.offset(offset).limit(limit))
    return list(result.scalars().all())


async def count_buildings(
    db: AsyncSession,
    name: str | None = None,
    is_archived: bool | None = None,
) -> int:
    """Return total count of buildings matching the optional name and is_archived filters."""
    q = select(func.count()).select_from(Building)
    if name is not None:
        q = q.where(func.lower(Building.name).contains(name.lower()))
    if is_archived is not None:
        q = q.where(Building.is_archived == is_archived)
    result = await db.execute(q)
    return result.scalar_one()


async def archive_building(building_id: uuid.UUID, db: AsyncSession) -> Building:
    """
    Archive a building and any lots whose persons do not appear in another non-archived building.
    """
    building = await get_building_or_404(building_id, db)

    if building.is_archived:
        raise HTTPException(status_code=409, detail="Building is already archived")

    building.is_archived = True

    # Find all active lots for this building
    owners_result = await db.execute(
        select(Lot).where(
            Lot.building_id == building_id,
            Lot.is_archived == False,  # noqa: E712
        )
    )
    owners = list(owners_result.scalars().all())

    # Batch-load person_ids for all lots to avoid O(N) queries (RR3-12).
    owner_ids = [o.id for o in owners]
    all_persons_result = await db.execute(
        select(lot_persons.c.lot_id, lot_persons.c.person_id).where(
            lot_persons.c.lot_id.in_(owner_ids)
        )
    )
    person_ids_by_lot: dict[uuid.UUID, list[uuid.UUID]] = {}
    all_person_ids_flat: list[uuid.UUID] = []
    for row in all_persons_result.all():
        person_ids_by_lot.setdefault(row[0], []).append(row[1])
        all_person_ids_flat.append(row[1])

    # Batch-check which persons appear in another non-archived building — one query
    # instead of one query per (owner, person) pair (RR3-12).
    persons_in_other_buildings: set[uuid.UUID] = set()
    if all_person_ids_flat:
        other_result = await db.execute(
            select(lot_persons.c.person_id)
            .join(Lot, lot_persons.c.lot_id == Lot.id)
            .join(Building, Lot.building_id == Building.id)
            .where(
                lot_persons.c.person_id.in_(all_person_ids_flat),
                Lot.building_id != building_id,
                Building.is_archived == False,  # noqa: E712
            )
        )
        persons_in_other_buildings = {row[0] for row in other_result.all()}

    for owner in owners:
        owner_person_ids = person_ids_by_lot.get(owner.id, [])
        found_in_other = any(pid in persons_in_other_buildings for pid in owner_person_ids)
        if not found_in_other:
            owner.is_archived = True

    await db.commit()
    result = await db.execute(
        select(Building).where(Building.id == building_id)
    )
    building = result.scalar_one()
    return building


async def update_building(
    building_id: uuid.UUID,
    data: BuildingUpdate,
    db: AsyncSession,
) -> Building:
    """Update name and/or manager_email on an existing building."""
    building = await get_building_or_404(building_id, db)
    if data.name is not None:
        building.name = data.name
    if data.manager_email is not None:
        building.manager_email = data.manager_email
    await db.commit()
    await db.refresh(building)
    return building


async def delete_building(building_id: uuid.UUID, db: AsyncSession) -> None:
    """Permanently delete an archived building and all its cascade data."""
    building = await get_building_or_404(building_id, db)
    if not building.is_archived:
        raise HTTPException(
            status_code=409,
            detail="Only archived buildings can be deleted",
        )
    await db.delete(building)
    await db.flush()
    # After cascade-delete of lots → lot_persons rows, clean up persons that are
    # no longer linked to any lot or proxy across all buildings.
    await db.execute(
        delete(Person).where(
            Person.id.not_in(select(lot_persons.c.person_id)),
            Person.id.not_in(select(LotProxy.person_id)),
        )
    )
    await db.commit()


# ---------------------------------------------------------------------------
# Lot owners
# ---------------------------------------------------------------------------


async def get_building_or_404(building_id: uuid.UUID, db: AsyncSession) -> Building:
    result = await db.execute(select(Building).where(Building.id == building_id))
    building = result.scalar_one_or_none()
    if building is None:
        raise HTTPException(status_code=404, detail="Building not found")
    return building


async def _get_proxy_info(lot_id: uuid.UUID, db: AsyncSession) -> dict | None:
    """Return {proxy_email, given_name, surname} for the lot's proxy via the persons join, or None."""
    proxy_result = await db.execute(
        select(Person.email, Person.given_name, Person.surname)
        .join(LotProxy, LotProxy.person_id == Person.id)
        .where(LotProxy.lot_id == lot_id)
    )
    row = proxy_result.first()
    if row is None:
        return None
    return {"proxy_email": row[0], "given_name": row[1], "surname": row[2]}


async def count_lot_owners(building_id: uuid.UUID, db: AsyncSession) -> int:
    result = await db.execute(
        select(func.count()).select_from(Lot).where(
            Lot.building_id == building_id,
        )
    )
    return result.scalar_one()


def _format_financial_position(fp: object) -> str:
    """Return the string value of a FinancialPosition enum or pass-through a string."""
    return fp.value if hasattr(fp, "value") else fp  # type: ignore[attr-defined]


def _person_to_dict(row: Person) -> dict:
    """Serialise a Person ORM row to the persons dict shape."""
    return {
        "id": row.id,
        "email": row.email,
        "given_name": row.given_name,
        "surname": row.surname,
        "phone_number": row.phone_number,
    }


async def _load_persons_for_one(lot_id: uuid.UUID, db: AsyncSession) -> list[dict]:
    """Return persons list for a single lot."""
    result = await db.execute(
        select(Person)
        .join(lot_persons, lot_persons.c.person_id == Person.id)
        .where(lot_persons.c.lot_id == lot_id)
    )
    return [_person_to_dict(row) for row in result.scalars().all()]


async def get_or_create_person(
    email: str,
    db: AsyncSession,
    given_name: str | None = None,
    surname: str | None = None,
) -> Person:
    """Look up a Person by email (case-insensitive). Create if not found.

    Fill-blanks policy: given_name/surname are only applied if the Person row
    currently has NULL for those fields (never overwrite existing names).
    """
    normalised = email.strip().lower()
    result = await db.execute(select(Person).where(Person.email == normalised))
    person = result.scalar_one_or_none()
    if person is None:
        person = Person(email=normalised, given_name=given_name, surname=surname)
        db.add(person)
        await db.flush()
    else:
        # Fill-blanks: only set if currently NULL
        if given_name is not None and person.given_name is None:
            person.given_name = given_name
        if surname is not None and person.surname is None:
            person.surname = surname
    return person


async def lookup_person(email: str, db: AsyncSession) -> Person | None:
    """Look up a Person by email. Returns None if not found."""
    normalised = email.strip().lower()
    result = await db.execute(select(Person).where(Person.email == normalised))
    return result.scalar_one_or_none()


async def update_person(
    person_id: uuid.UUID,
    data: UpdatePersonRequest,
    db: AsyncSession,
) -> Person:
    """Patch name/phone/email on a persons row. Returns updated Person. 404 if not found, 409 on email conflict."""
    result = await db.execute(select(Person).where(Person.id == person_id))
    person = result.scalar_one_or_none()
    if person is None:
        raise HTTPException(status_code=404, detail="Person not found")

    if "email" in data.model_fields_set and data.email is not None:
        normalised = data.email.strip().lower()
        conflict = await db.execute(
            select(Person).where(Person.email == normalised, Person.id != person_id)
        )
        if conflict.scalar_one_or_none() is not None:
            raise HTTPException(status_code=409, detail="Email already in use by another person")
        person.email = normalised

    if "given_name" in data.model_fields_set:
        raw = data.given_name
        person.given_name = bleach.clean(raw, tags=[], strip=True).strip() or None if raw else None

    if "surname" in data.model_fields_set:
        raw = data.surname
        person.surname = bleach.clean(raw, tags=[], strip=True).strip() or None if raw else None

    if "phone_number" in data.model_fields_set:
        person.phone_number = data.phone_number

    await db.commit()
    await db.refresh(person)
    return person


def _lot_owners_order_clause(sort_by: str | None, sort_dir: str | None):
    key = sort_by or "lot_number"
    descending = (sort_dir or "asc") == "desc"

    if key == "unit_entitlement":
        col = Lot.unit_entitlement
        return col.desc() if descending else col.asc()

    if key == "financial_position":
        col = Lot.financial_position
        return col.desc() if descending else col.asc()

    if key == "email":
        # Correlated subquery: min(lower(email)) for persons linked to this lot.
        # Lots with no person sort last regardless of direction (nullslast).
        from sqlalchemy import nullslast, nullsfirst
        email_subq = (
            select(func.min(func.lower(Person.email)))
            .join(lot_persons, lot_persons.c.person_id == Person.id)
            .where(lot_persons.c.lot_id == Lot.id)
            .correlate(Lot)
            .scalar_subquery()
        )
        if descending:
            return nullslast(email_subq.desc())
        return nullsfirst(email_subq.asc())

    if key == "proxy_email":
        # Correlated subquery: lower(proxy person email) for this lot's proxy.
        # Lots with no proxy sort last regardless of direction (nullslast).
        from sqlalchemy import nullslast, nullsfirst
        proxy_email_subq = (
            select(func.lower(Person.email))
            .join(LotProxy, LotProxy.person_id == Person.id)
            .where(LotProxy.lot_id == Lot.id)
            .correlate(Lot)
            .scalar_subquery()
        )
        if descending:
            return nullslast(proxy_email_subq.desc())
        return nullsfirst(proxy_email_subq.asc())

    # Default: lot_number — natural numeric sort by stripping non-digit suffix
    # and casting the leading numeric portion to integer, then falling back to
    # the raw string for lots whose numbers have no numeric prefix.
    numeric_prefix = func.cast(
        func.nullif(func.regexp_replace(Lot.lot_number, r"\D.*", "", "g"), ""),
        Integer,
    )
    col = (numeric_prefix, Lot.lot_number)
    if descending:
        return (col[0].desc(), col[1].desc())
    return (col[0].asc(), col[1].asc())


async def list_lot_owners(
    building_id: uuid.UUID,
    db: AsyncSession,
    limit: int = 20,
    offset: int = 0,
    sort_by: str | None = None,
    sort_dir: str | None = None,
) -> list[dict]:
    await get_building_or_404(building_id, db)
    order = _lot_owners_order_clause(sort_by, sort_dir)
    order_args = order if isinstance(order, tuple) else (order,)
    result = await db.execute(
        select(Lot)
        .where(Lot.building_id == building_id)
        .order_by(*order_args)
        .offset(offset)
        .limit(limit)
    )
    owners = list(result.scalars().all())

    if not owners:
        return []

    # Batch-load persons and proxies with IN queries — O(1) queries regardless of owner count.
    owner_ids = [o.id for o in owners]

    persons_result = await db.execute(
        select(lot_persons.c.lot_id, Person)
        .join(Person, lot_persons.c.person_id == Person.id)
        .where(lot_persons.c.lot_id.in_(owner_ids))
    )
    persons_by_lot: dict[uuid.UUID, list[dict]] = {}
    for row in persons_result.all():
        lot_id_val = row[0]
        person_obj = row[1]
        persons_by_lot.setdefault(lot_id_val, []).append(_person_to_dict(person_obj))

    proxies_result = await db.execute(
        select(LotProxy.lot_id, Person.email, Person.given_name, Person.surname)
        .join(Person, LotProxy.person_id == Person.id)
        .where(LotProxy.lot_id.in_(owner_ids))
    )
    proxy_by_lot: dict[uuid.UUID, dict] = {
        row[0]: {"proxy_email": row[1], "given_name": row[2], "surname": row[3]}
        for row in proxies_result.all()
    }

    out = []
    for owner in owners:
        proxy_info = proxy_by_lot.get(owner.id, {})
        out.append({
            "id": owner.id,
            "lot_number": owner.lot_number,
            "persons": persons_by_lot.get(owner.id, []),
            "unit_entitlement": owner.unit_entitlement,
            "financial_position": _format_financial_position(owner.financial_position),
            "proxy_email": proxy_info.get("proxy_email"),
            "proxy_given_name": proxy_info.get("given_name"),
            "proxy_surname": proxy_info.get("surname"),
        })
    return out


async def get_lot_owner(lot_owner_id: uuid.UUID, db: AsyncSession) -> dict:
    """Return a single lot by ID, including proxy info."""
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


_CSV_LOT_OWNER_ALIASES: dict[str, str] = {
    "lot#": "lot_number",
    "uoe2": "unit_entitlement",
    "phone": "phone_number",
}


def _normalise_lot_owner_fieldnames(fieldnames: list[str]) -> list[str]:
    """Map alternate header names (Lot#, UOE2, Phone) to canonical names."""
    return [_CSV_LOT_OWNER_ALIASES.get(f.strip().lower(), f.strip().lower()) for f in fieldnames]


def _normalise_phone_e164(raw: str) -> str | None:
    """Normalise a phone number string to E.164 format.

    Rules (applied after stripping spaces/dashes/brackets/dots):
    - Already starts with '+': keep as-is.
    - Starts with '04' (Australian mobile): replace leading '0' with '+61'.
    - Otherwise: store as-is (non-AU numbers without country code).
    - Blank / whitespace-only: return None.
    """
    import re
    stripped = re.sub(r"[\s\-\(\)\.]", "", raw).strip()
    if not stripped:
        return None
    if stripped.startswith("+"):
        return stripped
    if stripped.startswith("04"):
        return "+61" + stripped[1:]
    return stripped


def _parse_financial_position(raw: str) -> FinancialPosition:
    """Parse a financial position string from an import row.

    Accepted values (case-insensitive): 'normal', 'in arrear', 'in_arrear'.
    Blank/missing values default to 'normal'.
    """
    normalised = raw.strip().lower().replace(" ", "_")
    if normalised == "" or normalised == "normal":
        return FinancialPosition.normal
    if normalised == "in_arrear":
        return FinancialPosition.in_arrear
    raise ValueError(f"Invalid financial_position value: '{raw}'")


async def import_lot_owners_from_csv(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    """
    Parse a CSV of lot owners. Multiple rows with the same lot_number create one lot with
    multiple LotOwnerEmail records. Blank email is allowed (no email entry created).
    Returns {"imported": int, "emails": int}.
    """
    await get_building_or_404(building_id, db)

    text = content.decode("utf-8-sig")
    raw_reader = csv.DictReader(io.StringIO(text))

    if raw_reader.fieldnames is None:
        raise HTTPException(status_code=422, detail="CSV has no headers")

    normalised_fieldnames = _normalise_lot_owner_fieldnames(list(raw_reader.fieldnames))
    # Re-read with normalised fieldnames
    reader = csv.DictReader(io.StringIO(text), fieldnames=normalised_fieldnames)
    next(reader)  # skip original header row

    required_headers = {"lot_number", "unit_entitlement"}
    missing = required_headers - set(normalised_fieldnames)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required CSV headers: {sorted(missing)}",
        )

    rows = list(reader)

    # Determine name mode from available columns
    if "given_name" in normalised_fieldnames and "surname" in normalised_fieldnames:
        name_mode = "separate"
    elif "name" in normalised_fieldnames:
        name_mode = "name"
    else:
        name_mode = "none"

    errors: list[str] = []
    # Parse rows: group by lot_number
    lot_data: dict[str, dict] = {}  # lot_number -> {unit_entitlement, financial_position, email_entries: list}
    # Track row numbers seen per lot_number to detect duplicates (RR3-31)
    lot_number_rows: dict[str, list[int]] = {}

    has_phone_col = "phone_number" in normalised_fieldnames

    for i, row in enumerate(rows, start=2):
        lot_number = row.get("lot_number", "").strip()
        email = row.get("email", "").strip()
        unit_entitlement_raw = row.get("unit_entitlement", "").strip()
        financial_position_raw = row.get("financial_position", "").strip()
        phone_raw = row.get("phone_number", "").strip() if has_phone_col else ""

        row_errors = []

        if not lot_number:
            row_errors.append(f"Row {i}: lot_number is empty")

        unit_entitlement = None
        if not unit_entitlement_raw:
            row_errors.append(f"Row {i}: unit_entitlement is empty")
        else:
            try:
                unit_entitlement = int(unit_entitlement_raw)
                if unit_entitlement <= 0:
                    row_errors.append(
                        f"Row {i}: unit_entitlement must be > 0, got {unit_entitlement}"
                    )
            except ValueError:
                row_errors.append(
                    f"Row {i}: unit_entitlement must be an integer, got '{unit_entitlement_raw}'"
                )

        financial_position = FinancialPosition.normal
        try:
            financial_position = _parse_financial_position(financial_position_raw)
        except ValueError as e:
            row_errors.append(f"Row {i}: {e}")

        if row_errors:
            errors.extend(row_errors)
        else:
            if lot_number:
                lot_number_rows.setdefault(lot_number, []).append(i)

            # Resolve per-row name fields
            if name_mode == "separate":
                row_given_name = row.get("given_name", "").strip() or None
                row_surname = row.get("surname", "").strip() or None
            elif name_mode == "name":
                row_given_name, row_surname = _parse_name(row.get("name", ""))
            else:
                row_given_name, row_surname = None, None

            phone_number = _normalise_phone_e164(phone_raw) if has_phone_col else None

            if lot_number not in lot_data:
                lot_data[lot_number] = {
                    "unit_entitlement": unit_entitlement,
                    "financial_position": financial_position,
                    "email_entries": [],
                    "given_name": row_given_name,
                    "surname": row_surname,
                }
            for addr in email.split(";"):
                addr = addr.strip().lower()
                if addr:
                    # Deduplicate by email: last-row name/phone wins
                    existing_entry = next(
                        (e for e in lot_data[lot_number]["email_entries"] if e["email"] == addr),
                        None,
                    )
                    if existing_entry is not None:
                        existing_entry["given_name"] = row_given_name
                        existing_entry["surname"] = row_surname
                        existing_entry["phone_number"] = phone_number
                    else:
                        lot_data[lot_number]["email_entries"].append(
                            {
                                "email": addr,
                                "given_name": row_given_name,
                                "surname": row_surname,
                                "phone_number": phone_number,
                            }
                        )

    # Check for duplicate lot numbers and report them with row details (RR3-31)
    for lot_num, row_nums in lot_number_rows.items():
        if len(row_nums) > 1:
            rows_str = " and ".join(str(r) for r in row_nums)
            errors.append(f"Lot {lot_num} appears on rows {rows_str}")

    if errors:
        raise HTTPException(status_code=422, detail=errors)

    return await _upsert_lot_owners(building_id, lot_data, db)


async def import_lot_owners_from_excel(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    """
    Parse an Excel file of lot owners. Multiple rows with the same Lot# create one lot
    with multiple LotOwnerEmail records. Blank email is allowed.
    Returns {"imported": int, "emails": int}.
    """
    await get_building_or_404(building_id, db)

    # RR4-26: Narrow exception catch to file-format errors only; re-raise unexpected ones.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    if header_row is None or all(v is None for v in header_row):  # pragma: no cover  # openpyxl never yields None as a header row when the sheet has rows; StopIteration handles the empty case above
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    required_headers = {"lot#", "uoe2"}
    missing = required_headers - set(headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required Excel headers: {sorted(missing)}",
        )

    lot_idx = headers.index("lot#")
    uoe2_idx = headers.index("uoe2")
    email_idx = headers.index("email") if "email" in headers else None
    fp_idx = headers.index("financial position") if "financial position" in headers else (
        headers.index("financial_position") if "financial_position" in headers else None
    )
    given_name_idx = headers.index("given_name") if "given_name" in headers else None
    surname_idx = headers.index("surname") if "surname" in headers else None
    name_idx = headers.index("name") if "name" in headers else None
    phone_idx = headers.index("phone") if "phone" in headers else (
        headers.index("phone_number") if "phone_number" in headers else None
    )

    # Determine name mode from available columns
    if given_name_idx is not None and surname_idx is not None:
        name_mode = "separate"
    elif name_idx is not None:
        name_mode = "name"
    else:
        name_mode = "none"

    data_rows = list(rows_iter)
    wb.close()

    errors: list[str] = []
    lot_data: dict[str, dict] = {}  # lot_number -> {unit_entitlement, financial_position, email_entries: list}
    # Track row numbers seen per lot_number to detect duplicates (RR3-31)
    lot_number_rows: dict[str, list[int]] = {}

    row_num = 0
    for raw_row in data_rows:
        # Skip completely blank rows
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row_num += 1

        def _cell(idx: int) -> str:
            if idx is not None and idx < len(raw_row) and raw_row[idx] is not None:
                return str(raw_row[idx]).strip()
            return ""

        lot_number = _cell(lot_idx)
        email = _cell(email_idx) if email_idx is not None else ""
        unit_entitlement_raw = _cell(uoe2_idx)
        financial_position_raw = _cell(fp_idx) if fp_idx is not None else ""
        phone_raw = _cell(phone_idx) if phone_idx is not None else ""

        row_errors = []

        if not lot_number:
            row_errors.append(f"Row {row_num}: lot_number is empty")

        unit_entitlement = None
        if not unit_entitlement_raw:
            row_errors.append(f"Row {row_num}: unit_entitlement is empty")
        else:
            try:
                unit_entitlement = int(unit_entitlement_raw)
                if unit_entitlement <= 0:
                    row_errors.append(
                        f"Row {row_num}: unit_entitlement must be > 0, got {unit_entitlement}"
                    )
            except ValueError:
                row_errors.append(
                    f"Row {row_num}: unit_entitlement must be an integer, got '{unit_entitlement_raw}'"
                )

        financial_position = FinancialPosition.normal
        try:
            financial_position = _parse_financial_position(financial_position_raw)
        except ValueError as e:
            row_errors.append(f"Row {row_num}: {e}")

        if row_errors:
            errors.extend(row_errors)
        else:
            if lot_number:
                lot_number_rows.setdefault(lot_number, []).append(row_num)

            # Resolve per-row name fields
            if name_mode == "separate":
                row_given_name = _cell(given_name_idx).strip() or None
                row_surname = _cell(surname_idx).strip() or None
            elif name_mode == "name":
                row_given_name, row_surname = _parse_name(_cell(name_idx))
            else:
                row_given_name, row_surname = None, None

            phone_number = _normalise_phone_e164(phone_raw) if phone_idx is not None else None

            if lot_number not in lot_data:
                lot_data[lot_number] = {
                    "unit_entitlement": unit_entitlement,
                    "financial_position": financial_position,
                    "email_entries": [],
                    "given_name": row_given_name,
                    "surname": row_surname,
                }
            for addr in email.split(";"):
                addr = addr.strip().lower()
                if addr:
                    # Deduplicate by email: last-row name/phone wins
                    existing_entry = next(
                        (e for e in lot_data[lot_number]["email_entries"] if e["email"] == addr),
                        None,
                    )
                    if existing_entry is not None:
                        existing_entry["given_name"] = row_given_name
                        existing_entry["surname"] = row_surname
                        existing_entry["phone_number"] = phone_number
                    else:
                        lot_data[lot_number]["email_entries"].append(
                            {
                                "email": addr,
                                "given_name": row_given_name,
                                "surname": row_surname,
                                "phone_number": phone_number,
                            }
                        )

    # Check for duplicate lot numbers and report them with row details (RR3-31)
    for lot_num, row_nums in lot_number_rows.items():
        if len(row_nums) > 1:
            rows_str = " and ".join(str(r) for r in row_nums)
            errors.append(f"Lot {lot_num} appears on rows {rows_str}")

    if errors:
        raise HTTPException(status_code=422, detail=errors)

    return await _upsert_lot_owners(building_id, lot_data, db)


async def _resolve_or_create_person_for_import(
    email: str,
    given_name: str | None,
    surname: str | None,
    phone_number: str | None,
    db: AsyncSession,
) -> Person:
    """Look up or create a Person for an import row, applying fill-blanks policy."""
    normalised = email.strip().lower()
    result = await db.execute(select(Person).where(Person.email == normalised))
    person = result.scalar_one_or_none()
    if person is None:
        person = Person(
            email=normalised,
            given_name=given_name,
            surname=surname,
            phone_number=phone_number,
        )
        db.add(person)
        await db.flush()
    else:
        # Always overwrite on import (CSV is authoritative source of truth)
        if given_name is not None or surname is not None:
            person.given_name = given_name
            person.surname = surname
        if phone_number is not None:
            person.phone_number = phone_number
    return person


async def _upsert_lot_owners(
    building_id: uuid.UUID,
    lot_data: dict[str, dict],
    db: AsyncSession,
) -> dict[str, int]:
    """
    Upsert lots from parsed lot_data. Rebuild lot_persons links for the building.
    lot_data: {lot_number -> {unit_entitlement, financial_position, email_entries: list[dict]}}
    Returns {"imported": int, "emails": int}.
    """
    # Load existing lots keyed by lot_number to preserve IDs
    existing_result = await db.execute(
        select(Lot).where(Lot.building_id == building_id)
    )
    existing: dict[str, Lot] = {
        lo.lot_number: lo for lo in existing_result.scalars().all()
    }

    # Step 1: Upsert lots (never delete — would cascade-delete AGMLotWeight records)
    for lot_number, data in lot_data.items():
        if lot_number in existing:
            lo = existing[lot_number]
            lo.unit_entitlement = data["unit_entitlement"]
            lo.financial_position = data["financial_position"]
            await db.flush()
        else:
            new_lo = Lot(
                building_id=building_id,
                lot_number=lot_number,
                unit_entitlement=data["unit_entitlement"],
                financial_position=data["financial_position"],
            )
            db.add(new_lo)
            await db.flush()
            existing[lot_number] = new_lo

    # Step 2: Clear ALL lot_persons links for every lot in this building
    # (authoritative delete-and-rebuild — CSV is the complete source of truth)
    all_lot_ids = [lo.id for lo in existing.values()]
    if all_lot_ids:
        await db.execute(
            delete(lot_persons).where(lot_persons.c.lot_id.in_(all_lot_ids))
        )

    # Step 3: Rebuild lot_persons from CSV rows
    total_emails = 0
    for lot_number, data in lot_data.items():
        lo = existing[lot_number]
        for entry in data["email_entries"]:
            if entry["email"]:
                person = await _resolve_or_create_person_for_import(
                    entry["email"],
                    entry.get("given_name"),
                    entry.get("surname"),
                    entry.get("phone_number"),
                    db,
                )
                # on_conflict_do_nothing handles duplicate emails on same lot row
                from sqlalchemy.dialects.postgresql import insert as pg_insert
                await db.execute(
                    pg_insert(lot_persons)
                    .values(lot_id=lo.id, person_id=person.id)
                    .on_conflict_do_nothing()
                )
                total_emails += 1

    # Step 4: Delete orphaned persons — persons no longer linked to any lot or proxy.
    # After rebuilding lot_persons for this building, any person who has no remaining
    # lot_persons rows and no lot_proxies rows across ALL buildings is unreachable and
    # should be cleaned up to avoid stale data accumulation.
    await db.execute(
        delete(Person).where(
            Person.id.not_in(select(lot_persons.c.person_id)),
            Person.id.not_in(select(LotProxy.person_id)),
        )
    )

    await db.commit()

    return {"imported": len(lot_data), "emails": total_emails}


async def add_lot_owner(
    building_id: uuid.UUID,
    data: LotOwnerCreate,
    db: AsyncSession,
) -> dict:
    await get_building_or_404(building_id, db)

    # Check uniqueness within building
    result = await db.execute(
        select(Lot).where(
            Lot.building_id == building_id,
            Lot.lot_number == data.lot_number,
        )
    )
    existing = result.scalar_one_or_none()
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail=f"lot_number '{data.lot_number}' already exists in this building",
        )

    lot = Lot(
        building_id=building_id,
        lot_number=data.lot_number,
        unit_entitlement=data.unit_entitlement,
        financial_position=FinancialPosition(data.financial_position),
    )
    db.add(lot)
    await db.flush()

    for email in data.emails:
        if email.strip():
            person = await get_or_create_person(email.strip().lower(), db)
            from sqlalchemy.dialects.postgresql import insert as pg_insert
            await db.execute(
                pg_insert(lot_persons)
                .values(lot_id=lot.id, person_id=person.id)
                .on_conflict_do_nothing()
            )

    await db.commit()

    persons = await _load_persons_for_one(lot.id, db)
    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": None,
        "proxy_given_name": None,
        "proxy_surname": None,
    }


async def update_lot_owner(
    lot_owner_id: uuid.UUID,
    data: LotOwnerUpdate,
    db: AsyncSession,
) -> dict:
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    if data.unit_entitlement is not None:
        lot.unit_entitlement = data.unit_entitlement
    if data.financial_position is not None:
        lot.financial_position = FinancialPosition(data.financial_position)

    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)
    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def add_email_to_lot_owner(
    lot_owner_id: uuid.UUID,
    email: str,
    db: AsyncSession,
) -> dict:
    """Add an email (person) to a lot. Returns the updated lot dict."""
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    normalised = email.strip().lower()

    # Check if this person is already linked to this lot
    person = await get_or_create_person(normalised, db)
    existing_link = await db.execute(
        select(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person.id,
        )
    )
    if existing_link.first() is not None:
        raise HTTPException(status_code=409, detail="Email already exists for this lot owner")

    from sqlalchemy.dialects.postgresql import insert as pg_insert
    await db.execute(
        pg_insert(lot_persons)
        .values(lot_id=lot_owner_id, person_id=person.id)
        .on_conflict_do_nothing()
    )
    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def remove_email_from_lot_owner(
    lot_owner_id: uuid.UUID,
    email: str,
    db: AsyncSession,
) -> dict:
    """Remove a person from a lot by email. Returns the updated lot dict."""
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    person_result = await db.execute(
        select(Person).where(Person.email == email.strip().lower())
    )
    person = person_result.scalar_one_or_none()
    if person is None:
        raise HTTPException(status_code=404, detail="Email not found for this lot owner")

    link_result = await db.execute(
        select(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person.id,
        )
    )
    if link_result.first() is None:
        raise HTTPException(status_code=404, detail="Email not found for this lot owner")

    await db.execute(
        delete(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person.id,
        )
    )
    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def set_lot_owner_proxy(
    lot_owner_id: uuid.UUID,
    proxy_email: str,
    db: AsyncSession,
    given_name: str | None = None,
    surname: str | None = None,
) -> dict:
    """Create or replace the proxy nomination for a lot. Resolves/creates person by email."""
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    person = await get_or_create_person(proxy_email, db)

    # Update person's name fields if provided (fill-blanks policy: only set if currently NULL)
    if given_name is not None and person.given_name is None:
        person.given_name = bleach.clean(given_name, tags=[], strip=True).strip() or None
    if surname is not None and person.surname is None:
        person.surname = bleach.clean(surname, tags=[], strip=True).strip() or None

    proxy_result = await db.execute(
        select(LotProxy).where(LotProxy.lot_id == lot_owner_id)
    )
    existing_proxy = proxy_result.scalar_one_or_none()
    if existing_proxy is not None:
        existing_proxy.person_id = person.id
    else:
        db.add(LotProxy(
            lot_id=lot_owner_id,
            person_id=person.id,
        ))

    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def remove_lot_owner_proxy(
    lot_owner_id: uuid.UUID,
    db: AsyncSession,
) -> dict:
    """Remove the proxy nomination for a lot. 404 if no proxy is set."""
    result = await db.execute(
        select(Lot).where(Lot.id == lot_owner_id)
    )
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    proxy_result = await db.execute(
        select(LotProxy).where(LotProxy.lot_id == lot_owner_id)
    )
    existing_proxy = proxy_result.scalar_one_or_none()
    if existing_proxy is None:
        raise HTTPException(status_code=404, detail="No proxy nomination found for this lot owner")

    await db.delete(existing_proxy)
    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": None,
        "proxy_given_name": None,
        "proxy_surname": None,
    }


async def add_owner_email_to_lot_owner(
    lot_owner_id: uuid.UUID,
    email: str,
    db: AsyncSession,
    given_name: str | None = None,
    surname: str | None = None,
    phone_number: str | None = None,
) -> dict:
    """Add a person (by email) to a lot. Look up or create the person row. Returns updated lot dict."""
    result = await db.execute(select(Lot).where(Lot.id == lot_owner_id))
    lot = result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    normalised_email = email.strip().lower()
    person = await get_or_create_person(normalised_email, db)

    # Fill-blanks policy: only update name/phone if currently NULL on the person row
    if given_name is not None and person.given_name is None:
        person.given_name = bleach.clean(given_name, tags=[], strip=True).strip() or None
    if surname is not None and person.surname is None:
        person.surname = bleach.clean(surname, tags=[], strip=True).strip() or None
    if phone_number is not None and person.phone_number is None:
        person.phone_number = phone_number

    # Check if person already linked to this lot
    existing_link = await db.execute(
        select(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person.id,
        )
    )
    if existing_link.first() is not None:
        raise HTTPException(status_code=409, detail="Email already exists for this lot owner")

    from sqlalchemy.dialects.postgresql import insert as pg_insert
    await db.execute(
        pg_insert(lot_persons)
        .values(lot_id=lot_owner_id, person_id=person.id)
        .on_conflict_do_nothing()
    )
    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def remove_person_from_lot(
    lot_owner_id: uuid.UUID,
    person_id: uuid.UUID,
    db: AsyncSession,
) -> dict:
    """Remove a person from a lot by person_id. The persons row is NOT deleted."""
    lo_result = await db.execute(select(Lot).where(Lot.id == lot_owner_id))
    lot = lo_result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    link_result = await db.execute(
        select(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person_id,
        )
    )
    if link_result.first() is None:
        raise HTTPException(status_code=404, detail="Person not linked to this lot")

    await db.execute(
        delete(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == person_id,
        )
    )
    await db.commit()

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


# Kept for backward-compatibility with old endpoint paths
async def update_owner_email(
    lot_owner_id: uuid.UUID,
    email_id: uuid.UUID,
    email: str | None,
    given_name: str | None,
    surname: str | None,
    db: AsyncSession,
    phone_number: str | None = None,
    clear_phone: bool = False,
) -> dict:
    """Update a person row by person_id and return the updated lot dict.

    The old endpoint patched LotOwnerEmail rows; this now delegates to update_person.
    email_id is treated as person_id.

    Returns 404 if the person is not linked to the specified lot.
    """
    from app.schemas.admin import UpdatePersonRequest as _UPR

    # Verify the lot exists
    lo_result = await db.execute(select(Lot).where(Lot.id == lot_owner_id))
    lot = lo_result.scalar_one_or_none()
    if lot is None:
        raise HTTPException(status_code=404, detail="Lot owner not found")

    # Verify the person is linked to this lot via lot_persons
    link_result = await db.execute(
        select(lot_persons).where(
            lot_persons.c.lot_id == lot_owner_id,
            lot_persons.c.person_id == email_id,
        )
    )
    if link_result.first() is None:
        raise HTTPException(status_code=404, detail="Email record not found for this lot owner")

    fields: dict = {}
    if email is not None:
        fields["email"] = email
    if given_name is not None:
        fields["given_name"] = given_name
    if surname is not None:
        fields["surname"] = surname
    if phone_number is not None:
        fields["phone_number"] = phone_number
    elif clear_phone:
        fields["phone_number"] = None

    if fields:
        req = _UPR.model_construct(_fields_set=set(fields.keys()), **fields)
        await update_person(email_id, req, db)

    persons = await _load_persons_for_one(lot_owner_id, db)
    proxy_info = await _get_proxy_info(lot_owner_id, db)

    return {
        "id": lot.id,
        "lot_number": lot.lot_number,
        "persons": persons,
        "unit_entitlement": lot.unit_entitlement,
        "financial_position": _format_financial_position(lot.financial_position),
        "proxy_email": proxy_info["proxy_email"] if proxy_info else None,
        "proxy_given_name": proxy_info["given_name"] if proxy_info else None,
        "proxy_surname": proxy_info["surname"] if proxy_info else None,
    }


async def remove_owner_email_by_id(
    lot_owner_id: uuid.UUID,
    email_id: uuid.UUID,
    db: AsyncSession,
) -> dict:
    """Remove a person from a lot by person_id (email_id treated as person_id)."""
    return await remove_person_from_lot(lot_owner_id, email_id, db)


# ---------------------------------------------------------------------------
# General Meetings
# ---------------------------------------------------------------------------


async def create_general_meeting(data: GeneralMeetingCreate, db: AsyncSession) -> GeneralMeeting:
    # Validate building exists
    building = await get_building_or_404(data.building_id, db)

    # Check no open or pending General Meeting already exists for this building
    result = await db.execute(
        select(GeneralMeeting).where(
            GeneralMeeting.building_id == data.building_id,
            GeneralMeeting.status.in_([GeneralMeetingStatus.open, GeneralMeetingStatus.pending]),
        )
    )
    existing_open = result.scalar_one_or_none()
    if existing_open is not None:
        raise HTTPException(
            status_code=409,
            detail="An open or pending General Meeting already exists for this building.",
        )

    # Set initial status based on meeting_at
    initial_status = (
        GeneralMeetingStatus.pending
        if data.meeting_at > datetime.now(timezone.utc)
        else GeneralMeetingStatus.open
    )

    # Create General Meeting
    general_meeting = GeneralMeeting(
        building_id=data.building_id,
        title=data.title,
        status=initial_status,
        meeting_at=data.meeting_at,
        voting_closes_at=data.voting_closes_at,
    )
    db.add(general_meeting)
    await db.flush()  # get general_meeting.id

    # Create motions — sort by supplied display_order, then normalise to 1-based integers
    sorted_motions = sorted(data.motions, key=lambda m: m.display_order)
    # Validate no duplicate non-null motion_numbers within this meeting
    seen_motion_numbers: set[str] = set()
    for motion_data in sorted_motions:
        mn = motion_data.motion_number.strip() if motion_data.motion_number else None
        mn = mn if mn else None
        if mn is not None:
            if mn in seen_motion_numbers:
                raise HTTPException(
                    status_code=409,
                    detail=f"Duplicate motion number '{mn}' within the same General Meeting.",
                )
            seen_motion_numbers.add(mn)
    for position, motion_data in enumerate(sorted_motions, start=1):
        raw = (motion_data.motion_number or "").strip()
        motion_number = raw if raw else str(position)
        motion = Motion(
            general_meeting_id=general_meeting.id,
            title=motion_data.title,
            description=_sanitise_description(motion_data.description),
            display_order=position,
            motion_number=motion_number,
            motion_type=motion_data.motion_type,
            is_multi_choice=motion_data.is_multi_choice,
            option_limit=motion_data.option_limit if motion_data.is_multi_choice else None,
        )
        db.add(motion)
        await db.flush()  # get motion.id for option FK
        if motion_data.is_multi_choice:
            for opt in motion_data.options:
                db.add(MotionOption(
                    motion_id=motion.id,
                    text=_sanitise_option_text(opt.text),
                    display_order=opt.display_order,
                ))

    # Snapshot lot weights (include financial_position_snapshot)
    lot_owners_result = await db.execute(
        select(Lot).where(Lot.building_id == data.building_id)
    )
    lot_owners = list(lot_owners_result.scalars().all())

    for lot_owner in lot_owners:
        fp = lot_owner.financial_position
        fp_snapshot = FinancialPositionSnapshot(fp.value if hasattr(fp, "value") else fp)
        weight = GeneralMeetingLotWeight(
            general_meeting_id=general_meeting.id,
            lot_id=lot_owner.id,
            unit_entitlement_snapshot=lot_owner.unit_entitlement,
            financial_position_snapshot=fp_snapshot,
        )
        db.add(weight)

    await db.commit()
    await db.refresh(general_meeting)

    # Load motions explicitly (avoid lazy load on async session)
    motions_result = await db.execute(
        select(Motion).where(Motion.general_meeting_id == general_meeting.id).order_by(Motion.display_order)
    )
    loaded_motions = list(motions_result.scalars().all())

    # Load motion options for multi-choice motions
    multi_choice_motion_ids = [m.id for m in loaded_motions if m.is_multi_choice]
    options_by_motion: dict[uuid.UUID, list] = {}
    if multi_choice_motion_ids:
        opts_result = await db.execute(
            select(MotionOption)
            .where(MotionOption.motion_id.in_(multi_choice_motion_ids))
            .order_by(MotionOption.display_order)
        )
        for opt in opts_result.scalars().all():
            options_by_motion.setdefault(opt.motion_id, []).append(opt)

    _ = building  # used for 404 check

    # Return as dict to avoid triggering lazy relationship loads during serialization
    return {
        "id": general_meeting.id,
        "building_id": general_meeting.building_id,
        "title": general_meeting.title,
        "status": general_meeting.status.value if hasattr(general_meeting.status, "value") else general_meeting.status,
        "meeting_at": general_meeting.meeting_at,
        "voting_closes_at": general_meeting.voting_closes_at,
        "motions": [
            {
                "id": m.id,
                "title": m.title,
                "description": m.description,
                "display_order": m.display_order,
                "motion_number": m.motion_number,
                "motion_type": m.motion_type.value if hasattr(m.motion_type, "value") else m.motion_type,
                "is_multi_choice": m.is_multi_choice,
                "option_limit": m.option_limit,
                "options": [
                    {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
                    for opt in options_by_motion.get(m.id, [])
                ],
            }
            for m in loaded_motions
        ],
    }


_MEETINGS_TEXT_SORT_COLUMNS = {"title", "building_name"}
_MEETINGS_SORT_COLUMNS = {
    "title": GeneralMeeting.title,
    "created_at": GeneralMeeting.created_at,
    "meeting_at": GeneralMeeting.meeting_at,
    "voting_closes_at": GeneralMeeting.voting_closes_at,
    "status": GeneralMeeting.status,
    "building_name": Building.name,
}


def _effective_status_case():
    """SQL CASE expression mirroring get_effective_status() for use in WHERE clauses.

    Precedence (mirrors get_effective_status):
      1. stored status = 'closed'  → 'closed'
      2. voting_closes_at < NOW()  → 'closed'
      3. meeting_at > NOW()        → 'pending'
      4. otherwise                 → 'open'

    Returns a fresh expression on each call so func.now() is evaluated at
    query-execution time, not import time.
    """
    return case(
        (GeneralMeeting.status == GeneralMeetingStatus.closed.value, literal("closed")),
        (GeneralMeeting.voting_closes_at < func.now(), literal("closed")),
        (GeneralMeeting.meeting_at > func.now(), literal("pending")),
        else_=literal("open"),
    )


def _meetings_order_clause(sort_by: str | None, sort_dir: str | None):
    key = sort_by or "created_at"
    col = _MEETINGS_SORT_COLUMNS.get(key, GeneralMeeting.created_at)
    # Use func.lower() for text columns to make sorting case-insensitive
    effective_col = func.lower(col) if key in _MEETINGS_TEXT_SORT_COLUMNS else col
    if (sort_dir or "desc") == "asc":
        return effective_col.asc()
    return effective_col.desc()


async def list_general_meetings(
    db: AsyncSession,
    limit: int = 100,
    offset: int = 0,
    name: str | None = None,
    building_id: uuid.UUID | None = None,
    status: str | None = None,
    sort_by: str | None = None,
    sort_dir: str | None = None,
) -> list[dict]:
    q = (
        select(GeneralMeeting, Building.name.label("building_name"))
        .join(Building, GeneralMeeting.building_id == Building.id)
        .order_by(_meetings_order_clause(sort_by, sort_dir))
    )
    if name is not None:
        q = q.where(func.lower(GeneralMeeting.title).contains(name.lower()))
    if building_id is not None:
        q = q.where(GeneralMeeting.building_id == building_id)
    if status is not None:
        q = q.where(_effective_status_case() == status)
    result = await db.execute(q.offset(offset).limit(limit))
    rows = result.all()
    items = []
    for general_meeting, building_name in rows:
        effective = get_effective_status(general_meeting)
        effective_str = effective.value if hasattr(effective, "value") else effective
        items.append(
            {
                "id": general_meeting.id,
                "building_id": general_meeting.building_id,
                "building_name": building_name,
                "title": general_meeting.title,
                "status": effective_str,
                "meeting_at": general_meeting.meeting_at,
                "voting_closes_at": general_meeting.voting_closes_at,
                "created_at": general_meeting.created_at,
            }
        )
    return items


async def count_general_meetings(
    db: AsyncSession,
    name: str | None = None,
    building_id: uuid.UUID | None = None,
    status: str | None = None,
) -> int:
    """Return total count of general meetings matching the optional filters.

    RR5-09: When a status filter is applied, use a SQL CASE expression to derive
    effective status in the database rather than loading all rows into Python.
    The CASE expression mirrors get_effective_status():
      1. stored status = 'closed'  → 'closed'
      2. voting_closes_at < NOW()  → 'closed'
      3. meeting_at > NOW()        → 'pending'
      4. otherwise                 → 'open'
    """
    q = select(func.count()).select_from(GeneralMeeting)
    if name is not None:
        q = q.where(func.lower(GeneralMeeting.title).contains(name.lower()))
    if building_id is not None:
        q = q.where(GeneralMeeting.building_id == building_id)
    if status is not None:
        q = q.where(_effective_status_case() == status)
    result = await db.execute(q)
    return result.scalar_one()


async def get_general_meeting_detail(general_meeting_id: uuid.UUID, db: AsyncSession) -> dict:
    result = await db.execute(
        select(GeneralMeeting, Building.name.label("building_name"))
        .join(Building, GeneralMeeting.building_id == Building.id)
        .where(GeneralMeeting.id == general_meeting_id)
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")

    general_meeting, building_name = row

    # Load motions
    motions_result = await db.execute(
        select(Motion).where(Motion.general_meeting_id == general_meeting_id).order_by(Motion.display_order)
    )
    motions = list(motions_result.scalars().all())

    # Load options for multi-choice motions
    mc_motion_ids = [m.id for m in motions if m.is_multi_choice]
    motion_options_map: dict[uuid.UUID, list] = {}
    if mc_motion_ids:
        opts_result = await db.execute(
            select(MotionOption)
            .where(MotionOption.motion_id.in_(mc_motion_ids))
            .order_by(MotionOption.display_order)
        )
        for opt in opts_result.scalars().all():
            motion_options_map.setdefault(opt.motion_id, []).append(opt)

    # Load lot weights joined with lot to get lot numbers
    weights_result = await db.execute(
        select(GeneralMeetingLotWeight, Lot.lot_number.label("lot_number"))
        .join(Lot, GeneralMeetingLotWeight.lot_id == Lot.id)
        .where(GeneralMeetingLotWeight.general_meeting_id == general_meeting_id)
    )
    weight_rows = weights_result.all()

    # Build per-lot_id entitlement and lot info.
    # Batch-load emails for all lots in a single IN query to avoid O(N) queries (RR3-12).
    lot_entitlement: dict[uuid.UUID, int] = {}
    lot_info: dict[uuid.UUID, dict] = {}  # lot_id -> {lot_number, emails, entitlement}

    for w, lot_num in weight_rows:
        lot_entitlement[w.lot_id] = w.unit_entitlement_snapshot
        # Placeholder; emails filled by batch query below
        lot_info[w.lot_id] = {
            "lot_owner_id": w.lot_id,
            "lot_number": lot_num,
            "emails": [],
            "entitlement": w.unit_entitlement_snapshot,
        }

    # Maps (lot_id, email) -> display_name ("Given Surname" or None)
    lot_owner_email_to_name: dict[tuple, str | None] = {}

    if lot_entitlement:
        # Batch-load emails (with names) via lot_persons JOIN persons — single query (RR3-12)
        batch_emails_result = await db.execute(
            select(
                lot_persons.c.lot_id,
                Person.email,
                Person.given_name,
                Person.surname,
            ).join(
                Person, Person.id == lot_persons.c.person_id
            ).where(
                lot_persons.c.lot_id.in_(list(lot_entitlement.keys()))
            )
        )
        for row in batch_emails_result.all():
            if row[1] and row[0] in lot_info:
                lot_info[row[0]]["emails"].append(row[1])
            if row[1]:
                name_parts = [p for p in [row[2], row[3]] if p]
                lot_owner_email_to_name[(row[0], row[1])] = " ".join(name_parts).strip() or None

        # Populate proxy names from lot_proxies JOIN persons.
        proxy_rows = await db.execute(
            select(
                LotProxy.lot_id,
                Person.email,
                Person.given_name,
                Person.surname,
            ).join(
                Person, Person.id == LotProxy.person_id
            ).where(LotProxy.lot_id.in_(list(lot_entitlement.keys())))
        )
        for row in proxy_rows.all():
            name_parts = [p for p in [row[2], row[3]] if p]
            lot_owner_email_to_name[(row[0], row[1])] = " ".join(name_parts).strip() or None

    # Track whether the weight snapshot exists; SQL aggregation entitlement sums are only
    # accurate when the snapshot rows exist (they JOIN on GeneralMeetingLotWeight).
    has_weight_snapshot = bool(lot_entitlement)

    # Fallback: if snapshot is empty
    if not lot_entitlement:
        current_result = await db.execute(
            select(Lot).where(Lot.building_id == general_meeting.building_id)
        )
        fallback_owners = list(current_result.scalars().all())
        for lo in fallback_owners:
            lot_entitlement[lo.id] = lo.unit_entitlement
            lot_info[lo.id] = {
                "lot_owner_id": lo.id,
                "lot_number": lo.lot_number,
                "emails": [],
                "entitlement": lo.unit_entitlement,
            }
        if fallback_owners:
            fallback_lot_ids = [lo.id for lo in fallback_owners]
            fallback_emails_result = await db.execute(
                select(
                    lot_persons.c.lot_id,
                    Person.email,
                    Person.given_name,
                    Person.surname,
                ).join(
                    Person, Person.id == lot_persons.c.person_id
                ).where(
                    lot_persons.c.lot_id.in_(fallback_lot_ids)
                )
            )
            for row in fallback_emails_result.all():
                if row[1] and row[0] in lot_info:
                    lot_info[row[0]]["emails"].append(row[1])
                if row[1]:
                    name_parts = [p for p in [row[2], row[3]] if p]
                    lot_owner_email_to_name[(row[0], row[1])] = " ".join(name_parts).strip() or None

            # Populate proxy names from lot_proxies for fallback path.
            fallback_proxy_rows = await db.execute(
                select(
                    LotProxy.lot_id,
                    Person.email,
                    Person.given_name,
                    Person.surname,
                ).join(
                    Person, Person.id == LotProxy.person_id
                ).where(LotProxy.lot_id.in_(fallback_lot_ids))
            )
            for row in fallback_proxy_rows.all():
                name_parts = [p for p in [row[2], row[3]] if p]
                lot_owner_email_to_name[(row[0], row[1])] = " ".join(name_parts).strip() or None

    eligible_lot_owner_ids: set[uuid.UUID] = set(lot_entitlement.keys())
    total_eligible_voters = len(eligible_lot_owner_ids)

    # Load ballot submissions — separate actual votes (is_absent=False) from absent records
    submissions_result = await db.execute(
        select(BallotSubmission).where(BallotSubmission.general_meeting_id == general_meeting_id)
    )
    submissions = list(submissions_result.scalars().all())
    # Only real votes count toward submitted_lot_owner_ids and total_submitted
    voted_submissions = [s for s in submissions if not s.is_absent]
    submitted_lot_owner_ids: set[uuid.UUID] = {s.lot_owner_id for s in voted_submissions}
    total_submitted = len(submitted_lot_owner_ids)
    # Absent submissions keyed by lot_owner_id for email lookup
    absent_submissions: dict[uuid.UUID, BallotSubmission] = {
        s.lot_owner_id: s for s in submissions if s.is_absent
    }

    # SQL GROUP BY aggregation: compute tally counts and entitlement sums directly in the DB
    # instead of loading every Vote row into Python (eliminates O(V×M) in-memory scan).
    tally_agg_result = await db.execute(
        select(
            Vote.motion_id,
            Vote.choice,
            Vote.motion_option_id,
            func.count().label("voter_count"),
            func.coalesce(
                func.sum(GeneralMeetingLotWeight.unit_entitlement_snapshot), 0
            ).label("entitlement_sum"),
        )
        .join(
            GeneralMeetingLotWeight,
            (GeneralMeetingLotWeight.lot_id == Vote.lot_owner_id)
            & (GeneralMeetingLotWeight.general_meeting_id == Vote.general_meeting_id),
            isouter=True,
        )
        .where(
            Vote.general_meeting_id == general_meeting_id,
            Vote.status == VoteStatus.submitted,
        )
        .group_by(Vote.motion_id, Vote.choice, Vote.motion_option_id)
    )
    # tally_map: (motion_id, choice_str_or_None, option_id_or_None) -> (voter_count, entitlement_sum)
    tally_map: dict[tuple[uuid.UUID, str | None, uuid.UUID | None], tuple[int, int]] = {}
    for row in tally_agg_result.all():
        choice_str: str | None = row.choice.value if row.choice and hasattr(row.choice, "value") else row.choice
        tally_map[(row.motion_id, choice_str, row.motion_option_id)] = (row.voter_count, int(row.entitlement_sum))

    # Lightweight projection: load only the columns needed to build per-lot voter_lists
    # and compute the set of lot_owner_ids per motion (for implicit-abstained calculation).
    vote_proj_result = await db.execute(
        select(
            Vote.lot_owner_id,
            Vote.motion_id,
            Vote.choice,
            Vote.motion_option_id,
            Vote.voter_email,
        ).where(
            Vote.general_meeting_id == general_meeting_id,
            Vote.status == VoteStatus.submitted,
        )
    )
    vote_projections = vote_proj_result.all()

    def _tally(lot_owner_ids: set[uuid.UUID]) -> dict:
        return {
            "voter_count": len(lot_owner_ids),
            "entitlement_sum": sum(lot_entitlement.get(lid, 0) for lid in lot_owner_ids),
        }

    # Build lot_owner_id -> voter_email, proxy_email, ballot_hash, and submitted_by_admin from voted submissions
    lot_owner_to_email: dict[uuid.UUID, str] = {sub.lot_owner_id: sub.voter_email for sub in voted_submissions}
    lot_owner_to_proxy_email: dict[uuid.UUID, str | None] = {sub.lot_owner_id: sub.proxy_email for sub in voted_submissions}
    # US-VIL-03: expose ballot_hash for admin audit
    lot_owner_to_ballot_hash: dict[uuid.UUID, str | None] = {
        sub.lot_owner_id: sub.ballot_hash for sub in voted_submissions
    }
    # US-AVE-03: expose submitted_by_admin flag
    lot_owner_to_submitted_by_admin: dict[uuid.UUID, bool] = {
        sub.lot_owner_id: sub.submitted_by_admin for sub in voted_submissions
    }
    # RR4-27: expose submitted_by_admin_username for audit trail
    lot_owner_to_submitted_by_admin_username: dict[uuid.UUID, str | None] = {
        sub.lot_owner_id: sub.submitted_by_admin_username for sub in voted_submissions
    }
    # Expose submitted_at timestamp for per-motion CSV download
    lot_owner_to_submitted_at: dict[uuid.UUID, datetime] = {
        sub.lot_owner_id: sub.submitted_at for sub in voted_submissions
    }
    # Per-motion voter email: keyed on (lot_owner_id, motion_id) so that when a lot has
    # multiple voters (co-owners, proxy re-entry) each motion shows the email of the person
    # who actually submitted that specific Vote row, not the last BallotSubmission author.
    vote_voter_email_map: dict[tuple[uuid.UUID, uuid.UUID], str] = {}

    def _lots(lot_owner_ids: set[uuid.UUID], category: str, motion_id: uuid.UUID | None = None) -> list[dict]:
        result_list: list[dict] = []
        for lid in lot_owner_ids:
            info = lot_info.get(lid)
            if info:
                if category == "absent":
                    # For absent lots, read the snapshot recorded on close
                    absent_sub = absent_submissions.get(lid)
                    voter_email = absent_sub.voter_email if absent_sub else ""
                    proxy_email_val = None  # absent rows don't expose proxy separately in the list
                    ballot_hash_val = None  # absent lots have no ballot hash
                    submitted_by_admin_val = False
                    submitted_by_admin_username_val = None
                    submitted_at_val = absent_sub.submitted_at if absent_sub else None
                else:
                    # For voted categories, prefer the per-motion voter_email stamped on
                    # the Vote row (correct even when co-owners submit different motions),
                    # falling back to BallotSubmission.voter_email for absent/no-Vote cases.
                    if motion_id is not None:
                        voter_email = vote_voter_email_map.get((lid, motion_id), lot_owner_to_email.get(lid, ""))
                    else:  # pragma: no cover
                        voter_email = lot_owner_to_email.get(lid, "")
                    proxy_email_val = lot_owner_to_proxy_email.get(lid)
                    ballot_hash_val = lot_owner_to_ballot_hash.get(lid)
                    submitted_by_admin_val = lot_owner_to_submitted_by_admin.get(lid, False)
                    submitted_by_admin_username_val = lot_owner_to_submitted_by_admin_username.get(lid)
                    submitted_at_val = lot_owner_to_submitted_at.get(lid)
                voter_name = lot_owner_email_to_name.get((lid, voter_email))
                result_list.append({
                    "voter_email": voter_email,
                    "voter_name": voter_name,
                    "lot_number": info["lot_number"],
                    "entitlement": info["entitlement"],
                    "proxy_email": proxy_email_val,
                    "ballot_hash": ballot_hash_val,
                    "submitted_by_admin": submitted_by_admin_val,
                    "submitted_by_admin_username": submitted_by_admin_username_val,
                    "submitted_at": submitted_at_val,
                })
        return result_list

    is_closed = get_effective_status(general_meeting) == GeneralMeetingStatus.closed
    absent_ids_global: set[uuid.UUID] = set(absent_submissions.keys()) if is_closed else set()

    # Build per-motion indexes from the lightweight vote projection (for voter_lists and
    # implicit-abstained computation). Replaces the O(V×M) scan over full Vote objects.
    # votes_by_motion: motion_id -> list of (lot_owner_id, choice_str, option_id) tuples
    VoteRow = tuple[uuid.UUID, str | None, uuid.UUID | None]
    votes_by_motion: dict[uuid.UUID, list[VoteRow]] = {}
    for vp in vote_projections:
        if vp.lot_owner_id in submitted_lot_owner_ids:
            choice_str = vp.choice.value if vp.choice and hasattr(vp.choice, "value") else vp.choice
            votes_by_motion.setdefault(vp.motion_id, []).append(
                (vp.lot_owner_id, choice_str, vp.motion_option_id)
            )
        # Populate per-motion voter email map regardless of submission status so that
        # any Vote row's own voter_email is captured for the voter_list display.
        if vp.voter_email:
            vote_voter_email_map[(vp.lot_owner_id, vp.motion_id)] = vp.voter_email

    motion_details = []
    for motion in motions:
        motion_type_str = motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type
        motion_opts = motion_options_map.get(motion.id, [])

        if motion.is_multi_choice:
            # Multi-choice: per-option tallying using lightweight vote projection tuples
            # (lot_owner_id, choice_str, motion_option_id)
            motion_vote_rows = votes_by_motion.get(motion.id, [])

            # not_eligible lots
            not_eligible_ids: set[uuid.UUID] = {
                lot_id for lot_id, choice_s, _ in motion_vote_rows
                if choice_s == "not_eligible"
            }

            # abstained = submitted but no selected and not not_eligible
            # Derived from this motion's own vote rows to avoid incorrectly including
            # lots that simply didn't have a vote row for this motion (e.g. not yet submitted).
            abstained_ids: set[uuid.UUID] = {
                lot_id for lot_id, choice_s, _ in motion_vote_rows
                if choice_s == "abstained"
            }

            # Per-option tally — use SQL aggregation results via tally_map where available,
            # fall back to in-memory projection count (same as before for voter_lists).
            option_tallies = []
            option_for_voter_lists: dict[str, list] = {}
            option_against_voter_lists: dict[str, list] = {}
            option_abstained_voter_lists: dict[str, list] = {}
            for opt in motion_opts:
                opt_for_ids = {
                    lot_id for lot_id, choice_s, opt_id in motion_vote_rows
                    if opt_id == opt.id and choice_s == "selected"
                }
                opt_against_ids = {
                    lot_id for lot_id, choice_s, opt_id in motion_vote_rows
                    if opt_id == opt.id and choice_s == "against"
                }
                opt_abstained_ids = {
                    lot_id for lot_id, choice_s, opt_id in motion_vote_rows
                    if opt_id == opt.id and choice_s == "abstained"
                }
                # Use stored snapshot values when available (post-close), fall back to SQL aggregation
                for_vc, for_es = (int(opt.for_voter_count), int(opt.for_entitlement_sum)) if opt.for_voter_count else tally_map.get((motion.id, "selected", opt.id), (0, 0))  # type: ignore[assignment]
                against_vc, against_es = (int(opt.against_voter_count), int(opt.against_entitlement_sum)) if opt.against_voter_count else tally_map.get((motion.id, "against", opt.id), (0, 0))  # type: ignore[assignment]
                abstained_vc, abstained_es = (int(opt.abstained_voter_count), int(opt.abstained_entitlement_sum)) if opt.abstained_voter_count else tally_map.get((motion.id, "abstained", opt.id), (0, 0))  # type: ignore[assignment]
                option_tallies.append({
                    "option_id": opt.id,
                    "option_text": opt.text,
                    "display_order": opt.display_order,
                    "for_voter_count": for_vc,
                    "for_entitlement_sum": for_es,
                    "against_voter_count": against_vc,
                    "against_entitlement_sum": against_es,
                    "abstained_voter_count": abstained_vc,
                    "abstained_entitlement_sum": abstained_es,
                    # Backward-compatible aliases
                    "voter_count": for_vc,
                    "entitlement_sum": for_es,
                    "outcome": opt.outcome,
                })
                option_for_voter_lists[str(opt.id)] = _lots(opt_for_ids, "selected", motion.id)
                option_against_voter_lists[str(opt.id)] = _lots(opt_against_ids, "against", motion.id)
                option_abstained_voter_lists[str(opt.id)] = _lots(opt_abstained_ids, "abstained", motion.id)

            motion_details.append({
                "id": motion.id,
                "title": motion.title,
                "description": motion.description,
                "display_order": motion.display_order,
                "motion_number": motion.motion_number,
                "motion_type": motion_type_str,
                "is_multi_choice": motion.is_multi_choice,
                "is_visible": motion.is_visible,
                "option_limit": motion.option_limit,
                "voting_closed_at": motion.voting_closed_at,
                "options": [
                    {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
                    for opt in motion_opts
                ],
                "tally": {
                    "yes": {"voter_count": 0, "entitlement_sum": 0},
                    "no": {"voter_count": 0, "entitlement_sum": 0},
                    "abstained": _tally(abstained_ids),
                    "absent": _tally(absent_ids_global),
                    "not_eligible": _tally(not_eligible_ids),
                    "options": option_tallies,
                },
                "voter_lists": {
                    "yes": [],
                    "no": [],
                    "abstained": _lots(abstained_ids, "abstained", motion.id),
                    "absent": _lots(absent_ids_global, "absent"),
                    "not_eligible": _lots(not_eligible_ids, "not_eligible", motion.id),
                    "options_for": option_for_voter_lists,
                    "options_against": option_against_voter_lists,
                    "options_abstained": option_abstained_voter_lists,
                    # Backward-compatible alias
                    "options": option_for_voter_lists,
                },
            })
        else:
            # General / Special: use SQL aggregation for tally numbers; use lightweight
            # projection tuples (lot_owner_id, choice_str, option_id) for voter_lists.
            motion_vote_rows_standard = votes_by_motion.get(motion.id, [])
            # motion_votes_map: lot_owner_id -> choice_str (for voter_lists and implicit-abstained)
            motion_votes_map: dict[uuid.UUID, str] = {
                lot_id: (choice_s or "abstained")
                for lot_id, choice_s, _ in motion_vote_rows_standard
                if lot_id is not None
            }

            yes_ids: set[uuid.UUID] = set()
            no_ids: set[uuid.UUID] = set()
            abstained_ids: set[uuid.UUID] = set()
            not_eligible_ids: set[uuid.UUID] = set()

            for lot_id in submitted_lot_owner_ids:
                if lot_id not in motion_votes_map:
                    continue  # no real Vote row for this lot+motion — omit entirely
                choice = motion_votes_map[lot_id]
                if choice == "yes":
                    yes_ids.add(lot_id)
                elif choice in ("no", "against"):
                    # RR4-03: VoteChoice.against is semantically equivalent to "no" for
                    # General/Special motions. Map it to the no bucket so tallies are correct.
                    no_ids.add(lot_id)
                elif choice == "not_eligible":
                    not_eligible_ids.add(lot_id)
                else:
                    abstained_ids.add(lot_id)

            # Use SQL aggregation results for tally numbers (voter_count, entitlement_sum).
            # The lot_owner_id sets (yes_ids, no_ids, etc.) are still needed for voter_lists.
            # SQL entitlement sums are accurate only when the weight snapshot exists; fall back
            # to in-memory _tally() when the snapshot is absent (defensive edge case).
            if has_weight_snapshot:
                # "against" maps to "no" bucket — combine both for the tally lookup
                no_vc = (
                    tally_map.get((motion.id, "no", None), (0, 0))[0]
                    + tally_map.get((motion.id, "against", None), (0, 0))[0]
                )
                no_es = (
                    tally_map.get((motion.id, "no", None), (0, 0))[1]
                    + tally_map.get((motion.id, "against", None), (0, 0))[1]
                )
                yes_tally = {"voter_count": tally_map.get((motion.id, "yes", None), (0, 0))[0], "entitlement_sum": tally_map.get((motion.id, "yes", None), (0, 0))[1]}
                no_tally = {"voter_count": no_vc, "entitlement_sum": no_es}
                not_eligible_tally = {"voter_count": tally_map.get((motion.id, "not_eligible", None), (0, 0))[0], "entitlement_sum": tally_map.get((motion.id, "not_eligible", None), (0, 0))[1]}
            else:
                yes_tally = _tally(yes_ids)
                no_tally = _tally(no_ids)
                not_eligible_tally = _tally(not_eligible_ids)
            # abstained_ids contains only real Vote rows with choice=abstained.
            # Lots that submitted a ballot but have no Vote row for this motion are omitted
            # entirely (they are not inferred as abstained).  _tally() is always used because
            # SQL aggregation does not cover the abstained bucket (no separate tally_map entry).
            motion_details.append(
                {
                    "id": motion.id,
                    "title": motion.title,
                    "description": motion.description,
                    "display_order": motion.display_order,
                    "motion_number": motion.motion_number,
                    "motion_type": motion_type_str,
                    "is_multi_choice": motion.is_multi_choice,
                    "is_visible": motion.is_visible,
                    "option_limit": None,
                    "voting_closed_at": motion.voting_closed_at,
                    "options": [],
                    "tally": {
                        "yes": yes_tally,
                        "no": no_tally,
                        "abstained": _tally(abstained_ids),
                        "absent": _tally(absent_ids_global),
                        "not_eligible": not_eligible_tally,
                        "options": [],
                    },
                    "voter_lists": {
                        "yes": _lots(yes_ids, "yes", motion.id),
                        "no": _lots(no_ids, "no", motion.id),
                        "abstained": _lots(abstained_ids, "abstained", motion.id),
                        "absent": _lots(absent_ids_global, "absent"),
                        "not_eligible": _lots(not_eligible_ids, "not_eligible", motion.id),
                        "options": {},
                    },
                }
            )

    total_entitlement = sum(lot_entitlement.values())

    # Load email delivery record for this meeting
    email_delivery_result = await db.execute(
        select(EmailDelivery).where(EmailDelivery.general_meeting_id == general_meeting_id)
    )
    email_delivery_obj = email_delivery_result.scalar_one_or_none()

    effective = get_effective_status(general_meeting)
    return {
        "id": general_meeting.id,
        "building_id": general_meeting.building_id,
        "building_name": building_name,
        "title": general_meeting.title,
        "status": effective.value if hasattr(effective, "value") else effective,
        "meeting_at": general_meeting.meeting_at,
        "voting_closes_at": general_meeting.voting_closes_at,
        "closed_at": general_meeting.closed_at,
        "total_eligible_voters": total_eligible_voters,
        "total_submitted": total_submitted,
        "total_entitlement": total_entitlement,
        "motions": motion_details,
        "email_delivery": {
            "status": email_delivery_obj.status.value if hasattr(email_delivery_obj.status, "value") else email_delivery_obj.status,
            "last_error": email_delivery_obj.last_error,
        } if email_delivery_obj else None,
    }


async def toggle_motion_visibility(
    motion_id: uuid.UUID,
    is_visible: bool,
    db: AsyncSession,
) -> dict:
    """Toggle the visibility of a motion. Returns updated motion detail dict.

    Raises 404 if motion not found.
    Raises 409 if:
      - The meeting is closed
      - is_visible=False and the motion already has submitted Vote records
    """
    # Fetch motion
    result = await db.execute(select(Motion).where(Motion.id == motion_id))
    motion = result.scalar_one_or_none()
    if motion is None:
        raise HTTPException(status_code=404, detail="Motion not found")

    # Fetch meeting
    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == motion.general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:  # pragma: no cover
        raise HTTPException(status_code=404, detail="General Meeting not found")

    effective = get_effective_status(meeting)
    if effective == GeneralMeetingStatus.closed:
        raise HTTPException(status_code=409, detail="Cannot change visibility on a closed meeting")

    # Block hiding motions that already have votes or are individually closed
    if not is_visible:
        if motion.voting_closed_at is not None:
            raise HTTPException(
                status_code=409,
                detail="Cannot hide a closed motion",
            )
        vote_count_result = await db.execute(
            select(func.count()).select_from(Vote).where(
                Vote.motion_id == motion_id,
                Vote.status == VoteStatus.submitted,
            )
        )
        vote_count = vote_count_result.scalar_one()
        if vote_count > 0:
            raise HTTPException(
                status_code=409,
                detail="Cannot hide a motion that has received votes",
            )

    motion.is_visible = is_visible
    await db.flush()
    await db.commit()

    # Load options for this motion
    opts_result = await db.execute(
        select(MotionOption)
        .where(MotionOption.motion_id == motion.id)
        .order_by(MotionOption.display_order)
    )
    motion_options = list(opts_result.scalars().all())

    return {
        "id": motion.id,
        "title": motion.title,
        "description": motion.description,
        "display_order": motion.display_order,
        "motion_number": motion.motion_number,
        "motion_type": motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type,
        "is_multi_choice": motion.is_multi_choice,
        "is_visible": motion.is_visible,
        "option_limit": motion.option_limit,
        "voting_closed_at": motion.voting_closed_at,
        "options": [
            {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
            for opt in motion_options
        ],
        "tally": {
            "yes": {"voter_count": 0, "entitlement_sum": 0},
            "no": {"voter_count": 0, "entitlement_sum": 0},
            "abstained": {"voter_count": 0, "entitlement_sum": 0},
            "absent": {"voter_count": 0, "entitlement_sum": 0},
            "not_eligible": {"voter_count": 0, "entitlement_sum": 0},
            "options": [],
        },
        "voter_lists": {
            "yes": [],
            "no": [],
            "abstained": [],
            "absent": [],
            "not_eligible": [],
            "options": {},
        },
    }


async def close_motion(
    motion_id: uuid.UUID,
    db: AsyncSession,
) -> dict:
    """Close voting for a single motion. Returns updated motion detail dict.

    Raises 404 if motion not found.
    Raises 409 if:
      - motion is not visible (hidden)
      - voting_closed_at IS NOT NULL (already closed)
      - meeting effective_status != "open"
    """
    # RR4-21: Use SELECT FOR UPDATE so concurrent close-motion requests serialize
    # and exactly one request writes voting_closed_at.  Without the lock, two
    # concurrent requests could both read voting_closed_at=None and both proceed
    # to write a timestamp, resulting in a non-deterministic last-write-wins outcome.
    result = await db.execute(
        select(Motion).where(Motion.id == motion_id).with_for_update()
    )
    motion = result.scalar_one_or_none()
    if motion is None:
        raise HTTPException(status_code=404, detail="Motion not found")

    # Fetch meeting
    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == motion.general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:  # pragma: no cover
        raise HTTPException(status_code=404, detail="General Meeting not found")

    if not motion.is_visible:
        raise HTTPException(status_code=409, detail="Cannot close a hidden motion")

    if motion.voting_closed_at is not None:
        raise HTTPException(status_code=409, detail="Motion voting is already closed")

    effective = get_effective_status(meeting)
    if effective != GeneralMeetingStatus.open:
        raise HTTPException(status_code=409, detail="Cannot close motion on a meeting that is not open")

    # RR4-33: ensure voting_closed_at > meeting_at (defensive — meeting must already
    # have started for effective_status to be "open", but validate explicitly).
    close_time = datetime.now(UTC)
    if meeting.meeting_at is not None:
        starts_at = meeting.meeting_at
        # Normalise to UTC; DB columns use timezone=True so tzinfo is always set.
        starts_at = starts_at.astimezone(UTC)
        if close_time <= starts_at:
            raise HTTPException(
                status_code=422,
                detail="Voting close time must be after meeting start time",
            )

    motion.voting_closed_at = close_time
    await db.flush()
    await db.commit()

    # Load options for this motion
    opts_result = await db.execute(
        select(MotionOption)
        .where(MotionOption.motion_id == motion.id)
        .order_by(MotionOption.display_order)
    )
    motion_options = list(opts_result.scalars().all())

    return {
        "id": motion.id,
        "title": motion.title,
        "description": motion.description,
        "display_order": motion.display_order,
        "motion_number": motion.motion_number,
        "motion_type": motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type,
        "is_multi_choice": motion.is_multi_choice,
        "is_visible": motion.is_visible,
        "option_limit": motion.option_limit,
        "voting_closed_at": motion.voting_closed_at,
        "options": [
            {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
            for opt in motion_options
        ],
        "tally": {
            "yes": {"voter_count": 0, "entitlement_sum": 0},
            "no": {"voter_count": 0, "entitlement_sum": 0},
            "abstained": {"voter_count": 0, "entitlement_sum": 0},
            "absent": {"voter_count": 0, "entitlement_sum": 0},
            "not_eligible": {"voter_count": 0, "entitlement_sum": 0},
            "options": [],
        },
        "voter_lists": {
            "yes": [],
            "no": [],
            "abstained": [],
            "absent": [],
            "not_eligible": [],
            "options": {},
        },
    }


async def add_motion_to_meeting(
    general_meeting_id: uuid.UUID,
    data: MotionAddRequest,
    db: AsyncSession,
) -> dict:
    """Add a new motion to an existing General Meeting.

    Assigns display_order = MAX(existing) + 1 (starts at 0 if no motions).
    New motion is always created with is_visible=False.

    Raises 404 if meeting not found.
    Raises 409 if meeting is closed.
    """
    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")

    effective = get_effective_status(meeting)
    if effective == GeneralMeetingStatus.closed:
        raise HTTPException(status_code=409, detail="Cannot add a motion to a closed meeting")

    max_result = await db.execute(
        select(func.max(Motion.display_order)).where(
            Motion.general_meeting_id == general_meeting_id
        )
    )
    max_index = max_result.scalar_one_or_none()
    # display_order starts at 1 for the first motion (not 0)
    next_index = (max_index + 1) if max_index is not None else 1

    # Auto-assign motion_number from next display_order when the field is absent or blank.
    # The frontend may send null or "" when the user leaves the field empty — treat both
    # as "no explicit number supplied" and fall back to auto-assign.
    explicit_number = data.motion_number.strip() if data.motion_number is not None else ""

    if not explicit_number:
        # RR2-02: auto-assign from max(existing numeric motion_numbers) + 1 to avoid
        # conflicts with manually-set motion numbers that match display_order values.
        from sqlalchemy import Integer, cast as sa_cast
        numeric_max_result = await db.execute(
            select(func.max(func.cast(Motion.motion_number, Integer)))
            .where(Motion.general_meeting_id == general_meeting_id)
            .where(Motion.motion_number.regexp_match(r"^\d+$"))
        )
        max_numeric = numeric_max_result.scalar_one_or_none()
        if max_numeric is not None:
            assigned_motion_number = str(max_numeric + 1)
        else:
            # No existing numeric motion numbers — fall back to display_order
            assigned_motion_number = str(next_index)
    else:
        assigned_motion_number = explicit_number

    motion = Motion(
        general_meeting_id=general_meeting_id,
        title=data.title.strip(),
        description=_sanitise_description(data.description),
        display_order=next_index,
        motion_number=assigned_motion_number,
        motion_type=data.motion_type,
        is_multi_choice=data.is_multi_choice,
        option_limit=data.option_limit if data.is_multi_choice else None,
        is_visible=False,
    )
    db.add(motion)
    try:
        await db.flush()
    except IntegrityError as exc:
        await db.rollback()
        if "uq_motions_general_meeting_motion_number" in str(exc.orig):
            raise HTTPException(
                status_code=409,
                detail="A motion with this number already exists in this meeting",
            ) from exc
        raise  # pragma: no cover — re-raise unexpected integrity errors

    # Create motion options for multi-choice motions
    created_options = []
    if data.is_multi_choice:
        for opt in data.options:
            new_opt = MotionOption(
                motion_id=motion.id,
                text=_sanitise_option_text(opt.text),
                display_order=opt.display_order,
            )
            db.add(new_opt)
            created_options.append(new_opt)
        await db.flush()

    await db.commit()
    await db.refresh(motion)
    for opt in created_options:
        await db.refresh(opt)

    return {
        "id": motion.id,
        "title": motion.title,
        "description": motion.description,
        "display_order": motion.display_order,
        "motion_number": motion.motion_number,
        "motion_type": motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type,
        "is_multi_choice": motion.is_multi_choice,
        "is_visible": motion.is_visible,
        "option_limit": motion.option_limit,
        "options": [
            {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
            for opt in created_options
        ],
    }


async def update_motion(
    motion_id: uuid.UUID,
    data: MotionUpdateRequest,
    db: AsyncSession,
) -> dict:
    """Update title, description, or motion_type of a hidden motion.

    Raises 404 if motion not found.
    Raises 409 if motion is visible or meeting is closed.
    """
    result = await db.execute(select(Motion).where(Motion.id == motion_id))
    motion = result.scalar_one_or_none()
    if motion is None:
        raise HTTPException(status_code=404, detail="Motion not found")

    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == motion.general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:  # pragma: no cover
        raise HTTPException(status_code=404, detail="General Meeting not found")

    if get_effective_status(meeting) == GeneralMeetingStatus.closed:
        raise HTTPException(status_code=409, detail="Cannot edit a motion on a closed meeting")

    if motion.is_visible:
        raise HTTPException(
            status_code=409, detail="Cannot edit a visible motion. Hide it first."
        )

    if data.title is not None:
        motion.title = data.title.strip()
    if data.description is not None:
        motion.description = _sanitise_description(data.description)
    if data.motion_type is not None:
        motion.motion_type = data.motion_type
    if data.is_multi_choice is not None:
        # When changing away from multi_choice, clear options and option_limit
        if not data.is_multi_choice and motion.is_multi_choice:
            await db.execute(
                delete(MotionOption).where(MotionOption.motion_id == motion.id)
            )
            motion.option_limit = None
        motion.is_multi_choice = data.is_multi_choice
    if data.motion_number is not None:
        stripped = data.motion_number.strip()
        motion.motion_number = stripped if stripped else None
    if data.option_limit is not None:
        motion.option_limit = data.option_limit
    if data.options is not None:
        # Replace all existing options atomically
        await db.execute(
            delete(MotionOption).where(MotionOption.motion_id == motion.id)
        )
        await db.flush()
        for opt in data.options:
            db.add(MotionOption(
                motion_id=motion.id,
                text=_sanitise_option_text(opt.text),
                display_order=opt.display_order,
            ))

    await db.flush()
    await db.commit()
    await db.refresh(motion)

    # Load updated options
    opts_result = await db.execute(
        select(MotionOption)
        .where(MotionOption.motion_id == motion.id)
        .order_by(MotionOption.display_order)
    )
    updated_options = list(opts_result.scalars().all())

    return {
        "id": motion.id,
        "title": motion.title,
        "description": motion.description,
        "display_order": motion.display_order,
        "motion_number": motion.motion_number,
        "motion_type": motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type,
        "is_multi_choice": motion.is_multi_choice,
        "is_visible": motion.is_visible,
        "option_limit": motion.option_limit,
        "options": [
            {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
            for opt in updated_options
        ],
    }


async def delete_motion(
    motion_id: uuid.UUID,
    db: AsyncSession,
) -> None:
    """Delete a hidden motion permanently.

    Raises 404 if motion not found.
    Raises 409 if motion is visible or meeting is closed.
    """
    result = await db.execute(select(Motion).where(Motion.id == motion_id))
    motion = result.scalar_one_or_none()
    if motion is None:
        raise HTTPException(status_code=404, detail="Motion not found")

    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == motion.general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:  # pragma: no cover
        raise HTTPException(status_code=404, detail="General Meeting not found")

    if get_effective_status(meeting) == GeneralMeetingStatus.closed:
        raise HTTPException(status_code=409, detail="Cannot delete a motion on a closed meeting")

    if motion.is_visible:
        raise HTTPException(
            status_code=409, detail="Cannot delete a visible motion. Hide it first."
        )

    await db.delete(motion)
    await db.flush()
    await db.commit()


async def delete_motion_option(
    motion_id: uuid.UUID,
    option_id: uuid.UUID,
    db: AsyncSession,
) -> None:
    """Delete a single option from a multi-choice motion.

    Raises 404 if the motion or option does not exist.
    Raises 409 if any submitted Vote references this option (RESTRICT FK semantics).
    """
    from sqlalchemy.exc import IntegrityError

    # Verify option exists and belongs to the motion
    opt_result = await db.execute(
        select(MotionOption).where(
            MotionOption.id == option_id,
            MotionOption.motion_id == motion_id,
        )
    )
    option = opt_result.scalar_one_or_none()
    if option is None:
        raise HTTPException(status_code=404, detail="Motion option not found")

    # Check whether any submitted votes reference this option
    vote_count_result = await db.execute(
        select(func.count()).select_from(Vote).where(
            Vote.motion_option_id == option_id,
            Vote.status == VoteStatus.submitted,
        )
    )
    vote_count = vote_count_result.scalar_one()
    if vote_count > 0:
        raise HTTPException(
            status_code=409,
            detail="Cannot delete an option that has submitted votes",
        )

    await db.delete(option)
    try:
        await db.flush()
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Cannot delete an option that has submitted votes",
        )


async def start_general_meeting(general_meeting_id: uuid.UUID, db: AsyncSession) -> GeneralMeeting:
    """Manually start a pending General Meeting, setting status=open and meeting_at=now."""
    result = await db.execute(select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    general_meeting = result.scalar_one_or_none()
    if general_meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")
    if get_effective_status(general_meeting) != GeneralMeetingStatus.pending:
        raise HTTPException(status_code=409, detail="General Meeting is not in pending status")
    general_meeting.status = GeneralMeetingStatus.open
    general_meeting.meeting_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(general_meeting)
    return general_meeting


async def close_general_meeting(general_meeting_id: uuid.UUID, db: AsyncSession, background_tasks=None) -> GeneralMeeting:
    result = await db.execute(select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    general_meeting = result.scalar_one_or_none()
    if general_meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")
    if general_meeting.status == GeneralMeetingStatus.closed:
        raise HTTPException(status_code=409, detail="General Meeting is already closed")

    logger.info("meeting_close_initiated", agm_id=str(general_meeting_id))

    # Close the General Meeting
    now = datetime.now(UTC)
    general_meeting.status = GeneralMeetingStatus.closed
    general_meeting.closed_at = now

    # Update voting_closes_at to now if closing before the scheduled close time (US-PS05).
    # Only update when voting_closes_at is in the future AND meeting_at is in the past
    # (i.e., the meeting has started). If meeting_at is still in the future, preserving
    # voting_closes_at avoids violating the CHECK constraint (voting_closes_at > meeting_at).
    meeting_at_aware = general_meeting.meeting_at
    if meeting_at_aware is not None and meeting_at_aware.tzinfo is None:  # pragma: no cover — DB always returns tz-aware
        meeting_at_aware = meeting_at_aware.replace(tzinfo=UTC)  # pragma: no cover
    if (
        general_meeting.voting_closes_at is not None
        and general_meeting.voting_closes_at > now
        and (meeting_at_aware is None or meeting_at_aware <= now)
    ):
        general_meeting.voting_closes_at = now

    # Close all motions that have not yet been individually closed
    open_motions_result = await db.execute(
        select(Motion).where(
            Motion.general_meeting_id == general_meeting_id,
            Motion.voting_closed_at.is_(None),
        )
    )
    for open_motion in open_motions_result.scalars().all():
        open_motion.voting_closed_at = general_meeting.closed_at

    # Delete draft votes
    await db.execute(
        delete(Vote).where(
            Vote.general_meeting_id == general_meeting_id,
            Vote.status == VoteStatus.draft,
        )
    )

    # Create absent BallotSubmission records for lots that did not vote.
    # These capture contact emails as a snapshot at close time so the export has
    # the right emails even if lot owner emails are later changed.
    weights_result = await db.execute(
        select(GeneralMeetingLotWeight.lot_id).where(
            GeneralMeetingLotWeight.general_meeting_id == general_meeting_id
        )
    )
    eligible_lot_owner_ids: set[uuid.UUID] = {row[0] for row in weights_result.all()}

    if eligible_lot_owner_ids:
        subs_result = await db.execute(
            select(BallotSubmission.lot_owner_id).where(
                BallotSubmission.general_meeting_id == general_meeting_id,
                BallotSubmission.is_absent == False,  # noqa: E712
            )
        )
        voted_lot_owner_ids: set[uuid.UUID] = {row[0] for row in subs_result.all()}
        absent_lot_owner_ids = eligible_lot_owner_ids - voted_lot_owner_ids

        if absent_lot_owner_ids:
            # Batch-load owner emails for all absent lots via lot_persons JOIN persons
            emails_result = await db.execute(
                select(lot_persons.c.lot_id, Person.email).join(
                    Person, Person.id == lot_persons.c.person_id
                ).where(
                    lot_persons.c.lot_id.in_(absent_lot_owner_ids)
                )
            )
            emails_by_owner: dict[uuid.UUID, list[str]] = {}
            for row in emails_result.all():
                if row[1]:
                    emails_by_owner.setdefault(row[0], []).append(row[1])

            # Batch-load proxy emails for all absent lots via LotProxy JOIN persons
            proxies_result = await db.execute(
                select(LotProxy.lot_id, Person.email).join(
                    Person, Person.id == LotProxy.person_id
                ).where(
                    LotProxy.lot_id.in_(absent_lot_owner_ids)
                )
            )
            proxy_by_owner: dict[uuid.UUID, str] = {
                row[0]: row[1] for row in proxies_result.all() if row[1]
            }

            # Pre-load lot_number for absent lots to include in the warning (RR5-12)
            absent_lot_owners_result = await db.execute(
                select(Lot.id, Lot.lot_number).where(Lot.id.in_(absent_lot_owner_ids))
            )
            absent_lot_number_map: dict[uuid.UUID, str] = {row[0]: row[1] for row in absent_lot_owners_result.all()}

            for lid in absent_lot_owner_ids:
                owner_emails = emails_by_owner.get(lid, [])
                proxy_email_val = proxy_by_owner.get(lid)
                # Deduplicated union: owner emails + proxy email if not already included
                contact_emails = list(owner_emails)
                if proxy_email_val and proxy_email_val not in contact_emails:
                    contact_emails.append(proxy_email_val)
                # RR5-12: Warn when a lot has no contact emails at all
                if not contact_emails:
                    logger.warning(
                        "lot_no_contact_email",
                        lot_id=str(lid),
                        lot_number=absent_lot_number_map.get(lid, "unknown"),
                    )
                voter_email_str = ", ".join(contact_emails)
                db.add(BallotSubmission(
                    general_meeting_id=general_meeting_id,
                    lot_owner_id=lid,
                    voter_email=voter_email_str,
                    proxy_email=proxy_email_val,
                    is_absent=True,
                ))

    # Create EmailDelivery record
    email_delivery = EmailDelivery(
        general_meeting_id=general_meeting_id,
        status=EmailDeliveryStatus.pending,
        total_attempts=0,
    )
    db.add(email_delivery)

    # RR4-04: Commit ballot data BEFORE computing multi-choice outcomes so that
    # compute_multi_choice_outcomes reads fully-committed Vote rows, not
    # in-flight rows that may be rolled back.
    await db.commit()
    await db.refresh(general_meeting)

    # Compute multi-choice pass/fail outcomes (Slice 4) — runs after commit
    # so it reads the committed ballot data.
    await compute_multi_choice_outcomes(general_meeting_id, db)

    absent_count = len(absent_lot_owner_ids) if eligible_lot_owner_ids else 0
    lot_count = len(eligible_lot_owner_ids) if eligible_lot_owner_ids else 0
    logger.info(
        "meeting_closed",
        agm_id=str(general_meeting_id),
        lot_count=lot_count,
        absent_count=absent_count,
        email_triggered=True,
    )

    return general_meeting


async def compute_multi_choice_outcomes(general_meeting_id: uuid.UUID, db: AsyncSession) -> None:
    """Compute and persist pass/fail/tie outcomes for all multi-choice motions in the meeting.

    Algorithm (per motion):
    1. total_building_entitlement = sum of all AGMLotWeight.unit_entitlement_snapshot for the meeting.
    2. For each option:
       - for_entitlement_sum = sum of UOE for lots with Vote.choice = "selected" for this option.
       - against_entitlement_sum = sum of UOE for lots with Vote.choice = "against" for this option.
    3. Mark option as "fail" if against_entitlement_sum / total_building_entitlement > 0.50.
    4. Among remaining (non-failed) options, rank by for_entitlement_sum descending.
    5. Top option_limit ranked options: check for ties at the boundary.
       - If position option_limit and option_limit+1 have the same for_entitlement_sum,
         mark both (and all others at the boundary) as "tie".
       - Positions 1..option_limit without a tie boundary: mark "pass".
       - Positions after option_limit without tie: mark "fail".
    6. Persist outcome on each MotionOption row.
    """
    # Load motions for this meeting
    motions_result = await db.execute(
        select(Motion).where(
            Motion.general_meeting_id == general_meeting_id,
            Motion.is_multi_choice == True,  # noqa: E712
        )
    )
    mc_motions = list(motions_result.scalars().all())
    if not mc_motions:
        return

    # Total building entitlement for this meeting
    weights_result = await db.execute(
        select(func.sum(GeneralMeetingLotWeight.unit_entitlement_snapshot)).where(
            GeneralMeetingLotWeight.general_meeting_id == general_meeting_id
        )
    )
    total_entitlement = weights_result.scalar() or 0

    # RR4-06: Use a single SQL GROUP BY aggregate query to count and sum entitlements
    # per (option_id, choice), avoiding O(V) Python-side iteration over all votes.
    # This is O(1) in application memory regardless of vote count.
    agg_rows = await db.execute(
        select(
            Vote.motion_option_id,
            Vote.choice,
            func.count(Vote.id).label("voter_count"),
            func.coalesce(
                func.sum(GeneralMeetingLotWeight.unit_entitlement_snapshot), 0
            ).label("entitlement_sum"),
        )
        .join(
            GeneralMeetingLotWeight,
            (GeneralMeetingLotWeight.lot_id == Vote.lot_owner_id)
            & (GeneralMeetingLotWeight.general_meeting_id == general_meeting_id),
            isouter=True,
        )
        .where(
            Vote.general_meeting_id == general_meeting_id,
            Vote.status == VoteStatus.submitted,
            Vote.motion_option_id.is_not(None),
        )
        .group_by(Vote.motion_option_id, Vote.choice)
    )
    # Index aggregated results by (option_id, choice_str) for O(1) lookup
    agg_by_opt_choice: dict[tuple[uuid.UUID, str], tuple[int, int]] = {}
    for row in agg_rows.all():
        opt_id_key = row[0]
        choice_str = row[1].value if hasattr(row[1], "value") else row[1]
        agg_by_opt_choice[(opt_id_key, choice_str)] = (int(row[2]), int(row[3]))

    for motion in mc_motions:
        # Load options for this motion
        opts_result = await db.execute(
            select(MotionOption)
            .where(MotionOption.motion_id == motion.id)
            .order_by(MotionOption.display_order)
        )
        options = list(opts_result.scalars().all())
        if not options:
            continue

        option_limit = motion.option_limit or len(options)

        # Read per-option tallies from the pre-aggregated lookup
        for_voter_counts: dict[uuid.UUID, int] = {}
        for_sums: dict[uuid.UUID, int] = {}
        against_voter_counts: dict[uuid.UUID, int] = {}
        against_sums: dict[uuid.UUID, int] = {}
        abstained_voter_counts: dict[uuid.UUID, int] = {}
        abstained_sums: dict[uuid.UUID, int] = {}
        for opt in options:
            fc, fs = agg_by_opt_choice.get((opt.id, "selected"), (0, 0))
            ac, as_ = agg_by_opt_choice.get((opt.id, "against"), (0, 0))
            abc, abs_ = agg_by_opt_choice.get((opt.id, "abstained"), (0, 0))
            for_voter_counts[opt.id] = fc
            for_sums[opt.id] = fs
            against_voter_counts[opt.id] = ac
            against_sums[opt.id] = as_
            abstained_voter_counts[opt.id] = abc
            abstained_sums[opt.id] = abs_

        # Step 3: mark failed options (>50% against)
        failed_by_against: set[uuid.UUID] = set()
        if total_entitlement > 0:
            for opt in options:
                if against_sums.get(opt.id, 0) / total_entitlement > 0.50:
                    failed_by_against.add(opt.id)

        # Step 4 & 5: rank remaining options by for_entitlement_sum
        remaining = [opt for opt in options if opt.id not in failed_by_against]
        remaining.sort(key=lambda o: for_sums.get(o.id, 0), reverse=True)

        # Determine outcomes
        outcome_map: dict[uuid.UUID, str] = {}

        # All against-failed options are "fail"
        for opt_id in failed_by_against:
            outcome_map[opt_id] = "fail"

        if remaining:
            # Check for tie at boundary (position option_limit vs option_limit+1)
            if len(remaining) > option_limit:
                boundary_score = for_sums.get(remaining[option_limit - 1].id, 0)
                next_score = for_sums.get(remaining[option_limit].id, 0)
                if boundary_score == next_score:
                    # Tie: mark all options with boundary_score as "tie"
                    for opt in remaining:
                        if for_sums.get(opt.id, 0) == boundary_score:
                            outcome_map[opt.id] = "tie"
                        elif for_sums.get(opt.id, 0) > boundary_score:
                            outcome_map[opt.id] = "pass"
                        else:
                            outcome_map[opt.id] = "fail"
                else:
                    # No tie at boundary
                    for i, opt in enumerate(remaining):
                        outcome_map[opt.id] = "pass" if i < option_limit else "fail"
            else:
                # All remaining options fit within the limit — all pass
                for opt in remaining:
                    outcome_map[opt.id] = "pass"

        # Persist outcomes and all six tally snapshot fields
        for opt in options:
            opt.outcome = outcome_map.get(opt.id)
            opt.for_voter_count = for_voter_counts.get(opt.id, 0)
            opt.for_entitlement_sum = for_sums.get(opt.id, 0)
            opt.against_voter_count = against_voter_counts.get(opt.id, 0)
            opt.against_entitlement_sum = against_sums.get(opt.id, 0)
            opt.abstained_voter_count = abstained_voter_counts.get(opt.id, 0)
            opt.abstained_entitlement_sum = abstained_sums.get(opt.id, 0)

    await db.flush()


async def delete_general_meeting(general_meeting_id: uuid.UUID, db: AsyncSession) -> None:
    result = await db.execute(select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    meeting = result.scalar_one_or_none()
    if meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")
    if meeting.status == GeneralMeetingStatus.open:
        raise HTTPException(status_code=409, detail="Cannot delete an open General Meeting")
    # RR5-11: Block deletion of pending meetings that already have data (motions or lot weights),
    # to prevent accidental loss of configured but not-yet-started meetings.
    if meeting.status == GeneralMeetingStatus.pending:
        motion_count_result = await db.execute(
            select(func.count()).select_from(Motion).where(Motion.general_meeting_id == general_meeting_id)
        )
        motion_count = motion_count_result.scalar_one()
        if motion_count > 0:
            raise HTTPException(status_code=409, detail="Cannot delete a pending General Meeting that has motions or lot weights")
        weight_count_result = await db.execute(
            select(func.count()).select_from(GeneralMeetingLotWeight).where(
                GeneralMeetingLotWeight.general_meeting_id == general_meeting_id
            )
        )
        weight_count = weight_count_result.scalar_one()
        if weight_count > 0:
            raise HTTPException(status_code=409, detail="Cannot delete a pending General Meeting that has motions or lot weights")
    # Use a statement-level DELETE so PostgreSQL's ondelete=CASCADE FK constraints handle
    # child rows (votes, motions, ballot_submissions, etc.) at the DB level.  The ORM-level
    # db.delete() path requires all child collections to be eagerly loaded in the async
    # session; without that it raises a MissingGreenlet / lazy-load error at runtime.
    await db.execute(delete(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    await db.commit()


async def resend_report(general_meeting_id: uuid.UUID, db: AsyncSession) -> dict:
    result = await db.execute(select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    general_meeting = result.scalar_one_or_none()
    if general_meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")
    if general_meeting.status == GeneralMeetingStatus.open:
        raise HTTPException(status_code=409, detail="General Meeting is not closed")

    # Check EmailDelivery
    delivery_result = await db.execute(
        select(EmailDelivery).where(EmailDelivery.general_meeting_id == general_meeting_id)
    )
    delivery = delivery_result.scalar_one_or_none()
    if delivery is None:
        raise HTTPException(status_code=404, detail="Email delivery record not found")

    # Reset delivery regardless of current status to allow resend on demand
    delivery.status = EmailDeliveryStatus.pending
    delivery.total_attempts = 0
    delivery.last_error = None
    delivery.next_retry_at = None

    await db.commit()

    # Stub: log email delivery trigger
    logger.info("email_delivery_triggered", agm_id=str(general_meeting_id))

    return {"queued": True}


async def reorder_motions(
    general_meeting_id: uuid.UUID,
    request,  # MotionReorderRequest — imported at call site to avoid circular import
    db: AsyncSession,
) -> dict:
    """Bulk reorder motions for a general meeting.

    Validates:
    - Meeting exists (404 if not)
    - Meeting is not closed (403 if closed)
    - request.motions contains exactly the same set of IDs as the meeting's motions (422)
    - No duplicate display_order values in the request (422)

    Then normalises display_order to 1-based sequential integers sorted by the
    submitted display_order values and updates all motions atomically.

    Returns {"motions": [...]} sorted by display_order.
    """
    from fastapi import HTTPException as _HTTPException

    # Fetch meeting
    result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id)
    )
    meeting = result.scalar_one_or_none()
    if meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")

    # Check not closed
    effective = get_effective_status(meeting)
    if effective == GeneralMeetingStatus.closed:
        raise HTTPException(
            status_code=409,
            detail="Cannot reorder motions on a closed General Meeting",
        )

    # Load all motions for this meeting
    motions_result = await db.execute(
        select(Motion).where(Motion.general_meeting_id == general_meeting_id)
    )
    motions = list(motions_result.scalars().all())
    existing_ids = {m.id for m in motions}
    motion_map = {m.id: m for m in motions}

    # Validate request list
    submitted_ids = {item.motion_id for item in request.motions}

    if len(request.motions) == 0 or submitted_ids != existing_ids:
        raise HTTPException(
            status_code=422,
            detail="motion_order must contain exactly all motion IDs for this meeting",
        )

    # Check for duplicate display_order values in the request
    submitted_orders = [item.display_order for item in request.motions]
    if len(submitted_orders) != len(set(submitted_orders)):
        raise HTTPException(
            status_code=422,
            detail="Duplicate display_order values in request",
        )

    # Sort request items by submitted display_order, then assign normalised 1-based positions
    sorted_items = sorted(request.motions, key=lambda x: x.display_order)

    # Two-pass update to avoid unique constraint violations:
    # Pass 1: assign large temporary values
    offset = len(motions) + 1000
    for item in sorted_items:
        motion_map[item.motion_id].display_order = item.display_order + offset
    await db.flush()

    # Pass 2: assign final normalised values
    for position, item in enumerate(sorted_items, start=1):
        motion_map[item.motion_id].display_order = position
    await db.commit()

    # Reload sorted motions
    final_result = await db.execute(
        select(Motion)
        .where(Motion.general_meeting_id == general_meeting_id)
        .order_by(Motion.display_order)
    )
    final_motions = list(final_result.scalars().all())

    # Load options for multi-choice motions in the final result
    final_mc_ids = [m.id for m in final_motions if m.is_multi_choice]
    final_opts_map: dict[uuid.UUID, list] = {}
    if final_mc_ids:
        final_opts_result = await db.execute(
            select(MotionOption)
            .where(MotionOption.motion_id.in_(final_mc_ids))
            .order_by(MotionOption.display_order)
        )
        for opt in final_opts_result.scalars().all():
            final_opts_map.setdefault(opt.motion_id, []).append(opt)

    return {
        "motions": [
            {
                "id": m.id,
                "title": m.title,
                "description": m.description,
                "display_order": m.display_order,
                "motion_number": m.motion_number,
                "motion_type": m.motion_type.value if hasattr(m.motion_type, "value") else m.motion_type,
                "option_limit": m.option_limit,
                "options": [
                    {"id": opt.id, "text": opt.text, "display_order": opt.display_order}
                    for opt in final_opts_map.get(m.id, [])
                ],
            }
            for m in final_motions
        ]
    }


async def reset_general_meeting_ballots(general_meeting_id: uuid.UUID, db: AsyncSession) -> dict:
    """Delete all ballot submissions (and their associated submitted votes) for a General Meeting.

    Intended for E2E test setup only — clears submitted votes so the test
    suite can re-run the voting flow without hitting a 409 conflict.
    Returns the number of ballot submissions deleted.
    """
    result = await db.execute(select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id))
    general_meeting = result.scalar_one_or_none()
    if general_meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")

    # Count submissions before deleting so we can return the count.
    count_result = await db.execute(
        select(func.count(BallotSubmission.id)).where(BallotSubmission.general_meeting_id == general_meeting_id)
    )
    deleted_count = count_result.scalar_one()

    # Delete submitted votes for this General Meeting (must come before ballot submissions
    # to avoid any application-level FK issues, even though there is no DB-level
    # FK from votes -> ballot_submissions).
    await db.execute(
        delete(Vote).where(Vote.general_meeting_id == general_meeting_id, Vote.status == VoteStatus.submitted)
    )

    # Delete all ballot submissions for this General Meeting.
    await db.execute(
        delete(BallotSubmission).where(BallotSubmission.general_meeting_id == general_meeting_id)
    )

    await db.commit()

    return {"deleted": deleted_count}


# ---------------------------------------------------------------------------
# Proxy nomination import (US-PX02)
# ---------------------------------------------------------------------------


def _parse_proxy_csv_rows(content: bytes) -> list[dict]:
    """Parse CSV bytes into list of {lot_number, proxy_email} dicts.

    Raises HTTPException 422 on missing headers.
    """
    text = content.decode("utf-8-sig")
    raw_reader = csv.DictReader(io.StringIO(text))

    if raw_reader.fieldnames is None:
        raise HTTPException(status_code=422, detail="CSV has no headers")

    normalised = {f.strip().lower() for f in raw_reader.fieldnames}
    required = {"lot#", "proxy email"}
    missing = required - normalised
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required CSV headers: {sorted(missing)}",
        )

    rows = []
    for row in raw_reader:
        # Build a case-insensitive lookup
        row_lower = {k.strip().lower(): v for k, v in row.items()}
        lot_number = row_lower.get("lot#", "").strip()
        proxy_email = row_lower.get("proxy email", "").strip()
        given_name = row_lower.get("proxy_given_name", "").strip() or None
        surname = row_lower.get("proxy_surname", "").strip() or None
        rows.append({
            "lot_number": lot_number,
            "proxy_email": proxy_email,
            "given_name": given_name,
            "surname": surname,
        })
    return rows


def _parse_proxy_excel_rows(content: bytes) -> list[dict]:
    """Parse Excel bytes into list of {lot_number, proxy_email} dicts.

    Raises HTTPException 422 on invalid file or missing headers.
    """
    # RR4-26: Narrow exception catch to file-format errors only; re-raise unexpected ones.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    if header_row is None or all(v is None for v in header_row):  # pragma: no cover
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    required = {"lot#", "proxy email"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required Excel headers: {sorted(missing)}",
        )

    lot_idx = headers.index("lot#")
    proxy_idx = headers.index("proxy email")
    given_name_idx = headers.index("proxy_given_name") if "proxy_given_name" in headers else None
    surname_idx = headers.index("proxy_surname") if "proxy_surname" in headers else None

    data_rows = list(rows_iter)
    wb.close()

    rows = []
    for raw_row in data_rows:
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        def _cell(idx: int) -> str:
            if idx < len(raw_row) and raw_row[idx] is not None:
                return str(raw_row[idx]).strip()
            return ""

        rows.append({
            "lot_number": _cell(lot_idx),
            "proxy_email": _cell(proxy_idx),
            "given_name": _cell(given_name_idx) or None if given_name_idx is not None else None,
            "surname": _cell(surname_idx) or None if surname_idx is not None else None,
        })
    return rows


async def import_proxies(
    building_id: uuid.UUID,
    rows: list[dict],
    db: AsyncSession,
) -> dict[str, int]:
    """Upsert proxy nominations from parsed rows.

    Each row: {lot_number: str, proxy_email: str}.
    Blank proxy_email removes the nomination.
    Unknown lot_number is skipped with a warning.
    Returns {"upserted": N, "removed": N, "skipped": N}.
    """
    await get_building_or_404(building_id, db)

    # Load all lots for this building keyed by lot_number
    existing_result = await db.execute(
        select(Lot).where(Lot.building_id == building_id)
    )
    lot_map: dict[str, Lot] = {
        lo.lot_number: lo for lo in existing_result.scalars().all()
    }

    # Batch-load all existing proxies for this building's lots in a single query
    # to avoid N+1 SELECT per row (RR5-04).
    lot_ids = list({lo.id for lo in lot_map.values()})
    existing_proxies_result = await db.execute(
        select(LotProxy).where(LotProxy.lot_id.in_(lot_ids))
    )
    proxy_map: dict[uuid.UUID, LotProxy] = {
        p.lot_id: p for p in existing_proxies_result.scalars().all()
    }

    upserted = 0
    removed = 0
    skipped = 0

    for row in rows:
        lot_number = row["lot_number"]
        proxy_email = row["proxy_email"]

        lot = lot_map.get(lot_number)
        if lot is None:
            logger.warning(
                "Proxy import: lot_number %r not found in building %s — skipping",
                lot_number,
                building_id,
            )
            skipped += 1
            continue

        # Lookup existing proxy from pre-loaded dict (RR5-04: no per-row SELECT)
        existing_proxy = proxy_map.get(lot.id)

        given_name = row.get("given_name")
        surname = row.get("surname")

        if proxy_email == "":
            # Remove nomination
            if existing_proxy is not None:
                await db.delete(existing_proxy)
                removed += 1
        else:
            # Upsert nomination: resolve/create person, then set person_id on LotProxy
            person = await get_or_create_person(proxy_email, db, given_name=given_name, surname=surname)
            if existing_proxy is not None:
                existing_proxy.person_id = person.id
            else:
                db.add(LotProxy(
                    lot_id=lot.id,
                    person_id=person.id,
                ))
            upserted += 1

    await db.commit()
    return {"upserted": upserted, "removed": removed, "skipped": skipped}


async def import_proxies_from_csv(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    rows = _parse_proxy_csv_rows(content)
    return await import_proxies(building_id, rows, db)


async def import_proxies_from_excel(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    rows = _parse_proxy_excel_rows(content)
    return await import_proxies(building_id, rows, db)


# ---------------------------------------------------------------------------
# Financial position import (US-PX03)
# ---------------------------------------------------------------------------


def _parse_financial_position_import(raw: str) -> FinancialPosition | None:
    """Parse a financial position string from an import row.

    Accepted values (case-insensitive): 'Normal' -> normal, 'In Arrear' -> in_arrear.
    Returns None if empty (should be skipped or flagged).
    Raises ValueError on invalid value.
    """
    normalised = raw.strip().lower()
    if normalised == "normal":
        return FinancialPosition.normal
    if normalised in ("in arrear", "in_arrear"):
        return FinancialPosition.in_arrear
    raise ValueError(f"Invalid Financial Position value: '{raw}'")


def _parse_closing_balance(raw: str) -> FinancialPosition:
    """Parse a TOCS Closing Balance cell to FinancialPosition.

    - '$-' or '$ -' or empty → normal (zero balance)
    - Contains '(' → normal (credit/advance, negative balance)
    - Otherwise (positive number like $1,882.06) → in_arrear
    """
    cleaned = raw.strip().replace(" ", "")
    if not cleaned or cleaned in ("$-", "-"):
        return FinancialPosition.normal
    if "(" in cleaned:
        return FinancialPosition.normal
    return FinancialPosition.in_arrear


def _parse_tocs_financial_position_csv_rows(content: bytes) -> list[dict]:
    """Parse a TOCS Lot Positions Report CSV into {lot_number, financial_position_raw} dicts.

    The TOCS format has:
    - Header rows at the top (company name, address, etc.)
    - Multiple fund sections, each starting with a 'Lot#' header row
    - 9 columns: Lot#, Unit#, Owner Name, Opening Balance, Levied, Special Levy, Paid,
      Closing Balance, Interest Paid
    - Totals/Arrears/Advances summary rows at the end of each section
    - Blank rows between sections

    Uses worst-case logic: if a lot is in_arrear in ANY fund section → in_arrear.
    """
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    # Find section header rows (rows where first non-empty cell is 'lot#')
    section_starts: list[int] = []
    for i, row in enumerate(all_rows):
        first_cell = next((c.strip() for c in row if c.strip()), "")
        if first_cell.lower() == "lot#":
            section_starts.append(i)

    # Accumulate worst-case positions across all fund sections
    result: dict[str, FinancialPosition] = {}

    _stop_keywords = ("total", "arrear", "advance")

    for header_idx in section_starts:
        # Determine column indices from this section's header row
        header_row = all_rows[header_idx]
        # Find Lot# col (index 0 in standard TOCS) and Closing Balance col (index 7)
        lot_col = 0
        closing_col = 7
        for col_i, cell in enumerate(header_row):
            cell_lower = cell.strip().lower()
            if cell_lower == "lot#":
                lot_col = col_i
            elif cell_lower == "closing balance":
                closing_col = col_i

        # Read data rows until a stop condition
        for row in all_rows[header_idx + 1:]:
            # Blank row → end of section
            if not any(c.strip() for c in row):
                break

            first_cell = row[lot_col].strip() if lot_col < len(row) else ""

            # Summary rows (Totals/Arrears/Advances) — may be prefixed with fund name
            # e.g. "Administrative Fund Totals", "Maintenance Fund Arrears"
            if any(kw in first_cell.lower() for kw in _stop_keywords):
                break

            # Another Lot# header → end of section (shouldn't happen but be safe)
            if first_cell.lower() == "lot#":
                break

            # Skip rows with empty or non-lot lot# values
            if not first_cell:
                continue

            # Extract closing balance
            closing_raw = row[closing_col].strip() if closing_col < len(row) else ""
            position = _parse_closing_balance(closing_raw)

            # Worst-case: upgrade to in_arrear if either value is in_arrear
            existing = result.get(first_cell)
            if existing is None or position == FinancialPosition.in_arrear:
                result[first_cell] = position

    return [
        {"lot_number": lot_number, "financial_position_raw": fp.value}
        for lot_number, fp in result.items()
    ]


def _parse_simple_financial_position_csv_rows(content: bytes) -> list[dict]:
    """Parse simple template CSV bytes into list of {lot_number, financial_position_raw} dicts.

    Simple format: two columns — Lot#, Financial Position.
    Raises HTTPException 422 on missing headers.
    """
    text = content.decode("utf-8-sig")
    raw_reader = csv.DictReader(io.StringIO(text))

    if raw_reader.fieldnames is None:
        raise HTTPException(status_code=422, detail="CSV has no headers")

    normalised = {f.strip().lower() for f in raw_reader.fieldnames}
    required = {"lot#", "financial position"}
    missing = required - normalised
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required CSV headers: {sorted(missing)}",
        )

    rows = []
    for row in raw_reader:
        row_lower = {k.strip().lower(): v for k, v in row.items()}
        lot_number = row_lower.get("lot#", "").strip()
        fp_raw = row_lower.get("financial position", "").strip()
        rows.append({"lot_number": lot_number, "financial_position_raw": fp_raw})
    return rows


def _parse_financial_position_csv_rows(content: bytes) -> list[dict]:
    """Parse CSV bytes — auto-detects simple template vs TOCS Lot Positions Report format.

    Simple format starts with 'Lot#' as the first cell of the first line.
    TOCS format has header rows before the first 'Lot#' section.
    """
    text = content.decode("utf-8-sig")
    first_line = text.strip().split("\n")[0] if text.strip() else ""
    first_cell = first_line.split(",")[0].strip().strip('"').lower()

    if first_cell == "lot#":
        return _parse_simple_financial_position_csv_rows(content)
    else:
        return _parse_tocs_financial_position_csv_rows(content)


def _parse_closing_balance_numeric(val: int | float) -> FinancialPosition:
    """Classify a numeric closing balance cell value from an xlsx file.

    xlsx files exported from TOCS store Closing Balance as Python int/float, not
    currency strings.  The rule mirrors _parse_closing_balance for strings:

    - val <= 0  → normal (zero = paid up; negative = credit/overpaid)
    - val > 0   → in_arrear
    """
    return FinancialPosition.in_arrear if val > 0 else FinancialPosition.normal


def _parse_tocs_financial_position_excel_rows(all_rows: list[tuple]) -> list[dict]:
    """Parse a TOCS Lot Positions Report from already-loaded openpyxl rows.

    Mirrors _parse_tocs_financial_position_csv_rows but works on tuples of cell
    values rather than CSV text rows.  Each tuple element may be None (blank cell).

    The TOCS format has:
    - Header rows at the top (company name, address, etc.)
    - Multiple fund sections, each starting with a row whose first non-empty cell is 'Lot#'
    - 9 columns per section: Lot#, Unit#, Owner Name, Opening Balance, Levied,
      Special Levy, Paid, Closing Balance, Interest Paid
    - Totals/Arrears/Advances summary rows at the end of each section

    Closing Balance cells may be numeric (int/float from xlsx) or string-formatted
    currency (e.g. '$-', '$(190.77)', '$1,882.06' from CSV-sourced xlsx).  Both
    representations are handled.

    Uses worst-case logic: if a lot is in_arrear in ANY fund section → in_arrear.
    """

    def _str(val: object) -> str:
        return str(val).strip() if val is not None else ""

    # Find section header rows (rows whose first non-empty cell is 'lot#')
    section_starts: list[int] = []
    for i, row in enumerate(all_rows):
        first_cell = next((_str(c) for c in row if _str(c)), "")
        if first_cell.lower() == "lot#":
            section_starts.append(i)

    result: dict[str, FinancialPosition] = {}
    _stop_keywords = ("total", "arrear", "advance")

    for header_idx in section_starts:
        header_row = all_rows[header_idx]
        # Determine Lot# and Closing Balance column indices from this section header
        lot_col = 0
        closing_col = 7
        for col_i, cell in enumerate(header_row):
            cell_lower = _str(cell).lower()
            if cell_lower == "lot#":
                lot_col = col_i
            elif cell_lower == "closing balance":
                closing_col = col_i

        for row in all_rows[header_idx + 1:]:
            # Blank row → end of section
            if not any(_str(c) for c in row):
                break

            first_cell = _str(row[lot_col]) if lot_col < len(row) else ""

            # Summary rows (Totals/Arrears/Advances)
            if any(kw in first_cell.lower() for kw in _stop_keywords):
                break

            # Another Lot# header → end of section
            if first_cell.lower() == "lot#":
                break

            # Skip rows with empty lot# values
            if not first_cell:
                continue

            # Closing Balance may be numeric (xlsx native) or currency string (CSV-style)
            raw_closing = row[closing_col] if closing_col < len(row) else None
            if isinstance(raw_closing, (int, float)):
                position = _parse_closing_balance_numeric(raw_closing)
            else:
                position = _parse_closing_balance(_str(raw_closing) if raw_closing is not None else "")

            existing = result.get(first_cell)
            if existing is None or position == FinancialPosition.in_arrear:
                result[first_cell] = position

    return [
        {"lot_number": lot_number, "financial_position_raw": fp.value}
        for lot_number, fp in result.items()
    ]


def _parse_financial_position_excel_rows(content: bytes) -> list[dict]:
    """Parse Excel bytes into list of {lot_number, financial_position_raw} dicts.

    Auto-detects simple template vs TOCS Lot Positions Report format:
    - Simple format: first non-empty cell of first row is 'Lot#' AND the row also
      contains a 'Financial Position' column.
    - TOCS format: first row contains company/report header text — the Lot# rows
      appear later as section headers within multiple fund sections.

    Raises HTTPException 422 on invalid file or missing headers.
    """
    # RR4-26: Narrow exception catch to file-format errors only; re-raise unexpected ones.
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    first_row = all_rows[0]
    if first_row is None or all(v is None for v in first_row):  # pragma: no cover
        raise HTTPException(status_code=422, detail="Excel file has no headers")

    first_row_headers = [str(h).strip().lower() if h is not None else "" for h in first_row]
    first_non_empty = next((h for h in first_row_headers if h), "")

    # TOCS format detection: first row's first non-empty cell is NOT 'lot#'
    if first_non_empty != "lot#":
        return _parse_tocs_financial_position_excel_rows(all_rows)

    # Simple template format: require both 'lot#' and 'financial position' columns
    required = {"lot#", "financial position"}
    missing = required - set(first_row_headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required Excel headers: {sorted(missing)}",
        )

    lot_idx = first_row_headers.index("lot#")
    fp_idx = first_row_headers.index("financial position")

    rows = []
    for raw_row in all_rows[1:]:
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        def _cell(idx: int, row: tuple = raw_row) -> str:
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        rows.append({
            "lot_number": _cell(lot_idx),
            "financial_position_raw": _cell(fp_idx),
        })
    return rows


async def import_financial_positions(
    building_id: uuid.UUID,
    rows: list[dict],
    db: AsyncSession,
) -> dict[str, int]:
    """Update financial positions from parsed rows.

    Each row: {lot_number: str, financial_position_raw: str}.
    Unknown lot_number is skipped.
    Invalid financial_position_raw raises 422 with all offending rows listed.
    Returns {"updated": N, "skipped": N}.
    """
    await get_building_or_404(building_id, db)

    # Validate all rows first to collect errors
    errors: list[str] = []
    for i, row in enumerate(rows, start=2):
        raw = row["financial_position_raw"]
        if raw == "":
            errors.append(f"Row {i}: Financial Position is empty")
            continue
        try:
            _parse_financial_position_import(raw)
        except ValueError as exc:
            errors.append(f"Row {i}: {exc}")

    if errors:
        raise HTTPException(status_code=422, detail=errors)

    # Load all lots for this building keyed by lot_number
    existing_result = await db.execute(
        select(Lot).where(Lot.building_id == building_id)
    )
    lot_map: dict[str, Lot] = {
        lo.lot_number: lo for lo in existing_result.scalars().all()
    }

    updated = 0
    skipped = 0

    for row in rows:
        lot_number = row["lot_number"]
        fp_raw = row["financial_position_raw"]

        lot = lot_map.get(lot_number)
        if lot is None:
            logger.warning(
                "Financial position import: lot_number %r not found in building %s — skipping",
                lot_number,
                building_id,
            )
            skipped += 1
            continue

        fp = _parse_financial_position_import(fp_raw)
        lot.financial_position = fp  # type: ignore[assignment]
        updated += 1

    await db.commit()
    return {"updated": updated, "skipped": skipped}


async def import_financial_positions_from_csv(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    rows = _parse_financial_position_csv_rows(content)
    return await import_financial_positions(building_id, rows, db)


async def import_financial_positions_from_excel(
    building_id: uuid.UUID,
    content: bytes,
    db: AsyncSession,
) -> dict[str, int]:
    rows = _parse_financial_position_excel_rows(content)
    return await import_financial_positions(building_id, rows, db)


# ---------------------------------------------------------------------------
# Admin in-person vote entry (US-AVE-01, US-AVE-02, US-AVE-03)
# ---------------------------------------------------------------------------


async def enter_votes_for_meeting(
    general_meeting_id: uuid.UUID,
    request: AdminVoteEntryRequest,
    db: AsyncSession,
    admin_username: str | None = None,
) -> dict[str, int]:
    """
    Enter votes on behalf of in-person lot owners (US-AVE-01/02).

    Business rules:
    - Meeting must be open; returns 409 otherwise.
    - If a lot already has a real (non-absent) BallotSubmission, it is skipped
      (app-submitted ballots take precedence); skipped_count is incremented.
    - For each submitted lot, all visible motions are recorded:
        * in-arrear lots: not_eligible for general/multi_choice; normal for special
        * inline vote provided → use that choice
        * no inline vote provided → abstained
        * multi-choice options provided → validate option_ids and option_limit
        * no options provided for multi-choice → abstained
    - All created BallotSubmission rows have submitted_by_admin = True and
      submitted_by_admin_username = admin_username.
    - Returns {"submitted_count": N, "skipped_count": M}.
    """
    from sqlalchemy.exc import IntegrityError
    from app.models import (
        MotionType as _MotionType,
    )

    # Fetch and validate meeting
    meeting_result = await db.execute(
        select(GeneralMeeting).where(GeneralMeeting.id == general_meeting_id)
    )
    meeting = meeting_result.scalar_one_or_none()
    if meeting is None:
        raise HTTPException(status_code=404, detail="General Meeting not found")

    effective = get_effective_status(meeting)
    if effective != GeneralMeetingStatus.open:
        raise HTTPException(status_code=409, detail="Meeting is not open")

    # Collect all lot_owner_ids being entered
    lot_owner_ids = [e.lot_owner_id for e in request.entries]
    if not lot_owner_ids:
        return {"submitted_count": 0, "skipped_count": 0}

    # Validate all lot_owner_ids exist in the DB
    lo_result = await db.execute(
        select(Lot.id).where(Lot.id.in_(lot_owner_ids))
    )
    found_ids: set[uuid.UUID] = {row[0] for row in lo_result.all()}
    unknown = [str(lid) for lid in lot_owner_ids if lid not in found_ids]
    if unknown:
        raise HTTPException(
            status_code=422,
            detail=f"Unknown lot_owner_ids: {unknown}",
        )

    # Check existing real (non-absent) submissions for these lots.
    # Retained to distinguish "new submission" from "re-entry submission" in the
    # BallotSubmission creation step (Fix 4a).
    existing_result = await db.execute(
        select(BallotSubmission.lot_owner_id).where(
            BallotSubmission.general_meeting_id == general_meeting_id,
            BallotSubmission.lot_owner_id.in_(lot_owner_ids),
            BallotSubmission.is_absent == False,  # noqa: E712
        )
    )
    already_submitted: set[uuid.UUID] = {row[0] for row in existing_result.all()}

    # Load already-voted motion IDs per lot (single IN query).
    # Used by the per-motion skip logic to allow admin vote entry on partially-submitted
    # lots (Fix 4a: mirrors the re-entry logic in submit_ballot).
    all_voted_result = await db.execute(
        select(Vote.lot_owner_id, Vote.motion_id).where(
            Vote.general_meeting_id == general_meeting_id,
            Vote.lot_owner_id.in_(lot_owner_ids),
            Vote.status == VoteStatus.submitted,
        )
    )
    already_voted_by_lot: dict[uuid.UUID, set[uuid.UUID]] = {}
    for row in all_voted_result.all():
        already_voted_by_lot.setdefault(row[0], set()).add(row[1])

    # Load financial position snapshots for all lots
    weights_result = await db.execute(
        select(GeneralMeetingLotWeight).where(
            GeneralMeetingLotWeight.general_meeting_id == general_meeting_id,
            GeneralMeetingLotWeight.lot_id.in_(lot_owner_ids),
        )
    )
    weight_by_lot: dict[uuid.UUID, GeneralMeetingLotWeight] = {
        w.lot_id: w for w in weights_result.scalars().all()
    }

    # Load all visible motions for this meeting
    motions_result = await db.execute(
        select(Motion)
        .where(
            Motion.general_meeting_id == general_meeting_id,
            Motion.is_visible == True,  # noqa: E712
        )
        .order_by(Motion.display_order)
    )
    visible_motions = list(motions_result.scalars().all())
    valid_motion_ids = {m.id for m in visible_motions}

    # Load multi-choice options for validation
    mc_motion_ids = [m.id for m in visible_motions if m.is_multi_choice]
    mc_options_map: dict[uuid.UUID, set[uuid.UUID]] = {}
    if mc_motion_ids:
        opts_result = await db.execute(
            select(MotionOption).where(MotionOption.motion_id.in_(mc_motion_ids))
        )
        for opt in opts_result.scalars().all():
            mc_options_map.setdefault(opt.motion_id, set()).add(opt.id)

    submitted_count = 0
    skipped_count = 0

    for entry in request.entries:
        lot_owner_id = entry.lot_owner_id

        # Determine financial position
        weight = weight_by_lot.get(lot_owner_id)
        is_in_arrear = (
            weight is not None
            and weight.financial_position_snapshot == FinancialPositionSnapshot.in_arrear
        )

        # Build inline vote lookup: motion_id -> VoteChoice
        inline_lookup: dict[uuid.UUID, VoteChoice] = {}
        for v in entry.votes:
            mid = uuid.UUID(str(v["motion_id"])) if not isinstance(v["motion_id"], uuid.UUID) else v["motion_id"]
            choice_str = str(v["choice"]).lower()
            choice_map = {
                "yes": VoteChoice.yes,
                "no": VoteChoice.no,
                "abstained": VoteChoice.abstained,
                "for": VoteChoice.yes,
                "against": VoteChoice.no,
            }
            if choice_str not in choice_map:
                raise HTTPException(
                    status_code=422,
                    detail=f"Invalid choice '{v['choice']}' for motion {mid}",
                )
            if mid not in valid_motion_ids:
                raise HTTPException(
                    status_code=422,
                    detail=f"Unknown motion ID {mid}",
                )
            inline_lookup[mid] = choice_map[choice_str]

        # Build multi-choice vote lookup (US-AVE2-01):
        # motion_id -> list of (option_id, VoteChoice) pairs
        # Supports both new option_choices format and legacy option_ids format.
        mc_lookup: dict[uuid.UUID, list[tuple[uuid.UUID, VoteChoice]]] = {}
        for mv in entry.multi_choice_votes:
            mid = uuid.UUID(str(mv["motion_id"])) if not isinstance(mv["motion_id"], uuid.UUID) else mv["motion_id"]
            if mid not in valid_motion_ids:
                raise HTTPException(
                    status_code=422,
                    detail=f"Unknown motion ID {mid}",
                )
            valid_opts = mc_options_map.get(mid, set())

            # New format: option_choices takes precedence over option_ids
            raw_option_choices = mv.get("option_choices") or []
            raw_option_ids = mv.get("option_ids") or []

            option_choice_map = {
                "for": VoteChoice.selected,
                "against": VoteChoice.against,
                "abstained": VoteChoice.abstained,
            }

            if raw_option_choices:
                # New format: [{option_id, choice}]
                pairs: list[tuple[uuid.UUID, VoteChoice]] = []
                for oc in raw_option_choices:
                    oid = uuid.UUID(str(oc["option_id"])) if not isinstance(oc.get("option_id"), uuid.UUID) else oc["option_id"]
                    if oid not in valid_opts:
                        raise HTTPException(
                            status_code=422,
                            detail=f"Invalid option ID {oid} for motion {mid}",
                        )
                    choice_str = str(oc.get("choice", "")).lower()
                    if choice_str not in option_choice_map:
                        raise HTTPException(
                            status_code=422,
                            detail=f"Invalid choice '{oc.get('choice')}' for option {oid}",
                        )
                    pairs.append((oid, option_choice_map[choice_str]))
                mc_lookup[mid] = pairs
            else:
                # Legacy format: option_ids treated as all "for"
                opt_ids = [
                    uuid.UUID(str(oid)) if not isinstance(oid, uuid.UUID) else oid
                    for oid in raw_option_ids
                ]
                for oid in opt_ids:
                    if oid not in valid_opts:
                        raise HTTPException(
                            status_code=422,
                            detail=f"Invalid option ID {oid} for motion {mid}",
                        )
                mc_lookup[mid] = [(oid, VoteChoice.selected) for oid in opt_ids]

        # Build Vote rows for all visible motions.
        # Fix 4a: skip motions this lot has already voted on (per-motion check),
        # mirroring the re-entry logic in submit_ballot.  This allows admin vote entry
        # on partially-submitted lots (e.g. voter submitted M1 online; admin enters M2).
        already_voted_for_lot: set[uuid.UUID] = already_voted_by_lot.get(lot_owner_id, set())
        votes_to_add: list[Vote] = []
        for motion in visible_motions:
            if motion.id in already_voted_for_lot:
                continue  # already voted on this motion — skip

            motion_type = motion.motion_type.value if hasattr(motion.motion_type, "value") else motion.motion_type

            if motion.is_multi_choice:
                if is_in_arrear:
                    votes_to_add.append(Vote(
                        general_meeting_id=general_meeting_id,
                        motion_id=motion.id,
                        voter_email="admin",
                        lot_owner_id=lot_owner_id,
                        choice=VoteChoice.not_eligible,
                        status=VoteStatus.submitted,
                    ))
                    continue

                option_pairs = mc_lookup.get(motion.id, [])
                if not option_pairs:
                    # No options specified for this motion — skip it entirely.
                    # The frontend only sends motions the admin explicitly interacted with,
                    # so an absent motion means the admin made no choice and it should
                    # remain unrecorded for future entry.
                    continue
                else:
                    for opt_id, vote_choice in option_pairs:
                        votes_to_add.append(Vote(
                            general_meeting_id=general_meeting_id,
                            motion_id=motion.id,
                            voter_email="admin",
                            lot_owner_id=lot_owner_id,
                            choice=vote_choice,
                            motion_option_id=opt_id,
                            status=VoteStatus.submitted,
                        ))
                continue

            # Standard motion: not_eligible for in-arrear on general motions
            if is_in_arrear and motion_type == "general":
                votes_to_add.append(Vote(
                    general_meeting_id=general_meeting_id,
                    motion_id=motion.id,
                    voter_email="admin",
                    lot_owner_id=lot_owner_id,
                    choice=VoteChoice.not_eligible,
                    status=VoteStatus.submitted,
                ))
                continue

            if motion.id not in inline_lookup:
                # No explicit choice supplied for this motion — skip it entirely.
                # The frontend only sends motions the admin explicitly set, so an absent
                # motion means no choice was made and it should remain unrecorded for
                # future entry.
                continue
            choice = inline_lookup[motion.id]
            votes_to_add.append(Vote(
                general_meeting_id=general_meeting_id,
                motion_id=motion.id,
                voter_email="admin",
                lot_owner_id=lot_owner_id,
                choice=choice,
                status=VoteStatus.submitted,
            ))

        # If all motions for this lot are already voted, skip it entirely.
        if not votes_to_add:
            skipped_count += 1
            continue

        # Create BallotSubmission with submitted_by_admin=True (only when not already
        # submitted).  On re-entry (lot_owner_id in already_submitted), the existing
        # BallotSubmission is reused and only the new Vote rows are inserted.
        # RR4-05: Use a savepoint (begin_nested) so that only this lot's flush is
        # rolled back on IntegrityError.  Previously a full session rollback wiped
        # all successfully flushed lots that preceded the conflicting one.
        try:
            async with db.begin_nested():
                if lot_owner_id not in already_submitted:
                    submission = BallotSubmission(
                        general_meeting_id=general_meeting_id,
                        lot_owner_id=lot_owner_id,
                        voter_email="admin",
                        proxy_email=None,
                        submitted_by_admin=True,
                        submitted_by_admin_username=admin_username,
                    )
                    db.add(submission)
                for vote in votes_to_add:
                    db.add(vote)
                await db.flush()
        except IntegrityError:
            skipped_count += 1
            continue

        submitted_count += 1

    await db.commit()
    return {"submitted_count": submitted_count, "skipped_count": skipped_count}
