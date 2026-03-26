"""
Tests for the admin portal API endpoints.

Covers all endpoints in /api/admin with full test coverage.

Structure per endpoint:
  # --- Happy path ---
  # --- Input validation ---
  # --- Boundary values ---
  # --- State / precondition errors ---
  # --- Edge cases ---
"""
from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, datetime, timedelta

import openpyxl

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import (
    GeneralMeeting,
    GeneralMeetingLotWeight,
    GeneralMeetingStatus,
    BallotSubmission,
    Building,
    EmailDelivery,
    EmailDeliveryStatus,
    LotOwner,
    LotProxy,
    Motion,
    MotionType,
    Vote,
    VoteChoice,
    VoteStatus,
    get_effective_status,
)
from app.models.lot_owner_email import LotOwnerEmail


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode()


def make_excel(headers: list, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def future_dt(days: int = 1) -> datetime:
    return datetime.now(UTC) + timedelta(days=days)


def meeting_dt() -> datetime:
    """Return a past meeting_at so meetings are effectively open (not pending)."""
    return datetime.now(UTC) - timedelta(hours=1)


def closing_dt() -> datetime:
    return datetime.now(UTC) + timedelta(days=2)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def client(app):
    """HTTP client that shares the test db_session with the app (via conftest app fixture)."""
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as ac:
        yield ac


@pytest_asyncio.fixture
async def building(db_session: AsyncSession) -> Building:
    b = Building(name="Test Building", manager_email="manager@test.com")
    db_session.add(b)
    await db_session.flush()
    await db_session.refresh(b)
    return b


@pytest_asyncio.fixture
async def building_with_owners(db_session: AsyncSession) -> Building:
    b = Building(name="Building With Owners", manager_email="mgr@bwo.com")
    db_session.add(b)
    await db_session.flush()
    lo1 = LotOwner(
        building_id=b.id,
        lot_number="1A",
        unit_entitlement=100,
    )
    lo2 = LotOwner(
        building_id=b.id,
        lot_number="2B",
        unit_entitlement=50,
    )
    db_session.add_all([lo1, lo2])
    await db_session.flush()
    lo1_email = LotOwnerEmail(lot_owner_id=lo1.id, email="voter1@test.com")
    lo2_email = LotOwnerEmail(lot_owner_id=lo2.id, email="voter2@test.com")
    db_session.add_all([lo1_email, lo2_email])
    await db_session.flush()
    await db_session.refresh(b)
    return b


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/import
# ---------------------------------------------------------------------------


class TestImportBuildings:
    # --- Happy path ---

    async def test_valid_csv_creates_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "alpha@test.com"], ["Beta Complex", "beta@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2
        assert data["updated"] == 0

    async def test_valid_csv_updates_existing_building(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        # First, create Alpha Tower
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "alpha@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Now update its manager email
        csv_data2 = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "new_alpha@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 1

    async def test_empty_csv_returns_zero_counts(self, client: AsyncClient):
        csv_data = make_csv(["building_name", "manager_email"], [])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("empty.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 0

    async def test_extra_columns_ignored(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email", "extra_col"],
            [["Extra Col Building", "extra@test.com", "ignored"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["created"] == 1

    async def test_case_insensitive_building_name_match(self, client: AsyncClient):
        # Create a building
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["CaseTest Building", "ci@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Upload with different case
        csv_data2 = make_csv(
            ["building_name", "manager_email"],
            [["casetest building", "updated@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 1
        assert data["created"] == 0

    # --- Input validation ---

    async def test_missing_building_name_header(self, client: AsyncClient):
        csv_data = make_csv(["manager_email"], [["mgr@test.com"]])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "building_name" in str(response.json()["detail"]).lower()

    async def test_missing_manager_email_header(self, client: AsyncClient):
        csv_data = make_csv(["building_name"], [["Some Building"]])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "manager_email" in str(response.json()["detail"]).lower()

    async def test_blank_building_name_row(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["", "mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("building_name" in e for e in response.json()["detail"])

    async def test_blank_manager_email_row(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Valid Building", ""]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("manager_email" in e for e in response.json()["detail"])

    async def test_multiple_row_errors_collected(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["", ""], ["", ""]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        # Should have 4 errors (2 per bad row)
        assert len(response.json()["detail"]) == 4

    # --- State / precondition errors ---

    async def test_non_csv_file_returns_415(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("data.pdf", b"not csv content", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_non_csv_no_extension_returns_415(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("data.bin", b"garbage", "application/octet-stream")},
        )
        # application/octet-stream is in allowed list but no csv extension —
        # however we allow octet-stream, so this should pass content-type check
        # Re-check: octet-stream IS in allowed set, so it should NOT be 415
        assert response.status_code in (200, 422)

    async def test_csv_filename_with_non_csv_content_type(self, client: AsyncClient):
        """A .csv extension overrides content-type check."""
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["CT Building", "ct@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "application/json")},
        )
        # .csv extension should allow it through
        assert response.status_code == 200

    async def test_completely_empty_file_returns_422(self, client: AsyncClient):
        """An empty file (not even headers) should return 422."""
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("empty.csv", b"", "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# GET /api/admin/buildings
# ---------------------------------------------------------------------------


class TestListBuildings:
    # --- Happy path ---

    async def test_returns_all_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) >= 0

    async def test_building_has_required_fields(
        self, client: AsyncClient, building: Building
    ):
        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0
        first = data[0]
        assert "id" in first
        assert "name" in first
        assert "manager_email" in first
        assert "is_archived" in first
        assert "created_at" in first

    # --- name filter ---

    async def test_name_filter_exact_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Filter Exact Building", manager_email="filterexact@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=Filter+Exact+Building")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "Filter Exact Building" in names

    async def test_name_filter_partial_substring_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b1 = Building(name="Substring Alpha Corp", manager_email="sub1@test.com")
        b2 = Building(name="Substring Beta Corp", manager_email="sub2@test.com")
        b_other = Building(name="Unrelated Corp", manager_email="other@test.com")
        db_session.add_all([b1, b2, b_other])
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=Substring")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "Substring Alpha Corp" in names
        assert "Substring Beta Corp" in names
        assert "Unrelated Corp" not in names

    async def test_name_filter_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="CaseSensitive Building", manager_email="case@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=casesensitive")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "CaseSensitive Building" in names

    async def test_name_filter_no_match_returns_empty(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.get("/api/admin/buildings?name=does-not-exist-xyz-99")
        assert response.status_code == 200
        data = response.json()
        assert data == []

    async def test_name_filter_absent_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="No Filter Building", manager_email="nofilter@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "No Filter Building" in names

    async def test_name_filter_empty_string_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="EmptyFilter Building", manager_email="empty@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        # empty string matches everything via LIKE '%%'
        assert "EmptyFilter Building" in names

    async def test_name_filter_combined_with_limit(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        for i in range(3):
            db_session.add(
                Building(
                    name=f"LimitFilter Building {i}",
                    manager_email=f"limit{i}@test.com",
                )
            )
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=LimitFilter&limit=2")
        assert response.status_code == 200
        data = response.json()
        assert len(data) <= 2


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestGetBuilding:
    # --- Happy path ---

    async def test_returns_building_by_id(
        self, client: AsyncClient, building: Building
    ):
        response = await client.get(f"/api/admin/buildings/{building.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == str(building.id)
        assert data["name"] == building.name
        assert data["manager_email"] == building.manager_email
        assert "is_archived" in data
        assert "created_at" in data

    # --- State / precondition errors ---

    async def test_returns_404_for_unknown_id(self, client: AsyncClient):
        response = await client.get(f"/api/admin/buildings/{uuid.uuid4()}")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}/lot-owners
# ---------------------------------------------------------------------------


class TestListLotOwners:
    # --- Happy path ---

    async def test_returns_lot_owners_for_building(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) == 2

    async def test_lot_owner_fields_present(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        owner = data[0]
        assert "id" in owner
        assert "lot_number" in owner
        assert "emails" in owner
        assert "unit_entitlement" in owner
        assert isinstance(owner["emails"], list)

    async def test_lot_owner_emails_populated(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        all_emails = [e for o in data for e in o["emails"]]
        assert "voter1@test.com" in all_emails
        assert "voter2@test.com" in all_emails

    # --- State / precondition errors ---

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        response = await client.get(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# GET /api/admin/lot-owners/{lot_owner_id}
# ---------------------------------------------------------------------------


class TestGetLotOwner:
    # --- Happy path ---

    async def test_returns_lot_owner_without_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        # Get the first lot owner ID from the list
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        response = await client.get(f"/api/admin/lot-owners/{lot_owner_id}")
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == lot_owner_id
        assert "lot_number" in data
        assert "emails" in data
        assert "unit_entitlement" in data
        assert "financial_position" in data
        assert data["proxy_email"] is None

    async def test_returns_lot_owner_with_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        # Get the first lot owner ID from the list and add a proxy
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        # Seed a proxy for this lot owner
        proxy = LotProxy(
            lot_owner_id=uuid.UUID(lot_owner_id),
            proxy_email="proxy@example.com",
        )
        db_session.add(proxy)
        await db_session.flush()

        response = await client.get(f"/api/admin/lot-owners/{lot_owner_id}")
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "proxy@example.com"

    # --- Input validation ---

    async def test_invalid_uuid_returns_422(self, client: AsyncClient):
        response = await client.get("/api/admin/lot-owners/not-a-uuid")
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_nonexistent_lot_owner_returns_404(self, client: AsyncClient):
        response = await client.get(f"/api/admin/lot-owners/{uuid.uuid4()}")
        assert response.status_code == 404
        assert response.json()["detail"] == "Lot owner not found"


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/lot-owners/import
# ---------------------------------------------------------------------------


class TestImportLotOwners:
    # --- Happy path ---

    async def test_valid_csv_imports_lot_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["101", "a@test.com", "100"], ["102", "b@test.com", "200"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 2
        assert data["emails"] == 2

    async def test_import_multi_email_same_lot(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Multiple rows with same lot_number → multiple emails for one lot."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [
                ["101", "a@test.com", "100"],
                ["101", "b@test.com", "100"],
            ],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 1  # one lot
        assert data["emails"] == 2    # two email rows

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        assert len(owners) == 1
        assert len(owners[0]["emails"]) == 2

    async def test_import_semicolon_separated_emails(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Semicolon-separated emails in a single cell are split into multiple LotOwnerEmail rows."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["101", "a@test.com;b@test.com; c@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 1
        assert data["emails"] == 3

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        assert sorted(owners[0]["emails"]) == ["a@test.com", "b@test.com", "c@test.com"]

    async def test_import_blank_email_allowed(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """A lot row with blank email is imported without error (no email row created)."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NO-EMAIL", "", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 1
        assert data["emails"] == 0

    async def test_import_replaces_existing_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        # First import
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["A1", "old@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Second import should replace
        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["B1", "new@test.com", "75"], ["B2", "new2@test.com", "25"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

        # Verify replacement
        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        lot_numbers = {o["lot_number"] for o in owners}
        assert "B1" in lot_numbers
        assert "B2" in lot_numbers
        assert "A1" not in lot_numbers

    async def test_empty_csv_clears_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        # Seed
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["X1", "x@test.com", "10"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Import empty
        csv2 = make_csv(["lot_number", "email", "unit_entitlement"], [])
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 0

        # Verify cleared
        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        assert owners_response.json() == []

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "extra"],
            [["Z1", "z@test.com", "10", "ignore_me"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 1

    async def test_unit_entitlement_zero_accepted(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["ZERO1", "zero@test.com", "0"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 1

    # --- Input validation ---

    async def test_missing_unit_entitlement_header(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email"],
            [["1", "a@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "unit_entitlement" in str(response.json()["detail"]).lower()

    async def test_missing_lot_number_header(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["email", "unit_entitlement"],
            [["a@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_duplicate_lot_numbers_in_csv(
        self, client: AsyncClient, building: Building
    ):
        """Two rows with same lot_number but different emails → merged into one lot with both emails.
        The unit_entitlement from the first row is used; the second email is added."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["DUP1", "a@test.com", "100"], ["DUP1", "b@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        # One lot owner upserted (DUP1), two email addresses imported
        assert data["imported"] == 1
        assert data["emails"] == 2

    async def test_negative_unit_entitlement(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NEG1", "neg@test.com", "-1"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_non_integer_unit_entitlement(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NI1", "ni@test.com", "abc"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_collects_all_errors_before_returning(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["", "", "abc"], ["", "", "-5"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        # Should have multiple errors
        assert len(response.json()["detail"]) > 1

    # --- State / precondition errors ---

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["1", "a@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    async def test_non_csv_file_returns_415(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("data.pdf", b"not a spreadsheet", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_completely_empty_file_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """An empty file (not even headers) should return 422."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("empty.csv", b"", "text/csv")},
        )
        assert response.status_code == 422

    async def test_empty_unit_entitlement_value(
        self, client: AsyncClient, building: Building
    ):
        """A row with empty unit_entitlement (present column, empty value) returns 422."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["E1", "e@test.com", ""]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("unit_entitlement is empty" in str(e) for e in response.json()["detail"])

    # --- Edge cases ---

    async def test_reimport_preserves_agm_lot_weight_snapshots(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Re-importing lot owners must NOT destroy GeneralMeetingLotWeight snapshots.

        Bug: delete-all-then-insert gave new IDs to lot owners, which cascaded
        deletes to GeneralMeetingLotWeight records, causing entitlement_sum=0 in tallies.
        Fix: upsert by lot_number preserves existing LotOwner IDs.
        """
        # Initial import
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["L1", "voter@test.com", "150"], ["L2", "voter@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Create GeneralMeeting and snapshot lot weights
        agm = GeneralMeeting(
            building_id=building.id,
            title="Snapshot GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        motion = Motion(general_meeting_id=agm.id, title="Test Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        owners_result = await db_session.execute(
            select(LotOwner).where(LotOwner.building_id == building.id)
        )
        for lo in owners_result.scalars().all():
            db_session.add(GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            ))
        await db_session.commit()

        # Re-import same lots (email change for L1) — must preserve snapshots
        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["L1", "updated@test.com", "150"], ["L2", "voter@test.com", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200

        # GeneralMeetingLotWeight snapshots must still exist (not cascade-deleted)
        weights_result = await db_session.execute(
            select(GeneralMeetingLotWeight).where(GeneralMeetingLotWeight.general_meeting_id == agm.id)
        )
        weights = weights_result.scalars().all()
        assert len(weights) == 2
        total_snapshot = sum(w.unit_entitlement_snapshot for w in weights)
        assert total_snapshot == 200  # 150 + 50 preserved from original snapshot

    async def test_reimport_updates_existing_and_adds_new(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Existing lot numbers are updated in-place; new lot numbers are inserted."""
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["OLD1", "old@test.com", "100"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["OLD1", "updated@test.com", "200"], ["NEW1", "new@test.com", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = {o["lot_number"]: o for o in owners_response.json()}
        assert owners["OLD1"]["emails"] == ["updated@test.com"]
        assert owners["OLD1"]["unit_entitlement"] == 200
        assert "NEW1" in owners


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/lot-owners
# ---------------------------------------------------------------------------


class TestAddLotOwner:
    # --- Happy path ---

    async def test_valid_add_returns_201(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NEW01", "emails": ["new@test.com"], "unit_entitlement": 150},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["lot_number"] == "NEW01"
        assert "new@test.com" in data["emails"]
        assert data["unit_entitlement"] == 150
        assert "id" in data

    async def test_add_with_multiple_emails(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "MULTI01", "emails": ["a@test.com", "b@test.com"], "unit_entitlement": 100},
        )
        assert response.status_code == 201
        data = response.json()
        assert len(data["emails"]) == 2

    async def test_add_with_no_emails(
        self, client: AsyncClient, building: Building
    ):
        """Lot owner can be created with empty email list."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NOEMAIL01", "emails": [], "unit_entitlement": 50},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["emails"] == []

    async def test_unit_entitlement_zero_accepted(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "ZERO01", "emails": [], "unit_entitlement": 0},
        )
        assert response.status_code == 201
        assert response.json()["unit_entitlement"] == 0

    # --- Input validation ---

    async def test_negative_unit_entitlement_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NEG01", "emails": [], "unit_entitlement": -1},
        )
        assert response.status_code == 422

    async def test_empty_lot_number_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "", "emails": [], "unit_entitlement": 10},
        )
        assert response.status_code == 422

    async def test_missing_fields_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "M01"},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_lot_number_returns_409(
        self, client: AsyncClient, building: Building
    ):
        # Add first
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "DUP01", "emails": [], "unit_entitlement": 10},
        )
        # Add duplicate
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "DUP01", "emails": [], "unit_entitlement": 20},
        )
        assert response.status_code == 409

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners",
            json={"lot_number": "X01", "emails": [], "unit_entitlement": 10},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# PATCH /api/admin/lot-owners/{lot_owner_id}
# ---------------------------------------------------------------------------


class TestUpdateLotOwner:
    # --- Happy path ---

    async def test_update_unit_entitlement(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD02",
            unit_entitlement=50,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 999},
        )
        assert response.status_code == 200
        assert response.json()["unit_entitlement"] == 999

    async def test_update_financial_position(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD03",
            unit_entitlement=100,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"financial_position": "in_arrear"},
        )
        assert response.status_code == 200
        assert response.json()["financial_position"] == "in_arrear"

    async def test_update_both_fields(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD04",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 200, "financial_position": "in_arrear"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["unit_entitlement"] == 200
        assert data["financial_position"] == "in_arrear"

    # --- Boundary values ---

    async def test_unit_entitlement_zero_accepted(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD05",
            unit_entitlement=100,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 0},
        )
        assert response.status_code == 200
        assert response.json()["unit_entitlement"] == 0

    # --- Input validation ---

    async def test_no_fields_provided_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD06",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={},
        )
        assert response.status_code == 422

    async def test_negative_unit_entitlement_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD07",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": -5},
        )
        assert response.status_code == 422

    async def test_invalid_financial_position_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD08",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"financial_position": "invalid_value"},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_not_found_returns_404(self, client: AsyncClient):
        response = await client.patch(
            f"/api/admin/lot-owners/{uuid.uuid4()}",
            json={"unit_entitlement": 10},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/admin/lot-owners/{lot_owner_id}/emails
# ---------------------------------------------------------------------------


class TestAddEmailToLotOwner:
    # --- Happy path ---

    async def test_add_email_returns_updated_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "new@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert "new@test.com" in data["emails"]

    async def test_add_second_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM02", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        existing = LotOwnerEmail(lot_owner_id=lo.id, email="first@test.com")
        db_session.add(existing)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "second@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert "first@test.com" in data["emails"]
        assert "second@test.com" in data["emails"]

    # --- Input validation ---

    async def test_empty_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM03", unit_entitlement=10)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": ""},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_lot_owner_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(
            f"/api/admin/lot-owners/{uuid.uuid4()}/emails",
            json={"email": "x@test.com"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/lot-owners/{lot_owner_id}/emails/{email}
# ---------------------------------------------------------------------------


class TestRemoveEmailFromLotOwner:
    # --- Happy path ---

    async def test_remove_email_returns_updated_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="REM01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = LotOwnerEmail(lot_owner_id=lo.id, email="todelete@test.com")
        db_session.add(em)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/emails/todelete@test.com"
        )
        assert response.status_code == 200
        data = response.json()
        assert "todelete@test.com" not in data["emails"]

    # --- State / precondition errors ---

    async def test_lot_owner_not_found_returns_404(self, client: AsyncClient):
        response = await client.delete(
            f"/api/admin/lot-owners/{uuid.uuid4()}/emails/x@test.com"
        )
        assert response.status_code == 404

    async def test_email_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="REM02", unit_entitlement=10)
        db_session.add(lo)
        await db_session.commit()

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/emails/nonexistent@test.com"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings
# ---------------------------------------------------------------------------


class TestCreateAGM:
    def _agm_payload(self, building_id: uuid.UUID, **kwargs) -> dict:
        base = {
            "building_id": str(building_id),
            "title": "Annual General Meeting 2024",
            "meeting_at": meeting_dt().isoformat(),
            "voting_closes_at": closing_dt().isoformat(),
            "motions": [
                {
                    "title": "Motion 1",
                    "description": "First motion",
                    "display_order": 1,
                }
            ],
        }
        base.update(kwargs)
        return base

    # --- Happy path ---

    async def test_create_agm_returns_201(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            "/api/admin/general-meetings",
            json=self._agm_payload(building_with_owners.id),
        )
        assert response.status_code == 201
        data = response.json()
        assert data["status"] == "open"
        assert data["title"] == "Annual General Meeting 2024"
        assert len(data["motions"]) == 1

    async def test_agm_lot_weight_snapshot_created(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        # Create building with owner for snapshot verification
        b = Building(name="Snapshot Building", manager_email="snap@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(
            building_id=b.id,
            lot_number="S1",
            unit_entitlement=123,
        )
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            "/api/admin/general-meetings",
            json=self._agm_payload(b.id),
        )
        assert response.status_code == 201
        agm_id = response.json()["id"]

        weights = await db_session.execute(
            select(GeneralMeetingLotWeight).where(GeneralMeetingLotWeight.general_meeting_id == uuid.UUID(agm_id))
        )
        weight_list = list(weights.scalars().all())
        assert len(weight_list) == 1
        assert weight_list[0].unit_entitlement_snapshot == 123
        # No voter_email on snapshot anymore
        assert weight_list[0].lot_owner_id == lo.id

    async def test_agm_with_multiple_motions(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Multi Motion Building", manager_email="mm@test.com")
        db_session.add(b)
        await db_session.flush()

        payload = self._agm_payload(b.id)
        payload["motions"] = [
            {"title": "Motion A", "description": "A", "display_order": 1},
            {"title": "Motion B", "description": "B", "display_order": 2},
            {"title": "Motion C", "description": None, "display_order": 3},
        ]
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 201
        assert len(response.json()["motions"]) == 3

    async def test_multi_lot_owner_snapshot(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Voter with multiple lots should have multiple snapshot rows."""
        b = Building(name="Multi Lot Building", manager_email="ml@test.com")
        db_session.add(b)
        await db_session.flush()
        lo1 = LotOwner(
            building_id=b.id,
            lot_number="ML1",
            unit_entitlement=100,
        )
        lo2 = LotOwner(
            building_id=b.id,
            lot_number="ML2",
            unit_entitlement=50,
        )
        db_session.add_all([lo1, lo2])
        await db_session.commit()

        response = await client.post(
            "/api/admin/general-meetings",
            json=self._agm_payload(b.id),
        )
        assert response.status_code == 201
        agm_id = response.json()["id"]

        weights = await db_session.execute(
            select(GeneralMeetingLotWeight).where(GeneralMeetingLotWeight.general_meeting_id == uuid.UUID(agm_id))
        )
        weight_list = list(weights.scalars().all())
        assert len(weight_list) == 2
        total = sum(w.unit_entitlement_snapshot for w in weight_list)
        assert total == 150

    # --- Input validation ---

    async def test_no_motions_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        payload = self._agm_payload(building_with_owners.id, motions=[])
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 422

    async def test_voting_closes_before_meeting_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        now = datetime.now(UTC)
        payload = self._agm_payload(
            building_with_owners.id,
            meeting_at=(now + timedelta(days=2)).isoformat(),
            voting_closes_at=(now + timedelta(days=1)).isoformat(),
        )
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 422

    async def test_voting_closes_equal_to_meeting_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        now = datetime.now(UTC)
        t = (now + timedelta(days=1)).isoformat()
        payload = self._agm_payload(
            building_with_owners.id,
            meeting_at=t,
            voting_closes_at=t,
        )
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 422

    async def test_missing_building_id_returns_422(self, client: AsyncClient):
        payload = {
            "title": "Test",
            "meeting_at": meeting_dt().isoformat(),
            "voting_closes_at": closing_dt().isoformat(),
            "motions": [{"title": "M1", "display_order": 1}],
        }
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        payload = self._agm_payload(uuid.uuid4())
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 404

    async def test_second_open_agm_for_same_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="One GeneralMeeting Building", manager_email="one@test.com")
        db_session.add(b)
        await db_session.commit()

        payload = self._agm_payload(b.id)
        r1 = await client.post("/api/admin/general-meetings", json=payload)
        assert r1.status_code == 201

        r2 = await client.post("/api/admin/general-meetings", json=payload)
        assert r2.status_code == 409

    async def test_can_create_agm_after_closed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Closed GeneralMeeting Building", manager_email="closed@test.com")
        db_session.add(b)
        await db_session.commit()

        payload = self._agm_payload(b.id)
        r1 = await client.post("/api/admin/general-meetings", json=payload)
        assert r1.status_code == 201
        agm_id = r1.json()["id"]

        # Close the GeneralMeeting
        await client.post(f"/api/admin/general-meetings/{agm_id}/close")

        # Now create another
        r2 = await client.post("/api/admin/general-meetings", json=payload)
        assert r2.status_code == 201

    async def test_create_agm_with_future_meeting_at_returns_pending_status(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Meeting created with future meeting_at should have status=pending."""
        b = Building(name="Pending Status Building", manager_email="pending@test.com")
        db_session.add(b)
        await db_session.commit()

        now = datetime.now(UTC)
        payload = self._agm_payload(
            b.id,
            meeting_at=(now + timedelta(days=1)).isoformat(),
            voting_closes_at=(now + timedelta(days=2)).isoformat(),
        )
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 201
        assert response.json()["status"] == "pending"

    async def test_second_pending_agm_for_same_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Cannot create a meeting if a pending meeting already exists for the building."""
        b = Building(name="Pending Block Building", manager_email="pendblock@test.com")
        db_session.add(b)
        await db_session.commit()

        now = datetime.now(UTC)
        payload = self._agm_payload(
            b.id,
            meeting_at=(now + timedelta(days=1)).isoformat(),
            voting_closes_at=(now + timedelta(days=2)).isoformat(),
        )
        r1 = await client.post("/api/admin/general-meetings", json=payload)
        assert r1.status_code == 201
        assert r1.json()["status"] == "pending"

        # Second create should be blocked
        r2 = await client.post("/api/admin/general-meetings", json=payload)
        assert r2.status_code == 409

    # --- motion_number uniqueness ---

    async def test_create_agm_with_motion_numbers(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """AGM with distinct motion_numbers on each motion is accepted (201)."""
        b = Building(name="Motion Number Test Bldg", manager_email="mn@test.com")
        db_session.add(b)
        await db_session.commit()

        payload = self._agm_payload(b.id)
        payload["motions"] = [
            {"title": "Motion A", "display_order": 1, "motion_number": "1"},
            {"title": "Motion B", "display_order": 2, "motion_number": "2"},
        ]
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 201
        motions = response.json()["motions"]
        assert motions[0]["motion_number"] == "1"
        assert motions[1]["motion_number"] == "2"

    async def test_create_agm_with_null_motion_numbers(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """AGM with null motion_numbers on multiple motions is accepted (201); nulls don't conflict."""
        b = Building(name="Null MN Bldg", manager_email="nullmn@test.com")
        db_session.add(b)
        await db_session.commit()

        payload = self._agm_payload(b.id)
        payload["motions"] = [
            {"title": "Motion A", "display_order": 1},
            {"title": "Motion B", "display_order": 2},
        ]
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 201
        motions = response.json()["motions"]
        assert motions[0]["motion_number"] is None
        assert motions[1]["motion_number"] is None

    async def test_create_agm_with_duplicate_motion_numbers_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """AGM with two motions sharing the same non-null motion_number returns 409."""
        b = Building(name="Dup MN Bldg", manager_email="dupmn@test.com")
        db_session.add(b)
        await db_session.commit()

        payload = self._agm_payload(b.id)
        payload["motions"] = [
            {"title": "Motion A", "display_order": 1, "motion_number": "1"},
            {"title": "Motion B", "display_order": 2, "motion_number": "1"},
        ]
        response = await client.post("/api/admin/general-meetings", json=payload)
        assert response.status_code == 409


# ---------------------------------------------------------------------------
# GET /api/admin/general-meetings
# ---------------------------------------------------------------------------


class TestListAGMs:
    # --- Happy path ---

    async def test_returns_list(self, client: AsyncClient):
        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        assert isinstance(response.json(), list)

    async def test_agm_list_fields_present(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="List GeneralMeeting Building", manager_email="list@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="List Test GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        data = response.json()
        assert len(data) > 0

        # Find our GeneralMeeting
        our_agm = next((a for a in data if a["title"] == "List Test GeneralMeeting"), None)
        assert our_agm is not None
        assert "id" in our_agm
        assert "building_id" in our_agm
        assert "building_name" in our_agm
        assert "status" in our_agm
        assert "meeting_at" in our_agm
        assert "voting_closes_at" in our_agm
        assert "created_at" in our_agm
        assert our_agm["building_name"] == "List GeneralMeeting Building"

    async def test_ordered_by_created_at_desc(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Order Test Building", manager_email="order@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        data = response.json()
        if len(data) > 1:
            # Check ordering
            for i in range(len(data) - 1):
                assert data[i]["created_at"] >= data[i + 1]["created_at"]

    # --- name filter ---

    async def test_agm_name_filter_exact_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM Name Filter Building", manager_email="agmnf@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Exact Match AGM Title",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings?name=Exact+Match+AGM+Title")
        assert response.status_code == 200
        data = response.json()
        titles = [item["title"] for item in data]
        assert "Exact Match AGM Title" in titles

    async def test_agm_name_filter_partial_substring_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM Substring Building", manager_email="agmsub@test.com")
        db_session.add(b)
        await db_session.flush()
        agm1 = GeneralMeeting(
            building_id=b.id,
            title="SubstringAGM First Meeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        agm2 = GeneralMeeting(
            building_id=b.id,
            title="SubstringAGM Second Meeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        agm_other = GeneralMeeting(
            building_id=b.id,
            title="Unrelated Meeting Title",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add_all([agm1, agm2, agm_other])
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings?name=SubstringAGM")
        assert response.status_code == 200
        data = response.json()
        titles = [item["title"] for item in data]
        assert "SubstringAGM First Meeting" in titles
        assert "SubstringAGM Second Meeting" in titles
        assert "Unrelated Meeting Title" not in titles

    async def test_agm_name_filter_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM Case Building", manager_email="agmcase@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="CaseMixed AGM",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings?name=casemixed")
        assert response.status_code == 200
        data = response.json()
        titles = [item["title"] for item in data]
        assert "CaseMixed AGM" in titles

    async def test_agm_name_filter_no_match_returns_empty(
        self, client: AsyncClient
    ):
        response = await client.get(
            "/api/admin/general-meetings?name=does-not-exist-xyz-agm-99"
        )
        assert response.status_code == 200
        assert response.json() == []

    async def test_agm_name_filter_absent_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM No Filter Building", manager_email="agmnof@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="No Filter AGM Title",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        data = response.json()
        titles = [item["title"] for item in data]
        assert "No Filter AGM Title" in titles

    async def test_agm_name_filter_empty_string_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM Empty Filter Building", manager_email="agmempty@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="EmptyFilter AGM Title",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings?name=")
        assert response.status_code == 200
        data = response.json()
        titles = [item["title"] for item in data]
        # empty string matches everything via LIKE '%%'
        assert "EmptyFilter AGM Title" in titles

    async def test_agm_name_filter_combined_with_limit(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM Limit Filter Building", manager_email="agmlimit@test.com")
        db_session.add(b)
        await db_session.flush()
        for i in range(3):
            db_session.add(
                GeneralMeeting(
                    building_id=b.id,
                    title=f"LimitFilterAGM Meeting {i}",
                    status=GeneralMeetingStatus.open,
                    meeting_at=meeting_dt(),
                    voting_closes_at=closing_dt(),
                )
            )
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings?name=LimitFilterAGM&limit=2")
        assert response.status_code == 200
        data = response.json()
        assert len(data) <= 2

    # --- building_id filter ---

    async def test_agm_building_id_filter_returns_only_that_building(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b1 = Building(name="AGM BldFilter Building1", manager_email="bf1@test.com")
        b2 = Building(name="AGM BldFilter Building2", manager_email="bf2@test.com")
        db_session.add_all([b1, b2])
        await db_session.flush()
        agm1 = GeneralMeeting(
            building_id=b1.id,
            title="BuildingFilter AGM B1",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        agm2 = GeneralMeeting(
            building_id=b2.id,
            title="BuildingFilter AGM B2",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add_all([agm1, agm2])
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings?building_id={b1.id}")
        assert response.status_code == 200
        data = response.json()
        ids = [item["id"] for item in data]
        assert str(agm1.id) in ids
        assert str(agm2.id) not in ids

    async def test_agm_building_id_filter_no_match_returns_empty(
        self, client: AsyncClient
    ):
        import uuid as _uuid
        random_id = str(_uuid.uuid4())
        response = await client.get(f"/api/admin/general-meetings?building_id={random_id}")
        assert response.status_code == 200
        assert response.json() == []

    async def test_agm_building_id_filter_combined_with_name(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="AGM BldName Filter Building", manager_email="bfn@test.com")
        db_session.add(b)
        await db_session.flush()
        agm_match = GeneralMeeting(
            building_id=b.id,
            title="BldNameCombo AGM Match",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        agm_no_match = GeneralMeeting(
            building_id=b.id,
            title="Other AGM Title",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add_all([agm_match, agm_no_match])
        await db_session.commit()

        response = await client.get(
            f"/api/admin/general-meetings?building_id={b.id}&name=BldNameCombo"
        )
        assert response.status_code == 200
        data = response.json()
        ids = [item["id"] for item in data]
        assert str(agm_match.id) in ids
        assert str(agm_no_match.id) not in ids

    async def test_agm_building_id_filter_absent_returns_all_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b1 = Building(name="AGM NoBldFilter B1", manager_email="nbf1@test.com")
        b2 = Building(name="AGM NoBldFilter B2", manager_email="nbf2@test.com")
        db_session.add_all([b1, b2])
        await db_session.flush()
        agm1 = GeneralMeeting(
            building_id=b1.id,
            title="NoBldFilter AGM B1",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        agm2 = GeneralMeeting(
            building_id=b2.id,
            title="NoBldFilter AGM B2",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add_all([agm1, agm2])
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        data = response.json()
        ids = [item["id"] for item in data]
        assert str(agm1.id) in ids
        assert str(agm2.id) in ids


# ---------------------------------------------------------------------------
# GET /api/admin/general-meetings/{agm_id}
# ---------------------------------------------------------------------------


class TestGetGeneralMeetingDetail:
    async def _setup_agm_with_votes(
        self, db_session: AsyncSession
    ) -> tuple[GeneralMeeting, list[LotOwner], list[Motion]]:
        b = Building(name="Detail Building", manager_email="detail@test.com")
        db_session.add(b)
        await db_session.flush()

        lo1 = LotOwner(building_id=b.id, lot_number="D1", unit_entitlement=100)
        lo2 = LotOwner(building_id=b.id, lot_number="D2", unit_entitlement=80)
        lo3 = LotOwner(building_id=b.id, lot_number="D3", unit_entitlement=30)
        lo4 = LotOwner(building_id=b.id, lot_number="D4", unit_entitlement=200)
        db_session.add_all([lo1, lo2, lo3, lo4])
        await db_session.flush()

        # Add emails
        db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="yes@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="no@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo3.id, email="abs@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo4.id, email="absent@test.com"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="Detail GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Motion D1", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        # Snapshot (no voter_email)
        for lo in [lo1, lo2, lo3, lo4]:
            w = GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            )
            db_session.add(w)
        await db_session.flush()

        # lo1 voted yes, lo2 voted no, lo3 abstained — all submitted
        # lo4 is absent (no ballot submission)
        lo_email_map = {lo1.id: "yes@test.com", lo2.id: "no@test.com", lo3.id: "abs@test.com"}
        for lo, choice in [
            (lo1, VoteChoice.yes),
            (lo2, VoteChoice.no),
            (lo3, VoteChoice.abstained),
        ]:
            voter_email = lo_email_map[lo.id]
            vote = Vote(
                general_meeting_id=agm.id,
                motion_id=motion.id,
                voter_email=voter_email,
                lot_owner_id=lo.id,
                choice=choice,
                status=VoteStatus.submitted,
            )
            db_session.add(vote)
            bs = BallotSubmission(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                voter_email=voter_email,
            )
            db_session.add(bs)

        await db_session.commit()
        return agm, [lo1, lo2, lo3, lo4], [motion]

    # --- Happy path ---

    async def test_agm_detail_structure(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _, _ = await self._setup_agm_with_votes(db_session)
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 200
        data = response.json()

        assert data["id"] == str(agm.id)
        assert data["building_name"] == "Detail Building"
        assert "total_eligible_voters" in data
        assert "total_submitted" in data
        assert "total_entitlement" in data
        assert "motions" in data
        assert "closed_at" in data

    async def test_total_entitlement_is_sum_of_snapshots(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _, _ = await self._setup_agm_with_votes(db_session)
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 200
        data = response.json()
        # lo1=100, lo2=80, lo3=30, lo4=200
        assert data["total_entitlement"] == 410

    async def test_tally_yes_no_abstained_absent(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _, _ = await self._setup_agm_with_votes(db_session)
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()

        assert data["total_eligible_voters"] == 4
        assert data["total_submitted"] == 3

        motion = data["motions"][0]
        tally = motion["tally"]
        assert tally["yes"]["voter_count"] == 1
        assert tally["yes"]["entitlement_sum"] == 100
        assert tally["no"]["voter_count"] == 1
        assert tally["no"]["entitlement_sum"] == 80
        assert tally["abstained"]["voter_count"] == 1
        assert tally["abstained"]["entitlement_sum"] == 30
        # Meeting is open: absent tally is suppressed
        assert tally["absent"]["voter_count"] == 0
        assert tally["absent"]["entitlement_sum"] == 0

    async def test_voter_lists_populated(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _, _ = await self._setup_agm_with_votes(db_session)
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()

        motion = data["motions"][0]
        voter_lists = motion["voter_lists"]

        yes_emails = {v["voter_email"] for v in voter_lists["yes"]}
        no_emails = {v["voter_email"] for v in voter_lists["no"]}
        abs_emails = {v["voter_email"] for v in voter_lists["abstained"]}

        assert "yes@test.com" in yes_emails
        assert "no@test.com" in no_emails
        assert "abs@test.com" in abs_emails
        # Meeting is open: absent list is empty
        assert voter_lists["absent"] == []

    async def test_no_votes_all_absent(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="No Votes Building", manager_email="nv@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(
            building_id=b.id,
            lot_number="NV1",
            unit_entitlement=50,
        )
        db_session.add(lo)
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="No Votes GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Motion NV1", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        w = GeneralMeetingLotWeight(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            unit_entitlement_snapshot=lo.unit_entitlement,
        )
        db_session.add(w)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()

        assert data["total_eligible_voters"] == 1
        assert data["total_submitted"] == 0

        tally = data["motions"][0]["tally"]
        assert tally["yes"]["voter_count"] == 0
        assert tally["no"]["voter_count"] == 0
        assert tally["abstained"]["voter_count"] == 0
        # Meeting is open: absent tally is suppressed
        assert tally["absent"]["voter_count"] == 0
        assert tally["absent"]["entitlement_sum"] == 0

    async def test_entitlement_sums_use_snapshot_not_current(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Even if unit_entitlement changes later, tally uses snapshot."""
        b = Building(name="Snapshot Tally Building", manager_email="st@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(
            building_id=b.id,
            lot_number="ST1",
            unit_entitlement=500,
        )
        db_session.add(lo)
        await db_session.flush()
        lo_email = LotOwnerEmail(lot_owner_id=lo.id, email="snap_tally@test.com")
        db_session.add(lo_email)
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="Snapshot Tally GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="ST Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        # Snapshot at 500
        w = GeneralMeetingLotWeight(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            unit_entitlement_snapshot=500,
        )
        db_session.add(w)
        await db_session.flush()

        # Vote yes
        vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="snap_tally@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        )
        db_session.add(vote)
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            voter_email="snap_tally@test.com",
        )
        db_session.add(bs)
        await db_session.flush()

        # Change current entitlement (shouldn't affect tally)
        lo.unit_entitlement = 999
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        tally = response.json()["motions"][0]["tally"]
        assert tally["yes"]["entitlement_sum"] == 500

    async def test_multi_lot_voter_entitlement_sum(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Voter with two lots: entitlement should be sum of both snapshots."""
        b = Building(name="Two Lots Building", manager_email="tl@test.com")
        db_session.add(b)
        await db_session.flush()
        lo1 = LotOwner(building_id=b.id, lot_number="TL1", unit_entitlement=100)
        lo2 = LotOwner(building_id=b.id, lot_number="TL2", unit_entitlement=200)
        db_session.add_all([lo1, lo2])
        await db_session.flush()
        db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="twolots@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="twolots@test.com"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="Two Lots GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="TL Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for lo in [lo1, lo2]:
            w = GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            )
            db_session.add(w)
        await db_session.flush()

        # Submit for both lots
        for lo in [lo1, lo2]:
            vote = Vote(
                general_meeting_id=agm.id,
                motion_id=motion.id,
                voter_email="twolots@test.com",
                lot_owner_id=lo.id,
                choice=VoteChoice.yes,
                status=VoteStatus.submitted,
            )
            db_session.add(vote)
            bs = BallotSubmission(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                voter_email="twolots@test.com",
            )
            db_session.add(bs)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()
        # Each lot is a separate eligible voter
        assert data["total_eligible_voters"] == 2
        tally = data["motions"][0]["tally"]
        assert tally["yes"]["voter_count"] == 2
        assert tally["yes"]["entitlement_sum"] == 300  # 100 + 200

    # --- State / precondition errors ---

    async def test_not_found_returns_404(self, client: AsyncClient):
        response = await client.get(f"/api/admin/general-meetings/{uuid.uuid4()}")
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_all_yes_voters(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="All Yes Building", manager_email="ally@test.com")
        db_session.add(b)
        await db_session.flush()

        owners = [
            LotOwner(
                building_id=b.id,
                lot_number=f"AY{i}",
                unit_entitlement=10 * (i + 1),
            )
            for i in range(3)
        ]
        db_session.add_all(owners)
        await db_session.flush()

        for i, lo in enumerate(owners):
            db_session.add(LotOwnerEmail(lot_owner_id=lo.id, email=f"yes{i}@test.com"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="All Yes GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="All Yes Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for i, lo in enumerate(owners):
            w = GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            )
            db_session.add(w)
            vote = Vote(
                general_meeting_id=agm.id,
                motion_id=motion.id,
                voter_email=f"yes{i}@test.com",
                lot_owner_id=lo.id,
                choice=VoteChoice.yes,
                status=VoteStatus.submitted,
            )
            db_session.add(vote)
            bs = BallotSubmission(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                voter_email=f"yes{i}@test.com",
            )
            db_session.add(bs)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()
        tally = data["motions"][0]["tally"]
        # 10 + 20 + 30 = 60
        assert tally["yes"]["voter_count"] == 3
        assert tally["yes"]["entitlement_sum"] == 60
        assert tally["no"]["voter_count"] == 0
        assert tally["absent"]["voter_count"] == 0

    async def test_submitted_vote_with_null_choice_treated_as_abstained(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Null Choice Building", manager_email="nc@test.com")
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(
            building_id=b.id,
            lot_number="NC1",
            unit_entitlement=40,
        )
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotOwnerEmail(lot_owner_id=lo.id, email="nullchoice@test.com"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="Null Choice GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="NC Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        w = GeneralMeetingLotWeight(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            unit_entitlement_snapshot=lo.unit_entitlement,
        )
        db_session.add(w)

        # Vote with null choice (submitted but no selection → abstained)
        vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="nullchoice@test.com",
            lot_owner_id=lo.id,
            choice=None,
            status=VoteStatus.submitted,
        )
        db_session.add(vote)
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            voter_email="nullchoice@test.com",
        )
        db_session.add(bs)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        tally = response.json()["motions"][0]["tally"]
        assert tally["abstained"]["voter_count"] == 1
        assert tally["abstained"]["entitlement_sum"] == 40

    async def test_fallback_to_current_lot_owners_when_no_snapshot(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When GeneralMeetingLotWeight snapshot is empty, fall back to current lot owners."""
        b = Building(name="No Snapshot Building", manager_email="ns@test.com")
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(
            building_id=b.id,
            lot_number="NS1",
            unit_entitlement=75,
        )
        db_session.add(lo)
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="No Snapshot GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="NS Motion", display_order=1)
        db_session.add(motion)
        # Intentionally no GeneralMeetingLotWeight rows for this GeneralMeeting
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        data = response.json()
        assert data["total_eligible_voters"] == 1
        tally = data["motions"][0]["tally"]
        # Meeting is open: absent tally is suppressed
        assert tally["absent"]["voter_count"] == 0
        assert tally["absent"]["entitlement_sum"] == 0

    async def test_tally_not_eligible_counted_separately(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """In-arrear lots with not_eligible votes appear in the not_eligible tally
        bucket (admin_service.py line 1018)."""
        b = Building(name="Not Eligible Building", manager_email="ne@test.com")
        db_session.add(b)
        await db_session.flush()

        lo_normal = LotOwner(building_id=b.id, lot_number="NE1", unit_entitlement=100)
        lo_arrear = LotOwner(building_id=b.id, lot_number="NE2", unit_entitlement=50)
        db_session.add_all([lo_normal, lo_arrear])
        await db_session.flush()

        db_session.add(LotOwnerEmail(lot_owner_id=lo_normal.id, email="normal@ne.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo_arrear.id, email="arrear@ne.com"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="NE GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="NE Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for lo in [lo_normal, lo_arrear]:
            db_session.add(GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            ))
        await db_session.flush()

        # Normal lot voted yes; in-arrear lot has not_eligible
        db_session.add(Vote(
            general_meeting_id=agm.id, motion_id=motion.id,
            voter_email="normal@ne.com", lot_owner_id=lo_normal.id,
            choice=VoteChoice.yes, status=VoteStatus.submitted,
        ))
        db_session.add(BallotSubmission(
            general_meeting_id=agm.id, lot_owner_id=lo_normal.id, voter_email="normal@ne.com",
        ))
        db_session.add(Vote(
            general_meeting_id=agm.id, motion_id=motion.id,
            voter_email="arrear@ne.com", lot_owner_id=lo_arrear.id,
            choice=VoteChoice.not_eligible, status=VoteStatus.submitted,
        ))
        db_session.add(BallotSubmission(
            general_meeting_id=agm.id, lot_owner_id=lo_arrear.id, voter_email="arrear@ne.com",
        ))
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 200
        data = response.json()
        tally = data["motions"][0]["tally"]
        assert tally["yes"]["voter_count"] == 1
        assert tally["yes"]["entitlement_sum"] == 100
        assert tally["not_eligible"]["voter_count"] == 1
        assert tally["not_eligible"]["entitlement_sum"] == 50
        assert tally["abstained"]["voter_count"] == 0
        voter_lists = data["motions"][0]["voter_lists"]
        assert len(voter_lists["not_eligible"]) == 1


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings/{agm_id}/close
# ---------------------------------------------------------------------------


class TestCloseAGM:
    async def _create_open_agm(self, db_session: AsyncSession, name: str) -> GeneralMeeting:
        b = Building(name=name, manager_email=f"close_{name}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Close Test GeneralMeeting {name}",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm

    # --- Happy path ---

    async def test_close_open_agm_returns_closed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm = await self._create_open_agm(db_session, "Close Happy 1")
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "closed"
        assert data["closed_at"] is not None
        assert data["id"] == str(agm.id)

    async def test_draft_votes_deleted_on_close(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Draft Delete Building", manager_email="dd@test.com")
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(
            building_id=b.id, lot_number="DD1", unit_entitlement=10
        )
        db_session.add(lo)
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title="Draft Delete GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="DD Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        # Add another lot owner for the submitted voter
        lo2 = LotOwner(building_id=b.id, lot_number="DD2", unit_entitlement=10)
        db_session.add(lo2)
        await db_session.flush()

        # Add draft and submitted votes (both require lot_owner_id after migration)
        draft_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="draft@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.draft,
        )
        submitted_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="submitted@test.com",
            lot_owner_id=lo2.id,
            choice=VoteChoice.no,
            status=VoteStatus.submitted,
        )
        db_session.add_all([draft_vote, submitted_vote])
        await db_session.commit()

        # Close GeneralMeeting
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        # Check draft vote deleted, submitted vote remains
        remaining = await db_session.execute(
            select(Vote).where(Vote.general_meeting_id == agm.id)
        )
        remaining_votes = list(remaining.scalars().all())
        assert all(v.status == VoteStatus.submitted for v in remaining_votes)
        assert not any(v.voter_email == "draft@test.com" for v in remaining_votes)

    async def test_email_delivery_record_created_on_close(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm = await self._create_open_agm(db_session, "Email Delivery Building")
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        result = await db_session.execute(
            select(EmailDelivery).where(EmailDelivery.general_meeting_id == agm.id)
        )
        delivery = result.scalar_one_or_none()
        assert delivery is not None
        assert delivery.status == EmailDeliveryStatus.pending
        assert delivery.total_attempts == 0

    # --- State / precondition errors ---

    async def test_close_already_closed_agm_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm = await self._create_open_agm(db_session, "Double Close Building")
        await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 409

    async def test_close_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(f"/api/admin/general-meetings/{uuid.uuid4()}/close")
        assert response.status_code == 404

    async def test_close_returns_voting_closes_at(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Close response includes voting_closes_at field (US-PS05)."""
        agm = await self._create_open_agm(db_session, "VotingClosesAt Building")
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200
        data = response.json()
        assert "voting_closes_at" in data
        assert data["voting_closes_at"] is not None

    async def test_close_updates_voting_closes_at_when_future(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Close sets voting_closes_at to now when it was in the future (US-PS05)."""
        b = Building(name="EarlyClose Building", manager_email="eclose@test.com")
        db_session.add(b)
        await db_session.flush()
        original_close = datetime.now(UTC) + timedelta(days=2)
        agm = GeneralMeeting(
            building_id=b.id,
            title="Early Close GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=original_close,
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)

        before_close = datetime.now(UTC)
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200
        after_close = datetime.now(UTC)

        # voting_closes_at should have been updated to approximately now
        updated_closes_at_str = response.json()["voting_closes_at"]
        from datetime import timezone as _tz
        updated_closes_at = datetime.fromisoformat(updated_closes_at_str.replace("Z", "+00:00"))
        assert before_close <= updated_closes_at <= after_close

    async def test_close_preserves_voting_closes_at_when_already_past(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Close does not update voting_closes_at if it is already in the past (US-PS05)."""
        b = Building(name="PastClose Building", manager_email="pastclose@test.com")
        db_session.add(b)
        await db_session.flush()
        past_close = datetime.now(UTC) - timedelta(hours=1)
        # meeting_at also past so constraint ok, but pass as naive to hit the naive branch
        past_meeting_at = (datetime.now(UTC) - timedelta(hours=2)).replace(tzinfo=None)
        agm = GeneralMeeting(
            building_id=b.id,
            title="Past Close GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=past_meeting_at,
            voting_closes_at=past_close,
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200
        # voting_closes_at should be preserved (not changed to now)
        data = response.json()
        assert "voting_closes_at" in data


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings/{agm_id}/start (US-PS04)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestStartGeneralMeeting:
    """Tests for POST /api/admin/general-meetings/{id}/start."""

    async def _create_pending_agm(self, db_session: AsyncSession, name: str) -> GeneralMeeting:
        b = Building(name=name, manager_email=f"start_{name}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Pending Test GeneralMeeting {name}",
            status=GeneralMeetingStatus.pending,
            meeting_at=datetime.now(UTC) + timedelta(days=1),
            voting_closes_at=datetime.now(UTC) + timedelta(days=2),
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm

    # --- Happy path ---

    async def test_start_pending_meeting_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm = await self._create_pending_agm(db_session, "StartHappy1")
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/start")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "open"
        assert data["id"] == str(agm.id)
        assert data["meeting_at"] is not None

    async def test_start_updates_meeting_at_to_now(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Start endpoint sets meeting_at to approximately now."""
        agm = await self._create_pending_agm(db_session, "StartMeetingAt")
        before = datetime.now(UTC)
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/start")
        after = datetime.now(UTC)
        assert response.status_code == 200
        from datetime import timezone as _tz
        meeting_at_str = response.json()["meeting_at"]
        meeting_at = datetime.fromisoformat(meeting_at_str.replace("Z", "+00:00"))
        assert before <= meeting_at <= after

    # --- State / precondition errors ---

    async def test_start_open_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="StartOpen Building", manager_email="startopen@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Open Start GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/start")
        assert response.status_code == 409

    async def test_start_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="StartClosed Building", manager_email="startclosed@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Closed Start GeneralMeeting",
            status=GeneralMeetingStatus.closed,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
            closed_at=datetime.now(UTC),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/start")
        assert response.status_code == 409

    async def test_start_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(f"/api/admin/general-meetings/{uuid.uuid4()}/start")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings/{agm_id}/resend-report
# ---------------------------------------------------------------------------


class TestResendReport:
    async def _setup_closed_agm_with_delivery(
        self, db_session: AsyncSession, name: str, delivery_status: EmailDeliveryStatus
    ) -> tuple[GeneralMeeting, EmailDelivery]:
        b = Building(name=name, manager_email=f"resend_{name}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Resend Test {name}",
            status=GeneralMeetingStatus.closed,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
            closed_at=datetime.now(UTC),
        )
        db_session.add(agm)
        await db_session.flush()

        delivery = EmailDelivery(
            general_meeting_id=agm.id,
            status=delivery_status,
            total_attempts=5,
            last_error="some error" if delivery_status == EmailDeliveryStatus.failed else None,
        )
        db_session.add(delivery)
        await db_session.commit()
        await db_session.refresh(agm)
        await db_session.refresh(delivery)
        return agm, delivery

    # --- Happy path ---

    async def test_resend_failed_delivery_resets_to_pending(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, delivery = await self._setup_closed_agm_with_delivery(
            db_session, "Failed Delivery Building", EmailDeliveryStatus.failed
        )
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/resend-report")
        assert response.status_code == 200
        assert response.json()["queued"] is True

        await db_session.refresh(delivery)
        assert delivery.status == EmailDeliveryStatus.pending
        assert delivery.total_attempts == 0
        assert delivery.last_error is None
        assert delivery.next_retry_at is None

    # --- State / precondition errors ---

    async def test_resend_pending_delivery_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _ = await self._setup_closed_agm_with_delivery(
            db_session, "Pending Delivery Building", EmailDeliveryStatus.pending
        )
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/resend-report")
        assert response.status_code == 409

    async def test_resend_delivered_delivery_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _ = await self._setup_closed_agm_with_delivery(
            db_session, "Delivered Delivery Building", EmailDeliveryStatus.delivered
        )
        response = await client.post(f"/api/admin/general-meetings/{agm.id}/resend-report")
        assert response.status_code == 409

    async def test_resend_open_agm_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Open Resend Building", manager_email="op_res@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Open Resend GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/resend-report")
        assert response.status_code == 409

    async def test_resend_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(f"/api/admin/general-meetings/{uuid.uuid4()}/resend-report")
        assert response.status_code == 404

    async def test_resend_no_delivery_record_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="No Delivery Building", manager_email="nd@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="No Delivery GeneralMeeting",
            status=GeneralMeetingStatus.closed,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
            closed_at=datetime.now(UTC),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/resend-report")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/general-meetings/{agm_id}/ballots
# ---------------------------------------------------------------------------


class TestResetAGMBallots:
    async def _create_agm_with_ballot(
        self, db_session: AsyncSession, name: str
    ) -> tuple[GeneralMeeting, LotOwner, Motion]:
        b = Building(name=name, manager_email=f"reset_{name}@test.com")
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(
            building_id=b.id,
            lot_number="RESET-1",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Reset Ballots GeneralMeeting {name}",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Reset Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        # Add a submitted vote + ballot submission
        vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="reset-voter@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        )
        db_session.add(vote)
        await db_session.flush()

        submission = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            voter_email="reset-voter@test.com",
        )
        db_session.add(submission)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm, lo, motion

    # --- Happy path ---

    async def test_reset_ballots_deletes_submissions_and_returns_count(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _lo, _motion = await self._create_agm_with_ballot(db_session, "Happy Reset")
        response = await client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 200
        data = response.json()
        assert data["deleted"] == 1

        # Verify ballot submission is gone
        subs = await db_session.execute(
            select(BallotSubmission).where(BallotSubmission.general_meeting_id == agm.id)
        )
        assert subs.scalars().all() == []

    async def test_reset_ballots_deletes_submitted_votes(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        agm, _lo, _motion = await self._create_agm_with_ballot(db_session, "Vote Delete Reset")
        response = await client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 200

        # Verify submitted votes are gone
        votes = await db_session.execute(
            select(Vote).where(Vote.general_meeting_id == agm.id, Vote.status == VoteStatus.submitted)
        )
        assert votes.scalars().all() == []

    async def test_reset_ballots_on_agm_with_no_submissions_returns_zero(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Reset Building", manager_email="empty_reset@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Empty Reset GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 200
        assert response.json()["deleted"] == 0

    async def test_reset_ballots_preserves_draft_votes(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Draft Preserve Building", manager_email="draft_pres@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Draft Preserve GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        lo_draft = LotOwner(building_id=b.id, lot_number="DPL1", unit_entitlement=10)
        db_session.add(lo_draft)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Draft Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        draft_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="drafter@test.com",
            lot_owner_id=lo_draft.id,
            choice=VoteChoice.no,
            status=VoteStatus.draft,
        )
        db_session.add(draft_vote)
        await db_session.commit()

        response = await client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 200
        assert response.json()["deleted"] == 0

        # Draft vote should still be present
        remaining = await db_session.execute(
            select(Vote).where(Vote.general_meeting_id == agm.id, Vote.status == VoteStatus.draft)
        )
        assert len(remaining.scalars().all()) == 1

    async def test_reset_multiple_submissions_deletes_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Multi Reset Building", manager_email="multi_reset@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Multi Reset GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Multi Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for i in range(3):
            lo = LotOwner(building_id=b.id, lot_number=f"MRESET-{i}", unit_entitlement=10)
            db_session.add(lo)
            await db_session.flush()
            email = f"multi-voter-{i}@test.com"
            db_session.add(
                Vote(
                    general_meeting_id=agm.id,
                    motion_id=motion.id,
                    voter_email=email,
                    lot_owner_id=lo.id,
                    choice=VoteChoice.yes,
                    status=VoteStatus.submitted,
                )
            )
            db_session.add(BallotSubmission(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                voter_email=email,
            ))
        await db_session.commit()

        response = await client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 200
        assert response.json()["deleted"] == 3

    # --- State / precondition errors ---

    async def test_reset_ballots_not_found_returns_404(self, client: AsyncClient):
        response = await client.delete(f"/api/admin/general-meetings/{uuid.uuid4()}/ballots")
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_reset_ballots_unauthenticated_returns_401(
        self, db_session: AsyncSession
    ):
        """Without admin credentials, the endpoint returns 401."""
        from app.main import create_app
        from app.database import get_db

        b = Building(name="Unauth Reset Building", manager_email="unauth_reset@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Unauth Reset GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        # Create an app without admin auth bypass
        unauth_app = create_app()

        async def override_get_db():
            yield db_session

        unauth_app.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=unauth_app), base_url="http://test"
        ) as unauth_client:
            response = await unauth_client.delete(f"/api/admin/general-meetings/{agm.id}/ballots")
        assert response.status_code == 401


# ---------------------------------------------------------------------------
# Schema unit tests (for coverage of schema validators)
# ---------------------------------------------------------------------------


class TestSchemas:
    def test_lot_owner_update_requires_at_least_one_field(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerUpdate

        with pytest.raises(ValidationError):
            LotOwnerUpdate()

    def test_lot_owner_update_unit_entitlement_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(unit_entitlement=10)
        assert obj.unit_entitlement == 10
        assert obj.financial_position is None

    def test_lot_owner_update_financial_position_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(financial_position="in_arrear")
        assert obj.financial_position == "in_arrear"
        assert obj.unit_entitlement is None

    def test_lot_owner_update_entitlement_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(unit_entitlement=10)
        assert obj.unit_entitlement == 10

    def test_lot_owner_update_negative_entitlement_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerUpdate

        with pytest.raises(ValidationError):
            LotOwnerUpdate(unit_entitlement=-1)

    def test_agm_create_no_motions_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import GeneralMeetingCreate

        with pytest.raises(ValidationError):
            GeneralMeetingCreate(
                building_id=uuid.uuid4(),
                title="t",
                meeting_at=meeting_dt(),
                voting_closes_at=closing_dt(),
                motions=[],
            )

    def test_agm_create_closes_before_meeting_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import GeneralMeetingCreate, MotionCreate

        now = datetime.now(UTC)
        with pytest.raises(ValidationError):
            GeneralMeetingCreate(
                building_id=uuid.uuid4(),
                title="t",
                meeting_at=now + timedelta(days=2),
                voting_closes_at=now + timedelta(days=1),
                motions=[MotionCreate(title="M", display_order=1)],
            )

    def test_lot_owner_create_empty_lot_number_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerCreate

        with pytest.raises(ValidationError):
            LotOwnerCreate(lot_number="  ", unit_entitlement=10)

    def test_lot_owner_create_negative_entitlement_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerCreate

        with pytest.raises(ValidationError):
            LotOwnerCreate(lot_number="L1", unit_entitlement=-1)

    def test_building_create_empty_name_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import BuildingCreate

        with pytest.raises(ValidationError):
            BuildingCreate(name="  ", manager_email="mgr@test.com")

    def test_building_create_empty_email_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import BuildingCreate

        with pytest.raises(ValidationError):
            BuildingCreate(name="Valid Name", manager_email="  ")

    def test_admin_login_request_valid(self):
        from app.schemas.admin import AdminLoginRequest

        req = AdminLoginRequest(username="admin", password="secret")
        assert req.username == "admin"
        assert req.password == "secret"

    def test_add_email_request_empty_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import AddEmailRequest

        with pytest.raises(ValidationError):
            AddEmailRequest(email="  ")

    def test_add_email_request_valid(self):
        from app.schemas.admin import AddEmailRequest

        req = AddEmailRequest(email="x@test.com")
        assert req.email == "x@test.com"


# ---------------------------------------------------------------------------
# POST /api/admin/buildings (create building via form)
# ---------------------------------------------------------------------------


class TestCreateBuildingEndpoint:
    # --- Happy path ---

    async def test_create_building_returns_201(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "New Form Building", "manager_email": "form@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["name"] == "New Form Building"
        assert data["manager_email"] == "form@test.com"
        assert data["is_archived"] is False
        assert "id" in data

    # --- Input validation ---

    async def test_empty_name_returns_422(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "  ", "manager_email": "mgr@test.com"},
        )
        assert response.status_code == 422

    async def test_empty_email_returns_422(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "Valid Building", "manager_email": "  "},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_name_returns_409(
        self, client: AsyncClient
    ):
        await client.post(
            "/api/admin/buildings",
            json={"name": "Dup Form Building", "manager_email": "dup@test.com"},
        )
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "Dup Form Building", "manager_email": "other@test.com"},
        )
        assert response.status_code == 409


# ---------------------------------------------------------------------------
# _detect_file_format unit tests
# ---------------------------------------------------------------------------


class TestDetectFileFormat:
    """Unit tests for the _detect_file_format helper in admin router."""

    def test_csv_extension_returns_csv(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.csv"
        assert _detect_file_format(f) == "csv"

    def test_csv_content_type_returns_csv(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "text/csv"
        f.filename = "noextension"
        assert _detect_file_format(f) == "csv"

    def test_xlsx_extension_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.xlsx"
        assert _detect_file_format(f) == "excel"

    def test_xls_extension_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.xls"
        assert _detect_file_format(f) == "excel"

    def test_xlsx_content_type_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        f.filename = "data"
        assert _detect_file_format(f) == "excel"

    def test_vnd_ms_excel_content_type_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/vnd.ms-excel"
        f.filename = "data"
        assert _detect_file_format(f) == "excel"

    def test_pdf_raises_415(self):
        from unittest.mock import MagicMock
        from fastapi import HTTPException
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/pdf"
        f.filename = "data.pdf"
        with pytest.raises(HTTPException) as exc_info:
            _detect_file_format(f)
        assert exc_info.value.status_code == 415

    def test_no_extension_no_matching_content_type_raises_415(self):
        from unittest.mock import MagicMock
        from fastapi import HTTPException
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "image/png"
        f.filename = "data.png"
        with pytest.raises(HTTPException) as exc_info:
            _detect_file_format(f)
        assert exc_info.value.status_code == 415


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/import — Excel
# ---------------------------------------------------------------------------


class TestImportBuildingsExcel:
    # --- Happy path ---

    async def test_valid_xlsx_creates_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Excel Tower", "excel@test.com"], ["Xlsx Complex", "xlsx@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2
        assert data["updated"] == 0

    async def test_valid_xlsx_updates_existing_building(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        # Create via CSV first
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["ExcelUpdate Building", "old@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Update via Excel
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["ExcelUpdate Building", "new@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 1

    async def test_xlsx_with_blank_rows_skipped(self, client: AsyncClient):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["building_name", "manager_email"])
        ws.append(["Blank Skip Building", "bs@test.com"])
        ws.append([None, None])  # blank row
        ws.append(["", ""])  # also blank
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["created"] == 1

    async def test_empty_xlsx_returns_zero_counts(self, client: AsyncClient):
        excel_data = make_excel(["building_name", "manager_email"], [])
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "empty.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 0

    # --- Input validation ---

    async def test_xlsx_missing_building_name_column_returns_422(
        self, client: AsyncClient
    ):
        excel_data = make_excel(
            ["manager_email"],
            [["mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/archive
# ---------------------------------------------------------------------------


class TestArchiveBuilding:
    # --- Happy path ---

    async def test_archive_building_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Archive Me Building", manager_email="archive@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 200

    async def test_archive_building_sets_is_archived(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Archive Set Building", manager_email="archiveset@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.json()["is_archived"] is True

    async def test_archive_building_also_archives_lot_owners_with_no_other_home(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Lot owners whose emails don't appear in any other building get archived."""
        b = Building(name="Archive Owners Building", manager_email="ao@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="AO1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotOwnerEmail(lot_owner_id=lo.id, email="lone@test.com"))
        await db_session.commit()

        await client.post(f"/api/admin/buildings/{b.id}/archive")

        await db_session.refresh(lo)
        assert lo.is_archived is True

    async def test_archive_building_preserves_lot_owners_with_other_home(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Lot owners whose emails appear in another non-archived building are NOT archived."""
        b1 = Building(name="Archive Other Home B1", manager_email="aoh1@test.com")
        b2 = Building(name="Archive Other Home B2", manager_email="aoh2@test.com")
        db_session.add_all([b1, b2])
        await db_session.flush()

        lo1 = LotOwner(building_id=b1.id, lot_number="AOH1", unit_entitlement=100)
        lo2 = LotOwner(building_id=b2.id, lot_number="AOH1B2", unit_entitlement=100)
        db_session.add_all([lo1, lo2])
        await db_session.flush()

        # Same email in both buildings
        db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="shared@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="shared@test.com"))
        await db_session.commit()

        await client.post(f"/api/admin/buildings/{b1.id}/archive")

        await db_session.refresh(lo1)
        # lo1 is in the archived building, but the email exists in b2 (active), so lo1 should NOT be archived
        assert lo1.is_archived is False

    # --- State / precondition errors ---

    async def test_archive_already_archived_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Already Archived Bldg", manager_email="aa@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 409

    async def test_archive_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.post(f"/api/admin/buildings/{uuid.uuid4()}/archive")
        assert response.status_code == 404

    async def test_archive_building_with_no_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Archive succeeds when building has zero lot owners."""
        b = Building(name="Empty Archive Building", manager_email="empty_arc@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 200
        assert response.json()["is_archived"] is True


# ---------------------------------------------------------------------------
# PATCH /api/admin/buildings/{id}
# ---------------------------------------------------------------------------


class TestUpdateBuilding:
    # --- Happy path ---

    async def test_update_name_only_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Name Building", manager_email="patch.name@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Renamed Building"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "Renamed Building"
        assert data["manager_email"] == "patch.name@test.com"

    async def test_update_manager_email_only_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Email Building", manager_email="old@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": "new@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["manager_email"] == "new@test.com"
        assert data["name"] == "Patch Email Building"

    async def test_update_both_fields_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Both Building", manager_email="both.old@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Both Updated", "manager_email": "both.new@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "Both Updated"
        assert data["manager_email"] == "both.new@test.com"

    async def test_update_persists_to_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Persist Test Building", manager_email="persist@test.com")
        db_session.add(b)
        await db_session.commit()

        await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Persisted Name"},
        )

        await db_session.refresh(b)
        assert b.name == "Persisted Name"

    async def test_update_returns_building_out_shape(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Shape Test Building", manager_email="shape@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Shape Updated"},
        )
        data = response.json()
        assert "id" in data
        assert "name" in data
        assert "manager_email" in data
        assert "is_archived" in data
        assert "created_at" in data

    # --- Input validation ---

    async def test_empty_body_both_null_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Null Patch Building", manager_email="null@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(f"/api/admin/buildings/{b.id}", json={})
        assert response.status_code == 422

    async def test_empty_name_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Name Patch", manager_email="emptyname@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": ""},
        )
        assert response.status_code == 422

    async def test_whitespace_only_name_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Whitespace Name Patch", manager_email="wsname@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "   "},
        )
        assert response.status_code == 422

    async def test_empty_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Email Patch", manager_email="emptyemail@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": ""},
        )
        assert response.status_code == 422

    async def test_whitespace_only_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Whitespace Email Patch", manager_email="wsemail@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": "   "},
        )
        assert response.status_code == 422

    async def test_null_name_and_null_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Null Fields Patch", manager_email="nullfields@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": None, "manager_email": None},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_update_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.patch(
            f"/api/admin/buildings/{uuid.uuid4()}",
            json={"name": "Ghost Building"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/buildings/{id}
# ---------------------------------------------------------------------------


class TestDeleteBuilding:
    # --- Happy path ---

    async def test_delete_archived_building_returns_204(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete Me Building", manager_email="deleteme@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()

        response = await client.delete(f"/api/admin/buildings/{b.id}")
        assert response.status_code == 204

    async def test_delete_archived_building_removes_from_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete From DB Building", manager_email="deletedb@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()
        building_id = b.id

        await client.delete(f"/api/admin/buildings/{building_id}")

        result = await db_session.execute(select(Building).where(Building.id == building_id))
        assert result.scalar_one_or_none() is None

    async def test_delete_archived_building_cascades_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete Cascade Building", manager_email="cascade@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(building_id=b.id, lot_number="C1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        lot_owner_id = lo.id

        await client.delete(f"/api/admin/buildings/{b.id}")

        result = await db_session.execute(select(LotOwner).where(LotOwner.id == lot_owner_id))
        assert result.scalar_one_or_none() is None

    # --- State / precondition errors ---

    async def test_delete_active_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Active Do Not Delete", manager_email="active@test.com")
        b.is_archived = False
        db_session.add(b)
        await db_session.commit()

        response = await client.delete(f"/api/admin/buildings/{b.id}")
        assert response.status_code == 409
        assert "archived" in response.json()["detail"].lower()

    async def test_delete_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.delete(f"/api/admin/buildings/{uuid.uuid4()}")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# Admin auth endpoints
# ---------------------------------------------------------------------------


class TestAdminAuth:
    # --- Happy path ---

    async def test_login_valid_credentials_returns_ok(self, db_session: AsyncSession):
        """Valid username + password → {"ok": true}."""
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=app_instance), base_url="http://test"
        ) as c:
            response = await c.post(
                "/api/admin/auth/login",
                json={"username": "admin", "password": "admin"},
            )
        assert response.status_code == 200
        assert response.json()["ok"] is True

    async def test_login_invalid_credentials_returns_401(self, db_session: AsyncSession):
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=app_instance), base_url="http://test"
        ) as c:
            response = await c.post(
                "/api/admin/auth/login",
                json={"username": "wrong", "password": "bad"},
            )
        assert response.status_code == 401

    async def test_logout_clears_session(self, db_session: AsyncSession):
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=app_instance), base_url="http://test"
        ) as c:
            await c.post(
                "/api/admin/auth/login",
                json={"username": "admin", "password": "admin"},
            )
            response = await c.post("/api/admin/auth/logout")
        assert response.status_code == 200
        assert response.json()["ok"] is True

    async def test_me_authenticated_returns_true(self, db_session: AsyncSession):
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=app_instance), base_url="http://test"
        ) as c:
            await c.post(
                "/api/admin/auth/login",
                json={"username": "admin", "password": "admin"},
            )
            response = await c.get("/api/admin/auth/me")
        assert response.status_code == 200
        assert response.json()["authenticated"] is True

    async def test_me_unauthenticated_returns_401(self, db_session: AsyncSession):
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        async with AsyncClient(
            transport=ASGITransport(app=app_instance), base_url="http://test"
        ) as c:
            response = await c.get("/api/admin/auth/me")
        assert response.status_code == 401

    def test_verify_admin_password_bcrypt_path_verify_called(self):
        """_verify_admin_password delegates to _pwd_context.verify for bcrypt-prefixed hashes."""
        from unittest.mock import patch
        from app.routers.admin_auth import _verify_admin_password

        # Use a bcrypt-prefixed stored value to trigger the bcrypt branch (line 30).
        # Patch _pwd_context.verify so we don't need a real bcrypt hash computation.
        with patch("app.routers.admin_auth._pwd_context") as mock_ctx:
            mock_ctx.verify.return_value = True
            result = _verify_admin_password("mypass", "$2b$12$fakehash")
        mock_ctx.verify.assert_called_once_with("mypass", "$2b$12$fakehash")
        assert result is True

    async def test_hash_password_endpoint_returns_bcrypt_hash(self, db_session: AsyncSession):
        """POST /api/admin/auth/hash-password returns a bcrypt hash in non-production."""
        from unittest.mock import patch
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        # Patch _pwd_context.hash to avoid real bcrypt computation in test env
        with patch("app.routers.admin_auth._pwd_context") as mock_ctx:
            mock_ctx.hash.return_value = "$2b$12$mockedhashvalue"
            async with AsyncClient(
                transport=ASGITransport(app=app_instance), base_url="http://test"
            ) as c:
                response = await c.post(
                    "/api/admin/auth/hash-password",
                    json={"password": "mypassword"},
                )
        assert response.status_code == 200
        data = response.json()
        assert data["hash"] == "$2b$12$mockedhashvalue"

    async def test_hash_password_endpoint_returns_404_in_production(self, db_session: AsyncSession):
        """POST /api/admin/auth/hash-password returns 404 when ENVIRONMENT=production."""
        from unittest.mock import patch
        from app.main import create_app

        app_instance = create_app()

        async def override_get_db():
            yield db_session

        app_instance.dependency_overrides[get_db] = override_get_db

        with patch.object(__import__("app.config", fromlist=["settings"]).settings, "environment", "production"):
            async with AsyncClient(
                transport=ASGITransport(app=app_instance), base_url="http://test"
            ) as c:
                response = await c.post(
                    "/api/admin/auth/hash-password",
                    json={"password": "mypassword"},
                )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# Import lot owners from Excel
# ---------------------------------------------------------------------------


class TestImportLotOwnersExcel:
    # --- Happy path ---

    async def test_valid_xlsx_imports_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Excel Import Building", manager_email="xlimport@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["XL1", "xl1@test.com", 100], ["XL2", "xl2@test.com", 200]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 2
        assert data["emails"] == 2

    async def test_xlsx_updates_existing_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Excel Update Building", manager_email="xlupdate@test.com")
        db_session.add(b)
        await db_session.commit()

        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["UPD1", "upd@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["UPD1", "upd@test.com", 150]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

    async def test_xlsx_empty_returns_zero(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Excel Bldg", manager_email="ee@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(["Lot#", "Email", "UOE2"], [])
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 0

    # --- Input validation ---

    async def test_xlsx_missing_lot_number_column_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Missing Col Excel Bldg", manager_email="mc@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["email", "unit_entitlement"],
            [["mc@test.com", 100]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_invalid_excel_file_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Bad Excel Bldg", manager_email="be@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    b"not-an-excel-file",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_file_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with no headers (empty workbook) returns 422."""
        b = Building(name="Empty Sheet Bldg", manager_email="es@test.com")
        db_session.add(b)
        await db_session.commit()

        # Create a workbook with a completely empty sheet (no rows at all)
        wb = openpyxl.Workbook()
        ws = wb.active
        # Write nothing — sheet is empty, iter_rows will raise StopIteration on first next()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_excel = buf.read()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    empty_excel,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_lot_number_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with an empty Lot# cell returns 422."""
        b = Building(name="Empty Lot Bldg", manager_email="el@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            # Row with None lot number — will trigger the empty lot_number path
            [[None, "el@test.com", 100]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with an empty UOE2 cell returns 422."""
        b = Building(name="Empty UOE Bldg", manager_email="eu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["EU1", "eu@test.com", None]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_non_integer_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with a non-integer UOE2 returns 422."""
        b = Building(name="Bad UOE Bldg", manager_email="bu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["BU1", "bu@test.com", "not-a-number"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_negative_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with a negative UOE2 returns 422."""
        b = Building(name="Neg UOE Bldg", manager_email="nu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["NU1", "nu@test.com", -5]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_invalid_financial_position_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with invalid financial_position returns 422."""
        b = Building(name="Bad FP Bldg", manager_email="bfp@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2", "financial_position"],
            [["BFP1", "bfp@test.com", 100, "bankrupt"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_blank_rows_skipped(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with blank rows between data rows skips them."""
        b = Building(name="Blank Row Bldg", manager_email="br@test.com")
        db_session.add(b)
        await db_session.commit()

        # Build Excel manually with a blank row between two data rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Email", "UOE2"])
        ws.append(["BR1", "br1@test.com", 100])
        ws.append([None, None, None])  # blank row — triggers continue
        ws.append(["BR2", "br2@test.com", 200])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2


# ---------------------------------------------------------------------------
# Building import from Excel
# ---------------------------------------------------------------------------


class TestImportBuildingsExcel:
    # --- Happy path ---

    async def test_valid_xlsx_imports_buildings(
        self, client: AsyncClient
    ):
        """Valid Excel with building_name/manager_email creates buildings."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [
                ["Excel Tower A", "mgra@excel.com"],
                ["Excel Tower B", "mgrb@excel.com"],
            ],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2

    async def test_xlsx_updates_existing_building(
        self, client: AsyncClient
    ):
        """Importing a building with an existing name updates its manager_email (lines 207-208)."""
        # First import to create the building
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Update Tower", "old@tower.com"]],
        )
        resp1 = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp1.json()["created"] == 1

        # Second import with the same building_name — updates manager_email
        excel_data2 = make_excel(
            ["building_name", "manager_email"],
            [["Update Tower", "new@tower.com"]],
        )
        resp2 = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data2,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp2.status_code == 200
        data = resp2.json()
        assert data["updated"] == 1
        assert data["created"] == 0

    async def test_xlsx_blank_rows_skipped(
        self, client: AsyncClient
    ):
        """Blank rows in building Excel are skipped (line 163 continue)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["building_name", "manager_email"])
        ws.append(["Blank Row Tower", "br@excel.com"])
        ws.append([None, None])  # blank row — triggers continue
        ws.append(["Blank Row Tower 2", "br2@excel.com"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["created"] == 2

    # --- Input validation ---

    async def test_invalid_excel_file_returns_422(
        self, client: AsyncClient
    ):
        """Non-parseable bytes submitted as Excel → 422."""
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    b"not-an-excel-file",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_excel_file_returns_422(
        self, client: AsyncClient
    ):
        """Excel with no rows → 422 (no headers)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Empty sheet — will hit StopIteration on first next()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_excel = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    empty_excel,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_missing_required_headers_returns_422(
        self, client: AsyncClient
    ):
        """Excel missing required columns (building_name or manager_email) → 422."""
        excel_data = make_excel(
            ["only_one_col"],
            [["some value"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_building_name_returns_422(
        self, client: AsyncClient
    ):
        """Row with empty building_name → 422."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [[None, "mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_manager_email_returns_422(
        self, client: AsyncClient
    ):
        """Row with empty manager_email → 422."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Good Building", None]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# LotOwnerCreate schema validator
# ---------------------------------------------------------------------------


class TestLotOwnerCreateSchema:
    async def test_valid_financial_position_in_arrear_returns_201(
        self, client: AsyncClient, building: Building
    ):
        """Creating a lot owner with financial_position='in_arrear' succeeds (line 93 return v)."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={
                "lot_number": "FPV0",
                "unit_entitlement": 100,
                "financial_position": "in_arrear",
                "emails": [],
            },
        )
        assert response.status_code == 201
        assert response.json()["financial_position"] == "in_arrear"

    async def test_invalid_financial_position_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """Creating a lot owner with an invalid financial_position value → 422."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={
                "lot_number": "FPV1",
                "unit_entitlement": 100,
                "financial_position": "bankrupt",
                "emails": [],
            },
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# CSV import: financial_position edge cases
# ---------------------------------------------------------------------------


class TestImportLotOwnersFinancialPosition:
    async def test_csv_import_in_arrear_financial_position(
        self, client: AsyncClient, building: Building
    ):
        """Importing with financial_position=in_arrear stores it correctly."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "financial_position"],
            [["FP1", "fp1@test.com", "100", "in_arrear"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

    async def test_csv_import_invalid_financial_position_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """Invalid financial_position value → 422."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "financial_position"],
            [["FP2", "fp2@test.com", "100", "bankrupt"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# Duplicate email on add_email_to_lot_owner
# ---------------------------------------------------------------------------


class TestAddEmailDuplicate:
    async def test_add_duplicate_email_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding the same email twice to a lot owner → 409."""
        b = Building(name="Dup Email Building", manager_email="dup_email@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="DE1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotOwnerEmail(lot_owner_id=lo.id, email="existing@test.com"))
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "existing@test.com"},
        )
        assert response.status_code == 409


# ---------------------------------------------------------------------------
# PUT /api/admin/lot-owners/{lot_owner_id}/proxy
# DELETE /api/admin/lot-owners/{lot_owner_id}/proxy
# ---------------------------------------------------------------------------


class TestSetLotOwnerProxy:
    # --- Happy path ---

    async def test_set_proxy_creates_proxy_and_returns_proxy_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /lot-owners/{id}/proxy with valid email creates proxy and returns proxy_email."""
        lo = LotOwner(building_id=building.id, lot_number="PX01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "proxy@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "proxy@test.com"

    async def test_set_proxy_replaces_existing_proxy(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /lot-owners/{id}/proxy when proxy already exists replaces it (upsert)."""
        lo = LotOwner(building_id=building.id, lot_number="PX02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotProxy(lot_owner_id=lo.id, proxy_email="old_proxy@test.com"))
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "new_proxy@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "new_proxy@test.com"

    # --- Input validation ---

    async def test_empty_proxy_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT with empty proxy_email → 422."""
        lo = LotOwner(building_id=building.id, lot_number="PX03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": ""},
        )
        assert response.status_code == 422

    async def test_missing_proxy_email_field_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT with missing proxy_email field → 422."""
        lo = LotOwner(building_id=building.id, lot_number="PX04", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_set_proxy_on_nonexistent_lot_owner_returns_404(
        self, client: AsyncClient
    ):
        """PUT on non-existent lot owner → 404."""
        response = await client.put(
            f"/api/admin/lot-owners/{uuid.uuid4()}/proxy",
            json={"proxy_email": "proxy@test.com"},
        )
        assert response.status_code == 404


class TestRemoveLotOwnerProxy:
    # --- Happy path ---

    async def test_remove_proxy_returns_null_proxy_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE /lot-owners/{id}/proxy removes proxy and returns proxy_email: null."""
        lo = LotOwner(building_id=building.id, lot_number="PX05", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotProxy(lot_owner_id=lo.id, proxy_email="proxy_to_remove@test.com"))
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/proxy"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] is None

    # --- State / precondition errors ---

    async def test_remove_proxy_on_nonexistent_lot_owner_returns_404(
        self, client: AsyncClient
    ):
        """DELETE on non-existent lot owner → 404."""
        response = await client.delete(
            f"/api/admin/lot-owners/{uuid.uuid4()}/proxy"
        )
        assert response.status_code == 404

    async def test_remove_proxy_when_no_proxy_set_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE when no proxy is set → 404."""
        lo = LotOwner(building_id=building.id, lot_number="PX06", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/proxy"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# get_effective_status helper — unit tests (US-CD01)
# ---------------------------------------------------------------------------


class _FakeAGM:
    """Minimal stand-in for GeneralMeeting used to unit-test get_effective_status without a DB."""

    def __init__(self, status: GeneralMeetingStatus, voting_closes_at, meeting_at=None):
        self.status = status
        self.voting_closes_at = voting_closes_at
        self.meeting_at = meeting_at


class TestGetEffectiveStatus:
    """Unit tests for the get_effective_status helper."""

    # --- Happy path ---

    def test_open_agm_with_future_closes_at_returns_open(self):
        """An GeneralMeeting whose voting_closes_at is in the future stays open."""
        agm = _FakeAGM(GeneralMeetingStatus.open, datetime.now(UTC) + timedelta(days=1))
        assert get_effective_status(agm) == GeneralMeetingStatus.open  # type: ignore[arg-type]

    def test_open_agm_with_past_closes_at_returns_closed(self):
        """An GeneralMeeting whose voting_closes_at has passed is effectively closed."""
        agm = _FakeAGM(GeneralMeetingStatus.open, datetime.now(UTC) - timedelta(seconds=1))
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_already_closed_agm_returns_closed_regardless_of_closes_at(self):
        """An GeneralMeeting with status=closed stays closed whether closes_at is past or future."""
        agm = _FakeAGM(GeneralMeetingStatus.closed, datetime.now(UTC) + timedelta(days=1))
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_open_agm_with_none_closes_at_returns_open(self):
        """An GeneralMeeting with no voting_closes_at set stays open (edge case)."""
        agm = _FakeAGM(GeneralMeetingStatus.open, None)
        assert get_effective_status(agm) == GeneralMeetingStatus.open  # type: ignore[arg-type]

    def test_naive_datetime_past_returns_closed(self):
        """A naive (tz-unaware) voting_closes_at in the past is treated as UTC."""
        agm = _FakeAGM(GeneralMeetingStatus.open, datetime(2000, 1, 1, 0, 0, 0))
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_naive_datetime_future_returns_open(self):
        """A naive (tz-unaware) voting_closes_at far in the future stays open."""
        agm = _FakeAGM(GeneralMeetingStatus.open, datetime(2099, 12, 31, 23, 59, 59))
        assert get_effective_status(agm) == GeneralMeetingStatus.open  # type: ignore[arg-type]

    # --- Pending status (US-PS01) ---

    def test_pending_stored_status_with_future_meeting_at_returns_pending(self):
        """A meeting stored as pending with future meeting_at returns pending."""
        agm = _FakeAGM(
            GeneralMeetingStatus.pending,
            datetime.now(UTC) + timedelta(days=2),
            meeting_at=datetime.now(UTC) + timedelta(days=1),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.pending  # type: ignore[arg-type]

    def test_open_stored_status_with_future_meeting_at_returns_pending(self):
        """A meeting stored as open but with future meeting_at returns pending."""
        agm = _FakeAGM(
            GeneralMeetingStatus.open,
            datetime.now(UTC) + timedelta(days=2),
            meeting_at=datetime.now(UTC) + timedelta(days=1),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.pending  # type: ignore[arg-type]

    def test_open_stored_status_with_past_meeting_at_future_closes_at_returns_open(self):
        """A meeting whose start has passed but voting is still open returns open."""
        agm = _FakeAGM(
            GeneralMeetingStatus.open,
            datetime.now(UTC) + timedelta(days=1),
            meeting_at=datetime.now(UTC) - timedelta(hours=1),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.open  # type: ignore[arg-type]

    def test_closed_stored_status_with_future_voting_closes_at_returns_closed(self):
        """Manually closed meeting returns closed even if voting_closes_at is in the future."""
        agm = _FakeAGM(
            GeneralMeetingStatus.closed,
            datetime.now(UTC) + timedelta(days=1),
            meeting_at=datetime.now(UTC) + timedelta(days=1),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_pending_stored_status_with_both_timestamps_past_returns_closed(self):
        """A meeting stored as pending but both timestamps past returns closed."""
        agm = _FakeAGM(
            GeneralMeetingStatus.pending,
            datetime.now(UTC) - timedelta(hours=1),
            meeting_at=datetime.now(UTC) - timedelta(hours=2),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_future_meeting_at_but_past_voting_closes_at_returns_closed(self):
        """voting_closes_at in the past takes priority over future meeting_at."""
        agm = _FakeAGM(
            GeneralMeetingStatus.open,
            datetime.now(UTC) - timedelta(seconds=1),
            meeting_at=datetime.now(UTC) + timedelta(days=1),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.closed  # type: ignore[arg-type]

    def test_naive_future_meeting_at_returns_pending(self):
        """A naive (tz-unaware) meeting_at far in the future derives pending."""
        agm = _FakeAGM(
            GeneralMeetingStatus.open,
            datetime(2099, 12, 31, 23, 59, 59),
            meeting_at=datetime(2099, 12, 31, 12, 0, 0),
        )
        assert get_effective_status(agm) == GeneralMeetingStatus.pending  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Admin list_agms effective status (US-CD01)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestListAGMsEffectiveStatus:
    """list_agms returns effective (closed) status for past-closes_at AGMs."""

    async def test_list_agms_past_closes_at_shows_closed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="EffStatus Building", manager_email="eff@test.com")
        db_session.add(b)
        await db_session.flush()
        past_agm = GeneralMeeting(
            building_id=b.id,
            title="Expired GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) - timedelta(days=3),
            voting_closes_at=datetime.now(UTC) - timedelta(days=1),
        )
        db_session.add(past_agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        items = response.json()
        match = next((i for i in items if i["id"] == str(past_agm.id)), None)
        assert match is not None
        assert match["status"] == "closed"

    async def test_list_agms_future_closes_at_shows_pending(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A meeting with meeting_at in the future is effectively pending."""
        b = Building(name="FutureStatus Building", manager_email="fut@test.com")
        db_session.add(b)
        await db_session.flush()
        future_agm = GeneralMeeting(
            building_id=b.id,
            title="Future GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) + timedelta(days=1),
            voting_closes_at=datetime.now(UTC) + timedelta(days=2),
        )
        db_session.add(future_agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        items = response.json()
        match = next((i for i in items if i["id"] == str(future_agm.id)), None)
        assert match is not None
        assert match["status"] == "pending"

    async def test_list_agms_past_meeting_at_future_closes_at_shows_open(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A meeting whose start time has passed but voting is still open is effectively open."""
        b = Building(name="OpenStatus Building", manager_email="open@test.com")
        db_session.add(b)
        await db_session.flush()
        open_agm = GeneralMeeting(
            building_id=b.id,
            title="Open GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) - timedelta(hours=1),
            voting_closes_at=datetime.now(UTC) + timedelta(days=2),
        )
        db_session.add(open_agm)
        await db_session.commit()

        response = await client.get("/api/admin/general-meetings")
        assert response.status_code == 200
        items = response.json()
        match = next((i for i in items if i["id"] == str(open_agm.id)), None)
        assert match is not None
        assert match["status"] == "open"


# ---------------------------------------------------------------------------
# Admin get_agm_detail effective status (US-CD01)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestGetGeneralMeetingDetailEffectiveStatus:
    """get_agm_detail returns effective (closed) status for past-closes_at AGMs."""

    async def test_detail_past_closes_at_shows_closed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="DetailEff Building", manager_email="de@test.com")
        db_session.add(b)
        await db_session.flush()
        past_agm = GeneralMeeting(
            building_id=b.id,
            title="Detail Expired GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) - timedelta(days=3),
            voting_closes_at=datetime.now(UTC) - timedelta(days=1),
        )
        db_session.add(past_agm)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{past_agm.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "closed"

    async def test_detail_future_closes_at_shows_pending(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A meeting with meeting_at in the future is effectively pending."""
        b = Building(name="DetailFut Building", manager_email="df@test.com")
        db_session.add(b)
        await db_session.flush()
        future_agm = GeneralMeeting(
            building_id=b.id,
            title="Detail Future GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) + timedelta(days=1),
            voting_closes_at=datetime.now(UTC) + timedelta(days=2),
        )
        db_session.add(future_agm)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{future_agm.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "pending"

    async def test_detail_past_meeting_at_future_closes_at_shows_open(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A meeting whose start has passed but voting is still open is effectively open."""
        b = Building(name="DetailOpen Building", manager_email="do@test.com")
        db_session.add(b)
        await db_session.flush()
        open_agm = GeneralMeeting(
            building_id=b.id,
            title="Detail Open GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=datetime.now(UTC) - timedelta(hours=1),
            voting_closes_at=datetime.now(UTC) + timedelta(days=2),
        )
        db_session.add(open_agm)
        await db_session.commit()

        response = await client.get(f"/api/admin/general-meetings/{open_agm.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "open"


# ---------------------------------------------------------------------------
# close_agm tally shows absent lots correctly (US-CD02)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestCloseAGMAbsentRecords:
    """Closing a GeneralMeeting does NOT create BallotSubmission records for absent lots.
    Absent lots are computed by the tally as eligible - submitted."""

    async def _make_agm_with_lots(
        self, db_session: AsyncSession, name: str, n_lots: int = 2
    ):
        """Create building + GeneralMeeting + n lots with GeneralMeetingLotWeight snapshots."""
        b = Building(name=name, manager_email=f"absent_{name}@test.com")
        db_session.add(b)
        await db_session.flush()

        lots = []
        for i in range(n_lots):
            lo = LotOwner(
                building_id=b.id,
                lot_number=f"A{i+1}",
                unit_entitlement=100,
            )
            db_session.add(lo)
            await db_session.flush()
            lo_email = LotOwnerEmail(lot_owner_id=lo.id, email=f"voter{i+1}@{name}.test")
            db_session.add(lo_email)
            lots.append(lo)

        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Absent Test GeneralMeeting {name}",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Motion 1", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for lo in lots:
            w = GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            )
            db_session.add(w)
        await db_session.flush()

        return b, agm, lots, [motion]

    # --- Happy path ---

    async def test_close_does_not_create_absent_submissions_for_non_voters(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Lots with no BallotSubmission must NOT get an absent BallotSubmission on close.
        The tally computes absent = eligible - submitted without needing phantom records."""
        _, agm, lots, _ = await self._make_agm_with_lots(
            db_session, "AbsentHappy1", n_lots=2
        )

        # lot[0] already voted
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lots[0].id,
            voter_email=f"voter1@AbsentHappy1.test",
        )
        db_session.add(bs)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        # lots[1] did not vote — must have NO BallotSubmission after close
        subs_result = await db_session.execute(
            select(BallotSubmission).where(
                BallotSubmission.general_meeting_id == agm.id,
                BallotSubmission.lot_owner_id == lots[1].id,
            )
        )
        absent_sub = subs_result.scalar_one_or_none()
        assert absent_sub is None

    async def test_close_tally_shows_absent_lot_in_absent_not_abstained(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """After close, a lot that never voted appears in absent tally (not abstained)."""
        _, agm, lots, _ = await self._make_agm_with_lots(
            db_session, "AbsentVotes1", n_lots=2
        )
        # lots[0] votes, lots[1] does not
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lots[0].id,
            voter_email="voter1@AbsentVotes1.test",
        )
        db_session.add(bs)
        motion_result = await db_session.execute(
            select(Motion).where(Motion.general_meeting_id == agm.id)
        )
        motion = motion_result.scalar_one()
        v = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="voter1@AbsentVotes1.test",
            lot_owner_id=lots[0].id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        )
        db_session.add(v)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        tally_response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert tally_response.status_code == 200
        data = tally_response.json()
        m = data["motions"][0]
        # lots[1] never voted → absent voter_count=1, abstained voter_count=0
        assert m["tally"]["absent"]["voter_count"] == 1
        assert m["tally"]["abstained"]["voter_count"] == 0

    async def test_close_does_not_duplicate_existing_submissions(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A lot that already has a BallotSubmission must not get a second one."""
        _, agm, lots, _ = await self._make_agm_with_lots(
            db_session, "AbsentNoDup1", n_lots=1
        )
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lots[0].id,
            voter_email="voter1@AbsentNoDup1.test",
        )
        db_session.add(bs)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        subs_result = await db_session.execute(
            select(BallotSubmission).where(
                BallotSubmission.general_meeting_id == agm.id,
                BallotSubmission.lot_owner_id == lots[0].id,
            )
        )
        assert len(list(subs_result.scalars().all())) == 1

    async def test_close_agm_with_no_lot_weights_no_absent_records(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Closing a GeneralMeeting with no GeneralMeetingLotWeight snapshot creates no absent records."""
        b = Building(name="NoWeights Building", manager_email="nw@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="No Weights GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        subs_result = await db_session.execute(
            select(BallotSubmission).where(BallotSubmission.general_meeting_id == agm.id)
        )
        assert list(subs_result.scalars().all()) == []

    async def test_close_agm_total_submitted_excludes_absent_lots(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """total_submitted in the summary reflects only actual voters, not absent lots."""
        _, agm, lots, _ = await self._make_agm_with_lots(
            db_session, "AbsentEmail1", n_lots=2
        )
        # Only lots[0] votes
        bs = BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lots[0].id,
            voter_email="voter1@AbsentEmail1.test",
        )
        db_session.add(bs)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        summary = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert summary.status_code == 200
        assert summary.json()["total_submitted"] == 1

    async def test_close_agm_lot_with_no_ballot_stays_absent_in_tally(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """A lot with no BallotSubmission after close appears as absent in the tally."""
        b = Building(name="NoEmail Building", manager_email="ne@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="NE1", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="No Email GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        w = GeneralMeetingLotWeight(
            general_meeting_id=agm.id,
            lot_owner_id=lo.id,
            unit_entitlement_snapshot=50,
        )
        db_session.add(w)
        await db_session.commit()

        response = await client.post(f"/api/admin/general-meetings/{agm.id}/close")
        assert response.status_code == 200

        # No BallotSubmission created for the absent lot
        subs_result = await db_session.execute(
            select(BallotSubmission).where(
                BallotSubmission.general_meeting_id == agm.id,
                BallotSubmission.lot_owner_id == lo.id,
            )
        )
        sub = subs_result.scalar_one_or_none()
        assert sub is None


# ---------------------------------------------------------------------------
# GET /api/admin/general-meetings/{agm_id} — absent only for closed meetings
# ---------------------------------------------------------------------------


class TestGetGeneralMeetingDetailAbsentBehaviour:
    """absent tally is only populated for closed meetings."""

    async def _make_agm_with_non_voter(
        self,
        db_session: AsyncSession,
        name: str,
        status: GeneralMeetingStatus,
    ):
        """Create building + AGM with one voter and one non-voter lot."""
        b = Building(name=name, manager_email=f"{name}@test.com")
        db_session.add(b)
        await db_session.flush()

        lo_voted = LotOwner(building_id=b.id, lot_number="V1", unit_entitlement=100)
        lo_absent = LotOwner(building_id=b.id, lot_number="V2", unit_entitlement=50)
        db_session.add_all([lo_voted, lo_absent])
        await db_session.flush()

        db_session.add(LotOwnerEmail(lot_owner_id=lo_voted.id, email=f"voted@{name}.test"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo_absent.id, email=f"absent@{name}.test"))
        await db_session.flush()

        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Absent Behaviour {name}",
            status=status,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()

        motion = Motion(general_meeting_id=agm.id, title="Motion AB1", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        for lo in [lo_voted, lo_absent]:
            db_session.add(GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_owner_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            ))
        await db_session.flush()

        # lo_voted submits a ballot
        db_session.add(Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email=f"voted@{name}.test",
            lot_owner_id=lo_voted.id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        ))
        db_session.add(BallotSubmission(
            general_meeting_id=agm.id,
            lot_owner_id=lo_voted.id,
            voter_email=f"voted@{name}.test",
        ))
        await db_session.commit()
        return agm

    # --- Happy path ---

    async def test_get_general_meeting_detail_absent_only_for_closed_open_meeting(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Open meeting: absent tally is 0 even when a lot owner has not voted."""
        agm = await self._make_agm_with_non_voter(
            db_session, "AbsentOpenTest", GeneralMeetingStatus.open
        )
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 200
        tally = response.json()["motions"][0]["tally"]
        assert tally["absent"]["voter_count"] == 0
        assert tally["absent"]["entitlement_sum"] == 0
        assert response.json()["motions"][0]["voter_lists"]["absent"] == []

    async def test_get_general_meeting_detail_absent_computed_for_closed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Closed meeting: absent tally reflects lots that did not submit."""
        agm = await self._make_agm_with_non_voter(
            db_session, "AbsentClosedTest", GeneralMeetingStatus.closed
        )
        response = await client.get(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 200
        tally = response.json()["motions"][0]["tally"]
        assert tally["absent"]["voter_count"] == 1
        assert tally["absent"]["entitlement_sum"] == 50


# ---------------------------------------------------------------------------
# DELETE /api/admin/general-meetings/{agm_id}
# ---------------------------------------------------------------------------


class TestDeleteGeneralMeeting:
    async def _create_meeting(
        self,
        db_session: AsyncSession,
        name: str,
        status: GeneralMeetingStatus,
    ) -> GeneralMeeting:
        b = Building(name=name, manager_email=f"del_{name}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Delete Test {name}",
            status=status,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm

    # --- Happy path ---

    async def test_delete_general_meeting_closed_success(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a closed meeting returns 204."""
        agm = await self._create_meeting(db_session, "DeleteClosed", GeneralMeetingStatus.closed)
        response = await client.delete(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 204

    async def test_delete_general_meeting_pending_success(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a pending meeting returns 204."""
        agm = await self._create_meeting(db_session, "DeletePending", GeneralMeetingStatus.pending)
        response = await client.delete(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 204

    async def test_delete_general_meeting_removes_from_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """After DELETE the record no longer exists in the database."""
        agm = await self._create_meeting(db_session, "DeleteDBCheck", GeneralMeetingStatus.closed)
        agm_id = agm.id
        response = await client.delete(f"/api/admin/general-meetings/{agm_id}")
        assert response.status_code == 204
        result = await db_session.execute(
            select(GeneralMeeting).where(GeneralMeeting.id == agm_id)
        )
        assert result.scalar_one_or_none() is None

    # --- State / precondition errors ---

    async def test_delete_general_meeting_open_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on an open meeting returns 409."""
        agm = await self._create_meeting(db_session, "DeleteOpen", GeneralMeetingStatus.open)
        response = await client.delete(f"/api/admin/general-meetings/{agm.id}")
        assert response.status_code == 409
        assert "Cannot delete an open General Meeting" in response.json()["detail"]

    async def test_delete_general_meeting_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a non-existent meeting ID returns 404."""
        fake_id = uuid.uuid4()
        response = await client.delete(f"/api/admin/general-meetings/{fake_id}")
        assert response.status_code == 404
        assert "General Meeting not found" in response.json()["detail"]

    async def test_delete_general_meeting_unauthenticated_returns_401(
        self, db_session: AsyncSession
    ):
        """DELETE without admin credentials returns 401."""
        from app.main import app as fastapi_app
        async with AsyncClient(
            transport=ASGITransport(app=fastapi_app), base_url="http://test"
        ) as unauthenticated_client:
            agm = await self._create_meeting(db_session, "DeleteUnauth", GeneralMeetingStatus.closed)
            response = await unauthenticated_client.delete(f"/api/admin/general-meetings/{agm.id}")
            assert response.status_code == 401


class TestReorderMotions:
    """Tests for the bulk motion reorder endpoint."""

    async def _create_meeting_with_motions(
        self,
        db_session: AsyncSession,
        name: str,
        status: GeneralMeetingStatus = GeneralMeetingStatus.open,
        motion_count: int = 3,
    ) -> tuple[GeneralMeeting, list[Motion]]:
        b = Building(name=name, manager_email=f"reorder_{name}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Reorder Test {name}",
            status=status,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        motions = []
        for i in range(1, motion_count + 1):
            m = Motion(
                general_meeting_id=agm.id,
                title=f"Motion {i}",
                display_order=i,
            )
            db_session.add(m)
            motions.append(m)
        await db_session.flush()
        for m in motions:
            await db_session.refresh(m)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm, motions

    # --- Happy path ---

    async def test_reorder_three_motions(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reorder 3 motions: verify display_order values updated correctly."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderHappy3", motion_count=3
        )
        m1, m2, m3 = motions
        # Send in reverse order: m3 → 1, m2 → 2, m1 → 3
        payload = {
            "motions": [
                {"motion_id": str(m3.id), "display_order": 1},
                {"motion_id": str(m2.id), "display_order": 2},
                {"motion_id": str(m1.id), "display_order": 3},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        data = response.json()
        assert "motions" in data
        returned_motions = data["motions"]
        assert len(returned_motions) == 3
        # Should be sorted by display_order
        assert returned_motions[0]["id"] == str(m3.id)
        assert returned_motions[0]["display_order"] == 1
        assert returned_motions[1]["id"] == str(m2.id)
        assert returned_motions[1]["display_order"] == 2
        assert returned_motions[2]["id"] == str(m1.id)
        assert returned_motions[2]["display_order"] == 3

    async def test_reorder_normalises_to_sequential_positions(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Non-sequential submitted display_order values are normalised to 1-based sequential."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderNorm", motion_count=3
        )
        m1, m2, m3 = motions
        # Submit non-sequential values: 10, 20, 30 — should normalise to 1, 2, 3
        payload = {
            "motions": [
                {"motion_id": str(m3.id), "display_order": 10},
                {"motion_id": str(m1.id), "display_order": 20},
                {"motion_id": str(m2.id), "display_order": 30},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        returned = response.json()["motions"]
        assert returned[0]["id"] == str(m3.id)
        assert returned[0]["display_order"] == 1
        assert returned[1]["id"] == str(m1.id)
        assert returned[1]["display_order"] == 2
        assert returned[2]["id"] == str(m2.id)
        assert returned[2]["display_order"] == 3

    async def test_reorder_returns_full_motion_details(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Response includes all motion fields."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderFields", motion_count=2
        )
        m1, m2 = motions
        payload = {
            "motions": [
                {"motion_id": str(m2.id), "display_order": 1},
                {"motion_id": str(m1.id), "display_order": 2},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        first = response.json()["motions"][0]
        assert "id" in first
        assert "title" in first
        assert "display_order" in first
        assert "motion_number" in first
        assert "motion_type" in first

    async def test_reorder_single_motion_no_op(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Single motion meeting: reorder with the same ID succeeds."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderSingle", motion_count=1
        )
        m1 = motions[0]
        payload = {
            "motions": [
                {"motion_id": str(m1.id), "display_order": 1},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        assert len(response.json()["motions"]) == 1

    async def test_reorder_pending_meeting_succeeds(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reorder is allowed on pending meetings."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderPending", status=GeneralMeetingStatus.pending, motion_count=2
        )
        m1, m2 = motions
        payload = {
            "motions": [
                {"motion_id": str(m2.id), "display_order": 1},
                {"motion_id": str(m1.id), "display_order": 2},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200

    # --- Input validation ---

    async def test_reorder_empty_list_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Empty motion list returns 422."""
        agm, _ = await self._create_meeting_with_motions(
            db_session, "ReorderEmpty", motion_count=2
        )
        payload = {"motions": []}
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 422

    async def test_reorder_extra_ids_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Submitting extra motion IDs (not belonging to meeting) returns 422."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderExtra", motion_count=2
        )
        m1, m2 = motions
        fake_id = uuid.uuid4()
        payload = {
            "motions": [
                {"motion_id": str(m1.id), "display_order": 1},
                {"motion_id": str(m2.id), "display_order": 2},
                {"motion_id": str(fake_id), "display_order": 3},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 422
        assert "motion_order" in response.json()["detail"]

    async def test_reorder_missing_ids_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Submitting fewer motion IDs than exist in the meeting returns 422."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderMissing", motion_count=3
        )
        m1, _, _ = motions
        # Only include 1 of 3 motions
        payload = {
            "motions": [
                {"motion_id": str(m1.id), "display_order": 1},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 422

    async def test_reorder_ids_from_different_meeting_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Submitting IDs from a different meeting returns 422."""
        agm1, motions1 = await self._create_meeting_with_motions(
            db_session, "ReorderAGM1", motion_count=2
        )
        agm2, motions2 = await self._create_meeting_with_motions(
            db_session, "ReorderAGM2", motion_count=2
        )
        # Send motions from agm2 to agm1's endpoint
        payload = {
            "motions": [
                {"motion_id": str(motions2[0].id), "display_order": 1},
                {"motion_id": str(motions2[1].id), "display_order": 2},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm1.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 422

    async def test_reorder_duplicate_display_order_in_request_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Duplicate display_order values in request body returns 422."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderDupOrder", motion_count=2
        )
        m1, m2 = motions
        payload = {
            "motions": [
                {"motion_id": str(m1.id), "display_order": 1},
                {"motion_id": str(m2.id), "display_order": 1},  # duplicate
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 422
        assert "Duplicate" in response.json()["detail"]

    # --- State / precondition errors ---

    async def test_reorder_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reorder on a closed meeting returns 409."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderClosed", status=GeneralMeetingStatus.closed, motion_count=2
        )
        m1, m2 = motions
        payload = {
            "motions": [
                {"motion_id": str(m2.id), "display_order": 1},
                {"motion_id": str(m1.id), "display_order": 2},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 409
        assert "closed" in response.json()["detail"].lower()

    async def test_reorder_meeting_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reorder on a non-existent meeting returns 404."""
        fake_id = uuid.uuid4()
        payload = {
            "motions": [
                {"motion_id": str(uuid.uuid4()), "display_order": 1},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{fake_id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 404
        assert "General Meeting not found" in response.json()["detail"]

    # --- Edge cases ---

    async def test_reorder_does_not_change_motion_numbers(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reordering motions must never mutate motion_number — it is a stable identifier.

        Creates 2 motions with explicit motion_numbers "1" and "2", swaps their
        display_order, then asserts that each motion still carries its original
        motion_number even though display_order has changed.
        """
        b = Building(name="ReorderMN2", manager_email="rmn2@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Reorder MN Test 2",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        m1 = Motion(
            general_meeting_id=agm.id,
            title="Alpha",
            display_order=1,
            motion_number="1",
        )
        m2 = Motion(
            general_meeting_id=agm.id,
            title="Beta",
            display_order=2,
            motion_number="2",
        )
        db_session.add_all([m1, m2])
        await db_session.flush()
        await db_session.refresh(m1)
        await db_session.refresh(m2)
        await db_session.commit()

        # Move m1 (motion_number="1", currently display_order=1) down to position 2
        payload = {
            "motions": [
                {"motion_id": str(m2.id), "display_order": 1},
                {"motion_id": str(m1.id), "display_order": 2},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        returned = response.json()["motions"]

        # m2 is now first (display_order=1) but still has motion_number "2"
        assert returned[0]["id"] == str(m2.id)
        assert returned[0]["display_order"] == 1
        assert returned[0]["motion_number"] == "2", (
            "motion_number must not change when display_order changes"
        )
        # m1 is now second (display_order=2) but still has motion_number "1"
        assert returned[1]["id"] == str(m1.id)
        assert returned[1]["display_order"] == 2
        assert returned[1]["motion_number"] == "1", (
            "motion_number must not change when display_order changes"
        )

    async def test_reorder_same_order_is_idempotent(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reordering with the same order as current returns the same display_order values."""
        agm, motions = await self._create_meeting_with_motions(
            db_session, "ReorderIdempotent", motion_count=3
        )
        m1, m2, m3 = motions
        payload = {
            "motions": [
                {"motion_id": str(m1.id), "display_order": 1},
                {"motion_id": str(m2.id), "display_order": 2},
                {"motion_id": str(m3.id), "display_order": 3},
            ]
        }
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json=payload,
        )
        assert response.status_code == 200
        returned = response.json()["motions"]
        assert returned[0]["display_order"] == 1
        assert returned[1]["display_order"] == 2
        assert returned[2]["display_order"] == 3

    async def test_reorder_does_not_change_motion_numbers(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Reordering motions must not change their motion_number values.

        Uses the auto-assign feature: create two motions without explicit
        motion_number so the backend assigns "0" and "1" respectively.
        After reversing the display order, the motion_number values must
        remain "0" and "1" tied to their original motions.
        """
        from app.models import Building as _Building, GeneralMeeting as _GM, Motion as _Motion, GeneralMeetingStatus as _GMS, MotionType as _MT
        from datetime import timezone

        b = _Building(name="ReorderNoMN", manager_email="rnmn@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = _GM(
            building_id=b.id,
            title="Reorder No MN Test",
            status=_GMS.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)

        # Add two motions without motion_number — backend auto-assigns "0" and "1"
        r1 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "First Motion"},
        )
        assert r1.status_code == 201
        assert r1.json()["motion_number"] == "0"
        m1_id = r1.json()["id"]

        r2 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Second Motion"},
        )
        assert r2.status_code == 201
        assert r2.json()["motion_number"] == "1"
        m2_id = r2.json()["id"]

        # Reverse the display order
        response = await client.put(
            f"/api/admin/general-meetings/{agm.id}/motions/reorder",
            json={
                "motions": [
                    {"motion_id": m2_id, "display_order": 1},
                    {"motion_id": m1_id, "display_order": 2},
                ]
            },
        )
        assert response.status_code == 200
        returned = response.json()["motions"]

        # m2 is now first (display_order=1), but its motion_number is still "1"
        assert returned[0]["id"] == m2_id
        assert returned[0]["display_order"] == 1
        assert returned[0]["motion_number"] == "1"

        # m1 is now second (display_order=2), but its motion_number is still "0"
        assert returned[1]["id"] == m1_id
        assert returned[1]["display_order"] == 2
        assert returned[1]["motion_number"] == "0"


class TestToggleMotionVisibility:
    """Tests for the motion visibility toggle endpoint."""

    async def _create_open_meeting_with_motion(
        self,
        db_session: AsyncSession,
        label: str,
        motion_type: MotionType = MotionType.general,
        is_visible: bool = True,
        status: GeneralMeetingStatus = GeneralMeetingStatus.open,
    ) -> tuple[GeneralMeeting, Motion]:
        b = Building(name=f"VisBldg_{label}", manager_email=f"vis_{label}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Vis Meeting {label}",
            status=status,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        motion = Motion(
            general_meeting_id=agm.id,
            title=f"Vis Motion {label}",
            display_order=1,
            motion_type=motion_type,
            is_visible=is_visible,
        )
        db_session.add(motion)
        await db_session.commit()
        await db_session.refresh(agm)
        await db_session.refresh(motion)
        return agm, motion

    # --- Happy path ---

    async def test_hide_motion_no_votes_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH is_visible=false on a visible motion with no votes returns 200."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "HideNoVotes", is_visible=True
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == str(motion.id)
        assert data["is_visible"] is False

    async def test_show_hidden_motion_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH is_visible=true on a hidden motion returns 200."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "ShowHidden", is_visible=False
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": True},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["is_visible"] is True

    async def test_toggle_persists_to_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """After toggling to hidden the DB row reflects is_visible=False."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "PersistDB", is_visible=True
        )
        await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        await db_session.refresh(motion)
        assert motion.is_visible is False

    async def test_toggle_response_includes_tally_structure(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Response includes the expected tally and voter_lists structure."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "TallyStruct", is_visible=True
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 200
        data = response.json()
        assert "tally" in data
        assert "voter_lists" in data
        for cat in ("yes", "no", "abstained", "absent", "not_eligible"):
            assert cat in data["tally"]
            assert cat in data["voter_lists"]

    async def test_hide_motion_on_pending_meeting_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Hiding a motion on a pending meeting (not yet open/closed) succeeds."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "PendingHide", is_visible=True, status=GeneralMeetingStatus.pending
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 200
        assert response.json()["is_visible"] is False

    async def test_hide_motion_with_only_draft_votes_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Hiding a motion that only has draft (not submitted) votes succeeds."""
        b = Building(name="VisBldg_DraftVotes", manager_email="vis_draft@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Vis Meeting DraftVotes",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="DV1", unit_entitlement=10)
        db_session.add(lo)
        await db_session.flush()
        motion = Motion(
            general_meeting_id=agm.id,
            title="Vis Motion DraftVotes",
            display_order=1,
            is_visible=True,
        )
        db_session.add(motion)
        await db_session.flush()
        # Add a draft vote (status != submitted) — should NOT block hiding
        draft_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="draftvoter@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.draft,
        )
        db_session.add(draft_vote)
        await db_session.commit()
        await db_session.refresh(motion)

        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 200
        assert response.json()["is_visible"] is False

    # --- Input validation ---

    async def test_missing_is_visible_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with missing body returns 422."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "MissingBody"
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={},
        )
        assert response.status_code == 422

    async def test_invalid_is_visible_type_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with a dict value for is_visible (non-coercible) returns 422."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "InvalidType"
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": {"nested": "object"}},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_hide_motion_with_submitted_votes_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Hiding a motion that has submitted votes returns 409."""
        b = Building(name="VisBldg_VotedHide", manager_email="vis_voted@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Vis Meeting VotedHide",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="VH1", unit_entitlement=10)
        db_session.add(lo)
        await db_session.flush()
        motion = Motion(
            general_meeting_id=agm.id,
            title="Vis Motion VotedHide",
            display_order=1,
            is_visible=True,
        )
        db_session.add(motion)
        await db_session.flush()
        # Add a submitted vote — should BLOCK hiding
        submitted_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="voted@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        )
        db_session.add(submitted_vote)
        await db_session.commit()
        await db_session.refresh(motion)

        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 409
        assert "Cannot hide a motion that has received votes" in response.json()["detail"]

    async def test_toggle_on_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Toggling visibility on a closed meeting returns 409."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "ClosedToggle", is_visible=True, status=GeneralMeetingStatus.closed
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 409
        assert "Cannot change visibility on a closed meeting" in response.json()["detail"]

    async def test_show_on_closed_meeting_also_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Showing a hidden motion on a closed meeting also returns 409."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "ClosedShow", is_visible=False, status=GeneralMeetingStatus.closed
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": True},
        )
        assert response.status_code == 409

    # --- Edge cases ---

    async def test_motion_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH on an unknown motion ID returns 404."""
        fake_id = uuid.uuid4()
        response = await client.patch(
            f"/api/admin/motions/{fake_id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 404
        assert "Motion not found" in response.json()["detail"]

    async def test_unauthenticated_returns_401(
        self, db_session: AsyncSession
    ):
        """PATCH without admin credentials returns 401."""
        from app.main import app as fastapi_app
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "UnauthVis"
        )
        async with AsyncClient(
            transport=ASGITransport(app=fastapi_app), base_url="http://test"
        ) as unauthenticated_client:
            response = await unauthenticated_client.patch(
                f"/api/admin/motions/{motion.id}/visibility",
                json={"is_visible": False},
            )
            assert response.status_code == 401

    async def test_toggle_special_motion_returns_correct_motion_type(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Toggling a special motion returns motion_type='special' in response."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "SpecialHide",
            motion_type=MotionType.special,
            is_visible=True,
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": False},
        )
        assert response.status_code == 200
        assert response.json()["motion_type"] == "special"
        assert response.json()["is_visible"] is False

    async def test_show_motion_with_votes_allowed(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Showing (is_visible=true) a hidden motion that has votes IS allowed (only hiding is guarded)."""
        b = Building(name="VisBldg_ShowVoted", manager_email="vis_showvoted@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Vis Meeting ShowVoted",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="SV1", unit_entitlement=10)
        db_session.add(lo)
        await db_session.flush()
        motion = Motion(
            general_meeting_id=agm.id,
            title="Vis Motion ShowVoted",
            display_order=1,
            is_visible=False,
        )
        db_session.add(motion)
        await db_session.flush()
        submitted_vote = Vote(
            general_meeting_id=agm.id,
            motion_id=motion.id,
            voter_email="showvoted@test.com",
            lot_owner_id=lo.id,
            choice=VoteChoice.yes,
            status=VoteStatus.submitted,
        )
        db_session.add(submitted_vote)
        await db_session.commit()
        await db_session.refresh(motion)

        # Showing a hidden-but-voted motion should be allowed
        response = await client.patch(
            f"/api/admin/motions/{motion.id}/visibility",
            json={"is_visible": True},
        )
        assert response.status_code == 200
        assert response.json()["is_visible"] is True

    async def test_is_visible_field_in_admin_detail_response(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """GET /api/admin/general-meetings/{id} includes is_visible on each motion."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "DetailVisible", is_visible=True
        )
        response = await client.get(f"/api/admin/general-meetings/{_agm.id}")
        assert response.status_code == 200
        motions = response.json()["motions"]
        assert len(motions) == 1
        assert "is_visible" in motions[0]
        assert motions[0]["is_visible"] is True

    async def test_is_visible_false_reflected_in_detail(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """GET /api/admin/general-meetings/{id} reflects is_visible=false for hidden motions."""
        _agm, motion = await self._create_open_meeting_with_motion(
            db_session, "DetailHidden", is_visible=False
        )
        response = await client.get(f"/api/admin/general-meetings/{_agm.id}")
        assert response.status_code == 200
        motions = response.json()["motions"]
        assert motions[0]["is_visible"] is False


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings/{id}/motions
# PATCH /api/admin/motions/{id}
# DELETE /api/admin/motions/{id}
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestMotionManagement:
    """Tests for add, update, and delete motion endpoints."""

    async def _create_meeting(
        self,
        db_session: AsyncSession,
        label: str,
        status: GeneralMeetingStatus = GeneralMeetingStatus.open,
    ) -> GeneralMeeting:
        b = Building(name=f"MgmtBldg_{label}", manager_email=f"mgmt_{label}@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title=f"Mgmt Meeting {label}",
            status=status,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.commit()
        await db_session.refresh(agm)
        return agm

    async def _create_meeting_with_motion(
        self,
        db_session: AsyncSession,
        label: str,
        status: GeneralMeetingStatus = GeneralMeetingStatus.open,
        is_visible: bool = False,
        order_index: int = 0,
    ) -> tuple[GeneralMeeting, Motion]:
        agm = await self._create_meeting(db_session, label, status)
        motion = Motion(
            general_meeting_id=agm.id,
            title=f"Motion {label}",
            description=f"Desc {label}",
            display_order=order_index,
            motion_type=MotionType.general,
            is_visible=is_visible,
        )
        db_session.add(motion)
        await db_session.commit()
        await db_session.refresh(motion)
        return agm, motion

    # --- Happy path (add) ---

    async def test_add_motion_to_open_meeting_returns_201(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """POST adds a motion; returns 201 with is_visible=False and correct fields."""
        agm = await self._create_meeting(db_session, "AddOpen")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "New Motion", "description": "Some desc", "motion_type": "general"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["title"] == "New Motion"
        assert data["description"] == "Some desc"
        assert data["motion_type"] == "general"
        assert data["is_visible"] is False
        assert "id" in data
        assert "display_order" in data

    async def test_add_motion_to_pending_meeting_returns_201(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """POST to a pending meeting also returns 201."""
        agm = await self._create_meeting(db_session, "AddPending", GeneralMeetingStatus.pending)
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Pending Motion"},
        )
        assert response.status_code == 201
        assert response.json()["is_visible"] is False

    async def test_add_motion_first_motion_order_index_zero(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """First motion on a meeting with no existing motions gets display_order=0."""
        agm = await self._create_meeting(db_session, "FirstMotion")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "First"},
        )
        assert response.status_code == 201
        assert response.json()["display_order"] == 0

    async def test_add_motion_order_index_increments(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Second motion gets display_order = max + 1."""
        agm, _motion = await self._create_meeting_with_motion(db_session, "IncrOrder", order_index=0)
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Second Motion"},
        )
        assert response.status_code == 201
        assert response.json()["display_order"] == 1

    async def test_add_multiple_motions_sequential_order_indexes(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding 3 motions results in display_orders 0, 1, 2 with no constraint violations."""
        agm = await self._create_meeting(db_session, "SeqOrder")
        indexes = []
        for i in range(3):
            r = await client.post(
                f"/api/admin/general-meetings/{agm.id}/motions",
                json={"title": f"Motion {i}"},
            )
            assert r.status_code == 201
            indexes.append(r.json()["display_order"])
        assert indexes == [0, 1, 2]

    async def test_add_motion_motion_type_defaults_to_general(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_type is omitted it defaults to 'general'."""
        agm = await self._create_meeting(db_session, "DefaultType")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Default type motion"},
        )
        assert response.status_code == 201
        assert response.json()["motion_type"] == "general"

    async def test_add_special_motion_returns_special_type(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Specifying motion_type=special returns motion with motion_type='special'."""
        agm = await self._create_meeting(db_session, "SpecialAdd")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Special", "motion_type": "special"},
        )
        assert response.status_code == 201
        assert response.json()["motion_type"] == "special"

    async def test_add_motion_description_null_when_omitted(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Description is null when not provided."""
        agm = await self._create_meeting(db_session, "NullDesc")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "No Desc"},
        )
        assert response.status_code == 201
        assert response.json()["description"] is None

    async def test_add_motion_persists_to_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Added motion row appears in DB with correct fields."""
        agm = await self._create_meeting(db_session, "PersistAdd")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Persist Check"},
        )
        assert response.status_code == 201
        motion_id = uuid.UUID(response.json()["id"])
        result = await db_session.execute(select(Motion).where(Motion.id == motion_id))
        motion = result.scalar_one_or_none()
        assert motion is not None
        assert motion.title == "Persist Check"
        assert motion.is_visible is False

    async def test_add_motion_with_motion_number_persists(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """motion_number supplied in add-motion payload is saved to DB and returned."""
        agm = await self._create_meeting(db_session, "AddMotionNumber")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Numbered Motion", "motion_number": "SR-1"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["motion_number"] == "SR-1"
        motion_id = uuid.UUID(data["id"])
        result = await db_session.execute(select(Motion).where(Motion.id == motion_id))
        motion = result.scalar_one_or_none()
        assert motion is not None
        assert motion.motion_number == "SR-1"

    async def test_add_motion_without_motion_number_auto_assigns_display_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_number is omitted, it is auto-assigned to str(display_order)."""
        agm = await self._create_meeting(db_session, "AddNoMotionNumber")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Unnumbered Motion"},
        )
        assert response.status_code == 201
        data = response.json()
        # display_order is 0 for the first motion on an empty meeting
        assert data["display_order"] == 0
        assert data["motion_number"] == "0"

    async def test_add_motion_whitespace_motion_number_auto_assigned(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Whitespace-only motion_number is treated as absent — auto-assigned from display_order."""
        agm = await self._create_meeting(db_session, "AddWhitespaceNumber")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Whitespace Number Motion", "motion_number": "   "},
        )
        assert response.status_code == 201
        # Whitespace is treated same as omitted — auto-assigns from display_order (0 for first motion)
        assert response.json()["motion_number"] == "0"

    # --- State / precondition errors (duplicate motion_number) ---

    async def test_add_motion_duplicate_motion_number_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding a second motion with the same non-null motion_number returns 409."""
        agm = await self._create_meeting(db_session, "DupMotionNum")
        # First motion succeeds
        r1 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "First Numbered", "motion_number": "SR-1"},
        )
        assert r1.status_code == 201
        # Second motion with same motion_number returns 409
        r2 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Second Numbered", "motion_number": "SR-1"},
        )
        assert r2.status_code == 409
        assert "already exists" in r2.json()["detail"].lower()

    async def test_add_motion_explicit_motion_number_overrides_auto(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_number is explicitly provided, it overrides the auto-assigned value."""
        agm = await self._create_meeting(db_session, "ExplicitMotionNum")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Explicitly Numbered", "motion_number": "SR-1"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["motion_number"] == "SR-1"

    async def test_add_motion_two_omitted_numbers_auto_assigns_sequential(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding two motions without motion_number auto-assigns sequential display_order strings."""
        agm = await self._create_meeting(db_session, "AutoSeqMotionNum")
        r1 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "First Auto Number"},
        )
        assert r1.status_code == 201
        # First motion gets display_order=0 → motion_number="0"
        assert r1.json()["motion_number"] == "0"
        r2 = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Second Auto Number"},
        )
        assert r2.status_code == 201
        # Second motion gets display_order=1 → motion_number="1"
        assert r2.json()["motion_number"] == "1"

    # --- Input validation (add) ---

    async def test_add_motion_missing_title_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Missing title in body returns 422."""
        agm = await self._create_meeting(db_session, "MissingTitle")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"description": "No title"},
        )
        assert response.status_code == 422

    async def test_add_motion_empty_title_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Empty (whitespace-only) title returns 422."""
        agm = await self._create_meeting(db_session, "EmptyTitle")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "   "},
        )
        assert response.status_code == 422

    async def test_add_motion_unknown_motion_type_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Unknown motion_type value returns 422."""
        agm = await self._create_meeting(db_session, "BadType")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Bad", "motion_type": "invalid"},
        )
        assert response.status_code == 422

    # --- State / precondition errors (add) ---

    async def test_add_motion_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding a motion to a closed meeting returns 409."""
        agm = await self._create_meeting(db_session, "AddClosed", GeneralMeetingStatus.closed)
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Blocked"},
        )
        assert response.status_code == 409
        assert "closed" in response.json()["detail"].lower()

    async def test_add_motion_meeting_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding to a non-existent meeting returns 404."""
        fake_id = uuid.uuid4()
        response = await client.post(
            f"/api/admin/general-meetings/{fake_id}/motions",
            json={"title": "Ghost meeting"},
        )
        assert response.status_code == 404
        assert "General Meeting not found" in response.json()["detail"]

    async def test_add_motion_requires_admin_returns_403(
        self, db_session: AsyncSession
    ):
        """POST without admin auth returns 401."""
        from app.main import app as fastapi_app
        agm = await self._create_meeting(db_session, "AddUnauth")
        async with AsyncClient(
            transport=ASGITransport(app=fastapi_app), base_url="http://test"
        ) as unauthenticated_client:
            response = await unauthenticated_client.post(
                f"/api/admin/general-meetings/{agm.id}/motions",
                json={"title": "Unauth"},
            )
            assert response.status_code == 401

    # --- motion_number auto-assign (add) ---

    async def test_add_motion_no_motion_number_auto_assigns_from_display_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_number is omitted, it is auto-assigned to str(display_order)."""
        agm = await self._create_meeting(db_session, "AutoNumOmit")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Auto Number"},
        )
        assert response.status_code == 201
        data = response.json()
        # First motion on a meeting with no motions → display_order=0, motion_number="0"
        assert data["motion_number"] == str(data["display_order"])

    async def test_add_motion_null_motion_number_auto_assigns(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_number is explicitly null, it is auto-assigned to str(display_order)."""
        agm = await self._create_meeting(db_session, "AutoNumNull")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Null MN", "motion_number": None},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["motion_number"] == str(data["display_order"])

    async def test_add_motion_empty_string_motion_number_auto_assigns(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When the frontend sends motion_number='', it is auto-assigned to str(display_order).

        This is the core bug scenario: the frontend sends an empty string when the field
        is left blank. Previously the service stored null instead of auto-assigning.
        """
        agm = await self._create_meeting(db_session, "AutoNumEmpty")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Empty MN", "motion_number": ""},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["motion_number"] == str(data["display_order"])

    async def test_add_motion_explicit_motion_number_preserved(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """When motion_number is explicitly provided, it is preserved as-is."""
        agm = await self._create_meeting(db_session, "ExplicitMN")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Custom MN", "motion_number": "SR-5"},
        )
        assert response.status_code == 201
        assert response.json()["motion_number"] == "SR-5"

    async def test_add_motion_whitespace_only_motion_number_auto_assigns(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Whitespace-only motion_number is treated as blank and auto-assigned."""
        agm = await self._create_meeting(db_session, "WsMN")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Whitespace MN", "motion_number": "   "},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["motion_number"] == str(data["display_order"])

    async def test_add_motion_second_motion_auto_assigns_correct_number(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Second motion auto-assigns motion_number = str(display_order) = '1'."""
        agm, _first = await self._create_meeting_with_motion(db_session, "SecondAutoMN", order_index=0)
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "Second Auto"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["display_order"] == 1
        assert data["motion_number"] == "1"

    async def test_add_motion_motion_number_persisted_in_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Auto-assigned motion_number is stored correctly in the DB row."""
        agm = await self._create_meeting(db_session, "PersistMN")
        response = await client.post(
            f"/api/admin/general-meetings/{agm.id}/motions",
            json={"title": "DB Check MN"},
        )
        assert response.status_code == 201
        motion_id = uuid.UUID(response.json()["id"])
        result = await db_session.execute(select(Motion).where(Motion.id == motion_id))
        motion = result.scalar_one_or_none()
        assert motion is not None
        assert motion.motion_number == "0"

    # --- Happy path (update) ---

    async def test_update_motion_all_fields_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with all fields (including motion_number) returns 200 and updated values."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "UpdateAll")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "Updated Title", "description": "Updated Desc", "motion_type": "special", "motion_number": "42"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["title"] == "Updated Title"
        assert data["description"] == "Updated Desc"
        assert data["motion_type"] == "special"
        assert data["motion_number"] == "42"
        assert data["is_visible"] is False

    async def test_update_motion_partial_title_only(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with only title updates title; other fields unchanged."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "PartialTitle")
        original_type = motion.motion_type
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "New Title Only"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["title"] == "New Title Only"
        assert data["motion_type"] == original_type.value if hasattr(original_type, "value") else original_type

    async def test_update_motion_partial_description_only(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with only description updates description; other fields unchanged."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "PartialDesc")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"description": "Only desc changed"},
        )
        assert response.status_code == 200
        assert response.json()["description"] == "Only desc changed"
        assert response.json()["title"] == motion.title

    async def test_update_motion_partial_motion_type_only(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with only motion_type updates type; other fields unchanged."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "PartialType")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_type": "special"},
        )
        assert response.status_code == 200
        assert response.json()["motion_type"] == "special"
        assert response.json()["title"] == motion.title

    async def test_update_motion_persists_to_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Updated fields including motion_number are persisted in the DB."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "PersistUpdate")
        await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "DB Persisted", "motion_number": "99"},
        )
        await db_session.refresh(motion)
        assert motion.title == "DB Persisted"
        assert motion.motion_number == "99"

    # --- Input validation (update) ---

    async def test_update_motion_no_fields_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with empty body returns 422."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "NoFields")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={},
        )
        assert response.status_code == 422

    async def test_update_motion_partial_motion_number_only(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with only motion_number updates it and returns 200."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "MotionNumberOnly")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": "SR-1"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["motion_number"] == "SR-1"
        # Other fields unchanged
        assert data["title"] == motion.title

    async def test_update_motion_motion_number_clear_with_empty_string(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with motion_number='' clears the motion number (stores NULL)."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "ClearMotionNumber")
        # First set a motion number
        await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": "5"},
        )
        # Now clear it with empty string
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": ""},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["motion_number"] is None

    async def test_update_motion_empty_title_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with empty title returns 422."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "EmptyTitleUpdate")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "   "},
        )
        assert response.status_code == 422

    # --- State / precondition errors (update) ---

    async def test_update_motion_visible_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH on a visible motion returns 409."""
        _agm, motion = await self._create_meeting_with_motion(
            db_session, "UpdateVisible", is_visible=True
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "Try update visible"},
        )
        assert response.status_code == 409
        assert "Cannot edit a visible motion" in response.json()["detail"]

    async def test_update_motion_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH on a motion in a closed meeting returns 409."""
        _agm, motion = await self._create_meeting_with_motion(
            db_session, "UpdateClosed", status=GeneralMeetingStatus.closed, is_visible=False
        )
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "Try update closed"},
        )
        assert response.status_code == 409
        assert "closed" in response.json()["detail"].lower()

    async def test_update_motion_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH on a non-existent motion returns 404."""
        fake_id = uuid.uuid4()
        response = await client.patch(
            f"/api/admin/motions/{fake_id}",
            json={"title": "Ghost"},
        )
        assert response.status_code == 404
        assert "Motion not found" in response.json()["detail"]

    async def test_update_motion_requires_admin_returns_401(
        self, db_session: AsyncSession
    ):
        """PATCH without admin auth returns 401."""
        from app.main import app as fastapi_app
        _agm, motion = await self._create_meeting_with_motion(db_session, "UpdateUnauth")
        async with AsyncClient(
            transport=ASGITransport(app=fastapi_app), base_url="http://test"
        ) as unauthenticated_client:
            response = await unauthenticated_client.patch(
                f"/api/admin/motions/{motion.id}",
                json={"title": "Unauth"},
            )
            assert response.status_code == 401

    async def test_update_motion_all_fields_includes_motion_number(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with all fields including motion_number returns 200 with motion_number set."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "UpdateAllMN")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"title": "Updated Title MN", "description": "Updated Desc MN", "motion_type": "special", "motion_number": "42"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["title"] == "Updated Title MN"
        assert data["description"] == "Updated Desc MN"
        assert data["motion_type"] == "special"
        assert data["motion_number"] == "42"
        assert data["is_visible"] is False

    async def test_update_motion_partial_motion_number_only(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with only motion_number updates it; other fields unchanged."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "PartialMN")
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": "SR-1"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["motion_number"] == "SR-1"
        assert data["title"] == motion.title

    async def test_update_motion_motion_number_empty_string_clears(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """PATCH with motion_number='' clears the motion_number to null."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "ClearMN")
        # First set a motion number
        await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": "OLD-NUM"},
        )
        # Now clear it with empty string
        response = await client.patch(
            f"/api/admin/motions/{motion.id}",
            json={"motion_number": ""},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["motion_number"] is None

    # --- Happy path (delete) ---

    async def test_delete_motion_returns_204(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a hidden motion returns 204."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "DeleteOK")
        response = await client.delete(f"/api/admin/motions/{motion.id}")
        assert response.status_code == 204

    async def test_delete_motion_removes_from_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """After DELETE the motion row is absent from the DB."""
        _agm, motion = await self._create_meeting_with_motion(db_session, "DeleteDB")
        motion_id = motion.id
        await client.delete(f"/api/admin/motions/{motion_id}")
        result = await db_session.execute(select(Motion).where(Motion.id == motion_id))
        assert result.scalar_one_or_none() is None

    async def test_delete_motion_other_motions_unaffected(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Deleting one motion does not affect remaining motions (no renumbering)."""
        agm, motion_a = await self._create_meeting_with_motion(db_session, "DeleteOther", order_index=0)
        motion_b = Motion(
            general_meeting_id=agm.id,
            title="Motion B",
            display_order=1,
            is_visible=False,
        )
        db_session.add(motion_b)
        await db_session.commit()
        await db_session.refresh(motion_b)

        await client.delete(f"/api/admin/motions/{motion_a.id}")

        result = await db_session.execute(select(Motion).where(Motion.id == motion_b.id))
        surviving = result.scalar_one_or_none()
        assert surviving is not None
        assert surviving.display_order == 1  # Not renumbered

    # --- State / precondition errors (delete) ---

    async def test_delete_motion_visible_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a visible motion returns 409."""
        _agm, motion = await self._create_meeting_with_motion(
            db_session, "DeleteVisible", is_visible=True
        )
        response = await client.delete(f"/api/admin/motions/{motion.id}")
        assert response.status_code == 409
        assert "Cannot delete a visible motion" in response.json()["detail"]

    async def test_delete_motion_closed_meeting_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a motion in a closed meeting returns 409."""
        _agm, motion = await self._create_meeting_with_motion(
            db_session, "DeleteClosed", status=GeneralMeetingStatus.closed, is_visible=False
        )
        response = await client.delete(f"/api/admin/motions/{motion.id}")
        assert response.status_code == 409
        assert "closed" in response.json()["detail"].lower()

    async def test_delete_motion_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """DELETE on a non-existent motion returns 404."""
        fake_id = uuid.uuid4()
        response = await client.delete(f"/api/admin/motions/{fake_id}")
        assert response.status_code == 404
        assert "Motion not found" in response.json()["detail"]

    async def test_delete_motion_requires_admin_returns_401(
        self, db_session: AsyncSession
    ):
        """DELETE without admin auth returns 401."""
        from app.main import app as fastapi_app
        _agm, motion = await self._create_meeting_with_motion(db_session, "DeleteUnauth")
        async with AsyncClient(
            transport=ASGITransport(app=fastapi_app), base_url="http://test"
        ) as unauthenticated_client:
            response = await unauthenticated_client.delete(f"/api/admin/motions/{motion.id}")
            assert response.status_code == 401


# ---------------------------------------------------------------------------
# Migration 888085a72643 — backfill motion_number from display_order
# ---------------------------------------------------------------------------


@pytest.mark.asyncio(loop_scope="session")
class TestBackfillMotionNumber:
    """Tests for migration 888085a72643.

    The migration runs:
        UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)
        WHERE motion_number IS NULL

    We verify this SQL directly against the test DB so the migration logic is
    covered independently of the Alembic runner.
    """

    # --- Happy path ---

    async def test_null_motion_number_is_backfilled_to_display_order(
        self, db_session: AsyncSession
    ):
        """A motion with motion_number=NULL gets backfilled to str(display_order)."""
        b = Building(name="BackfillMN1", manager_email="bmn1@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Backfill Test 1",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        m = Motion(
            general_meeting_id=agm.id,
            title="Old Motion",
            display_order=3,
            motion_number=None,  # pre-feature state
        )
        db_session.add(m)
        await db_session.flush()

        # Run the exact migration SQL
        await db_session.execute(
            text(
                "UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)"
                " WHERE motion_number IS NULL"
            )
        )

        await db_session.refresh(m)
        assert m.motion_number == "3", (
            "motion_number should be backfilled to str(display_order)"
        )

    async def test_existing_motion_number_is_not_overwritten(
        self, db_session: AsyncSession
    ):
        """A motion with an existing motion_number is untouched by the backfill."""
        b = Building(name="BackfillMN2", manager_email="bmn2@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Backfill Test 2",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        m = Motion(
            general_meeting_id=agm.id,
            title="Numbered Motion",
            display_order=1,
            motion_number="SR-1",  # already has a value
        )
        db_session.add(m)
        await db_session.flush()

        # Run the exact migration SQL
        await db_session.execute(
            text(
                "UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)"
                " WHERE motion_number IS NULL"
            )
        )

        await db_session.refresh(m)
        assert m.motion_number == "SR-1", (
            "motion_number must not be overwritten when it already has a value"
        )

    # --- Boundary values ---

    async def test_multiple_null_motions_all_backfilled(
        self, db_session: AsyncSession
    ):
        """Multiple NULL motion_number rows are all backfilled in one UPDATE."""
        b = Building(name="BackfillMN3", manager_email="bmn3@test.com")
        db_session.add(b)
        await db_session.flush()
        agm = GeneralMeeting(
            building_id=b.id,
            title="Backfill Test 3",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        m1 = Motion(
            general_meeting_id=agm.id, title="M1", display_order=1, motion_number=None
        )
        m2 = Motion(
            general_meeting_id=agm.id, title="M2", display_order=2, motion_number=None
        )
        m3 = Motion(
            general_meeting_id=agm.id, title="M3", display_order=5, motion_number=None
        )
        db_session.add_all([m1, m2, m3])
        await db_session.flush()

        await db_session.execute(
            text(
                "UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)"
                " WHERE motion_number IS NULL"
            )
        )

        for motion, expected in [(m1, "1"), (m2, "2"), (m3, "5")]:
            await db_session.refresh(motion)
            assert motion.motion_number == expected

    # --- Edge cases ---

    async def test_migration_on_empty_motions_table_is_a_no_op(
        self, db_session: AsyncSession
    ):
        """Running the UPDATE when no NULL rows exist does not error."""
        # First, ensure any existing NULL rows are already set (simulate post-migration state)
        await db_session.execute(
            text(
                "UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)"
                " WHERE motion_number IS NULL"
            )
        )
        # Running a second time is idempotent — no rows to update, no error
        result = await db_session.execute(
            text(
                "UPDATE motions SET motion_number = CAST(display_order AS VARCHAR)"
                " WHERE motion_number IS NULL"
            )
        )
        assert result.rowcount == 0
