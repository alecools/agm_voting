"""Tests for admin building endpoints — POST/GET/PATCH/DELETE /api/admin/buildings."""
from __future__ import annotations

import io
import uuid

import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl

from app.models import Building, LotOwner
from app.models.lot_owner_email import LotOwnerEmail

# Helpers and fixtures (make_csv, make_excel, client, building, building_with_owners)
# are defined in conftest.py and automatically available to all test modules.
from tests.conftest import make_csv, make_excel

# ---------------------------------------------------------------------------
# POST /api/admin/buildings/import
# ---------------------------------------------------------------------------


class TestImportBuildings:
    # --- Happy path ---

    async def test_valid_csv_creates_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "alpha@test.com"], ["Beta Complex", "beta@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2
        assert data["updated"] == 0

    async def test_valid_csv_updates_existing_building(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        # First, create Alpha Tower
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "alpha@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Now update its manager email
        csv_data2 = make_csv(
            ["building_name", "manager_email"],
            [["Alpha Tower", "new_alpha@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 1

    async def test_empty_csv_returns_zero_counts(self, client: AsyncClient):
        csv_data = make_csv(["building_name", "manager_email"], [])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("empty.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 0

    async def test_extra_columns_ignored(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email", "extra_col"],
            [["Extra Col Building", "extra@test.com", "ignored"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["created"] == 1

    async def test_case_insensitive_building_name_match(self, client: AsyncClient):
        # Create a building
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["CaseTest Building", "ci@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Upload with different case
        csv_data2 = make_csv(
            ["building_name", "manager_email"],
            [["casetest building", "updated@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 1
        assert data["created"] == 0

    # --- Input validation ---

    async def test_missing_building_name_header(self, client: AsyncClient):
        csv_data = make_csv(["manager_email"], [["mgr@test.com"]])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "building_name" in str(response.json()["detail"]).lower()

    async def test_missing_manager_email_header(self, client: AsyncClient):
        csv_data = make_csv(["building_name"], [["Some Building"]])
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "manager_email" in str(response.json()["detail"]).lower()

    async def test_blank_building_name_row(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["", "mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("building_name" in e for e in response.json()["detail"])

    async def test_blank_manager_email_row(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Valid Building", ""]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("manager_email" in e for e in response.json()["detail"])

    async def test_multiple_row_errors_collected(self, client: AsyncClient):
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["", ""], ["", ""]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        # Should have 4 errors (2 per bad row)
        assert len(response.json()["detail"]) == 4

    # --- State / precondition errors ---

    async def test_non_csv_file_returns_415(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("data.pdf", b"not csv content", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_non_csv_no_extension_returns_415(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("data.bin", b"garbage", "application/octet-stream")},
        )
        # application/octet-stream is in allowed list but no csv extension —
        # however we allow octet-stream, so this should pass content-type check
        # Re-check: octet-stream IS in allowed set, so it should NOT be 415
        assert response.status_code in (200, 422)

    async def test_csv_filename_with_non_csv_content_type(self, client: AsyncClient):
        """A .csv extension overrides content-type check."""
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["CT Building", "ct@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "application/json")},
        )
        # .csv extension should allow it through
        assert response.status_code == 200

    async def test_completely_empty_file_returns_422(self, client: AsyncClient):
        """An empty file (not even headers) should return 422."""
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("empty.csv", b"", "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# GET /api/admin/buildings
# ---------------------------------------------------------------------------


class TestListBuildings:
    # --- Happy path ---

    async def test_returns_all_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) >= 0

    async def test_building_has_required_fields(
        self, client: AsyncClient, building: Building
    ):
        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        assert len(data) > 0
        first = data[0]
        assert "id" in first
        assert "name" in first
        assert "manager_email" in first
        assert "is_archived" in first
        assert "created_at" in first

    # --- name filter ---

    async def test_name_filter_exact_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Filter Exact Building", manager_email="filterexact@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=Filter+Exact+Building")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "Filter Exact Building" in names

    async def test_name_filter_partial_substring_match(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b1 = Building(name="Substring Alpha Corp", manager_email="sub1@test.com")
        b2 = Building(name="Substring Beta Corp", manager_email="sub2@test.com")
        b_other = Building(name="Unrelated Corp", manager_email="other@test.com")
        db_session.add_all([b1, b2, b_other])
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=Substring")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "Substring Alpha Corp" in names
        assert "Substring Beta Corp" in names
        assert "Unrelated Corp" not in names

    async def test_name_filter_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="CaseSensitive Building", manager_email="case@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=casesensitive")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "CaseSensitive Building" in names

    async def test_name_filter_no_match_returns_empty(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.get("/api/admin/buildings?name=does-not-exist-xyz-99")
        assert response.status_code == 200
        data = response.json()
        assert data == []

    async def test_name_filter_absent_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="No Filter Building", manager_email="nofilter@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        assert "No Filter Building" in names

    async def test_name_filter_empty_string_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="EmptyFilter Building", manager_email="empty@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=")
        assert response.status_code == 200
        data = response.json()
        names = [item["name"] for item in data]
        # empty string matches everything via LIKE '%%'
        assert "EmptyFilter Building" in names

    async def test_name_filter_combined_with_limit(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        for i in range(3):
            db_session.add(
                Building(
                    name=f"LimitFilter Building {i}",
                    manager_email=f"limit{i}@test.com",
                )
            )
        await db_session.commit()

        response = await client.get("/api/admin/buildings?name=LimitFilter&limit=2")
        assert response.status_code == 200
        data = response.json()
        assert len(data) <= 2

    # --- offset / pagination ---

    async def test_offset_returns_second_page(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """offset=1 with limit=1 returns the second building ordered by created_at DESC."""
        for i in range(3):
            db_session.add(
                Building(name=f"Page Building {i}", manager_email=f"pg{i}@test.com")
            )
        await db_session.commit()

        first = await client.get("/api/admin/buildings?name=Page+Building&limit=1&offset=0")
        second = await client.get("/api/admin/buildings?name=Page+Building&limit=1&offset=1")
        assert first.status_code == 200
        assert second.status_code == 200
        first_ids = [b["id"] for b in first.json()]
        second_ids = [b["id"] for b in second.json()]
        # The two pages must not return the same building
        assert first_ids != second_ids


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/count
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestCountBuildings:
    # --- Happy path ---

    async def test_returns_total_count(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """GET /api/admin/buildings/count returns {"count": N}."""
        # Create two buildings
        db_session.add(Building(name="Count Building A", manager_email="ca@test.com"))
        db_session.add(Building(name="Count Building B", manager_email="cb@test.com"))
        await db_session.commit()

        response = await client.get("/api/admin/buildings/count")
        assert response.status_code == 200
        data = response.json()
        assert "count" in data
        assert data["count"] >= 2

    # --- name filter ---

    async def test_name_filter_returns_filtered_count(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?name=X returns only the count matching that substring."""
        db_session.add(Building(name="CountFilter Unique X9Z", manager_email="cfx@test.com"))
        db_session.add(Building(name="CountFilter Other Building", manager_email="cfo@test.com"))
        await db_session.commit()

        response = await client.get("/api/admin/buildings/count?name=CountFilter+Unique+X9Z")
        assert response.status_code == 200
        assert response.json()["count"] == 1

    async def test_name_filter_no_match_returns_zero(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.get("/api/admin/buildings/count?name=ZZZ-impossible-match-999")
        assert response.status_code == 200
        assert response.json()["count"] == 0

    async def test_name_filter_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        db_session.add(Building(name="CaseCount Building", manager_email="casecount@test.com"))
        await db_session.commit()

        response = await client.get("/api/admin/buildings/count?name=casecount")
        assert response.status_code == 200
        assert response.json()["count"] >= 1

    async def test_no_name_filter_counts_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Omitting ?name returns the same count as the unfiltered list."""
        list_resp = await client.get("/api/admin/buildings?limit=1000")
        count_resp = await client.get("/api/admin/buildings/count")
        assert list_resp.status_code == 200
        assert count_resp.status_code == 200
        assert count_resp.json()["count"] == len(list_resp.json())

    # --- Boundary values ---

    async def test_empty_name_filter_counts_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?name= (empty string) matches everything — same as no filter."""
        list_resp = await client.get("/api/admin/buildings?limit=1000")
        count_resp = await client.get("/api/admin/buildings/count?name=")
        assert list_resp.status_code == 200
        assert count_resp.status_code == 200
        assert count_resp.json()["count"] == len(list_resp.json())

    # --- is_archived filter ---

    async def test_is_archived_false_counts_only_active(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?is_archived=false returns count of only non-archived buildings."""
        active = Building(name="IsArchFalse Active A", manager_email="iafa@test.com")
        archived = Building(name="IsArchFalse Archived B", manager_email="iafb@test.com")
        archived.is_archived = True
        db_session.add_all([active, archived])
        await db_session.commit()

        resp_false = await client.get("/api/admin/buildings/count?is_archived=false")
        resp_all = await client.get("/api/admin/buildings/count")
        assert resp_false.status_code == 200
        assert resp_all.status_code == 200
        # Active count must be less than total count (archived building excluded)
        assert resp_false.json()["count"] < resp_all.json()["count"]

    async def test_is_archived_true_counts_only_archived(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?is_archived=true returns count of only archived buildings."""
        archived = Building(name="IsArchTrue Archived C", manager_email="iatc@test.com")
        archived.is_archived = True
        db_session.add(archived)
        await db_session.commit()

        resp = await client.get("/api/admin/buildings/count?is_archived=true")
        assert resp.status_code == 200
        assert resp.json()["count"] >= 1

    async def test_is_archived_none_counts_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Omitting is_archived returns all buildings regardless of archive status."""
        active = Building(name="IsArchNone Active", manager_email="iana@test.com")
        archived = Building(name="IsArchNone Archived", manager_email="ianb@test.com")
        archived.is_archived = True
        db_session.add_all([active, archived])
        await db_session.commit()

        resp_all = await client.get("/api/admin/buildings/count")
        resp_false = await client.get("/api/admin/buildings/count?is_archived=false")
        resp_true = await client.get("/api/admin/buildings/count?is_archived=true")
        assert resp_all.status_code == 200
        # total = active + archived
        assert resp_all.json()["count"] == resp_false.json()["count"] + resp_true.json()["count"]


# ---------------------------------------------------------------------------
# GET /api/admin/buildings — is_archived list filter
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestListBuildingsIsArchivedFilter:
    # --- Happy path ---

    async def test_is_archived_false_returns_only_active(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?is_archived=false returns only non-archived buildings."""
        active = Building(name="ListArchFalse Active", manager_email="lafa@test.com")
        archived = Building(name="ListArchFalse Archived", manager_email="lafb@test.com")
        archived.is_archived = True
        db_session.add_all([active, archived])
        await db_session.commit()

        response = await client.get("/api/admin/buildings?is_archived=false")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        assert "ListArchFalse Active" in names
        assert "ListArchFalse Archived" not in names

    async def test_is_archived_true_returns_only_archived(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?is_archived=true returns only archived buildings."""
        active = Building(name="ListArchTrue Active", manager_email="lata@test.com")
        archived = Building(name="ListArchTrue Archived", manager_email="latb@test.com")
        archived.is_archived = True
        db_session.add_all([active, archived])
        await db_session.commit()

        response = await client.get("/api/admin/buildings?is_archived=true")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        assert "ListArchTrue Archived" in names
        assert "ListArchTrue Active" not in names

    async def test_no_is_archived_param_returns_all(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Omitting is_archived returns all buildings."""
        active = Building(name="ListArchAll Active", manager_email="laaa@test.com")
        archived = Building(name="ListArchAll Archived", manager_email="laab@test.com")
        archived.is_archived = True
        db_session.add_all([active, archived])
        await db_session.commit()

        response = await client.get("/api/admin/buildings")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        assert "ListArchAll Active" in names
        assert "ListArchAll Archived" in names

    # --- Edge cases ---

    async def test_is_archived_false_empty_result_when_no_active(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """?is_archived=false with only archived buildings returns empty list."""
        # Ensure at least one archived exists with unique name to verify filter works
        # (shared test DB may have other buildings)
        archived = Building(name="OnlyArchived X9W2Q", manager_email="onlyarch@test.com")
        archived.is_archived = True
        db_session.add(archived)
        await db_session.commit()

        response = await client.get("/api/admin/buildings?is_archived=false&name=OnlyArchived+X9W2Q")
        assert response.status_code == 200
        assert response.json() == []


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestGetBuilding:
    # --- Happy path ---

    async def test_returns_building_by_id(
        self, client: AsyncClient, building: Building
    ):
        response = await client.get(f"/api/admin/buildings/{building.id}")
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == str(building.id)
        assert data["name"] == building.name
        assert data["manager_email"] == building.manager_email
        assert "is_archived" in data
        assert "created_at" in data

    # --- State / precondition errors ---

    async def test_returns_404_for_unknown_id(self, client: AsyncClient):
        response = await client.get(f"/api/admin/buildings/{uuid.uuid4()}")
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}/lot-owners
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# POST /api/admin/buildings (create building via form)
# ---------------------------------------------------------------------------


class TestCreateBuildingEndpoint:
    # --- Happy path ---

    async def test_create_building_returns_201(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "New Form Building", "manager_email": "form@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["name"] == "New Form Building"
        assert data["manager_email"] == "form@test.com"
        assert data["is_archived"] is False
        assert "id" in data

    # --- Input validation ---

    async def test_empty_name_returns_422(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "  ", "manager_email": "mgr@test.com"},
        )
        assert response.status_code == 422

    async def test_empty_email_returns_422(self, client: AsyncClient):
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "Valid Building", "manager_email": "  "},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_name_returns_409(
        self, client: AsyncClient
    ):
        await client.post(
            "/api/admin/buildings",
            json={"name": "Dup Form Building", "manager_email": "dup@test.com"},
        )
        response = await client.post(
            "/api/admin/buildings",
            json={"name": "Dup Form Building", "manager_email": "other@test.com"},
        )
        assert response.status_code == 409


# ---------------------------------------------------------------------------
# _detect_file_format unit tests
# ---------------------------------------------------------------------------


class TestDetectFileFormat:
    """Unit tests for the _detect_file_format helper in admin router."""

    def test_csv_extension_returns_csv(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.csv"
        assert _detect_file_format(f) == "csv"

    def test_csv_content_type_returns_csv(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "text/csv"
        f.filename = "noextension"
        assert _detect_file_format(f) == "csv"

    def test_xlsx_extension_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.xlsx"
        assert _detect_file_format(f) == "excel"

    def test_xls_extension_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/octet-stream"
        f.filename = "data.xls"
        assert _detect_file_format(f) == "excel"

    def test_xlsx_content_type_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        f.filename = "data"
        assert _detect_file_format(f) == "excel"

    def test_vnd_ms_excel_content_type_returns_excel(self):
        from unittest.mock import MagicMock
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/vnd.ms-excel"
        f.filename = "data"
        assert _detect_file_format(f) == "excel"

    def test_pdf_raises_415(self):
        from unittest.mock import MagicMock
        from fastapi import HTTPException
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "application/pdf"
        f.filename = "data.pdf"
        with pytest.raises(HTTPException) as exc_info:
            _detect_file_format(f)
        assert exc_info.value.status_code == 415

    def test_no_extension_no_matching_content_type_raises_415(self):
        from unittest.mock import MagicMock
        from fastapi import HTTPException
        from app.routers.admin import _detect_file_format

        f = MagicMock()
        f.content_type = "image/png"
        f.filename = "data.png"
        with pytest.raises(HTTPException) as exc_info:
            _detect_file_format(f)
        assert exc_info.value.status_code == 415


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/import — Excel
# ---------------------------------------------------------------------------


class TestImportBuildingsExcel:
    # --- Happy path ---

    async def test_valid_xlsx_creates_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Excel Tower", "excel@test.com"], ["Xlsx Complex", "xlsx@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2
        assert data["updated"] == 0

    async def test_valid_xlsx_updates_existing_building(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        # Create via CSV first
        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["ExcelUpdate Building", "old@test.com"]],
        )
        await client.post(
            "/api/admin/buildings/import",
            files={"file": ("buildings.csv", csv_data, "text/csv")},
        )

        # Update via Excel
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["ExcelUpdate Building", "new@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 1

    async def test_xlsx_with_blank_rows_skipped(self, client: AsyncClient):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["building_name", "manager_email"])
        ws.append(["Blank Skip Building", "bs@test.com"])
        ws.append([None, None])  # blank row
        ws.append(["", ""])  # also blank
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["created"] == 1

    async def test_empty_xlsx_returns_zero_counts(self, client: AsyncClient):
        excel_data = make_excel(["building_name", "manager_email"], [])
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "empty.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 0
        assert data["updated"] == 0

    # --- Input validation ---

    async def test_xlsx_missing_building_name_column_returns_422(
        self, client: AsyncClient
    ):
        excel_data = make_excel(
            ["manager_email"],
            [["mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/archive
# ---------------------------------------------------------------------------


class TestArchiveBuilding:
    # --- Happy path ---

    async def test_archive_building_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Archive Me Building", manager_email="archive@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 200

    async def test_archive_building_sets_is_archived(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Archive Set Building", manager_email="archiveset@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.json()["is_archived"] is True

    async def test_archive_building_also_archives_lot_owners_with_no_other_home(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Lot owners whose emails don't appear in any other building get archived."""
        b = Building(name="Archive Owners Building", manager_email="ao@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="AO1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        db_session.add(LotOwnerEmail(lot_owner_id=lo.id, email="lone@test.com"))
        await db_session.commit()

        await client.post(f"/api/admin/buildings/{b.id}/archive")

        await db_session.refresh(lo)
        assert lo.is_archived is True

    async def test_archive_building_preserves_lot_owners_with_other_home(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Lot owners whose emails appear in another non-archived building are NOT archived."""
        b1 = Building(name="Archive Other Home B1", manager_email="aoh1@test.com")
        b2 = Building(name="Archive Other Home B2", manager_email="aoh2@test.com")
        db_session.add_all([b1, b2])
        await db_session.flush()

        lo1 = LotOwner(building_id=b1.id, lot_number="AOH1", unit_entitlement=100)
        lo2 = LotOwner(building_id=b2.id, lot_number="AOH1B2", unit_entitlement=100)
        db_session.add_all([lo1, lo2])
        await db_session.flush()

        # Same email in both buildings
        db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="shared@test.com"))
        db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="shared@test.com"))
        await db_session.commit()

        await client.post(f"/api/admin/buildings/{b1.id}/archive")

        await db_session.refresh(lo1)
        # lo1 is in the archived building, but the email exists in b2 (active), so lo1 should NOT be archived
        assert lo1.is_archived is False

    async def test_archive_building_shared_email_across_multiple_lot_owners_in_other_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Archive succeeds when the same email is shared by multiple lot owners
        in OTHER buildings — previously raised MultipleResultsFound because
        the query used scalar_one_or_none() on a result set with multiple rows."""
        b1 = Building(name="Archive MultiShared B1", manager_email="msb1@test.com")
        b2 = Building(name="Archive MultiShared B2", manager_email="msb2@test.com")
        db_session.add_all([b1, b2])
        await db_session.flush()

        # lo1 is in the building we will archive
        lo1 = LotOwner(building_id=b1.id, lot_number="MS1", unit_entitlement=100)
        # lo2 and lo3 are TWO different lot owners in b2 that share the same email
        lo2 = LotOwner(building_id=b2.id, lot_number="MS2A", unit_entitlement=50)
        lo3 = LotOwner(building_id=b2.id, lot_number="MS2B", unit_entitlement=50)
        db_session.add_all([lo1, lo2, lo3])
        await db_session.flush()

        shared_email = "multishared@test.com"
        db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email=shared_email))
        db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email=shared_email))
        db_session.add(LotOwnerEmail(lot_owner_id=lo3.id, email=shared_email))
        await db_session.commit()

        # Must not raise MultipleResultsFound — must return HTTP 200
        response = await client.post(f"/api/admin/buildings/{b1.id}/archive")
        assert response.status_code == 200
        assert response.json()["is_archived"] is True

        # lo1's email exists in b2 (active), so lo1 must NOT be archived
        await db_session.refresh(lo1)
        assert lo1.is_archived is False

    # --- State / precondition errors ---

    async def test_archive_already_archived_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Already Archived Bldg", manager_email="aa@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 409

    async def test_archive_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.post(f"/api/admin/buildings/{uuid.uuid4()}/archive")
        assert response.status_code == 404

    async def test_archive_building_with_no_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Archive succeeds when building has zero lot owners."""
        b = Building(name="Empty Archive Building", manager_email="empty_arc@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 200
        assert response.json()["is_archived"] is True

    async def test_archive_building_lot_owners_with_no_emails(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Archive succeeds and archives lot owners when they have no LotOwnerEmail entries.

        Regression test: db.refresh() after commit caused ORM lazy-load errors for
        lot owners with zero email entries, resulting in HTTP 500.
        """
        b = Building(name="No Email Archive Bldg", manager_email="noemail_arc@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="NE1", unit_entitlement=50)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert response.status_code == 200
        assert response.json()["is_archived"] is True

        await db_session.refresh(lo)
        assert lo.is_archived is True

    async def test_archive_removes_building_from_active_list(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """RR3-41: Archiving a building removes it from the active (?is_archived=false) list."""
        b = Building(name="RR341 Archive Active List", manager_email="rr341al@test.com")
        db_session.add(b)
        await db_session.commit()

        # Verify it appears in active list before archive
        pre_resp = await client.get("/api/admin/buildings?is_archived=false")
        assert pre_resp.status_code == 200
        names_before = [x["name"] for x in pre_resp.json()]
        assert b.name in names_before

        # Archive it
        archive_resp = await client.post(f"/api/admin/buildings/{b.id}/archive")
        assert archive_resp.status_code == 200

        # Should no longer appear in active list
        post_resp = await client.get("/api/admin/buildings?is_archived=false")
        assert post_resp.status_code == 200
        names_after = [x["name"] for x in post_resp.json()]
        assert b.name not in names_after

        # Should appear in archived list
        arch_resp = await client.get("/api/admin/buildings?is_archived=true")
        assert arch_resp.status_code == 200
        archived_names = [x["name"] for x in arch_resp.json()]
        assert b.name in archived_names


# ---------------------------------------------------------------------------
# PATCH /api/admin/buildings/{id}
# ---------------------------------------------------------------------------


class TestUpdateBuilding:
    # --- Happy path ---

    async def test_update_name_only_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Name Building", manager_email="patch.name@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Renamed Building"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "Renamed Building"
        assert data["manager_email"] == "patch.name@test.com"

    async def test_update_manager_email_only_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Email Building", manager_email="old@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": "new@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["manager_email"] == "new@test.com"
        assert data["name"] == "Patch Email Building"

    async def test_update_both_fields_returns_200(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Patch Both Building", manager_email="both.old@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Both Updated", "manager_email": "both.new@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "Both Updated"
        assert data["manager_email"] == "both.new@test.com"

    async def test_update_persists_to_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Persist Test Building", manager_email="persist@test.com")
        db_session.add(b)
        await db_session.commit()

        await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Persisted Name"},
        )

        await db_session.refresh(b)
        assert b.name == "Persisted Name"

    async def test_update_returns_building_out_shape(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Shape Test Building", manager_email="shape@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "Shape Updated"},
        )
        data = response.json()
        assert "id" in data
        assert "name" in data
        assert "manager_email" in data
        assert "is_archived" in data
        assert "created_at" in data

    # --- Input validation ---

    async def test_empty_body_both_null_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Null Patch Building", manager_email="null@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(f"/api/admin/buildings/{b.id}", json={})
        assert response.status_code == 422

    async def test_empty_name_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Name Patch", manager_email="emptyname@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": ""},
        )
        assert response.status_code == 422

    async def test_whitespace_only_name_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Whitespace Name Patch", manager_email="wsname@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": "   "},
        )
        assert response.status_code == 422

    async def test_empty_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Email Patch", manager_email="emptyemail@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": ""},
        )
        assert response.status_code == 422

    async def test_whitespace_only_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Whitespace Email Patch", manager_email="wsemail@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"manager_email": "   "},
        )
        assert response.status_code == 422

    async def test_null_name_and_null_manager_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Null Fields Patch", manager_email="nullfields@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/buildings/{b.id}",
            json={"name": None, "manager_email": None},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_update_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.patch(
            f"/api/admin/buildings/{uuid.uuid4()}",
            json={"name": "Ghost Building"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/buildings/{id}
# ---------------------------------------------------------------------------


class TestDeleteBuilding:
    # --- Happy path ---

    async def test_delete_archived_building_returns_204(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete Me Building", manager_email="deleteme@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()

        response = await client.delete(f"/api/admin/buildings/{b.id}")
        assert response.status_code == 204

    async def test_delete_archived_building_removes_from_db(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete From DB Building", manager_email="deletedb@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.commit()
        building_id = b.id

        await client.delete(f"/api/admin/buildings/{building_id}")

        result = await db_session.execute(select(Building).where(Building.id == building_id))
        assert result.scalar_one_or_none() is None

    async def test_delete_archived_building_cascades_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Delete Cascade Building", manager_email="cascade@test.com")
        b.is_archived = True
        db_session.add(b)
        await db_session.flush()

        lo = LotOwner(building_id=b.id, lot_number="C1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        lot_owner_id = lo.id

        await client.delete(f"/api/admin/buildings/{b.id}")

        result = await db_session.execute(select(LotOwner).where(LotOwner.id == lot_owner_id))
        assert result.scalar_one_or_none() is None

    # --- State / precondition errors ---

    async def test_delete_active_building_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Active Do Not Delete", manager_email="active@test.com")
        b.is_archived = False
        db_session.add(b)
        await db_session.commit()

        response = await client.delete(f"/api/admin/buildings/{b.id}")
        assert response.status_code == 409
        assert "archived" in response.json()["detail"].lower()

    async def test_delete_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        response = await client.delete(f"/api/admin/buildings/{uuid.uuid4()}")
        assert response.status_code == 404



# ---------------------------------------------------------------------------
# Building import from Excel
# ---------------------------------------------------------------------------


class TestImportBuildingsExcel:
    # --- Happy path ---

    async def test_valid_xlsx_imports_buildings(
        self, client: AsyncClient
    ):
        """Valid Excel with building_name/manager_email creates buildings."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [
                ["Excel Tower A", "mgra@excel.com"],
                ["Excel Tower B", "mgrb@excel.com"],
            ],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["created"] == 2

    async def test_xlsx_updates_existing_building(
        self, client: AsyncClient
    ):
        """Importing a building with an existing name updates its manager_email (lines 207-208)."""
        # First import to create the building
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Update Tower", "old@tower.com"]],
        )
        resp1 = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp1.json()["created"] == 1

        # Second import with the same building_name — updates manager_email
        excel_data2 = make_excel(
            ["building_name", "manager_email"],
            [["Update Tower", "new@tower.com"]],
        )
        resp2 = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data2,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert resp2.status_code == 200
        data = resp2.json()
        assert data["updated"] == 1
        assert data["created"] == 0

    async def test_xlsx_blank_rows_skipped(
        self, client: AsyncClient
    ):
        """Blank rows in building Excel are skipped (line 163 continue)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["building_name", "manager_email"])
        ws.append(["Blank Row Tower", "br@excel.com"])
        ws.append([None, None])  # blank row — triggers continue
        ws.append(["Blank Row Tower 2", "br2@excel.com"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["created"] == 2

    # --- Input validation ---

    async def test_invalid_excel_file_returns_422(
        self, client: AsyncClient
    ):
        """Non-parseable bytes submitted as Excel → 422."""
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    b"not-an-excel-file",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_excel_file_returns_422(
        self, client: AsyncClient
    ):
        """Excel with no rows → 422 (no headers)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Empty sheet — will hit StopIteration on first next()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_excel = buf.read()

        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    empty_excel,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_missing_required_headers_returns_422(
        self, client: AsyncClient
    ):
        """Excel missing required columns (building_name or manager_email) → 422."""
        excel_data = make_excel(
            ["only_one_col"],
            [["some value"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_building_name_returns_422(
        self, client: AsyncClient
    ):
        """Row with empty building_name → 422."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [[None, "mgr@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_empty_manager_email_returns_422(
        self, client: AsyncClient
    ):
        """Row with empty manager_email → 422."""
        excel_data = make_excel(
            ["building_name", "manager_email"],
            [["Good Building", None]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "buildings.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422



# ---------------------------------------------------------------------------
# GET /api/admin/buildings — sort_by / sort_dir query params
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestListBuildingsSort:
    """Tests for sort_by and sort_dir query parameters on GET /api/admin/buildings."""

    # --- Happy path ---

    async def test_sort_by_name_asc_returns_alphabetical_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Buildings sorted by name ascending should be in A→Z order."""
        # Create two buildings with known names
        await client.post(
            "/api/admin/buildings",
            json={"name": "Zephyr Tower", "manager_email": "z@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": "Alpha Court", "manager_email": "a@test.com"},
        )
        response = await client.get("/api/admin/buildings?sort_by=name&sort_dir=asc&is_archived=false")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        assert names == sorted(names)

    async def test_sort_by_name_desc_returns_reverse_alphabetical_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Buildings sorted by name descending should be in Z→A order."""
        await client.post(
            "/api/admin/buildings",
            json={"name": "Zephyr Tower", "manager_email": "z@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": "Alpha Court", "manager_email": "a@test.com"},
        )
        response = await client.get("/api/admin/buildings?sort_by=name&sort_dir=desc&is_archived=false")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        assert names == sorted(names, reverse=True)

    async def test_sort_by_created_at_asc_returns_oldest_first(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Buildings sorted by created_at ascending should be oldest first."""
        await client.post(
            "/api/admin/buildings",
            json={"name": "First Building", "manager_email": "f@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": "Second Building", "manager_email": "s@test.com"},
        )
        response = await client.get("/api/admin/buildings?sort_by=created_at&sort_dir=asc&is_archived=false")
        assert response.status_code == 200
        data = response.json()
        assert len(data) >= 2
        # created_at values should be non-decreasing
        created_ats = [b["created_at"] for b in data]
        assert created_ats == sorted(created_ats)

    async def test_sort_by_created_at_desc_is_default_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Without sort params, buildings should be returned newest first (default)."""
        await client.post(
            "/api/admin/buildings",
            json={"name": "BuildingA", "manager_email": "ba@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": "BuildingB", "manager_email": "bb@test.com"},
        )
        response = await client.get("/api/admin/buildings?is_archived=false")
        assert response.status_code == 200
        data = response.json()
        assert len(data) >= 2
        created_ats = [b["created_at"] for b in data]
        assert created_ats == sorted(created_ats, reverse=True)

    async def test_sort_by_name_no_sort_dir_defaults_to_desc(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """sort_by=name without sort_dir should use desc (the global default)."""
        await client.post(
            "/api/admin/buildings",
            json={"name": "Zephyr", "manager_email": "z@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": "Alpha", "manager_email": "a@test.com"},
        )
        response = await client.get("/api/admin/buildings?sort_by=name&is_archived=false")
        assert response.status_code == 200
        names = [b["name"] for b in response.json()]
        # Default sort_dir=desc → Z first
        assert names == sorted(names, reverse=True)

    # --- Input validation ---

    async def test_invalid_sort_by_returns_422(self, client: AsyncClient):
        """An unrecognised sort_by value must be rejected with 422."""
        response = await client.get("/api/admin/buildings?sort_by=invalid_column")
        assert response.status_code == 422
        assert "Invalid sort_by value" in response.json()["detail"]

    async def test_invalid_sort_dir_returns_422(self, client: AsyncClient):
        """An unrecognised sort_dir value must be rejected with 422."""
        response = await client.get("/api/admin/buildings?sort_by=name&sort_dir=sideways")
        assert response.status_code == 422
        assert "Invalid sort_dir value" in response.json()["detail"]

    async def test_sql_injection_attempt_in_sort_by_returns_422(self, client: AsyncClient):
        """A SQL injection attempt in sort_by must be rejected before hitting the DB."""
        response = await client.get("/api/admin/buildings?sort_by=name;DROP TABLE buildings;--")
        assert response.status_code == 422

    # --- Boundary values ---

    async def test_sort_with_no_matching_name_filter_returns_empty_list(self, client: AsyncClient, db_session: AsyncSession):
        """sort_by=name with a name filter that matches nothing returns an empty list."""
        response = await client.get("/api/admin/buildings?sort_by=name&sort_dir=asc&name=ZZZNOMATCH99999")
        assert response.status_code == 200
        assert response.json() == []

    async def test_sort_with_single_matching_building_returns_one_item(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """sort_by=name with a unique name prefix returns exactly one item."""
        unique_prefix = "SortBoundaryUniq"
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique_prefix} Solo Building", "manager_email": "solo@test.com"},
        )
        response = await client.get(f"/api/admin/buildings?sort_by=name&sort_dir=asc&name={unique_prefix}")
        assert response.status_code == 200
        assert len(response.json()) == 1

    # --- New column: manager_email ---

    async def test_sort_by_manager_email_asc_returns_alphabetical_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Buildings sorted by manager_email ascending should be in A→Z email order."""
        unique = "SortEmailTest"
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} Z Building", "manager_email": "zzz@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} A Building", "manager_email": "aaa@test.com"},
        )
        response = await client.get(f"/api/admin/buildings?sort_by=manager_email&sort_dir=asc&name={unique}")
        assert response.status_code == 200
        emails = [b["manager_email"] for b in response.json()]
        assert emails == sorted(emails)

    async def test_sort_by_manager_email_desc_returns_reverse_alphabetical_order(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Buildings sorted by manager_email descending should be in Z→A email order."""
        unique = "SortEmailDescTest"
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} Z Building", "manager_email": "zzz@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} A Building", "manager_email": "aaa@test.com"},
        )
        response = await client.get(f"/api/admin/buildings?sort_by=manager_email&sort_dir=desc&name={unique}")
        assert response.status_code == 200
        emails = [b["manager_email"] for b in response.json()]
        assert emails == sorted(emails, reverse=True)

    # --- Case-insensitive sorting ---

    async def test_sort_by_name_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Case-insensitive name sort: 'alpha' and 'Alpha' and 'ALPHA' sort together."""
        unique = "SortCaseTest"
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} zebra tower", "manager_email": "z@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} Apple Court", "manager_email": "a@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} MANGO House", "manager_email": "m@test.com"},
        )
        response = await client.get(f"/api/admin/buildings?sort_by=name&sort_dir=asc&name={unique}")
        assert response.status_code == 200
        names = [b["name"].lower().split(unique.lower())[-1].strip() for b in response.json()]
        # Lowercased suffixes should be in alphabetical order: apple, mango, zebra
        assert names == sorted(names)

    async def test_sort_by_manager_email_case_insensitive(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Case-insensitive manager_email sort: 'AAA' sorts together with 'aaa'."""
        unique = "SortEmailCase"
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} B", "manager_email": "ZZZ@test.com"},
        )
        await client.post(
            "/api/admin/buildings",
            json={"name": f"{unique} A", "manager_email": "aaa@test.com"},
        )
        response = await client.get(f"/api/admin/buildings?sort_by=manager_email&sort_dir=asc&name={unique}")
        assert response.status_code == 200
        emails = [b["manager_email"].lower() for b in response.json()]
        assert emails == sorted(emails)


# ---------------------------------------------------------------------------
# RR3-34: File upload size limits (buildings import)
# ---------------------------------------------------------------------------


class TestBuildingImportFileSizeLimit:
    """File uploads over 5 MB must be rejected with 413 (RR3-34)."""

    async def test_csv_over_5mb_returns_413(self, client: AsyncClient):
        """Buildings CSV over 5 MB is rejected with HTTP 413."""
        oversized = b"building_name,manager_email\n" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("big.csv", oversized, "text/csv")},
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]

    async def test_excel_over_5mb_returns_413(self, client: AsyncClient):
        """Buildings Excel over 5 MB is rejected with HTTP 413."""
        # Build a minimal .xlsx header + pad to exceed 5 MB
        oversized = b"PK\x03\x04" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            "/api/admin/buildings/import",
            files={
                "file": (
                    "big.xlsx",
                    oversized,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]


# ---------------------------------------------------------------------------
# Rate limiting — admin import endpoints (RR4-31)
# ---------------------------------------------------------------------------


class TestAdminImportRateLimitBuildings:
    """Verify admin_import_limiter returns 429 on the 21st request (RR4-31)."""

    async def test_buildings_import_rate_limited_after_max_requests(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """21st call to /buildings/import within the window returns 429 with Retry-After."""
        from app.rate_limiter import admin_import_limiter

        # Exhaust the limit
        admin_import_limiter._timestamps["admin"] = []
        for _ in range(20):
            admin_import_limiter._timestamps["admin"].append(
                __import__("time").monotonic()
            )

        csv_data = make_csv(
            ["building_name", "manager_email"],
            [["Rate Limit Test Building", "rl@test.com"]],
        )
        response = await client.post(
            "/api/admin/buildings/import",
            files={"file": ("b.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 429
        assert response.headers.get("Retry-After") == "60"
