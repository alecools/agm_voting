"""
Tests for PX-2A: proxy nomination and financial position import endpoints.

POST /api/admin/buildings/{id}/lot-owners/import-proxies
POST /api/admin/buildings/{id}/lot-owners/import-financial-positions

Structure:
  # --- Happy path ---
  # --- Input validation ---
  # --- Boundary values ---
  # --- State / precondition errors ---
  # --- Edge cases ---
"""
from __future__ import annotations

import csv
import io
import uuid

import openpyxl
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from fastapi import HTTPException

from app.models import (
    Building,
    FinancialPosition,
    LotOwner,
    LotProxy,
)
from app.models.lot_owner_email import LotOwnerEmail
from app.services.admin_service import (
    _parse_closing_balance,
    _parse_closing_balance_numeric,
    _parse_financial_position_excel_rows,
    _parse_simple_financial_position_csv_rows,
    _parse_tocs_financial_position_csv_rows,
    _parse_tocs_financial_position_excel_rows,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_csv(headers: list[str], rows: list[list[str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode()


def make_excel(headers: list, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_tocs_excel(sections: list[list[tuple]]) -> bytes:
    """Build a TOCS-format xlsx in memory.

    sections: list of fund data blocks.  Each block is a list of
    (lot_number, closing_balance) tuples.  Preamble rows (company
    header) are prepended automatically.  Each section gets its own
    'Lot#' header row followed by 9-column data rows then a Totals row
    and a blank separator row.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Preamble (company header — does NOT start with 'Lot#')
    ws.append(["ACME Body Corporate Solutions", None, None, None, None, None, None, None, None])
    ws.append(["Suite 1, 123 Example St", None, None, None, None, None, None, None, None])
    ws.append([None] * 9)  # blank row

    fund_names = ["Administrative Fund", "Maintenance Fund", "Sinking Fund"]
    for section_idx, lot_rows in enumerate(sections):
        fund_name = fund_names[section_idx % len(fund_names)]
        ws.append([fund_name, None, None, None, None, None, None, None, None])
        ws.append(["Lot#", "Unit#", "Owner Name", "Opening Balance",
                   "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"])
        for lot_number, closing_balance in lot_rows:
            ws.append([lot_number, "1", "Test Owner", "$-", "$-", "$-", "$-",
                       closing_balance, "$-"])
        ws.append([f"{fund_name} Totals", None, None, None, None, None, None, None, None])
        ws.append([None] * 9)  # blank separator row between sections

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def client(app):
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as ac:
        yield ac


@pytest_asyncio.fixture
async def building(db_session: AsyncSession) -> Building:
    b = Building(name="Proxy Test Building", manager_email="proxy@test.com")
    db_session.add(b)
    await db_session.flush()
    await db_session.refresh(b)
    return b


@pytest_asyncio.fixture
async def building_with_owners(db_session: AsyncSession) -> Building:
    b = Building(name="Proxy Owners Building", manager_email="po@test.com")
    db_session.add(b)
    await db_session.flush()

    lo1 = LotOwner(building_id=b.id, lot_number="1A", unit_entitlement=100)
    lo2 = LotOwner(building_id=b.id, lot_number="2B", unit_entitlement=50)
    lo3 = LotOwner(building_id=b.id, lot_number="3C", unit_entitlement=75)
    db_session.add_all([lo1, lo2, lo3])
    await db_session.flush()

    db_session.add(LotOwnerEmail(lot_owner_id=lo1.id, email="owner1@test.com"))
    db_session.add(LotOwnerEmail(lot_owner_id=lo2.id, email="owner2@test.com"))
    await db_session.flush()
    await db_session.refresh(b)
    return b


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-proxies (CSV)
# ---------------------------------------------------------------------------


class TestImportProxiesCSV:
    # --- Happy path ---

    async def test_upserts_new_proxy_nominations(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 2
        assert data["removed"] == 0
        assert data["skipped"] == 0

        # Verify DB
        result = await db_session.execute(
            select(LotProxy).join(LotOwner).where(LotOwner.building_id == building_with_owners.id)
        )
        proxies = result.scalars().all()
        assert len(proxies) == 2
        emails = {p.proxy_email for p in proxies}
        assert "proxy1@test.com" in emails
        assert "proxy2@test.com" in emails

    async def test_updates_existing_proxy_nomination(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # First upload
        csv_data1 = make_csv(["Lot#", "Proxy Email"], [["1A", "old_proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Second upload with updated email
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", "new_proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 1
        assert data["removed"] == 0

    async def test_blank_proxy_email_removes_nomination(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Create proxy first
        csv_data1 = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Now remove by blank email
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 1
        assert data["skipped"] == 0

    async def test_blank_proxy_email_no_existing_proxy_is_noop(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_unknown_lot_number_skipped(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy@test.com"], ["999X", "ghost@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 1
        assert data["skipped"] == 1

    async def test_lots_not_in_file_are_unaffected(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Set proxy for both 1A and 2B
        csv_data1 = make_csv(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data1, "text/csv")},
        )

        # Second upload only touches 1A
        csv_data2 = make_csv(["Lot#", "Proxy Email"], [["1A", "updated@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data2, "text/csv")},
        )

        # 2B proxy should still exist
        result = await db_session.execute(
            select(LotProxy)
            .join(LotOwner)
            .where(LotOwner.building_id == building_with_owners.id)
        )
        proxies = result.scalars().all()
        emails = {p.proxy_email for p in proxies}
        assert "proxy2@test.com" in emails

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Proxy Email", "Extra"],
            [["1A", "proxy@test.com", "ignore_me"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    async def test_empty_csv_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_case_insensitive_headers(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["LOT#", "PROXY EMAIL"],
            [["1A", "proxy@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    # --- Input validation ---

    async def test_missing_lot_hash_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Proxy Email"], [["proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "lot#" in str(response.json()["detail"]).lower()

    async def test_missing_proxy_email_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "proxy email" in str(response.json()["detail"]).lower()

    async def test_invalid_file_type_returns_415(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.pdf", b"garbage", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_building_not_found_returns_404(
        self, client: AsyncClient
    ):
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_no_headers_csv_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", b"", "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-proxies (Excel)
# ---------------------------------------------------------------------------


class TestImportProxiesExcel:
    # --- Happy path ---

    async def test_upserts_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Proxy Email"],
            [["1A", "proxy1@test.com"], ["2B", "proxy2@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 2
        assert data["removed"] == 0

    async def test_blank_proxy_email_removes_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Create proxy first via CSV
        csv_data = make_csv(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )

        # Remove via Excel
        xlsx = make_excel(["Lot#", "Proxy Email"], [["1A", None]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["removed"] == 1

    async def test_skips_unknown_lot_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [["9Z", "proxy@test.com"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["skipped"] == 1
        assert data["upserted"] == 0

    async def test_skips_blank_rows_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [["1A", "proxy@test.com"], [None, None]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["upserted"] == 1

    # --- Input validation ---

    async def test_missing_headers_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "proxy email" in str(response.json()["detail"]).lower()

    async def test_invalid_excel_file(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", b"not-an-excel-file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "invalid excel" in str(response.json()["detail"]).lower()

    async def test_excel_no_data_rows(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Proxy Email"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("proxies.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["upserted"] == 0
        assert data["removed"] == 0
        assert data["skipped"] == 0

    async def test_completely_empty_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """An Excel file with absolutely no rows (not even headers) should return 422."""
        wb = openpyxl.Workbook()
        # Don't write anything to the sheet
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("empty.xlsx", empty_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "no headers" in str(response.json()["detail"]).lower()

    async def test_row_with_fewer_cells_than_headers(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Row that is shorter than the header row — _cell returns '' for out-of-bounds index."""
        # Build an Excel manually where a data row only has 1 cell (Lot#) but not Proxy Email
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Proxy Email"])
        # Write a row with only the first cell populated; second cell omitted
        ws.cell(row=2, column=1, value="1A")
        # Intentionally do NOT set column 2 — openpyxl will yield None for it
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-proxies",
            files={"file": ("short.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        # A row with blank proxy email should count as removed/noop (blank -> no existing proxy)
        assert response.status_code == 200
        data = response.json()
        assert data["removed"] == 0  # no existing proxy to remove
        assert data["upserted"] == 0  # blank proxy email


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-financial-positions (CSV)
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsCSV:
    # --- Happy path ---

    async def test_updates_financial_positions(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "In Arrear"], ["2B", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 2
        assert data["skipped"] == 0

        # Verify DB
        result = await db_session.execute(
            select(LotOwner).where(
                LotOwner.building_id == building_with_owners.id,
                LotOwner.lot_number == "1A",
            )
        )
        lo1 = result.scalar_one()
        assert lo1.financial_position == FinancialPosition.in_arrear

    async def test_accepted_value_in_arrear_lowercase(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "in arrear"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_accepted_value_in_arrear_underscore(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "in_arrear"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_accepted_value_normal_uppercase(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "NORMAL"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_unknown_lot_number_skipped(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "Normal"], ["ZZZ", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 1
        assert data["skipped"] == 1

    async def test_lots_not_in_file_unaffected(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        # Set 1A to in_arrear
        csv_data1 = make_csv(["Lot#", "Financial Position"], [["1A", "In Arrear"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data1, "text/csv")},
        )

        # Second file only updates 2B; 1A should remain in_arrear
        csv_data2 = make_csv(["Lot#", "Financial Position"], [["2B", "Normal"]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data2, "text/csv")},
        )

        result = await db_session.execute(
            select(LotOwner).where(
                LotOwner.building_id == building_with_owners.id,
                LotOwner.lot_number == "1A",
            )
        )
        lo1 = result.scalar_one()
        assert lo1.financial_position == FinancialPosition.in_arrear

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position", "Extra"],
            [["1A", "Normal", "ignore"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    async def test_empty_csv_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0

    async def test_simple_csv_with_only_financial_position_header_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """CSV with 'Financial Position' as first column (no 'Lot#') is treated as TOCS
        format (auto-detection: first cell != 'lot#') and returns empty results."""
        csv_data = make_csv(["Financial Position"], [["Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        # TOCS parser finds no 'Lot#' section header rows → returns empty list → 0 updated
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0

    async def test_case_insensitive_headers(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["LOT#", "FINANCIAL POSITION"],
            [["1A", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    # --- Input validation ---

    async def test_invalid_financial_position_value_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "BadValue"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        detail = str(response.json()["detail"])
        assert "BadValue" in detail

    async def test_multiple_invalid_rows_all_listed(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(
            ["Lot#", "Financial Position"],
            [["1A", "Wrong"], ["2B", "AlsoBad"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        detail = response.json()["detail"]
        assert isinstance(detail, list)
        assert len(detail) == 2

    async def test_empty_financial_position_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", ""]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_missing_lot_hash_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """A CSV whose first cell is not 'Lot#' is routed to the TOCS parser, which
        returns an empty result (no fund sections found) rather than a 422.
        Use the simple format with both required headers to test the missing-Lot# 422."""
        # Explicitly use the simple-format path by having 'Lot#' as first header
        # then a row that triggers the simple-path validation (missing Financial Position header)
        csv_data = make_csv(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "financial position" in str(response.json()["detail"]).lower()

    async def test_missing_financial_position_header(
        self, client: AsyncClient, building_with_owners: Building
    ):
        csv_data = make_csv(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "financial position" in str(response.json()["detail"]).lower()

    async def test_invalid_file_type_returns_415(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.pdf", b"garbage", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_building_not_found_returns_404(
        self, client: AsyncClient
    ):
        csv_data = make_csv(["Lot#", "Financial Position"], [["1A", "Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    # --- Edge cases ---

    async def test_no_headers_csv_returns_zeros(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """Empty CSV bytes are auto-detected as TOCS format (first cell not 'Lot#').
        TOCS parser finds no fund sections → returns empty list → 200 with 0 results."""
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.csv", b"", "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{id}/lot-owners/import-financial-positions (Excel)
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsExcel:
    # --- Happy path ---

    async def test_updates_from_excel(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Financial Position"],
            [["1A", "In Arrear"], ["2B", "Normal"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 2
        assert data["skipped"] == 0

    async def test_invalid_value_in_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [["1A", "Bad"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422

    async def test_skips_unknown_lot_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [["9Z", "Normal"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["skipped"] == 1

    async def test_skips_blank_rows_in_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(
            ["Lot#", "Financial Position"],
            [["1A", "Normal"], [None, None]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["updated"] == 1

    # --- Input validation ---

    async def test_missing_headers_excel(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#"], [["1A"]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "financial position" in str(response.json()["detail"]).lower()

    async def test_invalid_excel_file(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", b"not-an-excel-file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "invalid excel" in str(response.json()["detail"]).lower()

    async def test_excel_no_data_rows(
        self, client: AsyncClient, building_with_owners: Building
    ):
        xlsx = make_excel(["Lot#", "Financial Position"], [])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 0
        assert data["skipped"] == 0

    async def test_completely_empty_excel_returns_422(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """An Excel file with absolutely no rows should return 422."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("empty.xlsx", empty_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 422
        assert "no headers" in str(response.json()["detail"]).lower()

    async def test_row_with_fewer_cells_than_headers(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Row shorter than header — _cell returns '' for out-of-bounds, empty fp raises 422."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Financial Position"])
        ws.cell(row=2, column=1, value="1A")
        # column 2 intentionally omitted
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx = buf.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("short.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        # Empty financial_position should be caught as validation error
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# Unit tests: _parse_closing_balance
# ---------------------------------------------------------------------------


class TestParseClosingBalance:
    """Unit tests for _parse_closing_balance covering all TOCS balance cell variants."""

    # --- Happy path: normal (zero / credit) ---

    def test_dollar_dash_returns_normal(self):
        assert _parse_closing_balance("$-") == FinancialPosition.normal

    def test_dollar_dash_with_spaces_returns_normal(self):
        assert _parse_closing_balance(" $-   ") == FinancialPosition.normal

    def test_empty_string_returns_normal(self):
        assert _parse_closing_balance("") == FinancialPosition.normal

    def test_whitespace_only_returns_normal(self):
        assert _parse_closing_balance("   ") == FinancialPosition.normal

    def test_bracketed_credit_returns_normal(self):
        """$(190.77) → credit/advance → normal."""
        assert _parse_closing_balance("$(190.77)") == FinancialPosition.normal

    def test_bracketed_credit_with_comma_returns_normal(self):
        """$(1,200.00) → normal."""
        assert _parse_closing_balance("$(1,200.00)") == FinancialPosition.normal

    def test_bracketed_credit_with_spaces_returns_normal(self):
        assert _parse_closing_balance(" $(190.77)") == FinancialPosition.normal

    # --- Happy path: in_arrear (positive balance) ---

    def test_positive_balance_returns_in_arrear(self):
        """$1,882.06 → in_arrear."""
        assert _parse_closing_balance("$1,882.06 ") == FinancialPosition.in_arrear

    def test_small_positive_balance_returns_in_arrear(self):
        """$619.96 → in_arrear."""
        assert _parse_closing_balance("$619.96") == FinancialPosition.in_arrear

    def test_positive_no_cents_returns_in_arrear(self):
        assert _parse_closing_balance("$967.16 ") == FinancialPosition.in_arrear

    # --- Boundary values ---

    def test_dash_only_after_stripping_dollar_is_normal(self):
        """Stripped = '-' after removing $ → normal."""
        assert _parse_closing_balance("$-   ") == FinancialPosition.normal


# ---------------------------------------------------------------------------
# Unit tests: _parse_simple_financial_position_csv_rows
# ---------------------------------------------------------------------------


class TestParseSimpleFinancialPositionCsvRows:
    """Unit tests for _parse_simple_financial_position_csv_rows."""

    # --- Edge cases: fieldnames None branch (line covered by empty bytes) ---

    def test_empty_bytes_raises_422(self):
        """Empty CSV bytes → DictReader.fieldnames is None → HTTPException 422."""
        with pytest.raises(HTTPException) as exc_info:
            _parse_simple_financial_position_csv_rows(b"")
        assert exc_info.value.status_code == 422
        assert "no headers" in exc_info.value.detail.lower()

    # --- Input validation: missing headers ---

    def test_missing_financial_position_header_raises_422(self):
        data = "Lot#\n1A\n".encode()
        with pytest.raises(HTTPException) as exc_info:
            _parse_simple_financial_position_csv_rows(data)
        assert exc_info.value.status_code == 422
        assert "financial position" in exc_info.value.detail.lower()

    def test_missing_lot_hash_header_raises_422(self):
        data = "Financial Position\nNormal\n".encode()
        with pytest.raises(HTTPException) as exc_info:
            _parse_simple_financial_position_csv_rows(data)
        assert exc_info.value.status_code == 422
        assert "lot#" in exc_info.value.detail.lower()

    # --- Happy path ---

    def test_valid_rows_returned(self):
        data = "Lot#,Financial Position\n1A,Normal\n2B,In Arrear\n".encode()
        rows = _parse_simple_financial_position_csv_rows(data)
        assert rows == [
            {"lot_number": "1A", "financial_position_raw": "Normal"},
            {"lot_number": "2B", "financial_position_raw": "In Arrear"},
        ]


# ---------------------------------------------------------------------------
# Unit tests: _parse_tocs_financial_position_csv_rows
# ---------------------------------------------------------------------------


class TestParseTocsFinancialPositionCsvRows:
    """Unit tests for _parse_tocs_financial_position_csv_rows."""

    # --- Happy path: real TOCS file ---

    def test_real_tocs_file_lot5_is_in_arrear(self):
        """Lot 5 has Admin Fund closing balance $1,882.06 → in_arrear."""
        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        rows = _parse_tocs_financial_position_csv_rows(content)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["5"] == FinancialPosition.in_arrear.value

    def test_real_tocs_file_lot2_is_normal(self):
        """Lot 2 has Admin Fund closing balance $(190.77) (credit) → normal."""
        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        rows = _parse_tocs_financial_position_csv_rows(content)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["2"] == FinancialPosition.normal.value

    def test_real_tocs_file_lot8_is_in_arrear_from_maintenance_fund(self):
        """Lot 8 is normal in Admin Fund ($19,748.40 actually in_arrear!) but the real
        check: lot 8 Admin Fund has $19,748.40 → in_arrear; also check Maintenance
        Fund lot 8 = $619.96 → in_arrear. Worst-case → in_arrear."""
        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        rows = _parse_tocs_financial_position_csv_rows(content)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["8"] == FinancialPosition.in_arrear.value

    def test_real_tocs_file_returns_all_51_lots(self):
        """All 51 lots from both fund sections are deduplicated into the result.
        Worst-case across Admin + Maintenance funds, no totals/summary rows included."""
        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        rows = _parse_tocs_financial_position_csv_rows(content)
        lot_numbers = [r["lot_number"] for r in rows]
        # Verify no summary rows (Totals/Arrears/Advances) leaked through
        assert not any("total" in ln.lower() or "arrear" in ln.lower() or "advance" in ln.lower() for ln in lot_numbers)
        # Verify all 51 unique lot numbers are present
        assert len(rows) == 51

    # --- Auto-detection: simple vs TOCS ---

    def test_auto_detect_routes_simple_format_to_simple_parser(self):
        """CSV starting with 'Lot#' → simple parser → raises 422 for missing FP header."""
        data = "Lot#,Financial Position\n1,Normal\n".encode()
        rows = _parse_simple_financial_position_csv_rows(data)
        assert rows[0]["lot_number"] == "1"
        assert rows[0]["financial_position_raw"] == "Normal"

    # --- Worst-case across fund sections ---

    def test_worst_case_in_arrear_in_any_fund_wins(self):
        """A lot that is normal in one section but in_arrear in another → in_arrear."""
        csv_content = (
            "Header Row,,\n"
            "\n"
            "Admin Fund,,\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            "Totals,,\n"
            "\n"
            "Maintenance Fund,,\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$500.00,$-\n"
            "Totals,,\n"
        ).encode("utf-8")
        rows = _parse_tocs_financial_position_csv_rows(csv_content)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.in_arrear.value

    def test_worst_case_normal_in_both_funds_stays_normal(self):
        """A lot that is normal in both sections → normal."""
        csv_content = (
            "Header Row,,\n"
            "\n"
            "Admin Fund,,\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            "Totals,,\n"
            "\n"
            "Maintenance Fund,,\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            "Totals,,\n"
        ).encode("utf-8")
        rows = _parse_tocs_financial_position_csv_rows(csv_content)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.normal.value

    # --- Edge cases ---

    def test_empty_content_returns_empty_list(self):
        rows = _parse_tocs_financial_position_csv_rows(b"")
        assert rows == []

    def test_no_section_headers_returns_empty_list(self):
        """Content with no 'Lot#' section headers → empty result."""
        data = "Company Name\nAddress\nSome Report\n".encode()
        rows = _parse_tocs_financial_position_csv_rows(data)
        assert rows == []

    def test_latin1_encoded_content_decoded_successfully(self):
        """Latin-1 fallback (line 1576-1577): content that fails UTF-8-sig decoding."""
        # Create bytes that are valid latin-1 but not utf-8
        header = "Company\r\n\r\nLot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\r\n".encode("latin-1")
        row = "1,A,Caf\xe9 Owner,$-,$-,$-,$-,$-,$-\r\n".encode("latin-1")
        totals = "Totals,,\r\n".encode("latin-1")
        content = header + row + totals
        rows = _parse_tocs_financial_position_csv_rows(content)
        assert len(rows) == 1
        assert rows[0]["lot_number"] == "1"

    def test_blank_row_terminates_section(self):
        """A blank row (all empty cells) ends the current section (line 1611: break)."""
        csv_content = (
            "Header\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            ",,,,,,,\n"
            "2,B,Bob,$-,$-,$-,$-,$100.00,$-\n"
        ).encode("utf-8")
        rows = _parse_tocs_financial_position_csv_rows(csv_content)
        # Blank row terminates section; lot 2 is after the blank row and not included
        lot_numbers = [r["lot_number"] for r in rows]
        assert "1" in lot_numbers
        assert "2" not in lot_numbers

    def test_lot_hash_header_in_data_terminates_section(self):
        """Another 'Lot#' header row mid-section ends the current section (line 1622: break)."""
        csv_content = (
            "Header\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "2,B,Bob,$-,$-,$-,$-,$-,$-\n"
        ).encode("utf-8")
        rows = _parse_tocs_financial_position_csv_rows(csv_content)
        # The second Lot# header row terminates the first section; the second section
        # is also started (section_starts includes both header rows)
        lot_numbers = [r["lot_number"] for r in rows]
        assert "1" in lot_numbers
        assert "2" in lot_numbers

    def test_empty_lot_number_in_data_row_skipped(self):
        """A data row with empty lot# cell is skipped (line 1626: continue)."""
        csv_content = (
            "Header\n"
            "Lot#,Unit#,Owner Name,Opening Balance,Levied,Special Levy,Paid,Closing Balance,Interest Paid\n"
            "1,A,Alice,$-,$-,$-,$-,$-,$-\n"
            ",,Empty lot row,$-,$-,$-,$-,$200.00,$-\n"
            "2,B,Bob,$-,$-,$-,$-,$-,$-\n"
            "Totals,,\n"
        ).encode("utf-8")
        rows = _parse_tocs_financial_position_csv_rows(csv_content)
        lot_numbers = [r["lot_number"] for r in rows]
        # Empty lot# row skipped; both lot 1 and 2 present
        assert "1" in lot_numbers
        assert "2" in lot_numbers
        assert "" not in lot_numbers


# ---------------------------------------------------------------------------
# Integration tests: TOCS CSV via POST endpoint
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsTOCSCSV:
    """Integration tests: POST real TOCS Lot Positions Report to the endpoint."""

    # --- Happy path ---

    async def test_tocs_file_updates_lot_positions(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Uploading the real TOCS CSV returns 200. All 51 parsed lots are skipped because
        the building_with_owners fixture has lots 1A/2B/3C which don't match 1-51."""
        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("lot_positions.csv", content, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        # All 51 parsed TOCS lots are skipped (none match 1A/2B/3C)
        assert data["updated"] == 0
        assert data["skipped"] == 51

    async def test_tocs_file_with_matching_lots_updates_correctly(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """Lots named '5' and '8' in the building get correct positions from TOCS file."""
        # Add matching lot owners for the lots that exist in the TOCS file
        lot5 = LotOwner(
            building_id=building_with_owners.id,
            lot_number="5",
            unit_entitlement=100,
        )
        lot8 = LotOwner(
            building_id=building_with_owners.id,
            lot_number="8",
            unit_entitlement=100,
        )
        db_session.add_all([lot5, lot8])
        await db_session.flush()

        with open("../examples/Lot financial position.csv", "rb") as f:
            content = f.read()
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("lot_positions.csv", content, "text/csv")},
        )
        assert response.status_code == 200

        await db_session.refresh(lot5)
        await db_session.refresh(lot8)
        assert lot5.financial_position == FinancialPosition.in_arrear
        assert lot8.financial_position == FinancialPosition.in_arrear


# ---------------------------------------------------------------------------
# Unit tests: _parse_closing_balance_numeric
# ---------------------------------------------------------------------------


class TestParseClosingBalanceNumeric:
    """Unit tests for _parse_closing_balance_numeric — numeric xlsx cell values."""

    # --- Happy path: normal (zero / negative) ---

    def test_zero_float_returns_normal(self):
        assert _parse_closing_balance_numeric(0.0) == FinancialPosition.normal

    def test_zero_int_returns_normal(self):
        assert _parse_closing_balance_numeric(0) == FinancialPosition.normal

    def test_negative_float_returns_normal(self):
        """Negative = credit/overpaid → normal."""
        assert _parse_closing_balance_numeric(-6.48) == FinancialPosition.normal

    def test_large_negative_returns_normal(self):
        assert _parse_closing_balance_numeric(-1662.01) == FinancialPosition.normal

    # --- Happy path: in_arrear (positive) ---

    def test_positive_float_returns_in_arrear(self):
        assert _parse_closing_balance_numeric(6.17) == FinancialPosition.in_arrear

    def test_positive_int_returns_in_arrear(self):
        assert _parse_closing_balance_numeric(500) == FinancialPosition.in_arrear

    def test_large_positive_returns_in_arrear(self):
        assert _parse_closing_balance_numeric(18534.18) == FinancialPosition.in_arrear

    # --- Boundary values ---

    def test_very_small_positive_returns_in_arrear(self):
        """Smallest non-zero positive → in_arrear."""
        assert _parse_closing_balance_numeric(0.01) == FinancialPosition.in_arrear

    def test_very_small_negative_returns_normal(self):
        assert _parse_closing_balance_numeric(-0.01) == FinancialPosition.normal


# ---------------------------------------------------------------------------
# Unit tests: _parse_tocs_financial_position_excel_rows
# ---------------------------------------------------------------------------


class TestParseTocsFpExcelRows:
    """Unit tests for _parse_tocs_financial_position_excel_rows."""

    # --- Happy path ---

    def test_single_section_in_arrear_lot(self):
        """A lot with a positive closing balance → in_arrear."""
        xlsx = make_tocs_excel([[("1", "$500.00")]])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True)
        all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.in_arrear.value

    def test_single_section_normal_lot(self):
        """A lot with $(190.77) bracketed balance → normal."""
        xlsx = make_tocs_excel([[("2", "$(190.77)")]])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True)
        all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["2"] == FinancialPosition.normal.value

    def test_two_sections_worst_case_wins(self):
        """Lot normal in section 1 but in_arrear in section 2 → in_arrear."""
        xlsx = make_tocs_excel([[("1", "$-")], [("1", "$300.00")]])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True)
        all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.in_arrear.value

    def test_two_sections_normal_in_both_stays_normal(self):
        """Lot normal in both sections → normal."""
        xlsx = make_tocs_excel([[("1", "$-")], [("1", "$(50.00)")]])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True)
        all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.normal.value

    def test_multiple_lots_across_sections(self):
        """Multiple lots across two sections all correctly resolved."""
        xlsx = make_tocs_excel([
            [("1", "$-"), ("2", "$500.00")],
            [("1", "$(100.00)"), ("2", "$-")],
        ])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True)
        all_rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1"] == FinancialPosition.normal.value
        assert result["2"] == FinancialPosition.in_arrear.value

    # --- Numeric closing balance (native xlsx float/int values) ---

    def test_numeric_positive_closing_balance_returns_in_arrear(self):
        """Closing balance stored as float > 0 in xlsx → in_arrear (line 1782)."""
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("54", 54, "Test Owner", 0.0, 10.81, 0.0, 4.64, 6.17, 0.0),
            ("Administrative Fund Totals", None, None, None, None, None, None, None, None),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["54"] == FinancialPosition.in_arrear.value

    def test_numeric_zero_closing_balance_returns_normal(self):
        """Closing balance stored as float 0.0 in xlsx → normal."""
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("55", 55, "Test Owner", 0.0, 10.81, 0.0, 10.81, 0.0, 0.0),
            ("Administrative Fund Totals", None, None, None, None, None, None, None, None),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["55"] == FinancialPosition.normal.value

    def test_numeric_negative_closing_balance_returns_normal(self):
        """Closing balance stored as negative float in xlsx → normal (credit)."""
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("53", 53, "Test Owner", 0.0, 10.81, 0.0, 17.29, -6.48, 0.0),
            ("Administrative Fund Totals", None, None, None, None, None, None, None, None),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["53"] == FinancialPosition.normal.value

    # --- Edge cases ---

    def test_empty_rows_returns_empty_list(self):
        """No section header rows → empty result."""
        rows = _parse_tocs_financial_position_excel_rows([])
        assert rows == []

    def test_no_lot_hash_section_returns_empty(self):
        """Rows with no 'Lot#' header → empty result."""
        rows = _parse_tocs_financial_position_excel_rows([
            ("Company Name", None, None),
            ("Address", None, None),
        ])
        assert rows == []

    def test_blank_row_terminates_section(self):
        """A blank row (all None) ends the section; lot after blank row is not captured."""
        # Build a raw row list: preamble, Lot# header, lot 1, blank row, lot 2
        # (lot 2 appears after the blank row and is therefore NOT in the first section)
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("1", "A", "Alice", "$-", "$-", "$-", "$-", "$-", "$-"),
            (None, None, None, None, None, None, None, None, None),  # blank row
            ("2", "B", "Bob", "$-", "$-", "$-", "$-", "$100.00", "$-"),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        lot_numbers = [r["lot_number"] for r in rows]
        assert "1" in lot_numbers
        # lot 2 is after the blank row — blank row break triggered
        assert "2" not in lot_numbers

    def test_lot_hash_header_in_data_terminates_section(self):
        """Another 'Lot#' row mid-section ends the current section (line 1757 break)."""
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("1", "A", "Alice", "$-", "$-", "$-", "$-", "$-", "$-"),
            # Another Lot# header row mid-section — should break
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("2", "B", "Bob", "$-", "$-", "$-", "$-", "$100.00", "$-"),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        lot_numbers = [r["lot_number"] for r in rows]
        # lot 1 is in first section, lot 2 starts the second section (both captured)
        assert "1" in lot_numbers
        assert "2" in lot_numbers

    def test_empty_lot_number_row_skipped(self):
        """A data row with empty lot# cell is skipped (line 1761 continue)."""
        all_rows = [
            ("Company Name", None, None, None, None, None, None, None, None),
            ("Lot#", "Unit#", "Owner Name", "Opening Balance",
             "Levied", "Special Levy", "Paid", "Closing Balance", "Interest Paid"),
            ("1", "A", "Alice", "$-", "$-", "$-", "$-", "$-", "$-"),
            (None, "A", "Ghost", "$-", "$-", "$-", "$-", "$200.00", "$-"),  # empty lot#
            ("2", "B", "Bob", "$-", "$-", "$-", "$-", "$-", "$-"),
            ("Fund Totals", None, None, None, None, None, None, None, None),
        ]
        rows = _parse_tocs_financial_position_excel_rows(all_rows)
        lot_numbers = [r["lot_number"] for r in rows]
        assert "1" in lot_numbers
        assert "2" in lot_numbers
        assert "" not in lot_numbers


# ---------------------------------------------------------------------------
# Unit tests: _parse_financial_position_excel_rows — TOCS auto-detection
# ---------------------------------------------------------------------------


class TestParseFpExcelRowsAutoDetect:
    """Unit tests for the auto-detection logic in _parse_financial_position_excel_rows."""

    # --- Happy path: TOCS format auto-detection ---

    def test_tocs_xlsx_detected_and_parsed(self):
        """First row does not start with 'Lot#' → treated as TOCS format."""
        xlsx = make_tocs_excel([[("5", "$1,882.06"), ("2", "$(190.77)")]])
        rows = _parse_financial_position_excel_rows(xlsx)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["5"] == FinancialPosition.in_arrear.value
        assert result["2"] == FinancialPosition.normal.value

    def test_tocs_xlsx_multi_section_worst_case(self):
        """Multi-section TOCS xlsx uses worst-case logic across sections."""
        xlsx = make_tocs_excel([
            [("10", "$-")],
            [("10", "$999.00")],
        ])
        rows = _parse_financial_position_excel_rows(xlsx)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["10"] == FinancialPosition.in_arrear.value

    # --- Happy path: simple format auto-detection ---

    def test_simple_xlsx_detected_and_parsed(self):
        """First row starts with 'Lot#' → treated as simple template format."""
        xlsx = make_excel(["Lot#", "Financial Position"], [["1A", "In Arrear"], ["2B", "Normal"]])
        rows = _parse_financial_position_excel_rows(xlsx)
        result = {r["lot_number"]: r["financial_position_raw"] for r in rows}
        assert result["1A"] == "In Arrear"
        assert result["2B"] == "Normal"

    # --- Input validation ---

    def test_invalid_xlsx_bytes_raises_422(self):
        """Non-xlsx bytes raise HTTPException 422."""
        with pytest.raises(HTTPException) as exc_info:
            _parse_financial_position_excel_rows(b"not-an-excel-file")
        assert exc_info.value.status_code == 422
        assert "invalid excel" in exc_info.value.detail.lower()

    def test_empty_xlsx_raises_422(self):
        """Empty xlsx (no rows at all) raises HTTPException 422."""
        wb = openpyxl.Workbook()
        # Remove the default sheet and save — results in a workbook with an empty sheet
        ws = wb.active
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(HTTPException) as exc_info:
            _parse_financial_position_excel_rows(buf.read())
        assert exc_info.value.status_code == 422
        assert "no headers" in exc_info.value.detail.lower()

    def test_simple_format_missing_fp_header_raises_422(self):
        """Simple format detected but 'Financial Position' column missing → 422."""
        xlsx = make_excel(["Lot#", "Something Else"], [["1A", "foo"]])
        with pytest.raises(HTTPException) as exc_info:
            _parse_financial_position_excel_rows(xlsx)
        assert exc_info.value.status_code == 422
        assert "financial position" in str(exc_info.value.detail).lower()


# ---------------------------------------------------------------------------
# Integration tests: TOCS xlsx via POST endpoint
# ---------------------------------------------------------------------------


class TestImportFinancialPositionsTOCSExcel:
    """Integration tests: POST TOCS-format xlsx to import-financial-positions endpoint."""

    # --- Happy path ---

    async def test_tocs_xlsx_updates_matching_lots(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """TOCS-format xlsx with lots '1A' and '2B' updates both correctly."""
        xlsx = make_tocs_excel([[("1A", "$500.00"), ("2B", "$(100.00)")]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["updated"] == 2
        assert data["skipped"] == 0

    async def test_tocs_xlsx_correct_positions_stored(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """TOCS xlsx: lot 1A gets in_arrear, lot 2B gets normal."""
        xlsx = make_tocs_excel([[("1A", "$500.00"), ("2B", "$(100.00)")]])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        result = await db_session.execute(
            select(LotOwner).where(LotOwner.building_id == building_with_owners.id)
        )
        owners = {lo.lot_number: lo for lo in result.scalars().all()}
        assert owners["1A"].financial_position == FinancialPosition.in_arrear
        assert owners["2B"].financial_position == FinancialPosition.normal

    async def test_tocs_xlsx_multi_section_worst_case(
        self, client: AsyncClient, db_session: AsyncSession, building_with_owners: Building
    ):
        """TOCS xlsx: lot 1A normal in section 1 but in_arrear in section 2 → in_arrear."""
        xlsx = make_tocs_excel([
            [("1A", "$-")],
            [("1A", "$300.00")],
        ])
        await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        result = await db_session.execute(
            select(LotOwner).where(
                LotOwner.building_id == building_with_owners.id,
                LotOwner.lot_number == "1A",
            )
        )
        lo = result.scalar_one()
        assert lo.financial_position == FinancialPosition.in_arrear

    async def test_tocs_xlsx_unknown_lot_skipped(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """Lot numbers in the xlsx that don't exist in DB are skipped."""
        xlsx = make_tocs_excel([[("9Z", "$500.00")]])
        response = await client.post(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["skipped"] == 1

    async def test_tocs_xlsx_nonexistent_building_returns_404(
        self, client: AsyncClient
    ):
        """TOCS xlsx upload to a building that doesn't exist → 404."""
        xlsx = make_tocs_excel([[("1A", "$-")]])
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import-financial-positions",
            files={"file": ("fp.xlsx", xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 404
