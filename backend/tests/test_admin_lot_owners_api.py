"""Tests for admin lot owner endpoints — /api/admin/buildings/{id}/lot-owners and /api/admin/lot-owners."""
from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime, timedelta

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Building,
    GeneralMeeting,
    GeneralMeetingLotWeight,
    GeneralMeetingStatus,
    LotOwner,
    LotProxy,
    Motion,
)
from app.models.lot import Lot
from app.models.lot_person import lot_persons
from app.models.person import Person

# Helpers and fixtures (make_csv, make_excel, meeting_dt, closing_dt, client, building,
# building_with_owners) are defined in conftest.py and automatically available.
from tests.conftest import make_csv, make_excel, meeting_dt, closing_dt, add_person_to_lot, get_or_create_person

# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}/lot-owners
# ---------------------------------------------------------------------------


class TestListLotOwners:
    # --- Happy path ---

    async def test_returns_lot_owners_for_building(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) == 2

    async def test_lot_owner_fields_present(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        owner = data[0]
        assert "id" in owner
        assert "lot_number" in owner
        assert "emails" in owner
        assert "unit_entitlement" in owner
        assert isinstance(owner["emails"], list)

    async def test_proxy_given_name_and_surname_null_for_lot_without_proxy(
        self, client: AsyncClient, building_with_owners: Building
    ):
        """proxy_given_name and proxy_surname are null for a lot with no proxy."""
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        owner = data[0]
        assert owner["proxy_email"] is None
        assert owner["proxy_given_name"] is None
        assert owner["proxy_surname"] is None

    async def test_proxy_given_name_and_surname_returned_for_lot_with_named_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        """proxy_given_name and proxy_surname are returned from GET list when proxy has names."""
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        _named_p = await get_or_create_person(db_session, "named@proxy.com", given_name="Alice", surname="Brown")
        db_session.add(LotProxy(lot_id=uuid.UUID(lot_owner_id), person_id=_named_p.id))
        await db_session.flush()

        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        owner = next(o for o in data if o["id"] == lot_owner_id)
        assert owner["proxy_email"] == "named@proxy.com"
        assert owner["proxy_given_name"] == "Alice"
        assert owner["proxy_surname"] == "Brown"

    async def test_lot_owner_emails_populated(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        data = response.json()
        all_emails = [e for o in data for e in o["emails"]]
        assert "voter1@test.com" in all_emails
        assert "voter2@test.com" in all_emails

    async def test_pagination_limit_and_offset(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """limit and offset params page through results correctly."""
        for i in range(5):
            lo = LotOwner(building_id=building.id, lot_number=f"PG{i:02d}", unit_entitlement=10)
            db_session.add(lo)
        await db_session.commit()

        resp_p1 = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?limit=3&offset=0"
        )
        assert resp_p1.status_code == 200
        page1 = resp_p1.json()
        assert len(page1) == 3

        resp_p2 = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?limit=3&offset=3"
        )
        assert resp_p2.status_code == 200
        page2 = resp_p2.json()
        assert len(page2) == 2  # only 2 remaining

        # No overlap between pages
        ids_p1 = {o["id"] for o in page1}
        ids_p2 = {o["id"] for o in page2}
        assert ids_p1.isdisjoint(ids_p2)

    async def test_default_limit_is_20(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Default limit is 20 — requesting without limit returns at most 20."""
        for i in range(25):
            lo = LotOwner(building_id=building.id, lot_number=f"DL{i:02d}", unit_entitlement=10)
            db_session.add(lo)
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        assert response.status_code == 200
        assert len(response.json()) == 20

    # --- State / precondition errors ---

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        response = await client.get(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners"
        )
        assert response.status_code == 404

    # --- Server-side sort ---

    async def test_sort_by_lot_number_asc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=lot_number&sort_dir=asc returns lots in numeric-ascending order."""
        for lot_num in ("10", "2", "1"):
            db_session.add(LotOwner(building_id=building.id, lot_number=lot_num, unit_entitlement=10))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=lot_number&sort_dir=asc"
        )
        assert response.status_code == 200
        data = response.json()
        lot_numbers = [o["lot_number"] for o in data]
        # Natural numeric order: 1, 2, 10
        assert lot_numbers == ["1", "2", "10"]

    async def test_sort_by_lot_number_desc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=lot_number&sort_dir=desc returns lots in numeric-descending order."""
        for lot_num in ("1", "2", "10"):
            db_session.add(LotOwner(building_id=building.id, lot_number=lot_num, unit_entitlement=10))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=lot_number&sort_dir=desc"
        )
        assert response.status_code == 200
        data = response.json()
        lot_numbers = [o["lot_number"] for o in data]
        # Natural numeric order descending: 10, 2, 1
        assert lot_numbers == ["10", "2", "1"]

    async def test_sort_by_unit_entitlement_asc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=unit_entitlement&sort_dir=asc returns lots ordered by entitlement ascending."""
        for entitlement, lot_num in ((300, "A"), (100, "B"), (200, "C")):
            db_session.add(LotOwner(building_id=building.id, lot_number=lot_num, unit_entitlement=entitlement))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=unit_entitlement&sort_dir=asc"
        )
        assert response.status_code == 200
        data = response.json()
        entitlements = [o["unit_entitlement"] for o in data]
        assert entitlements == [100, 200, 300]

    async def test_sort_by_unit_entitlement_desc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=unit_entitlement&sort_dir=desc returns lots ordered by entitlement descending."""
        for entitlement, lot_num in ((100, "A"), (300, "B"), (200, "C")):
            db_session.add(LotOwner(building_id=building.id, lot_number=lot_num, unit_entitlement=entitlement))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=unit_entitlement&sort_dir=desc"
        )
        assert response.status_code == 200
        data = response.json()
        entitlements = [o["unit_entitlement"] for o in data]
        assert entitlements == [300, 200, 100]

    async def test_sort_by_financial_position_asc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=financial_position asc: PostgreSQL enum declaration order is normal < in_arrear,
        so ascending sort puts 'normal' before 'in_arrear'."""
        from app.models import FinancialPosition
        db_session.add(LotOwner(building_id=building.id, lot_number="X1", unit_entitlement=10,
                                financial_position=FinancialPosition.in_arrear))
        db_session.add(LotOwner(building_id=building.id, lot_number="X2", unit_entitlement=10,
                                financial_position=FinancialPosition.normal))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=financial_position&sort_dir=asc"
        )
        assert response.status_code == 200
        data = response.json()
        # Enum declaration order: normal (1st) < in_arrear (2nd)
        positions = [o["financial_position"] for o in data]
        assert positions[0] == "normal"
        assert positions[-1] == "in_arrear"

    async def test_sort_by_financial_position_desc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """sort_by=financial_position desc: 'in_arrear' comes first (declared last)."""
        from app.models import FinancialPosition
        db_session.add(LotOwner(building_id=building.id, lot_number="Y1", unit_entitlement=10,
                                financial_position=FinancialPosition.normal))
        db_session.add(LotOwner(building_id=building.id, lot_number="Y2", unit_entitlement=10,
                                financial_position=FinancialPosition.in_arrear))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=financial_position&sort_dir=desc"
        )
        assert response.status_code == 200
        data = response.json()
        # Descending: in_arrear (2nd in declaration) comes first
        positions = [o["financial_position"] for o in data]
        assert positions[0] == "in_arrear"
        assert positions[-1] == "normal"

    async def test_invalid_sort_by_returns_400(self, client: AsyncClient, building: Building):
        """An unrecognised sort_by value returns 400."""
        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_by=invalid_column"
        )
        assert response.status_code == 400

    async def test_invalid_sort_dir_returns_400(self, client: AsyncClient, building: Building):
        """An unrecognised sort_dir value returns 400."""
        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners?sort_dir=sideways"
        )
        assert response.status_code == 400

    async def test_default_sort_is_lot_number_asc(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Without sort params the results are ordered by lot_number ascending (natural numeric)."""
        for lot_num in ("10", "2", "1"):
            db_session.add(LotOwner(building_id=building.id, lot_number=lot_num, unit_entitlement=10))
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        assert response.status_code == 200
        data = response.json()
        lot_numbers = [o["lot_number"] for o in data]
        assert lot_numbers == ["1", "2", "10"]


# ---------------------------------------------------------------------------
# GET /api/admin/buildings/{building_id}/lot-owners/count
# ---------------------------------------------------------------------------


class TestCountLotOwners:
    # --- Happy path ---

    async def test_returns_count_for_building(
        self, client: AsyncClient, building_with_owners: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners/count"
        )
        assert response.status_code == 200
        data = response.json()
        assert data == {"count": 2}

    async def test_returns_zero_for_empty_building(
        self, client: AsyncClient, building: Building
    ):
        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners/count"
        )
        assert response.status_code == 200
        assert response.json() == {"count": 0}

    async def test_count_reflects_added_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        for i in range(7):
            lo = LotOwner(building_id=building.id, lot_number=f"CNT{i:02d}", unit_entitlement=10)
            db_session.add(lo)
        await db_session.commit()

        response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners/count"
        )
        assert response.status_code == 200
        assert response.json()["count"] == 7

    async def test_count_is_independent_of_other_buildings(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Count for one building is not affected by lot owners in another."""
        b1 = Building(name="Cnt Bldg A", manager_email="cnta@test.com")
        b2 = Building(name="Cnt Bldg B", manager_email="cntb@test.com")
        db_session.add(b1)
        db_session.add(b2)
        await db_session.flush()
        for i in range(3):
            db_session.add(LotOwner(building_id=b1.id, lot_number=f"CA{i}", unit_entitlement=10))
        for i in range(5):
            db_session.add(LotOwner(building_id=b2.id, lot_number=f"CB{i}", unit_entitlement=10))
        await db_session.commit()

        r1 = await client.get(f"/api/admin/buildings/{b1.id}/lot-owners/count")
        r2 = await client.get(f"/api/admin/buildings/{b2.id}/lot-owners/count")
        assert r1.json()["count"] == 3
        assert r2.json()["count"] == 5

    # --- Input validation ---

    async def test_invalid_building_uuid_returns_422(self, client: AsyncClient):
        response = await client.get(
            "/api/admin/buildings/not-a-uuid/lot-owners/count"
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# GET /api/admin/lot-owners/{lot_owner_id}
# ---------------------------------------------------------------------------


class TestGetLotOwner:
    # --- Happy path ---

    async def test_returns_lot_owner_without_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        # Get the first lot owner ID from the list
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        response = await client.get(f"/api/admin/lot-owners/{lot_owner_id}")
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == lot_owner_id
        assert "lot_number" in data
        assert "emails" in data
        assert "unit_entitlement" in data
        assert "financial_position" in data
        assert data["proxy_email"] is None
        assert data["proxy_given_name"] is None
        assert data["proxy_surname"] is None

    async def test_returns_lot_owner_with_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        # Get the first lot owner ID from the list and add a proxy
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        # Seed a proxy for this lot owner
        _proxy_p = await get_or_create_person(db_session, "proxy@example.com")
        proxy = LotProxy(lot_id=uuid.UUID(lot_owner_id), person_id=_proxy_p.id)
        db_session.add(proxy)
        await db_session.flush()

        response = await client.get(f"/api/admin/lot-owners/{lot_owner_id}")
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "proxy@example.com"
        assert data["proxy_given_name"] is None
        assert data["proxy_surname"] is None

    async def test_returns_lot_owner_with_named_proxy(
        self, client: AsyncClient, building_with_owners: Building, db_session: AsyncSession
    ):
        """GET /lot-owners/{id} returns proxy_given_name and proxy_surname when proxy has names."""
        list_response = await client.get(
            f"/api/admin/buildings/{building_with_owners.id}/lot-owners"
        )
        owners = list_response.json()
        lot_owner_id = owners[0]["id"]

        _named_jane = await get_or_create_person(db_session, "named@proxy.com", given_name="Jane", surname="Doe")
        db_session.add(LotProxy(lot_id=uuid.UUID(lot_owner_id), person_id=_named_jane.id))
        await db_session.flush()

        response = await client.get(f"/api/admin/lot-owners/{lot_owner_id}")
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "named@proxy.com"
        assert data["proxy_given_name"] == "Jane"
        assert data["proxy_surname"] == "Doe"

    # --- Input validation ---

    async def test_invalid_uuid_returns_422(self, client: AsyncClient):
        response = await client.get("/api/admin/lot-owners/not-a-uuid")
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_nonexistent_lot_owner_returns_404(self, client: AsyncClient):
        response = await client.get(f"/api/admin/lot-owners/{uuid.uuid4()}")
        assert response.status_code == 404
        assert response.json()["detail"] == "Lot owner not found"


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/lot-owners/import
# ---------------------------------------------------------------------------


class TestImportLotOwners:
    # --- Happy path ---

    async def test_valid_csv_imports_lot_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["101", "a@test.com", "100"], ["102", "b@test.com", "200"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 2
        assert data["emails"] == 2

    async def test_import_multi_email_same_lot_returns_422_since_rr3_31(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Multiple rows with same lot_number → 422 with duplicate row detail (RR3-31).

        Use semicolons to specify multiple emails for one lot in a single row.
        """
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [
                ["101", "a@test.com", "100"],
                ["101", "b@test.com", "100"],
            ],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        # Since RR3-31, duplicate lot numbers are an error (use semicolons for multi-email)
        assert response.status_code == 422
        detail = response.json()["detail"]
        assert isinstance(detail, list)
        assert any("101" in e for e in detail)

    async def test_import_semicolon_separated_emails(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Semicolon-separated emails in a single cell are split into multiple LotOwnerEmail rows."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["101", "a@test.com;b@test.com; c@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 1
        assert data["emails"] == 3

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        assert sorted(owners[0]["emails"]) == ["a@test.com", "b@test.com", "c@test.com"]

    async def test_import_blank_email_allowed(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """A lot row with blank email is imported without error (no email row created)."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NO-EMAIL", "", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 1
        assert data["emails"] == 0

    async def test_import_email_normalised_to_lowercase(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Emails imported in mixed case are stored as lowercase."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["LOWER-CSV", "UPPER@TEST.COM", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        emails = [e for o in owners for e in o["emails"]]
        assert "upper@test.com" in emails
        assert "UPPER@TEST.COM" not in emails

    async def test_import_adds_new_owners_without_removing_absent_lots(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Pure upsert: existing lots absent from a re-import are preserved, not deleted.

        This guards the architecture decision that import never deletes lots — doing so
        would cascade-delete AGMLotWeight records and destroy historical vote tallies.
        """
        # First import
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["A1", "old@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Second import introduces new lots; A1 is absent but must NOT be deleted
        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["B1", "new@test.com", "75"], ["B2", "new2@test.com", "25"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = owners_response.json()
        lot_numbers = {o["lot_number"] for o in owners}
        assert "B1" in lot_numbers
        assert "B2" in lot_numbers
        # A1 was absent from csv2 — pure upsert must preserve it
        assert "A1" in lot_numbers, "Absent lot deleted by re-import (pure-upsert violated)"

    async def test_empty_csv_import_preserves_existing_owners(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """An empty re-import must not clear existing lots (pure upsert, never delete)."""
        # Seed
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["X1", "x@test.com", "10"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Import empty file
        csv2 = make_csv(["lot_number", "email", "unit_entitlement"], [])
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 0

        # X1 must still exist — empty import must not clear existing lots
        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        lot_numbers = {o["lot_number"] for o in owners_response.json()}
        assert "X1" in lot_numbers, "Empty import deleted existing lots (pure-upsert violated)"

    async def test_extra_columns_ignored(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "extra"],
            [["Z1", "z@test.com", "10", "ignore_me"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 1

    async def test_unit_entitlement_minimum_one_accepted(
        self, client: AsyncClient, building: Building
    ):
        """RR3-37: unit_entitlement must be > 0; minimum valid value is 1."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["MINONE1", "minone@test.com", "1"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 1

    async def test_unit_entitlement_zero_rejected(
        self, client: AsyncClient, building: Building
    ):
        """RR3-37: unit_entitlement=0 is rejected (must be > 0)."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["ZERO1", "zero@test.com", "0"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("> 0" in str(e) or "must be" in str(e) for e in response.json()["detail"])

    # --- Input validation ---

    async def test_missing_unit_entitlement_header(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email"],
            [["1", "a@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert "unit_entitlement" in str(response.json()["detail"]).lower()

    async def test_missing_lot_number_header(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["email", "unit_entitlement"],
            [["a@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_duplicate_lot_numbers_in_csv_returns_422_with_row_detail(
        self, client: AsyncClient, building: Building
    ):
        """Two rows with the same lot_number → 422 with row-level detail (RR3-31).
        The error detail must identify the duplicate lot number and both row numbers."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["DUP1", "a@test.com", "100"], ["DUP1", "b@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        detail = response.json()["detail"]
        # detail is a list of error strings; find the duplicate error
        assert isinstance(detail, list)
        duplicate_errors = [e for e in detail if "DUP1" in e]
        assert len(duplicate_errors) == 1
        assert "rows" in duplicate_errors[0].lower() or "row" in duplicate_errors[0].lower()

    async def test_negative_unit_entitlement(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NEG1", "neg@test.com", "-1"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_non_integer_unit_entitlement(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["NI1", "ni@test.com", "abc"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422

    async def test_collects_all_errors_before_returning(
        self, client: AsyncClient, building: Building
    ):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["", "", "abc"], ["", "", "-5"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        # Should have multiple errors
        assert len(response.json()["detail"]) > 1

    # --- State / precondition errors ---

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["1", "a@test.com", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 404

    async def test_non_csv_file_returns_415(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("data.pdf", b"not a spreadsheet", "application/pdf")},
        )
        assert response.status_code == 415

    async def test_completely_empty_file_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """An empty file (not even headers) should return 422."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("empty.csv", b"", "text/csv")},
        )
        assert response.status_code == 422

    async def test_empty_unit_entitlement_value(
        self, client: AsyncClient, building: Building
    ):
        """A row with empty unit_entitlement (present column, empty value) returns 422."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["E1", "e@test.com", ""]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422
        assert any("unit_entitlement is empty" in str(e) for e in response.json()["detail"])

    # --- Edge cases ---

    async def test_reimport_preserves_agm_lot_weight_snapshots(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Re-importing lot owners must NOT destroy GeneralMeetingLotWeight snapshots.

        Bug: delete-all-then-insert gave new IDs to lot owners, which cascaded
        deletes to GeneralMeetingLotWeight records, causing entitlement_sum=0 in tallies.
        Fix: upsert by lot_number preserves existing LotOwner IDs.
        """
        # Initial import
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["L1", "voter@test.com", "150"], ["L2", "voter@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Create GeneralMeeting and snapshot lot weights
        agm = GeneralMeeting(
            building_id=building.id,
            title="Snapshot GeneralMeeting",
            status=GeneralMeetingStatus.open,
            meeting_at=meeting_dt(),
            voting_closes_at=closing_dt(),
        )
        db_session.add(agm)
        await db_session.flush()
        motion = Motion(general_meeting_id=agm.id, title="Test Motion", display_order=1)
        db_session.add(motion)
        await db_session.flush()

        owners_result = await db_session.execute(
            select(LotOwner).where(LotOwner.building_id == building.id)
        )
        for lo in owners_result.scalars().all():
            db_session.add(GeneralMeetingLotWeight(
                general_meeting_id=agm.id,
                lot_id=lo.id,
                unit_entitlement_snapshot=lo.unit_entitlement,
            ))
        await db_session.commit()

        # Re-import same lots (email change for L1) — must preserve snapshots
        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["L1", "updated@test.com", "150"], ["L2", "voter@test.com", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200

        # GeneralMeetingLotWeight snapshots must still exist (not cascade-deleted)
        weights_result = await db_session.execute(
            select(GeneralMeetingLotWeight).where(GeneralMeetingLotWeight.general_meeting_id == agm.id)
        )
        weights = weights_result.scalars().all()
        assert len(weights) == 2
        total_snapshot = sum(w.unit_entitlement_snapshot for w in weights)
        assert total_snapshot == 200  # 150 + 50 preserved from original snapshot

    async def test_reimport_updates_existing_and_adds_new(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Existing lot numbers are updated in-place; new lot numbers are inserted."""
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["OLD1", "old@test.com", "100"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["OLD1", "updated@test.com", "200"], ["NEW1", "new@test.com", "50"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = {o["lot_number"]: o for o in owners_response.json()}
        assert owners["OLD1"]["emails"] == ["updated@test.com"]
        assert owners["OLD1"]["unit_entitlement"] == 200
        assert "NEW1" in owners

    async def test_partial_reimport_does_not_delete_absent_lots(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """HIGH-1: A partial re-import must never delete lot owners absent from the file.

        Deleting absent lots would cascade-delete AGMLotWeight records and destroy
        historical vote tallies for existing AGMs.
        """
        # Initial import: two lots
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["KEEP1", "keep@test.com", "100"], ["KEEP2", "keep2@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        # Partial re-import: only one lot — KEEP2 is absent from the file
        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["KEEP1", "updated@test.com", "200"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200

        owners_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = {o["lot_number"]: o for o in owners_response.json()}
        # KEEP1 updated in-place
        assert owners["KEEP1"]["emails"] == ["updated@test.com"]
        assert owners["KEEP1"]["unit_entitlement"] == 200
        # KEEP2 must still exist — partial import must not delete absent lots
        assert "KEEP2" in owners, "Absent lot was deleted by a partial import (HIGH-1)"

    async def test_import_with_headers_but_zero_data_rows(
        self, client: AsyncClient, building: Building
    ):
        """RR3-41: CSV with all required headers but zero data rows returns 200 with
        imported=0 and emails=0 (not an error — empty import is valid)."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [],  # zero data rows
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 0
        assert data["emails"] == 0


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/lot-owners/import — owner name columns
# ---------------------------------------------------------------------------


class TestImportLotOwnersNameColumns:
    """Tests for given_name/surname/name column support during lot owner import."""

    # --- CSV with separate given_name / surname columns ---

    async def test_csv_separate_given_name_surname_stored_on_email_record(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV with given_name and surname columns → names stored on LotOwnerEmail records."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "given_name", "surname"],
            [["101", "alice@test.com", "100", "Alice", "Smith"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "alice@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Alice"
        assert email_row.surname == "Smith"

    # --- CSV with Name column ---

    async def test_csv_name_column_parsed_and_stored_on_email_record(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV with Name column: last token → surname, rest → given_name, stored on LotOwnerEmail."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "name"],
            [["102", "bob@test.com", "50", "Robert James Brown"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "bob@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Robert James"
        assert email_row.surname == "Brown"

    async def test_csv_name_column_single_token_stored_as_surname(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Single-token Name (company name) → given_name is None, surname is the token."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "name"],
            [["103", "corp@test.com", "75", "ACME"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "corp@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name is None
        assert email_row.surname == "ACME"

    # --- CSV with no name columns ---

    async def test_csv_no_name_columns_email_records_have_null_names(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV with no name columns → LotOwnerEmail.given_name and surname are None."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["104", "noname@test.com", "30"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "noname@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name is None
        assert email_row.surname is None

    # --- Deduplication: same email appears twice for a lot ---

    async def test_csv_dedup_same_email_last_name_wins(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """When same email address appears twice in semicolon list (same row), last-row name wins."""
        # Two rows same lot — would normally be a 422 (duplicate lot_number).
        # Use the semicolons-in-one-row approach: can't trigger the dedup path with
        # a single row. Use reimport scenario: first import sets name, second import
        # with updated name overwrites.
        csv1 = make_csv(
            ["lot_number", "email", "unit_entitlement", "name"],
            [["DEDUP1", "dedup@test.com", "100", "First Name"]],
        )
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv1, "text/csv")},
        )

        csv2 = make_csv(
            ["lot_number", "email", "unit_entitlement", "name"],
            [["DEDUP1", "dedup@test.com", "100", "Second Name"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv2, "text/csv")},
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "dedup@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Second"
        assert email_row.surname == "Name"

    # --- Excel with Name column ---

    async def test_excel_name_column_parsed_and_stored(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel import with Name column: parsed into given_name/surname on LotOwnerEmail."""
        b = Building(name="Excel Name Bldg", manager_email="en@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2", "Name"],
            [["201", "excel.name@test.com", 80, "Carol Anne White"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "excel.name@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Carol Anne"
        assert email_row.surname == "White"

    async def test_excel_no_name_column_email_records_have_null_names(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel import with no name columns → LotOwnerEmail.given_name and surname are None."""
        b = Building(name="Excel No Name Bldg", manager_email="enn@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["202", "noname.xl@test.com", 60]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "noname.xl@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name is None
        assert email_row.surname is None

    # --- Real-world: Owners.csv has a Name column ---

    async def test_csv_owners_file_name_column_stores_names(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Subset of the real Owners.csv format: Name column maps correctly to email record."""
        # Simulate a typical row from examples/Owners.csv: Name column with full name,
        # Email column with email address, UOE2 as unit_entitlement (mapped via Lot# alias).
        # Using a small synthetic slice that mirrors the real file structure.
        csv_data = make_csv(
            ["Lot#", "Email", "UOE2", "Name"],
            [
                ["53", "ntassell@outlook.com", "1", "Nicholas Warren Tassell"],
                ["55", "sbtunit5@gmail.com", "1", "Nicole Anne Seils"],
            ],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("Owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

        result = await db_session.execute(
            select(Person).where(Person.email == "ntassell@outlook.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Nicholas Warren"
        assert email_row.surname == "Tassell"

        result2 = await db_session.execute(
            select(Person).where(Person.email == "sbtunit5@gmail.com")
        )
        email_row2 = result2.scalar_one_or_none()
        assert email_row2 is not None
        assert email_row2.given_name == "Nicole Anne"
        assert email_row2.surname == "Seils"

    # --- CSV dedup: same email appears twice in semicolon-separated list ---

    async def test_csv_duplicate_email_in_semicolon_list_last_name_wins(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """Same email appearing twice in semicolons: last-seen name wins on the email record."""
        # "dup@test.com;dup@test.com" — second occurrence overwrites name fields.
        # Name mode is "name" so we get the parsed name.
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "name"],
            [["DUPCSV", "dup.csv@test.com;dup.csv@test.com", "100", "Final Name"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200
        # Only one LotOwnerEmail created (dedup)
        result = await db_session.execute(
            select(Person).where(Person.email == "dup.csv@test.com")
        )
        rows = result.scalars().all()
        assert len(rows) == 1
        assert rows[0].surname == "Name"

    # --- Excel separate given_name / surname columns ---

    async def test_excel_separate_given_name_surname_stored_on_email_record(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel with given_name and surname columns → names stored on LotOwnerEmail records."""
        b = Building(name="Excel Sep Name Bldg", manager_email="esn@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2", "given_name", "surname"],
            [["301", "sep.name@test.com", 90, "Diana", "Prince"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

        result = await db_session.execute(
            select(Person).where(Person.email == "sep.name@test.com")
        )
        email_row = result.scalar_one_or_none()
        assert email_row is not None
        assert email_row.given_name == "Diana"
        assert email_row.surname == "Prince"

    # --- Excel dedup: same email appears twice in semicolon-separated list ---

    async def test_excel_duplicate_email_in_semicolon_list_last_name_wins(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel: same email appearing twice in semicolons — last-seen name wins."""
        b = Building(name="Excel Dedup Bldg", manager_email="edp@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2", "Name"],
            [["DUPXL", "dup.xl@test.com;dup.xl@test.com", 50, "Excel Final Name"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        result = await db_session.execute(
            select(Person).where(Person.email == "dup.xl@test.com")
        )
        rows = result.scalars().all()
        assert len(rows) == 1
        assert rows[0].surname == "Name"


# ---------------------------------------------------------------------------
# POST /api/admin/buildings/{building_id}/lot-owners
# ---------------------------------------------------------------------------


class TestAddLotOwner:
    # --- Happy path ---

    async def test_valid_add_returns_201(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NEW01", "emails": ["new@test.com"], "unit_entitlement": 150},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["lot_number"] == "NEW01"
        assert "new@test.com" in data["emails"]
        assert data["unit_entitlement"] == 150
        assert "id" in data

    async def test_add_with_multiple_emails(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "MULTI01", "emails": ["a@test.com", "b@test.com"], "unit_entitlement": 100},
        )
        assert response.status_code == 201
        data = response.json()
        assert len(data["emails"]) == 2

    async def test_add_with_no_emails(
        self, client: AsyncClient, building: Building
    ):
        """Lot owner can be created with empty email list."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NOEMAIL01", "emails": [], "unit_entitlement": 50},
        )
        assert response.status_code == 201
        data = response.json()
        assert data["emails"] == []

    async def test_unit_entitlement_zero_rejected(
        self, client: AsyncClient, building: Building
    ):
        """RR3-37: unit_entitlement=0 is rejected; must be > 0."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "ZERO01", "emails": [], "unit_entitlement": 0},
        )
        assert response.status_code == 422

    async def test_unit_entitlement_one_accepted(
        self, client: AsyncClient, building: Building
    ):
        """RR3-37: unit_entitlement=1 (minimum positive) is accepted."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "ONE01", "emails": [], "unit_entitlement": 1},
        )
        assert response.status_code == 201
        assert response.json()["unit_entitlement"] == 1

    async def test_email_normalised_to_lowercase(
        self, client: AsyncClient, building: Building
    ):
        """Email submitted in mixed case is stored as lowercase."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "LOWER01", "emails": ["UPPER@TEST.COM"], "unit_entitlement": 10},
        )
        assert response.status_code == 201
        data = response.json()
        assert "upper@test.com" in data["emails"]
        assert "UPPER@TEST.COM" not in data["emails"]

    # --- Input validation ---

    async def test_negative_unit_entitlement_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "NEG01", "emails": [], "unit_entitlement": -1},
        )
        assert response.status_code == 422

    async def test_empty_lot_number_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "", "emails": [], "unit_entitlement": 10},
        )
        assert response.status_code == 422

    async def test_missing_fields_returns_422(
        self, client: AsyncClient, building: Building
    ):
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "M01"},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_lot_number_returns_409(
        self, client: AsyncClient, building: Building
    ):
        # Add first
        await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "DUP01", "emails": [], "unit_entitlement": 10},
        )
        # Add duplicate
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "DUP01", "emails": [], "unit_entitlement": 20},
        )
        assert response.status_code == 409

    async def test_building_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(
            f"/api/admin/buildings/{uuid.uuid4()}/lot-owners",
            json={"lot_number": "X01", "emails": [], "unit_entitlement": 10},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# PATCH /api/admin/lot-owners/{lot_owner_id}
# ---------------------------------------------------------------------------


class TestUpdateLotOwner:
    # --- Happy path ---

    async def test_update_unit_entitlement(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD02",
            unit_entitlement=50,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 999},
        )
        assert response.status_code == 200
        assert response.json()["unit_entitlement"] == 999

    async def test_update_financial_position(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD03",
            unit_entitlement=100,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"financial_position": "in_arrear"},
        )
        assert response.status_code == 200
        assert response.json()["financial_position"] == "in_arrear"

    async def test_update_both_fields(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD04",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 200, "financial_position": "in_arrear"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["unit_entitlement"] == 200
        assert data["financial_position"] == "in_arrear"

    # --- Boundary values ---

    async def test_unit_entitlement_zero_rejected(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """RR3-37: unit_entitlement=0 is rejected on update (must be > 0)."""
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD05",
            unit_entitlement=100,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 0},
        )
        assert response.status_code == 422

    async def test_unit_entitlement_one_accepted(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """RR3-37: unit_entitlement=1 (minimum positive) is accepted on update."""
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD05B",
            unit_entitlement=100,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": 1},
        )
        assert response.status_code == 200
        assert response.json()["unit_entitlement"] == 1

    # --- Input validation ---

    async def test_no_fields_provided_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD06",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={},
        )
        assert response.status_code == 422

    async def test_negative_unit_entitlement_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD07",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"unit_entitlement": -5},
        )
        assert response.status_code == 422

    async def test_invalid_financial_position_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(
            building_id=building.id,
            lot_number="UPD08",
            unit_entitlement=10,
        )
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}",
            json={"financial_position": "invalid_value"},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_not_found_returns_404(self, client: AsyncClient):
        response = await client.patch(
            f"/api/admin/lot-owners/{uuid.uuid4()}",
            json={"unit_entitlement": 10},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/admin/lot-owners/{lot_owner_id}/emails
# ---------------------------------------------------------------------------


class TestAddEmailToLotOwner:
    # --- Happy path ---

    async def test_add_email_returns_updated_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "new@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert "new@test.com" in data["emails"]

    async def test_add_second_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM02", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "first@test.com")
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "second@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        assert "first@test.com" in data["emails"]
        assert "second@test.com" in data["emails"]

    # --- Input validation ---

    async def test_empty_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="EM03", unit_entitlement=10)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": ""},
        )
        assert response.status_code == 422

    async def test_add_email_normalised_to_lowercase(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """Email submitted in mixed case is stored as lowercase."""
        lo = LotOwner(building_id=building.id, lot_number="EM04", unit_entitlement=10)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "UPPER@TEST.COM"},
        )
        assert response.status_code == 201
        data = response.json()
        assert "upper@test.com" in data["emails"]
        assert "UPPER@TEST.COM" not in data["emails"]

    # --- State / precondition errors ---

    async def test_lot_owner_not_found_returns_404(self, client: AsyncClient):
        response = await client.post(
            f"/api/admin/lot-owners/{uuid.uuid4()}/emails",
            json={"email": "x@test.com"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/lot-owners/{lot_owner_id}/emails/{email}
# ---------------------------------------------------------------------------


class TestRemoveEmailFromLotOwner:
    # --- Happy path ---

    async def test_remove_email_returns_updated_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="REM01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "todelete@test.com")
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/emails/todelete@test.com"
        )
        assert response.status_code == 200
        data = response.json()
        assert "todelete@test.com" not in data["emails"]

    # --- State / precondition errors ---

    async def test_lot_owner_not_found_returns_404(self, client: AsyncClient):
        response = await client.delete(
            f"/api/admin/lot-owners/{uuid.uuid4()}/emails/x@test.com"
        )
        assert response.status_code == 404

    async def test_email_not_found_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="REM02", unit_entitlement=10)
        db_session.add(lo)
        await db_session.commit()

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/emails/nonexistent@test.com"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/admin/general-meetings
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# Schema unit tests (for coverage of schema validators)
# ---------------------------------------------------------------------------


class TestSchemas:
    def test_lot_owner_update_requires_at_least_one_field(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerUpdate

        with pytest.raises(ValidationError):
            LotOwnerUpdate()

    def test_lot_owner_update_unit_entitlement_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(unit_entitlement=10)
        assert obj.unit_entitlement == 10
        assert obj.financial_position is None

    def test_lot_owner_update_financial_position_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(financial_position="in_arrear")
        assert obj.financial_position == "in_arrear"
        assert obj.unit_entitlement is None

    def test_lot_owner_update_entitlement_only(self):
        from app.schemas.admin import LotOwnerUpdate

        obj = LotOwnerUpdate(unit_entitlement=10)
        assert obj.unit_entitlement == 10

    def test_lot_owner_update_negative_entitlement_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerUpdate

        with pytest.raises(ValidationError):
            LotOwnerUpdate(unit_entitlement=-1)

    def test_agm_create_no_motions_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import GeneralMeetingCreate

        with pytest.raises(ValidationError):
            GeneralMeetingCreate(
                building_id=uuid.uuid4(),
                title="t",
                meeting_at=meeting_dt(),
                voting_closes_at=closing_dt(),
                motions=[],
            )

    def test_agm_create_closes_before_meeting_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import GeneralMeetingCreate, MotionCreate

        now = datetime.now(UTC)
        with pytest.raises(ValidationError):
            GeneralMeetingCreate(
                building_id=uuid.uuid4(),
                title="t",
                meeting_at=now + timedelta(days=2),
                voting_closes_at=now + timedelta(days=1),
                motions=[MotionCreate(title="M", display_order=1)],
            )

    def test_lot_owner_create_empty_lot_number_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerCreate

        with pytest.raises(ValidationError):
            LotOwnerCreate(lot_number="  ", unit_entitlement=10)

    def test_lot_owner_create_negative_entitlement_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import LotOwnerCreate

        with pytest.raises(ValidationError):
            LotOwnerCreate(lot_number="L1", unit_entitlement=-1)

    def test_building_create_empty_name_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import BuildingCreate

        with pytest.raises(ValidationError):
            BuildingCreate(name="  ", manager_email="mgr@test.com")

    def test_building_create_empty_email_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import BuildingCreate

        with pytest.raises(ValidationError):
            BuildingCreate(name="Valid Name", manager_email="  ")

    def test_admin_login_request_valid(self):
        from app.schemas.admin import AdminLoginRequest

        req = AdminLoginRequest(username="admin", password="secret")
        assert req.username == "admin"
        assert req.password == "secret"

    def test_add_email_request_empty_raises(self):
        from pydantic import ValidationError

        from app.schemas.admin import AddEmailRequest

        with pytest.raises(ValidationError):
            AddEmailRequest(email="  ")

    def test_add_email_request_valid(self):
        from app.schemas.admin import AddEmailRequest

        req = AddEmailRequest(email="x@test.com")
        assert req.email == "x@test.com"


# ---------------------------------------------------------------------------
# Import lot owners from Excel
# ---------------------------------------------------------------------------


class TestImportLotOwnersExcel:
    # --- Happy path ---

    async def test_valid_xlsx_imports_lot_owners(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Excel Import Building", manager_email="xlimport@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["XL1", "xl1@test.com", 100], ["XL2", "xl2@test.com", 200]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["imported"] == 2
        assert data["emails"] == 2

    async def test_xlsx_updates_existing_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Excel Update Building", manager_email="xlupdate@test.com")
        db_session.add(b)
        await db_session.commit()

        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement"],
            [["UPD1", "upd@test.com", "50"]],
        )
        await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["UPD1", "upd@test.com", 150]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

    async def test_xlsx_empty_returns_zero(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Empty Excel Bldg", manager_email="ee@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(["Lot#", "Email", "UOE2"], [])
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 0

    # --- Input validation ---

    async def test_xlsx_missing_lot_number_column_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Missing Col Excel Bldg", manager_email="mc@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["email", "unit_entitlement"],
            [["mc@test.com", 100]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_invalid_excel_file_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        b = Building(name="Bad Excel Bldg", manager_email="be@test.com")
        db_session.add(b)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    b"not-an-excel-file",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_file_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with no headers (empty workbook) returns 422."""
        b = Building(name="Empty Sheet Bldg", manager_email="es@test.com")
        db_session.add(b)
        await db_session.commit()

        # Create a workbook with a completely empty sheet (no rows at all)
        wb = openpyxl.Workbook()
        ws = wb.active
        # Write nothing — sheet is empty, iter_rows will raise StopIteration on first next()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        empty_excel = buf.read()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    empty_excel,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_lot_number_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with an empty Lot# cell returns 422."""
        b = Building(name="Empty Lot Bldg", manager_email="el@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            # Row with None lot number — will trigger the empty lot_number path
            [[None, "el@test.com", 100]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_empty_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with an empty UOE2 cell returns 422."""
        b = Building(name="Empty UOE Bldg", manager_email="eu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["EU1", "eu@test.com", None]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_non_integer_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with a non-integer UOE2 returns 422."""
        b = Building(name="Bad UOE Bldg", manager_email="bu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["BU1", "bu@test.com", "not-a-number"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_negative_uoe2_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with a negative UOE2 returns 422."""
        b = Building(name="Neg UOE Bldg", manager_email="nu@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2"],
            [["NU1", "nu@test.com", -5]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_invalid_financial_position_returns_422(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with invalid financial_position returns 422."""
        b = Building(name="Bad FP Bldg", manager_email="bfp@test.com")
        db_session.add(b)
        await db_session.commit()

        excel_data = make_excel(
            ["Lot#", "Email", "UOE2", "financial_position"],
            [["BFP1", "bfp@test.com", 100, "bankrupt"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422

    async def test_xlsx_lot_owner_blank_rows_skipped(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel lot owner import with blank rows between data rows skips them."""
        b = Building(name="Blank Row Bldg", manager_email="br@test.com")
        db_session.add(b)
        await db_session.commit()

        # Build Excel manually with a blank row between two data rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Email", "UOE2"])
        ws.append(["BR1", "br1@test.com", 100])
        ws.append([None, None, None])  # blank row — triggers continue
        ws.append(["BR2", "br2@test.com", 200])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 2

    async def test_xlsx_duplicate_lot_numbers_returns_422_with_row_detail(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Excel import with duplicate lot numbers → 422 with row detail (RR3-31)."""
        b = Building(name="Dup Excel Bldg", manager_email="dupxl@test.com")
        db_session.add(b)
        await db_session.commit()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Lot#", "Email", "UOE2"])
        ws.append(["DUP-XL1", "a@test.com", 100])
        ws.append(["DUP-XL1", "b@test.com", 100])  # duplicate row
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_data = buf.read()

        response = await client.post(
            f"/api/admin/buildings/{b.id}/lot-owners/import",
            files={
                "file": (
                    "owners.xlsx",
                    excel_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 422
        detail = response.json()["detail"]
        assert isinstance(detail, list)
        # Must identify the duplicate lot number and the rows it appeared on
        dup_errors = [e for e in detail if "DUP-XL1" in e]
        assert len(dup_errors) == 1
        assert "row" in dup_errors[0].lower()




# ---------------------------------------------------------------------------
# LotOwnerCreate schema validator
# ---------------------------------------------------------------------------


class TestLotOwnerCreateSchema:
    async def test_valid_financial_position_in_arrear_returns_201(
        self, client: AsyncClient, building: Building
    ):
        """Creating a lot owner with financial_position='in_arrear' succeeds (line 93 return v)."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={
                "lot_number": "FPV0",
                "unit_entitlement": 100,
                "financial_position": "in_arrear",
                "emails": [],
            },
        )
        assert response.status_code == 201
        assert response.json()["financial_position"] == "in_arrear"

    async def test_invalid_financial_position_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """Creating a lot owner with an invalid financial_position value → 422."""
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={
                "lot_number": "FPV1",
                "unit_entitlement": 100,
                "financial_position": "bankrupt",
                "emails": [],
            },
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# CSV import: financial_position edge cases
# ---------------------------------------------------------------------------


class TestImportLotOwnersFinancialPosition:
    async def test_csv_import_in_arrear_financial_position(
        self, client: AsyncClient, building: Building
    ):
        """Importing with financial_position=in_arrear stores it correctly."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "financial_position"],
            [["FP1", "fp1@test.com", "100", "in_arrear"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

    async def test_csv_import_invalid_financial_position_returns_422(
        self, client: AsyncClient, building: Building
    ):
        """Invalid financial_position value → 422."""
        csv_data = make_csv(
            ["lot_number", "email", "unit_entitlement", "financial_position"],
            [["FP2", "fp2@test.com", "100", "bankrupt"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 422


# ---------------------------------------------------------------------------
# Duplicate email on add_email_to_lot_owner
# ---------------------------------------------------------------------------


class TestAddEmailDuplicate:
    async def test_add_duplicate_email_returns_409(
        self, client: AsyncClient, db_session: AsyncSession
    ):
        """Adding the same email twice to a lot owner → 409."""
        b = Building(name="Dup Email Building", manager_email="dup_email@test.com")
        db_session.add(b)
        await db_session.flush()
        lo = LotOwner(building_id=b.id, lot_number="DE1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "existing@test.com")
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/emails",
            json={"email": "existing@test.com"},
        )
        assert response.status_code == 409


# ---------------------------------------------------------------------------
# PUT /api/admin/lot-owners/{lot_owner_id}/proxy
# DELETE /api/admin/lot-owners/{lot_owner_id}/proxy
# ---------------------------------------------------------------------------


class TestSetLotOwnerProxy:
    # --- Happy path ---

    async def test_set_proxy_creates_proxy_and_returns_proxy_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /lot-owners/{id}/proxy with valid email creates proxy and returns proxy_email."""
        lo = LotOwner(building_id=building.id, lot_number="PX01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "proxy@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "proxy@test.com"
        # With omitted name fields, both proxy name fields are null
        assert data["proxy_given_name"] is None
        assert data["proxy_surname"] is None

    async def test_set_proxy_with_names_returns_proxy_given_name_and_proxy_surname(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /proxy with given_name/surname returns proxy_given_name/proxy_surname in response."""
        lo = LotOwner(building_id=building.id, lot_number="PX01B", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "named@test.com", "given_name": "Alice", "surname": "Brown"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "named@test.com"
        assert data["proxy_given_name"] == "Alice"
        assert data["proxy_surname"] == "Brown"

    async def test_set_proxy_replaces_existing_proxy(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /lot-owners/{id}/proxy when proxy already exists replaces it (upsert)."""
        lo = LotOwner(building_id=building.id, lot_number="PX02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        _proxy_p = await get_or_create_person(db_session, "old_proxy@test.com")
        db_session.add(LotProxy(lot_id=lo.id, person_id=_proxy_p.id))
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "new_proxy@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] == "new_proxy@test.com"

    # --- Input validation ---

    async def test_empty_proxy_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT with empty proxy_email → 422."""
        lo = LotOwner(building_id=building.id, lot_number="PX03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": ""},
        )
        assert response.status_code == 422

    async def test_missing_proxy_email_field_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT with missing proxy_email field → 422."""
        lo = LotOwner(building_id=building.id, lot_number="PX04", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_set_proxy_on_nonexistent_lot_owner_returns_404(
        self, client: AsyncClient
    ):
        """PUT on non-existent lot owner → 404."""
        response = await client.put(
            f"/api/admin/lot-owners/{uuid.uuid4()}/proxy",
            json={"proxy_email": "proxy@test.com"},
        )
        assert response.status_code == 404


class TestRemoveLotOwnerProxy:
    # --- Happy path ---

    async def test_remove_proxy_returns_null_proxy_email(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE /lot-owners/{id}/proxy removes proxy and returns proxy_email: null."""
        lo = LotOwner(building_id=building.id, lot_number="PX05", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        _proxy_p = await get_or_create_person(db_session, "proxy_to_remove@test.com")
        db_session.add(LotProxy(lot_id=lo.id, person_id=_proxy_p.id))
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/proxy"
        )
        assert response.status_code == 200
        data = response.json()
        assert data["proxy_email"] is None
        assert data["proxy_given_name"] is None
        assert data["proxy_surname"] is None

    # --- State / precondition errors ---

    async def test_remove_proxy_on_nonexistent_lot_owner_returns_404(
        self, client: AsyncClient
    ):
        """DELETE on non-existent lot owner → 404."""
        response = await client.delete(
            f"/api/admin/lot-owners/{uuid.uuid4()}/proxy"
        )
        assert response.status_code == 404

    async def test_remove_proxy_when_no_proxy_set_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE when no proxy is set → 404."""
        lo = LotOwner(building_id=building.id, lot_number="PX06", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/proxy"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# get_effective_status helper — unit tests (US-CD01)
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# RR3-34: File upload size limits (lot owner, proxy, financial position imports)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestImportFileSizeLimits:
    """Import endpoints reject files over 5 MB with 413 (RR3-34)."""

    async def test_lot_owner_csv_over_5mb_returns_413(
        self, client: AsyncClient, building: Building
    ):
        """Lot owner CSV over 5 MB is rejected with HTTP 413."""
        oversized = b"lot_number,unit_entitlement\n" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("big.csv", oversized, "text/csv")},
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]

    async def test_proxy_csv_over_5mb_returns_413(
        self, client: AsyncClient, building: Building
    ):
        """Proxy import CSV over 5 MB is rejected with HTTP 413."""
        oversized = b"lot_number,proxy_email\n" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import-proxies",
            files={"file": ("big.csv", oversized, "text/csv")},
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]

    async def test_financial_position_csv_over_5mb_returns_413(
        self, client: AsyncClient, building: Building
    ):
        """Financial position CSV over 5 MB is rejected with HTTP 413."""
        oversized = b"Lot#,Closing Balance\n" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import-financial-positions",
            files={"file": ("big.csv", oversized, "text/csv")},
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]

    async def test_lot_owner_excel_over_5mb_returns_413(
        self, client: AsyncClient, building: Building
    ):
        """Lot owner Excel over 5 MB is rejected with HTTP 413."""
        oversized = b"PK\x03\x04" + b"x" * (5 * 1024 * 1024 + 1)
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={
                "file": (
                    "big.xlsx",
                    oversized,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 413
        assert "5 MB" in response.json()["detail"]


# ---------------------------------------------------------------------------
# US-LON-01 / US-LON-02: Lot Owner Names (given_name / surname)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestLotOwnerNames:
    """Tests for given_name and surname fields on lot owners and proxies."""

    async def test_add_lot_owner_with_name(
        self, client: AsyncClient, building: Building
    ):
        """Add email to lot with given_name and surname; both are persisted on the Person."""
        # First create the lot
        create_resp = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "LON01", "emails": [], "unit_entitlement": 100},
        )
        assert create_resp.status_code == 201
        lot_id = create_resp.json()["id"]

        # Add email with name via POST /owner-emails
        response = await client.post(
            f"/api/admin/lot-owners/{lot_id}/owner-emails",
            json={"email": "lon01@test.com", "given_name": "Alice", "surname": "Smith"},
        )
        assert response.status_code == 201
        data = response.json()
        entry = next((e for e in data["owner_emails"] if e["email"] == "lon01@test.com"), None)
        assert entry is not None
        assert entry["given_name"] == "Alice"
        assert entry["surname"] == "Smith"

    async def test_add_lot_owner_without_name(
        self, client: AsyncClient, building: Building
    ):
        """Add email to lot without given_name/surname; name fields are null."""
        create_resp = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners",
            json={"lot_number": "LON02", "emails": [], "unit_entitlement": 50},
        )
        assert create_resp.status_code == 201
        lot_id = create_resp.json()["id"]

        response = await client.post(
            f"/api/admin/lot-owners/{lot_id}/owner-emails",
            json={"email": "lon02@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        entry = data["owner_emails"][0]
        assert entry["given_name"] is None
        assert entry["surname"] is None

    async def test_update_lot_owner_given_name_and_surname(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PATCH /owner-emails/{id} with given_name and surname persists and returns them."""
        lo = LotOwner(building_id=building.id, lot_number="LON03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "lon03@test.com")
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"given_name": "Bob", "surname": "Jones"},
        )
        assert response.status_code == 200
        data = response.json()
        entry = next((e for e in data["owner_emails"] if str(e["id"]) == str(em.id)), None)
        assert entry is not None
        assert entry["given_name"] == "Bob"
        assert entry["surname"] == "Jones"

    async def test_lot_owner_names_returned_in_list(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """List lot owners includes given_name and surname in owner_emails entries."""
        lo = LotOwner(building_id=building.id, lot_number="LON04", unit_entitlement=80)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "lon04@test.com", given_name="Carol", surname="White")
        await db_session.commit()

        response = await client.get(f"/api/admin/buildings/{building.id}/lot-owners")
        assert response.status_code == 200
        owners = response.json()
        named = [o for o in owners if o["lot_number"] == "LON04"]
        assert len(named) == 1
        entry = named[0]["owner_emails"][0]
        assert entry["given_name"] == "Carol"
        assert entry["surname"] == "White"

    async def test_csv_import_with_name_columns(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV import with given_name/surname and email stores names on Person row."""
        csv_data = make_csv(
            ["lot_number", "unit_entitlement", "email", "given_name", "surname"],
            [["LON05", "100", "lon05@test.com", "Dave", "Brown"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        # Verify name was stored on the Person
        list_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = list_response.json()
        lon05 = next((o for o in owners if o["lot_number"] == "LON05"), None)
        assert lon05 is not None
        entry = next((e for e in lon05["owner_emails"] if e["email"] == "lon05@test.com"), None)
        assert entry is not None
        assert entry["given_name"] == "Dave"
        assert entry["surname"] == "Brown"

    async def test_csv_import_without_name_columns_names_are_null(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV import without name columns: lot created without any persons (no email col)."""
        csv_data = make_csv(
            ["lot_number", "unit_entitlement"],
            [["LON06", "100"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        list_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = list_response.json()
        lon06 = next((o for o in owners if o["lot_number"] == "LON06"), None)
        assert lon06 is not None
        # No owner_emails since no email column in CSV
        assert lon06["owner_emails"] == []

    async def test_csv_import_updates_names_on_existing_lot(
        self, client: AsyncClient, building: Building, db_session: AsyncSession
    ):
        """CSV import with name and email columns creates/updates Person with name."""
        lo = LotOwner(building_id=building.id, lot_number="LON07", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        csv_data = make_csv(
            ["lot_number", "unit_entitlement", "email", "given_name", "surname"],
            [["LON07", "100", "lon07@test.com", "Eve", "Green"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("owners.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        list_response = await client.get(
            f"/api/admin/buildings/{building.id}/lot-owners"
        )
        owners = list_response.json()
        lon07 = next((o for o in owners if o["lot_number"] == "LON07"), None)
        assert lon07 is not None
        entry = next((e for e in lon07["owner_emails"] if e["email"] == "lon07@test.com"), None)
        assert entry is not None
        assert entry["given_name"] == "Eve"
        assert entry["surname"] == "Green"

    async def test_set_proxy_with_name(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /proxy with given_name and surname persists them on LotProxy."""
        lo = LotOwner(building_id=building.id, lot_number="LON08", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "proxy@test.com", "given_name": "Frank", "surname": "Black"},
        )
        assert response.status_code == 200

        # Verify LotProxy has the names
        from sqlalchemy.orm import selectinload
        proxy_result = await db_session.execute(
            select(LotProxy).where(LotProxy.lot_id == lo.id)
            .options(selectinload(LotProxy.person))
        )
        proxy = proxy_result.scalar_one()
        assert proxy.person.given_name == "Frank"
        assert proxy.person.surname == "Black"

    async def test_set_proxy_with_name_updates_existing_proxy(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PUT /proxy with given_name/surname updates existing LotProxy names."""
        lo = LotOwner(building_id=building.id, lot_number="LON09", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        _proxy_p = await get_or_create_person(db_session, "old@test.com")
        db_session.add(LotProxy(lot_id=lo.id, person_id=_proxy_p.id))
        await db_session.commit()
        await db_session.refresh(lo)

        response = await client.put(
            f"/api/admin/lot-owners/{lo.id}/proxy",
            json={"proxy_email": "new@test.com", "given_name": "Grace", "surname": "Hall"},
        )
        assert response.status_code == 200

        from sqlalchemy.orm import selectinload
        proxy_result = await db_session.execute(
            select(LotProxy).where(LotProxy.lot_id == lo.id)
            .options(selectinload(LotProxy.person))
        )
        proxy = proxy_result.scalar_one()
        assert proxy.person.given_name == "Grace"
        assert proxy.person.surname == "Hall"
        assert proxy.person.email == "new@test.com"

    async def test_proxy_csv_import_with_names_on_existing_proxy(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """Proxy CSV import with proxy_given_name/proxy_surname updates existing proxy names."""
        lo = LotOwner(building_id=building.id, lot_number="LON10", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        _proxy_p = await get_or_create_person(db_session, "old@test.com")
        db_session.add(LotProxy(lot_id=lo.id, person_id=_proxy_p.id))
        await db_session.commit()

        csv_data = make_csv(
            ["Lot#", "Proxy Email", "proxy_given_name", "proxy_surname"],
            [["LON10", "new@test.com", "Henry", "King"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import-proxies",
            files={"file": ("proxies.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 200

        from sqlalchemy.orm import selectinload
        proxy_result = await db_session.execute(
            select(LotProxy).where(LotProxy.lot_id == lo.id)
            .options(selectinload(LotProxy.person))
        )
        proxy = proxy_result.scalar_one()
        assert proxy.person.given_name == "Henry"
        assert proxy.person.surname == "King"
        assert proxy.person.email == "new@test.com"


# ---------------------------------------------------------------------------
# Rate limiting — admin lot owner import endpoint (RR4-31)
# ---------------------------------------------------------------------------


class TestAdminImportRateLimitLotOwners:
    """Verify admin_import_limiter returns 429 on the 21st lot owner import (RR4-31)."""

    async def test_lot_owners_import_rate_limited_after_max_requests(
        self, client: AsyncClient, building: "Building", db_session: "AsyncSession"
    ):
        """21st call to lot-owners/import within the window returns 429 with Retry-After."""
        from app.rate_limiter import admin_import_limiter
        import time

        admin_import_limiter._timestamps["admin"] = [time.monotonic() for _ in range(20)]

        csv_data = make_csv(
            ["Lot#", "UOE2", "Email"],
            [["99", "100", "rl_test@test.com"]],
        )
        response = await client.post(
            f"/api/admin/buildings/{building.id}/lot-owners/import",
            files={"file": ("lo.csv", csv_data, "text/csv")},
        )
        assert response.status_code == 429
        assert response.headers.get("Retry-After") == "60"


# ---------------------------------------------------------------------------
# New: owner_emails field in list/get responses
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestOwnerEmailsField:
    """LotOwnerOut.owner_emails is populated for list and get endpoints."""

    async def test_list_returns_owner_emails_with_name_fields(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /lot-owners returns owner_emails list with id, email, given_name, surname."""
        lo = LotOwner(building_id=building.id, lot_number="OE01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "oe1@test.com", given_name="Jane", surname="Doe")
        await db_session.commit()

        response = await client.get(f"/api/admin/buildings/{building.id}/lot-owners")
        assert response.status_code == 200
        owners = response.json()
        oe01 = next((o for o in owners if o["lot_number"] == "OE01"), None)
        assert oe01 is not None
        assert "owner_emails" in oe01
        assert len(oe01["owner_emails"]) == 1
        entry = oe01["owner_emails"][0]
        assert entry["email"] == "oe1@test.com"
        assert entry["given_name"] == "Jane"
        assert entry["surname"] == "Doe"
        assert "id" in entry

    async def test_get_returns_owner_emails_with_null_names(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /lot-owners/{id} returns owner_emails with null names when not set."""
        lo = LotOwner(building_id=building.id, lot_number="OE02", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "oe2@test.com")
        await db_session.commit()

        response = await client.get(f"/api/admin/lot-owners/{lo.id}")
        assert response.status_code == 200
        data = response.json()
        assert "owner_emails" in data
        entry = data["owner_emails"][0]
        assert entry["email"] == "oe2@test.com"
        assert entry["given_name"] is None
        assert entry["surname"] is None

    async def test_backward_compat_emails_field_still_present(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /lot-owners/{id} response still includes computed emails: list[str]."""
        lo = LotOwner(building_id=building.id, lot_number="OE03", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "oe3@test.com")
        await db_session.commit()

        response = await client.get(f"/api/admin/lot-owners/{lo.id}")
        assert response.status_code == 200
        data = response.json()
        assert "emails" in data
        assert "oe3@test.com" in data["emails"]


# ---------------------------------------------------------------------------
# POST /api/admin/lot-owners/{lot_owner_id}/owner-emails
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestAddOwnerEmail:
    """Tests for POST /lot-owners/{id}/owner-emails (name + email)."""

    # --- Happy path ---

    async def test_add_with_name_and_email_returns_201(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "aoe1@test.com", "given_name": "Alice", "surname": "Smith"},
        )
        assert response.status_code == 201
        data = response.json()
        entry = next((e for e in data["owner_emails"] if e["email"] == "aoe1@test.com"), None)
        assert entry is not None
        assert entry["given_name"] == "Alice"
        assert entry["surname"] == "Smith"

    async def test_add_without_name_returns_201_with_null_names(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "aoe2@test.com"},
        )
        assert response.status_code == 201
        data = response.json()
        entry = next((e for e in data["owner_emails"] if e["email"] == "aoe2@test.com"), None)
        assert entry is not None
        assert entry["given_name"] is None
        assert entry["surname"] is None

    async def test_add_returns_updated_owner_emails_list(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE03", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "existing@test.com")
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "aoe3@test.com", "given_name": "Bob"},
        )
        assert response.status_code == 201
        data = response.json()
        emails_in_response = [e["email"] for e in data["owner_emails"]]
        assert "existing@test.com" in emails_in_response
        assert "aoe3@test.com" in emails_in_response

    async def test_email_normalised_to_lowercase(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE04", unit_entitlement=50)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "UPPER@TEST.COM"},
        )
        assert response.status_code == 201
        data = response.json()
        emails = [e["email"] for e in data["owner_emails"]]
        assert "upper@test.com" in emails
        assert "UPPER@TEST.COM" not in emails

    # --- Input validation ---

    async def test_empty_email_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE05", unit_entitlement=50)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": ""},
        )
        assert response.status_code == 422

    async def test_given_name_max_length_enforced(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE06", unit_entitlement=50)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "aoe6@test.com", "given_name": "A" * 256},
        )
        assert response.status_code == 422

    async def test_surname_max_length_enforced(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE07", unit_entitlement=50)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "aoe7@test.com", "surname": "S" * 256},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_email_returns_409(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="AOE08", unit_entitlement=50)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "dup@test.com")
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "dup@test.com"},
        )
        assert response.status_code == 409

    async def test_nonexistent_lot_owner_returns_404(self, client: AsyncClient):
        response = await client.post(
            f"/api/admin/lot-owners/{uuid.uuid4()}/owner-emails",
            json={"email": "x@test.com"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# PATCH /api/admin/lot-owners/{lot_owner_id}/owner-emails/{email_id}
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestUpdateOwnerEmail:
    """Tests for PATCH /lot-owners/{id}/owner-emails/{emailId}."""

    # --- Happy path ---

    async def test_update_given_name_only(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "uoe1@test.com", given_name="Old", surname="Name")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"given_name": "Jane"},
        )
        assert response.status_code == 200
        data = response.json()
        entry = next((e for e in data["owner_emails"] if str(e["id"]) == str(em.id)), None)
        assert entry is not None
        assert entry["given_name"] == "Jane"
        assert entry["surname"] == "Name"  # unchanged

    async def test_update_email_address(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "old@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"email": "new@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        emails = [e["email"] for e in data["owner_emails"]]
        assert "new@test.com" in emails
        assert "old@test.com" not in emails

    async def test_update_surname_only(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "uoe3@test.com", given_name="John", surname="Old")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"surname": "Smith"},
        )
        assert response.status_code == 200
        data = response.json()
        entry = next((e for e in data["owner_emails"] if str(e["id"]) == str(em.id)), None)
        assert entry is not None
        assert entry["surname"] == "Smith"
        assert entry["given_name"] == "John"  # unchanged

    async def test_update_all_three_fields(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE04", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "uoe4@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"email": "updated@test.com", "given_name": "Alice", "surname": "Jones"},
        )
        assert response.status_code == 200
        data = response.json()
        entry = next((e for e in data["owner_emails"] if str(e["id"]) == str(em.id)), None)
        assert entry is not None
        assert entry["email"] == "updated@test.com"
        assert entry["given_name"] == "Alice"
        assert entry["surname"] == "Jones"

    # --- Input validation ---

    async def test_no_fields_provided_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE05", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "uoe5@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={},
        )
        assert response.status_code == 422

    async def test_given_name_max_length_enforced(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE06", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "uoe6@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"given_name": "A" * 256},
        )
        assert response.status_code == 422

    # --- State / precondition errors ---

    async def test_duplicate_email_returns_409(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE07", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em1 = await add_person_to_lot(db_session, lo, "first@test.com")
        em2 = await add_person_to_lot(db_session, lo, "second@test.com")
        await db_session.commit()

        # Try to change em1's email to second@test.com — duplicate
        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em1.id}",
            json={"email": "second@test.com"},
        )
        assert response.status_code == 409

    async def test_email_record_belonging_to_different_lot_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo1 = LotOwner(building_id=building.id, lot_number="UOE08A", unit_entitlement=100)
        lo2 = LotOwner(building_id=building.id, lot_number="UOE08B", unit_entitlement=100)
        db_session.add(lo1)
        db_session.add(lo2)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo2, "other@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        # Try to update lo2's email via lo1's URL
        response = await client.patch(
            f"/api/admin/lot-owners/{lo1.id}/owner-emails/{em.id}",
            json={"given_name": "Jane"},
        )
        assert response.status_code == 404

    async def test_nonexistent_email_record_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="UOE09", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{uuid.uuid4()}",
            json={"given_name": "Jane"},
        )
        assert response.status_code == 404

    async def test_nonexistent_lot_owner_returns_404(self, client: AsyncClient):
        response = await client.patch(
            f"/api/admin/lot-owners/{uuid.uuid4()}/owner-emails/{uuid.uuid4()}",
            json={"given_name": "Jane"},
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# DELETE /api/admin/lot-owners/{lot_owner_id}/owner-emails/{email_id}
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestRemoveOwnerEmailById:
    """Tests for DELETE /lot-owners/{id}/owner-emails/{emailId}."""

    # --- Happy path ---

    async def test_delete_owner_email_returns_200(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="DOE01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "doe1@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}"
        )
        assert response.status_code == 200

    async def test_deleted_email_not_in_response(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="DOE02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em1 = await add_person_to_lot(db_session, lo, "keep@test.com")
        em2 = await add_person_to_lot(db_session, lo, "delete@test.com")
        await db_session.commit()

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em2.id}"
        )
        assert response.status_code == 200
        data = response.json()
        emails = [e["email"] for e in data["owner_emails"]]
        assert "delete@test.com" not in emails
        assert "keep@test.com" in emails

    async def test_backward_compat_emails_field_updated(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="DOE03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "gone@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}"
        )
        assert response.status_code == 200
        data = response.json()
        assert "gone@test.com" not in data["emails"]

    # --- State / precondition errors ---

    async def test_nonexistent_email_record_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo = LotOwner(building_id=building.id, lot_number="DOE04", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{uuid.uuid4()}"
        )
        assert response.status_code == 404

    async def test_email_belonging_to_different_lot_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        lo1 = LotOwner(building_id=building.id, lot_number="DOE05A", unit_entitlement=100)
        lo2 = LotOwner(building_id=building.id, lot_number="DOE05B", unit_entitlement=100)
        db_session.add(lo1)
        db_session.add(lo2)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo2, "other@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.delete(
            f"/api/admin/lot-owners/{lo1.id}/owner-emails/{em.id}"
        )
        assert response.status_code == 404

    async def test_nonexistent_lot_owner_returns_404(self, client: AsyncClient):
        response = await client.delete(
            f"/api/admin/lot-owners/{uuid.uuid4()}/owner-emails/{uuid.uuid4()}"
        )
        assert response.status_code == 404


# ---------------------------------------------------------------------------
# AddOwnerEmailRequest / UpdateOwnerEmailRequest schema unit tests
# ---------------------------------------------------------------------------


class TestOwnerEmailSchemas:
    def test_add_owner_email_empty_email_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import AddOwnerEmailRequest

        with pytest.raises(ValidationError):
            AddOwnerEmailRequest(email="  ")

    def test_add_owner_email_valid_with_names(self):
        from app.schemas.admin import AddOwnerEmailRequest

        req = AddOwnerEmailRequest(email="x@test.com", given_name="Alice", surname="Smith")
        assert req.email == "x@test.com"
        assert req.given_name == "Alice"
        assert req.surname == "Smith"

    def test_add_owner_email_valid_no_names(self):
        from app.schemas.admin import AddOwnerEmailRequest

        req = AddOwnerEmailRequest(email="x@test.com")
        assert req.email == "x@test.com"
        assert req.given_name is None
        assert req.surname is None

    def test_add_owner_email_given_name_too_long_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import AddOwnerEmailRequest

        with pytest.raises(ValidationError):
            AddOwnerEmailRequest(email="x@test.com", given_name="A" * 256)

    def test_update_owner_email_no_fields_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import UpdateOwnerEmailRequest

        with pytest.raises(ValidationError):
            UpdateOwnerEmailRequest()

    def test_update_owner_email_email_only_valid(self):
        from app.schemas.admin import UpdateOwnerEmailRequest

        req = UpdateOwnerEmailRequest(email="new@test.com")
        assert req.email == "new@test.com"
        assert req.given_name is None
        assert req.surname is None

    def test_update_owner_email_given_name_only_valid(self):
        from app.schemas.admin import UpdateOwnerEmailRequest

        req = UpdateOwnerEmailRequest(given_name="Jane")
        assert req.given_name == "Jane"
        assert req.email is None

    def test_update_owner_email_all_fields_valid(self):
        from app.schemas.admin import UpdateOwnerEmailRequest

        req = UpdateOwnerEmailRequest(email="e@test.com", given_name="A", surname="B")
        assert req.email == "e@test.com"
        assert req.given_name == "A"
        assert req.surname == "B"

    def test_update_owner_email_surname_too_long_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import UpdateOwnerEmailRequest

        with pytest.raises(ValidationError):
            UpdateOwnerEmailRequest(surname="S" * 256)

    def test_update_owner_email_phone_number_only_valid(self):
        from app.schemas.admin import UpdateOwnerEmailRequest

        req = UpdateOwnerEmailRequest(phone_number="+61412345678")
        assert req.phone_number == "+61412345678"
        assert req.email is None

    def test_update_owner_email_phone_number_too_long_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import UpdateOwnerEmailRequest

        with pytest.raises(ValidationError):
            UpdateOwnerEmailRequest(phone_number="0" * 21)

    def test_add_owner_email_phone_number_accepted(self):
        from app.schemas.admin import AddOwnerEmailRequest

        req = AddOwnerEmailRequest(email="x@test.com", phone_number="+61400000000")
        assert req.phone_number == "+61400000000"

    def test_add_owner_email_phone_number_too_long_raises(self):
        from pydantic import ValidationError
        from app.schemas.admin import AddOwnerEmailRequest

        with pytest.raises(ValidationError):
            AddOwnerEmailRequest(email="x@test.com", phone_number="0" * 21)


# ---------------------------------------------------------------------------
# Phone number per LotOwnerEmail (per-contact)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
class TestOwnerEmailPhoneNumber:
    """phone_number lives on LotOwnerEmail, not LotOwner."""

    # --- Happy path: owner_emails response includes phone_number field ---

    async def test_owner_emails_response_includes_phone_number_field(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /lot-owners/{id} returns phone_number in each owner_emails entry."""
        lo = LotOwner(building_id=building.id, lot_number="PHN01", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "phn1@test.com", phone_number="+61412345678",)
        await db_session.commit()

        response = await client.get(f"/api/admin/lot-owners/{lo.id}")
        assert response.status_code == 200
        data = response.json()
        entry = data["owner_emails"][0]
        assert "phone_number" in entry
        assert entry["phone_number"] == "+61412345678"

    async def test_owner_emails_phone_number_null_when_not_set(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """owner_emails.phone_number is null when not set."""
        lo = LotOwner(building_id=building.id, lot_number="PHN02", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "phn2@test.com")
        await db_session.commit()

        response = await client.get(f"/api/admin/lot-owners/{lo.id}")
        assert response.status_code == 200
        entry = response.json()["owner_emails"][0]
        assert entry["phone_number"] is None

    async def test_add_owner_email_with_phone_number(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """POST /owner-emails with phone_number stores it on the email row."""
        lo = LotOwner(building_id=building.id, lot_number="PHN03", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "phn3@test.com", "phone_number": "+61499000000"},
        )
        assert response.status_code == 201
        entry = next(
            (e for e in response.json()["owner_emails"] if e["email"] == "phn3@test.com"),
            None,
        )
        assert entry is not None
        assert entry["phone_number"] == "+61499000000"

    async def test_add_owner_email_without_phone_number_defaults_to_null(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """POST /owner-emails without phone_number: phone_number is null."""
        lo = LotOwner(building_id=building.id, lot_number="PHN04", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.post(
            f"/api/admin/lot-owners/{lo.id}/owner-emails",
            json={"email": "phn4@test.com"},
        )
        assert response.status_code == 201
        entry = response.json()["owner_emails"][0]
        assert entry["phone_number"] is None

    async def test_update_owner_email_sets_phone_number(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PATCH /owner-emails/{id} with phone_number updates it."""
        lo = LotOwner(building_id=building.id, lot_number="PHN05", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "phn5@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"phone_number": "+61411111111"},
        )
        assert response.status_code == 200
        entry = next(
            (e for e in response.json()["owner_emails"] if str(e["id"]) == str(em.id)),
            None,
        )
        assert entry is not None
        assert entry["phone_number"] == "+61411111111"

    async def test_update_owner_email_clears_phone_number(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PATCH /owner-emails/{id} with phone_number=null clears it."""
        lo = LotOwner(building_id=building.id, lot_number="PHN06", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "phn6@test.com", phone_number="+61400000000")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"phone_number": None},
        )
        assert response.status_code == 200
        entry = next(
            (e for e in response.json()["owner_emails"] if str(e["id"]) == str(em.id)),
            None,
        )
        assert entry is not None
        assert entry["phone_number"] is None

    # --- Input validation ---

    async def test_update_owner_email_phone_number_too_long_returns_422(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PATCH /owner-emails/{id} with phone_number > 20 chars returns 422."""
        lo = LotOwner(building_id=building.id, lot_number="PHN07", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        em = await add_person_to_lot(db_session, lo, "phn7@test.com")
        await db_session.commit()
        await db_session.refresh(em)

        response = await client.patch(
            f"/api/admin/lot-owners/{lo.id}/owner-emails/{em.id}",
            json={"phone_number": "0" * 21},
        )
        assert response.status_code == 422

    # --- Migration correctness: phone on email row, not lot owner ---

    async def test_lot_owner_has_no_phone_number_field(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /lot-owners/{id} response does not contain a top-level phone_number field."""
        lo = LotOwner(building_id=building.id, lot_number="PHN08", unit_entitlement=100)
        db_session.add(lo)
        await db_session.commit()

        response = await client.get(f"/api/admin/lot-owners/{lo.id}")
        assert response.status_code == 200
        assert "phone_number" not in response.json()


# ---------------------------------------------------------------------------
# Coverage: get_or_create_person fill-blanks, lookup_person, update_person 404,
# remove_person_from_lot 404
# ---------------------------------------------------------------------------


@pytest.mark.asyncio(loop_scope="session")
class TestPersonServiceCoverage:
    """Targeted tests to reach uncovered branches in admin_service."""

    # --- get_or_create_person fill-blanks: existing person with NULL names ---

    async def test_get_or_create_person_fills_blank_given_name(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """get_or_create_person fills given_name when it is currently NULL."""
        from app.services.admin_service import get_or_create_person as svc_goc
        # Create person with no name
        p = await svc_goc("fillblank_gn@test.com", db_session)
        assert p.given_name is None
        # Call again with a name — should fill in
        p2 = await svc_goc("fillblank_gn@test.com", db_session, given_name="Alice")
        assert p2.given_name == "Alice"
        assert p2.id == p.id

    async def test_get_or_create_person_fills_blank_surname(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """get_or_create_person fills surname when it is currently NULL."""
        from app.services.admin_service import get_or_create_person as svc_goc
        p = await svc_goc("fillblank_sn@test.com", db_session)
        assert p.surname is None
        p2 = await svc_goc("fillblank_sn@test.com", db_session, surname="Smith")
        assert p2.surname == "Smith"
        assert p2.id == p.id

    # --- lookup_person endpoint ---

    async def test_lookup_person_returns_person_when_found(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """GET /api/admin/persons/lookup returns person when email matches."""
        lo = LotOwner(building_id=building.id, lot_number="LKUP1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        await add_person_to_lot(db_session, lo, "lookup_found@test.com")
        await db_session.commit()

        response = await client.get(
            "/api/admin/persons/lookup",
            params={"email": "lookup_found@test.com"},
        )
        assert response.status_code == 200
        data = response.json()
        assert data["email"] == "lookup_found@test.com"

    async def test_lookup_person_returns_404_when_not_found(
        self, client: AsyncClient
    ):
        """GET /api/admin/persons/lookup returns 404 when email is not in DB."""
        response = await client.get(
            "/api/admin/persons/lookup",
            params={"email": "nobody_here@test.com"},
        )
        assert response.status_code == 404

    # --- update_person 404 ---

    async def test_update_person_returns_404_for_nonexistent_id(
        self, client: AsyncClient
    ):
        """PATCH /api/admin/persons/{id} returns 404 for an unknown person ID."""
        response = await client.patch(
            f"/api/admin/persons/{uuid.uuid4()}",
            json={"given_name": "Ghost"},
        )
        assert response.status_code == 404

    # --- remove_person_from_lot 404: person not linked to lot ---

    async def test_remove_owner_email_from_wrong_lot_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE /lot-owners/{id}/owner-emails/{eid} returns 404 when person not linked to lot."""
        lo1 = LotOwner(building_id=building.id, lot_number="RMWL1", unit_entitlement=100)
        lo2 = LotOwner(building_id=building.id, lot_number="RMWL2", unit_entitlement=100)
        db_session.add(lo1)
        db_session.add(lo2)
        await db_session.flush()
        em2 = await add_person_to_lot(db_session, lo2, "rmwl_other@test.com")
        await db_session.commit()

        # Try to delete lo2's person via lo1's URL
        response = await client.delete(
            f"/api/admin/lot-owners/{lo1.id}/owner-emails/{em2.id}",
        )
        assert response.status_code == 404

    # --- remove_email_from_lot_owner 404: email exists as Person but not linked to lot ---

    async def test_remove_email_by_string_not_linked_returns_404(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE /lot-owners/{id}/emails/{email} returns 404 when email exists but is not linked to lot."""
        lo1 = LotOwner(building_id=building.id, lot_number="RMEL1", unit_entitlement=100)
        lo2 = LotOwner(building_id=building.id, lot_number="RMEL2", unit_entitlement=100)
        db_session.add(lo1)
        db_session.add(lo2)
        await db_session.flush()
        # Email is linked to lo2, NOT lo1
        await add_person_to_lot(db_session, lo2, "rmel_other@test.com")
        await db_session.commit()

        # Try to remove lo2's email via lo1's URL
        response = await client.delete(
            f"/api/admin/lot-owners/{lo1.id}/emails/rmel_other@test.com",
        )
        assert response.status_code == 404

    # --- update_person happy path (covers router line 770) ---

    async def test_update_person_returns_updated_person(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """PATCH /api/admin/persons/{id} updates the person and returns PersonOut."""
        from app.models.person import Person

        person = Person(email="upd_person_hp@test.com", given_name="Old", surname="Name")
        db_session.add(person)
        await db_session.commit()
        await db_session.refresh(person)

        response = await client.patch(
            f"/api/admin/persons/{person.id}",
            json={"given_name": "New", "surname": "Updated"},
        )
        assert response.status_code == 200
        body = response.json()
        assert body["given_name"] == "New"
        assert body["surname"] == "Updated"

    # --- remove_person_from_lot happy path (covers router lines 783-784) ---

    async def test_remove_person_from_lot_returns_lot_owner(
        self, client: AsyncClient, db_session: AsyncSession, building: Building
    ):
        """DELETE /lot-owners/{id}/persons/{pid} removes the person link and returns LotOwnerOut."""
        lo = LotOwner(building_id=building.id, lot_number="RPLHP1", unit_entitlement=100)
        db_session.add(lo)
        await db_session.flush()
        person = await add_person_to_lot(db_session, lo, "rplhp_voter@test.com")
        await db_session.commit()

        response = await client.delete(
            f"/api/admin/lot-owners/{lo.id}/persons/{person.id}",
        )
        assert response.status_code == 200
        body = response.json()
        assert body["id"] == str(lo.id)
        # Person was removed — persons list should be empty
        assert body["persons"] == []
